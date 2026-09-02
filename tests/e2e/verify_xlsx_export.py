#!/usr/bin/env python3
"""Excel-Export Ende-zu-Ende (zurueck am 02.09.2026, Kundenwunsch via Michael).

Grafik-Projekt mit zwei fiktiven Bildern, Alt-Texte von Hand (inkl.
Formel-Injection-Kandidat), dann: Einzel-Export (XLSX direkt), Alle-Dokumente
(ZIP mit zwei XLSX), Kopfzeilen/Werte per openpyxl, X-Export-Credits 10,
Buchung xlsx_export, ohne Anmeldung 401. Laeuft IM Container: BASE,
INKLUDOCS_E2E_MAIL, INKLUDOCS_E2E_PW. Alle Daten fiktiv; Projekt und
Test-Buchungen werden am Ende geloescht."""
import http.cookiejar, io, json, os, sqlite3, sys, time, urllib.error, urllib.request, uuid

BASE = os.environ["BASE"]; MAIL = os.environ["INKLUDOCS_E2E_MAIL"]; PW = os.environ["INKLUDOCS_E2E_PW"]
DB = os.environ.get("INKLUDOCS_DB", "/app/data/inkludocs.db")
cj = http.cookiejar.CookieJar(); op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
ok = fehler = 0

def check(name, cond, info=""):
    global ok, fehler
    if cond: ok += 1; print("  OK ", name)
    else: fehler += 1; print("  FEHLT", name, "--", str(info)[:300])

def req(m, path, data=None, files=None, roh=False, ohne_cookies=False):
    """JSON-Request; roh=True liefert (status, bytes, headers) fuer Binaer-Antworten."""
    if files:
        b = uuid.uuid4().hex; body = io.BytesIO()
        for k, v in (data or {}).items():
            body.write(f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
        for k, (fn, content) in files.items():
            body.write(f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"; filename=\"{fn}\"\r\nContent-Type: application/octet-stream\r\n\r\n".encode())
            body.write(content); body.write(b"\r\n")
        body.write(f"--{b}--\r\n".encode())
        r = urllib.request.Request(BASE + path, data=body.getvalue(), method=m, headers={"Content-Type": f"multipart/form-data; boundary={b}"})
    else:
        r = urllib.request.Request(BASE + path, data=json.dumps(data).encode() if data is not None else None, method=m, headers={"Content-Type": "application/json"})
    opener = urllib.request.build_opener() if ohne_cookies else op
    try:
        with opener.open(r, timeout=120) as resp:
            c = resp.read()
            if roh: return resp.status, c, dict(resp.headers)
            return resp.status, (json.loads(c) if c else {})
    except urllib.error.HTTPError as e:
        c = e.read()
        if roh: return e.code, c, dict(e.headers)
        try: return e.code, json.loads(c)
        except Exception: return e.code, {"raw": c[:300].decode(errors="replace")}

from PIL import Image
from openpyxl import load_workbook

def hget(hdr, name):
    """Header-Wert unabhaengig von Gross-/Kleinschreibung (der Proxy schreibt klein)."""
    return next((v for k, v in hdr.items() if k.lower() == name.lower()), None)

def png_bytes(farbe):
    buf = io.BytesIO(); Image.new("RGB", (240, 160), farbe).save(buf, format="PNG"); return buf.getvalue()

print("== A. Login + Grafik-Projekt mit zwei Bildern ==")
s, _ = req("POST", "/api/login", {"email": MAIL, "password": PW}); check("Login", s == 200)
s, b = req("POST", "/api/projects", {"name": "Excel-Export-Test (fiktiv) " + time.strftime("%H:%M:%S"), "tool": "grafik"})
pid = b.get("id") or b.get("project_id"); check("Grafik-Projekt angelegt", s == 200 and pid, b)
s, b = req("POST", "/api/upload", {"project_id": pid}, {"file": ("testbild_eins_fiktiv.png", png_bytes((30, 90, 160)))})
check("Bild 1 hochgeladen", s == 200, b)
s, b = req("POST", "/api/upload", {"project_id": pid}, {"file": ("testbild_zwei_fiktiv.png", png_bytes((160, 40, 40)))})
check("Bild 2 hochgeladen", s == 200, b)
imgs = []
for _ in range(20):
    s, p = req("GET", f"/api/projects/{pid}"); imgs = p.get("images") or []
    if len(imgs) >= 2: break
    time.sleep(0.5)
check("Beide Bilder in der Projektansicht", len(imgs) == 2, imgs)
# Grafik-Projekte haben KEINE Dokument-Zeilen -> _load_export_units_for_table
# liefert genau eine Einheit, der Export ist immer die direkte XLSX (kein ZIP).
# Der ZIP-Zweig ist formatunabhaengig und durch die CSV/JSON-Suiten abgedeckt.

conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
uid = conn.execute("SELECT id FROM users WHERE email=?", (MAIL,)).fetchone()["id"]
START_MAX_ID = conn.execute("SELECT COALESCE(MAX(id), 0) FROM usage_events").fetchone()[0]
# Alt-Texte von Hand setzen (fiktiv) — Bild 1 normal, Bild 2 Formel-Injection-Kandidat.
conn.execute("UPDATE images SET alt_text=?, langbeschreibung=?, status='done' WHERE id=?",
             ("Blaues Testrechteck (fiktiv)", "Ein einfarbig blaues Rechteck als Testbild.", imgs[0]["id"]))
conn.execute("UPDATE images SET alt_text=?, langbeschreibung=?, status='done' WHERE id=?",
             ('=HYPERLINK("http://boese.example";"klick")', "+SUMME(A1:A9)", imgs[1]["id"]))
conn.commit()

print("== B. Export -> direkte XLSX mit beiden Bildern ==")
s, roh, hdr = req("POST", f"/api/projects/{pid}/export/xlsx", {}, roh=True)
check("Status 200", s == 200, s)
check("Content-Type spreadsheetml", "spreadsheetml" in (hget(hdr, "Content-Type") or ""), hget(hdr, "Content-Type"))
check("X-Export-Credits: 10", hget(hdr, "X-Export-Credits") == "10", hget(hdr, "X-Export-Credits"))
check("Dateiname .xlsx im Header", ".xlsx" in (hget(hdr, "Content-Disposition") or ""), hget(hdr, "Content-Disposition"))
ws = load_workbook(io.BytesIO(roh)).active
check("Blatt 'Alt-Texte' mit Kopfzeile", ws.title == "Alt-Texte" and [ws["A1"].value, ws["B1"].value, ws["C1"].value] == ["Bild", "Alt-Text", "Langbeschreibung"], (ws.title, ws["A1"].value))
check("Alt-Text in B2", ws["B2"].value == "Blaues Testrechteck (fiktiv)", ws["B2"].value)
check("Langbeschreibung in C2", ws["C2"].value == "Ein einfarbig blaues Rechteck als Testbild.", ws["C2"].value)
check("Formel-Injection entschaerft (B3 beginnt mit ')", str(ws["B3"].value).startswith("'="), ws["B3"].value)
check("Langbeschreibung entschaerft (C3 beginnt mit ')", str(ws["C3"].value).startswith("'+"), ws["C3"].value)
ev = conn.execute("SELECT quelle, aktion, credits FROM usage_events WHERE konto_user_id=? AND id>? ORDER BY id DESC LIMIT 1", (uid, START_MAX_ID)).fetchone()
check("Buchung: quelle export, aktion xlsx_export, 10 Credits", ev and ev["quelle"] == "export" and ev["aktion"] == "xlsx_export" and ev["credits"] == 10, dict(ev) if ev else None)

print("== C. Zweiter Export -> eigener Vorgang, eigene Buchung ==")
s, roh, hdr = req("POST", f"/api/projects/{pid}/export/xlsx", {}, roh=True)
check("Status 200", s == 200 and roh[:2] == b"PK", (s, roh[:8]))
anz = conn.execute("SELECT COUNT(*) FROM usage_events WHERE konto_user_id=? AND id>? AND aktion='xlsx_export'", (uid, START_MAX_ID)).fetchone()[0]
check("Zwei Buchungen nach zwei Exporten", anz == 2, anz)

print("== D. Zugriffsschutz ==")
s, b = req("POST", f"/api/projects/{pid}/export/xlsx", {}, ohne_cookies=True)
check("Ohne Anmeldung kein Export (401)", s == 401, (s, b))
s, b = req("POST", f"/api/projects/{pid + 100000}/export/xlsx", {})
check("Fremdes/unbekanntes Projekt: 404", s == 404, (s, b))

print("== E. Aufraeumen ==")
s, _ = req("DELETE", f"/api/projects/{pid}"); check("Testprojekt geloescht", s == 200, s)
conn.execute("DELETE FROM usage_events WHERE konto_user_id=? AND id>? AND aktion='xlsx_export'", (uid, START_MAX_ID))
conn.commit(); conn.close()
print(f"Ergebnis: {ok} OK, {fehler} FEHLER")
sys.exit(1 if fehler else 0)
