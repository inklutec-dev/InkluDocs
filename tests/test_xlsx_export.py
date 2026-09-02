# -*- coding: utf-8 -*-
"""Excel-Export (zurueck am 02.09.2026, Kundenwunsch via Michael): _build_xlsx_bytes.

Struktur (Kopfzeile, eine Zeile je Bild), Formel-Injection-Schutz (_csv_safe),
Fehlertext-Filter fuer die Langbeschreibung. Laeuft im Container:
python3 -m unittest tests/test_xlsx_export.py
"""
import io
import os
import sys
import unittest

sys.path.insert(0, "/app")
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
import main  # noqa: E402


def _lade(unit):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(main._build_xlsx_bytes(unit)))


class XlsxExport(unittest.TestCase):
    def test_struktur_und_inhalte(self):
        unit = {"images": [
            {"image_path": "/nirgends/bild_eins.png", "alt_text": "Ein rotes Haus", "status": "done",
             "alt_text_edited": None, "original_alt": None,
             "langbeschreibung": "Ein rotes Haus mit zwei Fenstern."},
            {"image_path": "/nirgends/bild_zwei.png", "alt_text": "Blauer Kreis", "status": "done",
             "alt_text_edited": None, "original_alt": None,
             "langbeschreibung": ""},
        ]}
        ws = _lade(unit).active
        self.assertEqual(ws.title, "Alt-Texte")
        self.assertEqual([ws["A1"].value, ws["B1"].value, ws["C1"].value],
                         ["Bild", "Alt-Text", "Langbeschreibung"])
        self.assertEqual(ws["A2"].value, "bild_eins.png")
        self.assertEqual(ws["B2"].value, "Ein rotes Haus")
        self.assertEqual(ws["C2"].value, "Ein rotes Haus mit zwei Fenstern.")
        self.assertEqual(ws["B3"].value, "Blauer Kreis")

    def test_formel_injection_entschaerft(self):
        unit = {"images": [
            {"image_path": "/nirgends/boese.png", "status": "done", "alt_text_edited": None, "original_alt": None,
             "alt_text": "=HYPERLINK(\"http://boese.example\";\"klick\")",
             "langbeschreibung": "+SUMME(A1:A9)"},
        ]}
        ws = _lade(unit).active
        self.assertTrue(str(ws["B2"].value).startswith("'="), ws["B2"].value)
        self.assertTrue(str(ws["C2"].value).startswith("'+"), ws["C2"].value)

    def test_geleertes_bild_bleibt_leer(self):
        # Geleerter Alt-Text ('' seit 01.09.2026 = NULL-Regel; _ausgabe_alt_text
        # liefert dann nichts) -> leere Zelle, kein Fehlertext.
        unit = {"images": [
            {"image_path": "/nirgends/leer.png", "alt_text": None, "status": "done",
             "alt_text_edited": None, "original_alt": None,
             "langbeschreibung": None},
        ]}
        ws = _lade(unit).active
        self.assertIn(ws["B2"].value, (None, ""))
        self.assertIn(ws["C2"].value, (None, ""))

    def test_fehlertext_geht_nicht_nach_aussen(self):
        unit = {"images": [
            {"image_path": "/nirgends/kaputt.png", "status": "done",
             "alt_text_edited": None, "original_alt": None,
             "alt_text": "Pipeline-Fehler: Bedrock nicht erreichbar",
             "langbeschreibung": "Traceback (most recent call last): kaputt"},
        ]}
        ws = _lade(unit).active
        self.assertIn(ws["B2"].value, (None, ""))
        self.assertIn(ws["C2"].value, (None, ""))


if __name__ == "__main__":
    unittest.main()
