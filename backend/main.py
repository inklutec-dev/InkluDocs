import os
import urllib.parse  # Content-Disposition nach RFC 6266 (_content_disposition)
import re
import shutil
import json
import csv
import asyncio
import html  # Review-Befund 7 (31.07.2026): Nutzereingaben in Mail-HTML escapen
import secrets
import sqlite3
import threading  # Abo-Tageslauf (Erinnerungen/Verlaengerungen) ohne Cron
import time
import io
from collections import defaultdict
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from urllib.parse import urljoin, urlparse
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import httpx
from bs4 import BeautifulSoup

import logging

import absender  # Sicherheitspaket 04.08.2026: echte Besucher-IP hinter dem Proxy (X-Forwarded-For von RECHTS)
import billing  # Abo-/Credit-System Etappe 1: zentrale Verbrauchs-Zaehlung (backend/ABRECHNUNG.md)
import demo as demo_mod  # Demo-Modus (oeffentliche Kostprobe ohne Anmeldung) — nur aktiv bei DEMO_MODE=on
import stripe_zahlung  # Online-Zahlung (06.08.2026): inert ohne STRIPE_SECRET_KEY

# Abo-System (06.08.2026): Fehler in Mail-/Tageslauf-Pfaden landen im Log,
# nie beim Nutzer — gleiche Nie-Crashen-Philosophie wie billing.
logger = logging.getLogger("inkludocs")
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request, Form
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from jose import jwt, JWTError

from database import (
    init_db, get_db, create_user, verify_user, get_user_by_email, get_user_by_id,
    create_password_reset_token, verify_reset_token, reset_password,
    list_all_users, update_user_active, delete_user_data, admin_reset_password,
    list_admins, count_full_admins, set_user_admin,
    create_api_key, verify_api_key, list_api_keys, delete_api_key, rename_api_key,
    log_api_usage, get_api_usage_stats,
    create_email_change_token, confirm_email_change,
    create_api_result, get_api_result, update_api_result,
    create_email_verification_token, mark_email_verified, resend_verification_token,
    get_daily_image_count, get_daily_api_count,
    set_user_language,
)
from pdf_processor import extract_images_from_pdf, generate_alt_text, generate_alt_text_for_image, clear_project_cache
# WORD-WERKZEUG (26.08.2026): .docx lesen (Bilder + Kontext) und Alt-Texte zurueckschreiben
from docx_processor import extract_images_from_docx, extract_docx, validiere_docx, DocxFehler
import docx_export
# QUICKINFO-WERKZEUG (27.08.2026): PDF-Formularfelder lesen/schreiben, eigener Router
import formular_api
from formular_processor import validiere_formular, FormularFehler
import sharing  # Gastzugang / Projekt-Freigabe (19.06.2026)
from i18n import get_templates, detect_language, template_context, get_gettext, SUPPORTED_LANGUAGES
from typing import Optional  # frueh, da ab get_optional_user in Typannotationen genutzt
from tools import TOOLS, is_valid_tool_key

# --- Temperatur-Politik der Alt-Text-Generierung (30.06.2026) ---
# Standardbetrieb (Bulk / "alle generieren" / erste Generierung): leicht erhoeht
# fuer etwas lebendigere Formulierung, aber bewusst niedrig, damit das Modell
# keine Details dazuerfindet. Einzel-"neu generieren": hoeher, weil der Nutzer
# dort jeden Vorschlag einzeln prueft und Variation ausdruecklich wuenscht.
# Klassifikation bleibt immer deterministisch (temp 0, im Orchestrator).
GENERATION_TEMPERATURE = 0.3   # Normalbetrieb; jederzeit verstellbar (0.0-0.5)
REGENERATE_TEMPERATURE = 0.5   # Einzel-Neu-Generieren; seit 05.07.2026 gesenkt,
# weil die Variation nicht mehr allein am Zufall haengt: der bisherige Alt-Text
# wird als previous_alt mitgereicht und der Orchestrator haengt eine gezielte
# Variations-Anweisung an (_variation_suffix) — Verschiedenheit per Auftrag,
# Fakten bleiben stabil.

# Generate a persistent SECRET_KEY if not set
SECRET_KEY_FILE = "/app/data/.secret_key"
def _get_secret_key():
    env_key = os.environ.get("SECRET_KEY", "")
    if env_key and env_key != "inkludocs-production-key-2025":
        return env_key
    if os.path.exists(SECRET_KEY_FILE):
        return open(SECRET_KEY_FILE).read().strip()
    key = secrets.token_hex(32)
    os.makedirs(os.path.dirname(SECRET_KEY_FILE), exist_ok=True)
    with open(SECRET_KEY_FILE, "w") as f:
        f.write(key)
    return key

SECRET_KEY = _get_secret_key()
ALGORITHM = "HS256"
TOKEN_EXPIRE_HOURS = 24
UPLOAD_DIR = "/app/data/uploads"
RESULTS_DIR = "/app/data/results"
BASE_URL = os.environ.get("BASE_URL", "https://inkludocs.inklutec.de")
MAX_UPLOAD_SIZE = 50 * 1024 * 1024  # 50 MB

# SMTP configuration for email sending
SMTP_SERVER = os.environ.get("SMTP_SERVER", "w01ccfc3.kasserver.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "kontakt@inklutec.de")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
SMTP_FROM = os.environ.get("SMTP_FROM", "kontakt@inklutec.de")
NOTIFICATION_EMAIL = os.environ.get("NOTIFICATION_EMAIL", SMTP_FROM)
# Technisches Tageslimit je Konto. Steve 24.08.2026: von 100 auf 500
# erhoeht — der Missbrauchsschutz laeuft jetzt ueber die Credits; das
# Limit ist nur noch Schutzventil gegen AWS-Drosselung und Fehlerschleifen.
# Einzelne Grosskunden bekommen ueber users.api_tageslimit mehr.
DAILY_IMAGE_LIMIT = int(os.environ.get("DAILY_IMAGE_LIMIT", "500"))


def effektives_api_tageslimit(user_row) -> int:
    """Missbrauchsbremse der Bild-Generierung: Tageslimit pro Konto.

    users.api_tageslimit: NULL = globaler Standard DAILY_IMAGE_LIMIT,
    0 = Generierung fuer dieses Konto gesperrt. Das Abo-Guthaben gilt
    IMMER zusaetzlich — diese Bremse faengt nur Amok-Laeufe ab und ist
    fuer Betreiber-Konten (unbegrenzte Credits) die einzige Grenze.
    """
    try:
        wert = user_row["api_tageslimit"]
    except (KeyError, IndexError, TypeError):
        wert = None
    return DAILY_IMAGE_LIMIT if wert is None else int(wert)


def send_email(to_email: str, subject: str, html_body: str, bcc_admin: bool = True, attachment_path: str = None, reply_to: str = None, from_name: str = None) -> bool:
    """Send an email via SMTP with optional file attachment.

    Args:
        bcc_admin: If True, send BCC copy to admin. Set False for
                   sensitive emails like password resets.
        attachment_path: Optional path to a file to attach (e.g. image).
    """
    if not SMTP_PASS:
        print(f"E-Mail nicht gesendet (kein SMTP-Passwort konfiguriert): {subject} an {to_email}")
        return False
    # Testlaeufe (06.08.2026): Adressen auf der reservierten .invalid-Domain
    # (RFC 2606) werden NIE versendet — sonst fluten Unzustellbarkeits-
    # Meldungen das support-Postfach (Steve-Befund nach den E2E-Laeufen).
    if (to_email or "").strip().lower().rsplit(".", 1)[-1] == "invalid":
        print(f"E-Mail unterdrueckt (Testadresse .invalid): {subject} an {to_email}")
        return True
    try:
        from email.mime.base import MIMEBase
        from email import encoders

        msg = MIMEMultipart("mixed")
        # Kopfzeilen saeubern (Selbst-Review 08.08.2026): Ein Zeilenumbruch in
        # Betreff oder Empfaenger kann zusaetzliche Mail-Kopfzeilen einschleusen
        # (etwa ein heimliches Bcc). Heute kommen alle Werte aus geprueften
        # Quellen — aber das gilt nur, solange niemand einen neuen Aufrufer
        # ergaenzt. Darum hier zentral, wo jede Mail durchmuss.
        def _kopfzeile(wert: str) -> str:
            return " ".join((wert or "").replace("\r", " ").replace("\n", " ").split())

        subject = _kopfzeile(subject)
        to_email = _kopfzeile(to_email)
        if "staging" in BASE_URL:
            subject = f"[STAGING] {subject}"
        msg["Subject"] = subject
        from email.utils import formataddr
        # Nur den Anzeigenamen kodieren, Adresse bleibt gueltig (sonst lehnt Gmail mit 550 5.7.1 ab).
        msg["From"] = formataddr((from_name, SMTP_FROM), charset="utf-8") if from_name else f"InkluDocs <{SMTP_FROM}>"
        if reply_to:
            msg["Reply-To"] = reply_to
        msg["To"] = to_email

        html_part = MIMEMultipart("alternative")
        html_part.attach(MIMEText(html_body, "html", "utf-8"))
        msg.attach(html_part)

        if attachment_path and os.path.exists(attachment_path):
            try:
                with open(attachment_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                filename = os.path.basename(attachment_path)
                part.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(part)
            except Exception as e:
                print(f"Anhang konnte nicht hinzugefügt werden: {e}")

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=15)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        recipients = [to_email]
        if bcc_admin and to_email != NOTIFICATION_EMAIL:
            msg["Bcc"] = NOTIFICATION_EMAIL
            recipients.append(NOTIFICATION_EMAIL)
        server.sendmail(SMTP_FROM, recipients, msg.as_string())
        server.quit()
        print(f"E-Mail gesendet an {to_email}: {subject}")
        return True
    except Exception as e:
        print(f"E-Mail-Fehler ({to_email}): {e}")
        return False


# Allowed image extensions for direct upload
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".heif", ".bmp", ".tiff", ".tif", ".svg"}
# SVG-Hinweis (17.07.2026): PIL/Pillow kann SVG nicht oeffnen (frueherer Crash-Grund,
# deshalb war SVG hier ausgeschlossen). SVGs werden jetzt direkt bei der Annahme mit
# cairosvg nach PNG konvertiert (_svg_to_png_bytes) und laufen ab dort den normalen
# Bild-Weg — hinter den Upload-Endpunkten existiert nie eine SVG-Datei.


def _svg_to_png_bytes(svg_bytes: bytes) -> bytes:
    """Konvertiert eine SVG-Datei (Bytes) nach PNG (Bytes) — cairosvg.

    Weisser Hintergrund ist Pflicht: transparente SVGs wuerden beim spaeteren
    RGB-Convert sonst schwarz — gleiche Falle wie der Transparenz-Fix in
    pdf_processor._resize_image_for_model (Befund App-Durchlauf 17.07.2026).
    output_width=1200 wie bisher im Web-Scan-Pfad: genug Detail fuer das
    Modell, haelt Riesen-PNGs aus Vektor-Rendering im Zaum (Seitenverhaeltnis
    bleibt erhalten). Wirft bei kaputtem SVG eine Exception — die Aufrufer
    fangen sie und antworten mit 400."""
    import cairosvg
    return cairosvg.svg2png(bytestring=svg_bytes, output_width=1200,
                            background_color="white")

# Rate limiting for login
# Sicherheitspaket 04.08.2026 (Befund 2 vom 02.08.): Schluessel ist jetzt die
# ECHTE Besucher-IP via absender.limit_schluessel — vorher zaehlte
# request.client.host (hinter NPM immer dieselbe Proxy-IP) alle Nutzer in
# EINEN Topf, ein Angreifer konnte damit alle echten Nutzer aussperren.
_login_attempts = defaultdict(list)
MAX_LOGIN_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300  # 5 minutes

# Anmelde-Limit pro Absender (Punkt 2 des Abo-Umbaus, Steves Mail an Michael
# 03.08.: "Anmelde-Limits pro Absender gelten zusaetzlich"). Gezaehlt werden
# nur ERFOLGREICHE Konto-Anlagen — Tippfehler kosten keinen Versuch.
_register_attempts = defaultdict(list)
MAX_REGISTRATIONS_PER_WINDOW = 5
REGISTER_WINDOW_SECONDS = 3600  # 1 Stunde

# Einladungs-Bremse (Sicherheits-Review 06.08.): Team-Einladungen versenden
# Mails an frei waehlbare Adressen — ohne Bremse waere das eine Spam-Kanone
# mit gueltiger Absender-Signatur. Gezaehlt werden nur ERFOLGREICH versandte
# Einladungen, pro Inhaber-Konto.
_einladung_attempts = defaultdict(list)
MAX_EINLADUNGEN_PER_WINDOW = 10
EINLADUNG_WINDOW_SECONDS = 3600  # 1 Stunde

# Fassung der Widerrufsbelehrung, der die Kundschaft beim Buchen zustimmt
# (08.08.2026). Wird mitprotokolliert und in der Bestaetigungs-Mail genannt:
# Aendert sich der Text, muss nachvollziehbar bleiben, WELCHE Fassung galt.
# Bei jeder inhaltlichen Aenderung von templates/widerruf.html hochzaehlen
# (bis 25.08.2026 lag der Text in frontend/widerruf.html).
WIDERRUFSBELEHRUNG_FASSUNG = "2026-08-24"

# Loesch-Bremse (Selbst-Review 08.08.): Der Loeschweg prueft das Passwort —
# mit einer gekaperten Sitzung koennte man es sonst unbegrenzt raten (und die
# Anmelde-Bremse greift dort gar nicht, weil man ja schon angemeldet ist).
# Gezaehlt werden nur FEHLVERSUCHE, pro Konto.
_loesch_attempts = defaultdict(list)
MAX_LOESCH_VERSUCHE = 5
LOESCH_WINDOW_SECONDS = 900  # 15 Minuten

# Schlanke E-Mail-Formatpruefung (Sicherheits-Review 06.08.): kein RFC-Parser,
# nur das Noetigste gegen Steuerzeichen/Leerzeichen und fehlende Domain —
# Konten mit unzustellbaren Adressen sollen gar nicht erst entstehen.
_EMAIL_RE = re.compile(r"^[^\s@\x00-\x1f\x7f]+@[^\s@\x00-\x1f\x7f]+\.[^\s@\x00-\x1f\x7f]+$")


def _email_plausibel(email: str) -> bool:
    return bool(email) and len(email) <= 254 and bool(_EMAIL_RE.match(email))


# Rate limiting for API (per API key)
_api_rate_minute = defaultdict(list)  # key_id -> [timestamps]
_api_rate_day = defaultdict(list)     # key_id -> [timestamps]
API_RATE_LIMIT_MINUTE = 60
API_RATE_LIMIT_DAY = 1000


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)
    # Clean up stale API temp files from previous runs/crashes
    api_tmp = os.path.join(UPLOAD_DIR, "api_tmp")
    if os.path.exists(api_tmp):
        cutoff = time.time() - 3600
        for f in os.listdir(api_tmp):
            fp = os.path.join(api_tmp, f)
            try:
                if os.path.isfile(fp) and os.path.getmtime(fp) < cutoff:
                    os.remove(fp)
            except Exception:
                pass
    # Create default admin user only if NO users exist at all (fresh install)
    conn = get_db()
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()
    if user_count == 0:
        # 18.08.2026 (Fable 5): KEIN fest eingebautes Passwort mehr im Quellcode.
        # Frueher wurde hier ein Admin mit dem oeffentlich im Code stehenden Passwort
        # "inkludocs2025" angelegt -> auf jeder Frischinstallation dieselben bekannten
        # Zugangsdaten (Sicherheitsluecke, oeffentlich einsehbar). Jetzt drei Wege:
        #   - CREATE_DEFAULT_ADMIN=false  -> es wird GAR KEIN Admin angelegt
        #     (so faehrt die oeffentliche Demo: sie braucht keinen Admin-Login).
        #   - INITIAL_ADMIN_PASSWORD gesetzt -> dieses Passwort wird verwendet.
        #   - sonst -> ein Zufallspasswort wird erzeugt und EINMALIG ins Log
        #     geschrieben. Damit existiert nie ein bekanntes Standard-Passwort.
        if os.getenv("CREATE_DEFAULT_ADMIN", "true").lower() not in ("false", "0", "no"):
            _admin_email = os.getenv("INITIAL_ADMIN_EMAIL", "kontakt@inklutec.de")
            _admin_pw_env = os.getenv("INITIAL_ADMIN_PASSWORD")
            _admin_pw = _admin_pw_env or secrets.token_urlsafe(18)
            try:
                create_user(_admin_email, _admin_pw, "Administrator", is_admin=1)
                if _admin_pw_env:
                    print(f"Default-Admin '{_admin_email}' angelegt (Passwort aus INITIAL_ADMIN_PASSWORD).")
                else:
                    print(f"Default-Admin '{_admin_email}' angelegt mit ZUFALLSPASSWORT: {_admin_pw}")
                    print("  -> bitte notieren und nach dem ersten Login aendern.")
            except Exception as e:
                print(f"Default-Admin-Anlage fehlgeschlagen: {e}")
        else:
            print("Kein Default-Admin angelegt (CREATE_DEFAULT_ADMIN=false).")

    # ── Demo-Modus: periodischer Aufraeumer fuer abgelaufene Sitzungen ──
    # Laeuft NUR bei DEMO_MODE=on (eigene Demo-Instanz). Entfernt Wegwerf-Sitzungen
    # (Nutzer + Projekte + Bilder + Dateien), die laenger als das gleitende TTL
    # untaetig waren. In Production/Staging wird der Task gar nicht erst gestartet.
    if demo_mod.demo_enabled():
        demo_mod.ensure_demo_schema()  # Zaehler-Tabelle der Demo anlegen (idempotent)
        async def _demo_cleanup_loop():
            while True:
                try:
                    n = await asyncio.get_event_loop().run_in_executor(
                        None, demo_mod.cleanup_expired_sessions, UPLOAD_DIR, RESULTS_DIR
                    )
                    if n:
                        print(f"[demo] {n} abgelaufene Sitzung(en) aufgeraeumt")
                except Exception as e:
                    print(f"[demo] Aufraeum-Fehler: {e}")
                await asyncio.sleep(600)  # alle 10 Minuten
        asyncio.create_task(_demo_cleanup_loop())

    yield


app = FastAPI(title="InkluDocs", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://inkludocs.inklutec.de", "https://staging.inkludocs.inklutec.de"],
    allow_methods=["GET", "POST", "PATCH", "DELETE"],
    allow_headers=["Content-Type", "X-API-Key"],
    allow_credentials=True,
    expose_headers=["X-Export-Warnings", "X-Export-Tagged", "X-Export-Total",
                    "X-RateLimit-Remaining-Minute", "X-RateLimit-Remaining-Day",
                    "X-RateLimit-Limit", "X-RateLimit-Remaining", "X-RateLimit-Reset", "Retry-After"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="/app/frontend"), name="static")

# Jinja2-Templates (fuer i18n-Seiten, inkrementell migriert)
templates = get_templates()


def create_token(user_id: int, email: str, is_admin: int = 0) -> str:
    expire = datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    return jwt.encode(
        {"sub": str(user_id), "email": email, "is_admin": is_admin, "exp": expire},
        SECRET_KEY, algorithm=ALGORITHM,
    )


def get_current_user(request: Request) -> dict:
    token = request.cookies.get("token")
    if not token:
        raise HTTPException(status_code=401, detail="Nicht angemeldet")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return {"id": int(payload["sub"]), "email": payload["email"], "is_admin": payload.get("is_admin", 0)}
    except JWTError:
        raise HTTPException(status_code=401, detail="Token ungueltig")


def get_optional_user(request: Request) -> Optional[dict]:
    """Wie get_current_user, aber ohne 401: liefert None statt Fehler, wenn
    kein (gueltiges) Login vorliegt. Fuer Routen, die fuer Gast UND
    eingeloggten Nutzer funktionieren muessen (z.B. die Sprachwahl).
    """
    try:
        return get_current_user(request)
    except HTTPException:
        return None


def require_admin(request: Request) -> dict:
    """Jeder Admin (Voll-Admin oder Nur-Einsicht) – lesender Zugriff."""
    user = get_current_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Nur fuer Administratoren")
    return user


def require_full_admin(request: Request) -> dict:
    """Nur Voll-Admins – fuer veraendernde Aktionen (Nutzer sperren, loeschen,
    Passwoerter zuruecksetzen, Konten anlegen, Admins verwalten).

    Liest die aktuelle Rechte-Stufe frisch aus der Datenbank, damit ein
    entzogenes oder herabgestuftes Recht sofort greift – nicht erst nach
    Ablauf des Login-Tokens.
    """
    user = get_current_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Nur fuer Administratoren")
    db_user = get_user_by_id(user["id"])
    if not db_user or not db_user.get("is_admin") or db_user.get("admin_level") != "full":
        raise HTTPException(status_code=403, detail="Nur fuer Voll-Administratoren")
    return user


# --- Gastzugang / Projekt-Freigabe (19.06.2026) ---
# Gast-Sitzung = eigenes guest_token-Cookie, an GENAU EINEN Freigabe-Token
# gebunden. Spiegelt das bestehende JWT-Auth-Muster (create_token / get_current_user),
# kein Eigenbau. Der eingeloggte Pfad bleibt voellig unberuehrt.
def create_guest_token(share_token, guest_email):
    """JWT fuer eine bestaetigte Gast-Sitzung (nach erfolgreichem E-Mail-Gate)."""
    expire = datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    return jwt.encode(
        {"type": "guest", "share_token": share_token, "guest": guest_email, "exp": expire},
        SECRET_KEY, algorithm=ALGORITHM,
    )


def get_guest_session(request, share_token):
    """Liest das guest_token-Cookie und liefert die Gast-Sitzung NUR, wenn sie zu
    diesem share_token gehoert. Sonst None (kein Zugriff)."""
    gt = request.cookies.get("guest_token")
    if not gt:
        return None
    try:
        payload = jwt.decode(gt, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        return None
    if payload.get("type") != "guest" or payload.get("share_token") != share_token:
        return None
    return {"guest": payload.get("guest"), "share_token": share_token}


def get_api_user(request: Request) -> dict:
    """Authenticate via X-API-Key header for public API endpoints."""
    api_key = request.headers.get("X-API-Key", "")
    if not api_key:
        raise HTTPException(status_code=401, detail="API-Key fehlt. Bitte X-API-Key Header setzen.")
    key_info = verify_api_key(api_key)
    if not key_info:
        raise HTTPException(status_code=401, detail="API-Key ungueltig oder deaktiviert.")
    return {"id": key_info["user_id"], "email": key_info["email"], "is_admin": 0, "api_key_id": key_info["id"]}


def check_api_rate_limit(api_key_id: int):
    """Check and enforce rate limits for an API key. Raises 429 if exceeded."""
    now = time.time()

    # Clean old entries and check minute limit
    _api_rate_minute[api_key_id] = [t for t in _api_rate_minute[api_key_id] if now - t < 60]
    if len(_api_rate_minute[api_key_id]) >= API_RATE_LIMIT_MINUTE:
        raise HTTPException(
            status_code=429,
            detail=f"Rate-Limit ueberschritten: max. {API_RATE_LIMIT_MINUTE} Anfragen pro Minute.",
            headers={
                "X-RateLimit-Limit": str(API_RATE_LIMIT_MINUTE),
                "X-RateLimit-Remaining": "0",
                "X-RateLimit-Reset": str(int(min(_api_rate_minute[api_key_id]) + 60)),
                "Retry-After": "60",
            }
        )

    # Clean old entries and check daily limit
    _api_rate_day[api_key_id] = [t for t in _api_rate_day[api_key_id] if now - t < 86400]
    if len(_api_rate_day[api_key_id]) >= API_RATE_LIMIT_DAY:
        raise HTTPException(
            status_code=429,
            detail=f"Tageslimit ueberschritten: max. {API_RATE_LIMIT_DAY} Anfragen pro Tag.",
            headers={
                "X-RateLimit-Limit": str(API_RATE_LIMIT_DAY),
                "X-RateLimit-Remaining": "0",
                "Retry-After": "3600",
            }
        )

    # Record this request
    _api_rate_minute[api_key_id].append(now)
    _api_rate_day[api_key_id].append(now)

    return {
        "minute_remaining": API_RATE_LIMIT_MINUTE - len(_api_rate_minute[api_key_id]),
        "day_remaining": API_RATE_LIMIT_DAY - len(_api_rate_day[api_key_id]),
    }


# ─── Auth Routes ─────────────────────────────────────────────

@app.post("/api/login")
async def login(request: Request):
    # Rate limiting — pro ECHTER Besucher-IP (Sicherheitspaket 04.08.2026):
    # request.client.host war hinter NPM immer die Proxy-IP, die Bremse
    # zaehlte global und liess sich als Aussperr-Angriff missbrauchen.
    client_ip = absender.limit_schluessel(request)
    now = time.time()
    _login_attempts[client_ip] = [t for t in _login_attempts[client_ip] if now - t < LOGIN_WINDOW_SECONDS]
    if len(_login_attempts[client_ip]) >= MAX_LOGIN_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Zu viele Anmeldeversuche. Bitte 5 Minuten warten.")

    data = await request.json()
    email = data.get("email", "").strip()
    password = data.get("password", "")
    user = verify_user(email, password)
    if not user:
        _login_attempts[client_ip].append(now)
        raise HTTPException(status_code=401, detail="E-Mail oder Passwort falsch")

    # Check email verification
    if not user.get("email_verified", 1):
        # Steve 08.06.2026: echte Umlaute auch im Backend-Detail-String. Das Frontend
        # erkennt den Fehlerpfad ueber msg.includes('nicht bestätigt') — bleibt damit
        # synchron mit der Umlaut-Umstellung in frontend/index.html und backend/templates/index.html.
        raise HTTPException(status_code=403, detail="E-Mail-Adresse noch nicht bestätigt. Bitte prüfen Sie Ihr Postfach oder fordern Sie einen neuen Bestätigungslink an.")

    # Update last_login
    conn = get_db()
    conn.execute("UPDATE users SET last_login = datetime('now') WHERE id = ?", (user["id"],))
    conn.commit()
    conn.close()
    token = create_token(user["id"], user["email"], user["is_admin"])
    response = JSONResponse({
        "ok": True,
        "email": user["email"],
        "display_name": user["display_name"],
        "is_admin": user["is_admin"],
    })
    response.set_cookie("token", token, httponly=True, secure=True, samesite="strict", max_age=TOKEN_EXPIRE_HOURS * 3600)
    return response


def _normiere_display_name(name) -> str:
    """Anzeigenamen auf EINE Zeile normieren und auf 100 Zeichen begrenzen.

    Review-Befund 7 (31.07.2026): display_name landet in System-Mails
    (HTML-Body, teils Betreff). Zeilenumbrueche und Steuerzeichen wuerden
    dort Header-/HTML-Injection ermoeglichen — darum laeuft JEDE Setz-Stelle
    (Registrierung, Admin-Anlage, Team-Einladung, Namensaenderung) durch
    diesen Filter. HTML-Escaping passiert zusaetzlich an der Ausgabe-Stelle.
    """
    name = re.sub(r"[\x00-\x1f\x7f]+", " ", str(name or ""))
    name = re.sub(r"\s{2,}", " ", name).strip()
    return name[:100]


@app.post("/api/register")
async def register(request: Request):
    # Registration can be disabled via environment variable
    if os.getenv("REGISTRATION_ENABLED", "true").lower() in ("false", "0", "no"):
        raise HTTPException(status_code=403, detail="Die Registrierung ist derzeit geschlossen. Bitte wende dich an kontakt@inklutec.de, um einen Testzugang zu erhalten.")
    data = await request.json()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    display_name = _normiere_display_name(data.get("display_name", ""))  # Befund 7: eine Zeile, max. 100

    if not _email_plausibel(email):
        raise HTTPException(status_code=400, detail="Bitte eine gueltige E-Mail-Adresse eingeben")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Passwort muss mindestens 8 Zeichen lang sein")
    if not display_name:
        raise HTTPException(status_code=400, detail="Bitte einen Namen eingeben")

    # Wegwerf-Schutz (Punkt 2 des Abo-Umbaus, 04.08.2026): Wegwerf-Adressen
    # hebeln die Free-Domain-Buendelung aus — ehrliche Ablehnung statt
    # stiller Drossel.
    email_domain = email.rsplit("@", 1)[-1]
    if email_domain in billing.WEGWERF_DOMAINS:
        raise HTTPException(status_code=400,
                            detail="Wegwerf-E-Mail-Adressen koennen nicht registriert werden. "
                                   "Bitte eine dauerhafte E-Mail-Adresse verwenden")

    # Anmelde-Limit pro Absender (echte IP dank absender.py): bremst
    # Konten-Farming, ohne ehrliche Einzel-Registrierungen zu stoeren.
    reg_ip = absender.limit_schluessel(request)
    now = time.time()
    _register_attempts[reg_ip] = [t for t in _register_attempts[reg_ip]
                                  if now - t < REGISTER_WINDOW_SECONDS]
    if len(_register_attempts[reg_ip]) >= MAX_REGISTRATIONS_PER_WINDOW:
        raise HTTPException(status_code=429,
                            detail="Zu viele Registrierungen von dieser Adresse. Bitte spaeter erneut versuchen")

    existing = get_user_by_email(email)
    if existing:
        raise HTTPException(status_code=409, detail="Diese E-Mail-Adresse ist bereits registriert")

    try:
        user_id = create_user(email, password, display_name, email_verified=0)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Registrierung fehlgeschlagen")
    _register_attempts[reg_ip].append(now)

    # Neue Nutzer starten bei den Neuigkeiten "auf dem aktuellen Stand". Bestandsnutzer
    # haben news_seen_until NULL und sehen daher beim naechsten Login alles Bisherige.
    try:
        _c = get_db()
        _c.execute("UPDATE users SET news_seen_until = ? WHERE id = ?", (_newest_news_key(), user_id))
        _c.commit()
        _c.close()
    except Exception:
        pass

    # Browser-Sprache als Konto-Sprache uebernehmen (03.07.2026): ein daenischer
    # Registrierer bekommt UI und Mails ab der ersten Minute in seiner Sprache.
    reg_lang = resolve_ui_language(request)
    if reg_lang not in SUPPORTED_LANGUAGES:
        reg_lang = "de"
    try:
        set_user_language(user_id, reg_lang)
    except Exception:
        reg_lang = "de"

    # Send verification email
    verify_token = create_email_verification_token(user_id, email)
    verify_url = f"{BASE_URL}/api/verify-email?token={verify_token}"
    _subj, email_body = _verification_mail_full(reg_lang, display_name, verify_url)
    send_email(email, _subj, email_body, bcc_admin=False)

    # Notify admin about new registration
    admin_body = f"""<!DOCTYPE html>
<html lang="de"><head><meta charset="utf-8"></head><body style="font-family:sans-serif;color:#1e293b;max-width:600px;">
<h2>Neue Registrierung bei InkluDocs</h2>
<p><strong>Name:</strong> {display_name}</p>
<p><strong>E-Mail:</strong> {email}</p>
<p><strong>Zeitpunkt:</strong> {datetime.utcnow().strftime("%d.%m.%Y %H:%M")} UTC</p>
<p style="color:#64748b;font-size:0.9rem;">Die E-Mail-Adresse wurde noch nicht bestätigt.</p>
</body></html>"""
    send_email(NOTIFICATION_EMAIL, "InkluDocs: Neue Registrierung", admin_body, bcc_admin=False)

    return JSONResponse({"ok": True, "message": "Bestätigungslink wurde gesendet. Bitte prüfen Sie Ihr Postfach."})


@app.post("/api/logout")
async def logout():
    response = JSONResponse({"ok": True})
    response.delete_cookie("token")
    return response


@app.get("/api/verify-email")
async def verify_email_registration(request: Request, token: str = ""):
    """Verify email address from registration link."""
    lang = resolve_ui_language(request)
    _ = get_gettext(lang)
    zur_anmeldung = f'<p><a href="/">{_("Zur Anmeldung")}</a></p>'
    if not token:
        return _auth_notice_page(lang,
            f'<p style="color:#dc2626;margin:2rem 0;">{_("Ungültiger Bestätigungslink.")}</p>' + zur_anmeldung,
            status_code=400)

    result = mark_email_verified(token)
    if not result:
        return _auth_notice_page(lang,
            f'<p style="color:#dc2626;margin:2rem 0;">{_("Dieser Bestätigungslink ist ungültig oder abgelaufen.")}</p>' + zur_anmeldung,
            status_code=400)

    # Notify admin that email was confirmed
    admin_body = f"""<!DOCTYPE html>
<html lang="de"><head><meta charset="utf-8"></head><body style="font-family:sans-serif;color:#1e293b;max-width:600px;">
<h2 style="color:#16a34a;">E-Mail bestaetigt</h2>
<p><strong>E-Mail:</strong> {result['email']}</p>
<p>Der Benutzer kann sich jetzt anmelden.</p>
</body></html>"""
    send_email(NOTIFICATION_EMAIL, "InkluDocs: E-Mail bestaetigt", admin_body, bcc_admin=False)

    return _auth_notice_page(lang,
        f'''<p style="color:#16a34a;font-size:1.2rem;margin:2rem 0;font-weight:600;">&#10003; {_("E-Mail-Adresse erfolgreich bestätigt!")}</p>
    <p>{_("Ihr Konto ist jetzt aktiv. Sie können sich jetzt anmelden.")}</p>
    <p style="margin-top:1.5rem;"><a href="/" style="display:inline-block;background:#e87722;color:white;padding:0.75rem 1.5rem;border-radius:6px;text-decoration:none;font-weight:600;">{_("Jetzt anmelden")}</a></p>''',
        title_suffix=_("E-Mail bestätigt"))


@app.post("/api/resend-verification")
async def resend_verification(request: Request):
    """Resend verification email for unverified accounts."""
    data = await request.json()
    email = data.get("email", "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-Mail-Adresse erforderlich")

    token = resend_verification_token(email)
    if not token:
        # Don't reveal whether the email exists or is already verified
        return JSONResponse({"ok": True, "message": "Falls ein unverifiziertes Konto existiert, wurde eine neue Bestaetigungsmail gesendet."})

    user = get_user_by_email(email)
    display_name = user["display_name"] if user else ""
    verify_url = f"{BASE_URL}/api/verify-email?token={token}"
    _subj, email_body = _verification_mail_short(_mail_lang(user), display_name, verify_url)
    send_email(email, _subj, email_body, bcc_admin=False)

    return JSONResponse({"ok": True, "message": "Falls ein unverifiziertes Konto existiert, wurde eine neue Bestaetigungsmail gesendet."})


@app.get("/api/me")
async def me(user: dict = Depends(get_current_user)):
    db_user = get_user_by_id(user["id"])
    if not db_user:
        raise HTTPException(status_code=401, detail="User nicht gefunden")
    daily_used = get_daily_image_count(db_user["id"])

    # Abo-Tageslauf (06.08.2026): Erinnerungen + Auto-Verlaengerungen ohne
    # Cron — laeuft hoechstens 1x/Tag, im Hintergrund-Thread (Mails!).
    # Prozess-Merker zuerst: im Normalfall (heute schon gelaufen) entsteht
    # hier weder ein Thread noch ein DB-Zugriff (Review-Befund 06.08.).
    try:
        if _abo_tageslauf_noetig():
            threading.Thread(target=_abo_tageslauf, daemon=True).start()
    except Exception:
        pass

    # Abo & Verbrauch: Zahlen IMMER fuer den AKTIVEN Topf (Kontext-
    # Umschalter, 06.08.2026) — ein Mitglied, das gerade im Team-Kontext
    # arbeitet, sieht den gemeinsamen Topf; im eigenen Kontext das eigene
    # Abo. Quelle billing.pruefe_kontingent (eine Wahrheit fuer Anzeige UND
    # Sperre) — pruefe_kontingent loest den Kontext selbst auf.
    abo_info = billing.pruefe_kontingent(db_user["id"])
    konto_id = billing._konto_fuer(db_user["id"])

    # Verfuegbare Kontexte: eigenes Abo + jede Team-/Enterprise-
    # Mitgliedschaft mit AKTIVEM Plan des Inhabers (abgelaufene Teams
    # werden nicht angeboten — dort kann man nicht arbeiten).
    eigener_plan = billing.effektiver_plan(db_user)
    kontexte = [{"topf": "eigen", "plan": eigener_plan}]
    conn = get_db()
    try:
        _rows = conn.execute(
            "SELECT u.id, u.display_name, u.team_name, u.plan, u.plan_gueltig_bis, "
            "u.geplanter_plan, COALESCE(u.auto_verlaengerung, 1) AS auto_verlaengerung "
            "FROM team_mitgliedschaften m JOIN users u ON u.id = m.inhaber_id "
            "WHERE m.mitglied_id = ? ORDER BY u.display_name COLLATE NOCASE",
            (db_user["id"],)).fetchall()
    finally:
        conn.close()
    aktives_team = None
    for _r in _rows:
        _plan = billing.effektiver_plan(_r)
        if _plan not in billing.PLAN_SITZE:
            continue
        # Endet der gemeinsame Topf absehbar (Inhaber hat gekuendigt oder
        # einen Downgrade ohne Sitze vorgemerkt), sagen wir das Datum
        # ehrlich dazu — Steve 07.08.: „das sollte jeder im Team sehen".
        _endet_am = None
        _gp = (_r["geplanter_plan"] or "").strip()
        if not int(_r["auto_verlaengerung"] or 0) or (_gp and _gp not in billing.PLAN_SITZE):
            _endet_am = (_r["plan_gueltig_bis"] or "")[:10] or None
        _eintrag = {
            "topf": _r["id"],
            "plan": _plan,
            "team_name": _r["team_name"] or "",
            "inhaber_name": _r["display_name"],
            "endet_am": _endet_am,
        }
        kontexte.append(_eintrag)
        if konto_id == _r["id"]:
            aktives_team = _eintrag

    _verfuegbar = abo_info.get("verfuegbar_monat")
    _verbraucht = abo_info.get("verbraucht", 0)
    _av = db_user.get("auto_verlaengerung")

    return {
        "ok": True,
        "user": {
            "id": db_user["id"],
            "email": db_user["email"],
            "display_name": db_user["display_name"],
            "is_admin": db_user["is_admin"],
            "admin_level": db_user.get("admin_level", "full"),
        },
        # deprecated: altes Tages-Limit — bleibt bis zur Frontend-Umstellung
        # auf den "abo"-Block mitgeliefert, danach entfernen.
        "daily_limit": {
            "used": daily_used,
            "limit": effektives_api_tageslimit(db_user),
            "remaining": max(0, effektives_api_tageslimit(db_user) - daily_used),
        },
        "abo": {
            # Plan des AKTIVEN Topfs (danach richten sich die Credit-Zeilen).
            "plan": abo_info.get("plan", "free"),
            # Kontext-Umschalter: was ist gerade aktiv, was ist waehlbar.
            "aktiver_topf": "eigen" if aktives_team is None else aktives_team["topf"],
            "aktives_team": aktives_team,
            "kontexte": kontexte,
            # Eckdaten des EIGENEN Plans — Kuendigung/Laufzeit betreffen immer
            # das eigene Abo, nie ein Team, in dem man nur Mitglied ist.
            "eigen": {
                "plan": eigener_plan,
                "gueltig_bis": (db_user.get("plan_gueltig_bis") or "")[:10] or None,
                "laufzeit_monate": db_user.get("plan_laufzeit_monate"),
                # Treue-Monatsabo (AGB Ziffer 7): laeuft das Abo bereits in
                # der Anschluss-Kondition? Steuert die Anzeige auf /abo.
                "treue": bool(db_user.get("plan_treue")),
                "auto_verlaengerung": bool(1 if _av is None else _av),
                "ist_topf_inhaber": eigener_plan in billing.PLAN_SITZE,
                "team_name": db_user.get("team_name") or "",
                "quelle": db_user.get("plan_quelle") or None,
                # Vorgemerkter Plan-Wechsel (Downgrade) — die Oberflaeche
                # zeigt „ab TT.MM. gilt Plan X" statt eines Wechsel-Angebots.
                "geplanter_plan": db_user.get("geplanter_plan") or None,
                "geplante_laufzeit": db_user.get("geplante_laufzeit"),
            },
            # Stripe-Buchungsweg (06.08. nachts): steuert, ob die Abo-Seite
            # echte Buchen-Knoepfe oder den E-Mail-Hinweis zeigt.
            "stripe_aktiv": stripe_zahlung.AKTIV,
            "hat_stripe_kunde": bool(db_user.get("stripe_customer_id")),
            # kompatibel zu Dashboard/App-Shell (Team-Zeile):
            "ist_team_mitglied": aktives_team is not None,
            "team_inhaber": aktives_team["inhaber_name"] if aktives_team else None,
            "kontingent": abo_info.get("kontingent"),
            "uebertrag": abo_info.get("uebertrag", 0),
            "verfuegbar_monat": _verfuegbar,
            "verbraucht": _verbraucht,
            "rest": None if _verfuegbar is None else max(0, _verfuegbar - _verbraucht),
            "pakete_rest": abo_info.get("pakete_rest", 0),
            "zeitraum_ende": abo_info.get("zeitraum_ende"),
            # Punkt 2 (04.08.2026): Free-Volumen ist bei Firmen-Domains
            # gebuendelt — die Oberflaeche sagt dann ehrlich, dass der
            # Verbrauch fuer die ganze Domain gilt.
            "domain_pool": abo_info.get("domain_pool"),
            "enforcement": billing.ABO_ENFORCEMENT,
            # Betreiber-Konten (Admins) werden gezaehlt, aber nie gesperrt.
            # Ohne dieses Kennzeichen zeigte die Oberflaeche ihnen ein
            # Kontingent an, das fuer sie gar nicht gilt (Steve-Befund 31.07.).
            "ist_betreiber": bool(db_user["is_admin"]),
        },
    }


@app.post("/api/change-password")
async def change_password(request: Request, user: dict = Depends(get_current_user)):
    data = await request.json()
    old_password = data.get("old_password", "")
    new_password = data.get("new_password", "")

    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="Neues Passwort muss mindestens 8 Zeichen lang sein")

    db_user = get_user_by_id(user["id"])
    verified = verify_user(db_user["email"], old_password)
    if not verified:
        raise HTTPException(status_code=401, detail="Aktuelles Passwort ist falsch")

    admin_reset_password(user["id"], new_password)
    return {"ok": True, "message": "Passwort wurde geaendert"}


@app.post("/api/konto/loeschen")
async def konto_loeschen(request: Request, user: dict = Depends(get_current_user)):
    """Konto-Selbstloeschung durch den Kunden (08.08.2026, DSGVO).

    Bis hierher konnten nur Admins loeschen. Reihenfolge ist wichtig:
    1. Passwort verifizieren (Schutz gegen fremden Zugriff auf offene Sitzung),
    2. Betreiber-Konten ablehnen (sonst sperrt sich der Betrieb aus),
    3. laufende Stripe-Subscription SOFORT beenden — sonst bucht Stripe
       weiter gegen ein Konto, das es nicht mehr gibt,
    4. Team-Mitglieder informieren (ihr gemeinsamer Topf faellt weg),
    5. Abschieds-Mail + Admin-Benachrichtigung (VOR dem Loeschen verschicken,
       danach sind Adresse und Name weg),
    6. Daten loeschen (delete_user_data raeumt Projekte, Dateien,
       Mitgliedschaften, Einladungen, API-Schluessel mit ab),
    7. Sitzungs-Cookie entfernen.
    """
    data = await request.json()
    passwort = data.get("password", "")
    db_user = get_user_by_id(user["id"])
    if not db_user:
        raise HTTPException(status_code=401, detail="Konto nicht gefunden")
    jetzt = time.time()
    _loesch_attempts[user["id"]] = [t for t in _loesch_attempts[user["id"]]
                                    if jetzt - t < LOESCH_WINDOW_SECONDS]
    if len(_loesch_attempts[user["id"]]) >= MAX_LOESCH_VERSUCHE:
        raise HTTPException(status_code=429,
                            detail="Zu viele Fehlversuche. Bitte 15 Minuten warten.")
    if not passwort or not verify_user(db_user["email"], passwort):
        _loesch_attempts[user["id"]].append(jetzt)
        raise HTTPException(status_code=401, detail="Passwort ist falsch")
    if db_user.get("is_admin"):
        raise HTTPException(
            status_code=403,
            detail="Betreiber-Konten können sich nicht selbst löschen. Bitte an "
                   "support@inkludocs.de wenden.")

    # 3. Stripe: laufendes Abo sofort beenden (nicht erst zum Periodenende —
    #    das Konto verschwindet ja jetzt).
    sub_id = (db_user.get("stripe_subscription_id") or "").strip()
    kunde_id = (db_user.get("stripe_customer_id") or "").strip()
    # Nur ein WIRKLICH laufendes Stripe-Abo blockiert. Eine blosse
    # stripe_customer_id (jemand hat mal etwas gekauft) darf niemanden
    # aussperren — sonst koennte sich auf einer Instanz ohne Stripe-Schluessel
    # kein Altkunde mehr loeschen.
    hat_stripe_abo = bool(sub_id) or (db_user.get("plan_quelle") == "stripe"
                                      and (db_user.get("plan") or "free") != "free")
    if hat_stripe_abo and not stripe_zahlung.AKTIV:
        # Darf nicht still passieren: die Subscription liefe bei Stripe weiter,
        # waehrend das Konto hier verschwindet. Lieber abbrechen und melden.
        logger.error("Loeschung abgelehnt: Stripe-Abo %s, aber Stripe ist nicht "
                     "konfiguriert (user=%s)", sub_id or "(ohne id)", db_user["id"])
        raise HTTPException(
            status_code=503,
            detail="Dein Konto hat ein laufendes Abo, das wir gerade nicht beenden können. "
                   "Bitte melde dich kurz bei support@inkludocs.de — wir erledigen das von Hand.")
    if kunde_id and stripe_zahlung.AKTIV:
        # Ueber den KUNDEN aufraeumen, nicht nur ueber die gespeicherte
        # Subscription: sonst ueberlebt ein Abo, dessen Webhook nie ankam.
        try:
            beendet = stripe_zahlung.beende_alle_subscriptions(kunde_id)
            logger.info("Konto-Loeschung: %s Stripe-Abo(s) beendet (user=%s)",
                        beendet, db_user["id"])
        except Exception:
            logger.exception("Stripe-Abo beim Loeschen nicht beendet (user=%s, kunde=%s)",
                             db_user["id"], kunde_id)
            raise HTTPException(
                status_code=502,
                detail="Dein laufendes Abo konnte gerade nicht beendet werden. Bitte in ein "
                       "paar Minuten erneut versuchen — wir wollen vermeiden, dass dir nach "
                       "der Löschung noch etwas berechnet wird.")

    elif sub_id and stripe_zahlung.AKTIV:
        try:
            stripe_zahlung.loese_schedule(sub_id)
            stripe_zahlung.beende_subscription_sofort(sub_id)
        except Exception:
            logger.exception("Stripe-Abo beim Loeschen nicht beendet (user=%s, sub=%s)",
                             db_user["id"], sub_id)
            raise HTTPException(
                status_code=502,
                detail="Dein laufendes Abo konnte gerade nicht beendet werden. Bitte in ein "
                       "paar Minuten erneut versuchen — wir wollen vermeiden, dass dir nach "
                       "der Löschung noch etwas berechnet wird.")

    # 4. Team-Mitglieder informieren, solange die Daten noch da sind.
    conn = get_db()
    try:
        mitglieder = [dict(r) for r in conn.execute(
            "SELECT u.email, u.display_name FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.mitglied_id WHERE m.inhaber_id = ?",
            (db_user["id"],)).fetchall()]
        inhaber = [dict(r) for r in conn.execute(
            "SELECT u.email, u.display_name, u.team_name FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.inhaber_id WHERE m.mitglied_id = ?",
            (db_user["id"],)).fetchall()]
    finally:
        conn.close()
    # Umgekehrter Fall: Wer nur MITGLIED ist, gibt einen bezahlten Platz frei —
    # der Inhaber soll das erfahren, sonst wundert er sich ueber die Liste.
    team_label = html.escape(db_user.get("team_name") or "") or f"das Team von {html.escape(db_user['display_name'])}"
    for m in mitglieder:
        try:
            _abo_mail(m["email"], "InkluDocs: Euer gemeinsames Guthaben endet",
                      "Team aufgelöst",
                      [f"Hallo {html.escape(m['display_name'] or '')},",
                       f"{html.escape(db_user['display_name'])} hat sein InkluDocs-Konto gelöscht; "
                       f"damit entfällt {team_label}.",
                       "Ab sofort arbeitest du wieder mit deinem eigenen Guthaben "
                       "(oder dem kostenlosen Free-Plan). Dein Konto und alle deine Projekte "
                       "bleiben unverändert erhalten."])
        except Exception:
            logger.exception("Team-Info beim Loeschen fehlgeschlagen (%s)", m["email"])

    for inh in inhaber:
        try:
            _abo_mail(inh["email"], "InkluDocs: Ein Team-Mitglied hat sein Konto gelöscht",
                      "Ein Platz ist wieder frei",
                      [f"Hallo {html.escape(inh['display_name'] or '')},",
                       f"{html.escape(db_user['display_name'])} "
                       f"({html.escape(db_user['email'])}) hat das eigene InkluDocs-Konto "
                       f"gelöscht und ist damit nicht mehr in "
                       f"{html.escape(inh.get('team_name') or 'deinem Team')}.",
                       "Der Platz ist wieder frei — du kannst jemand anderen einladen. "
                       "An deinem Abo und deinem Guthaben ändert sich nichts."])
        except Exception:
            logger.exception("Inhaber-Info beim Loeschen fehlgeschlagen (%s)", inh["email"])

    # 5. Abschied + Admin-Info (danach sind die Daten weg).
    mail, name = db_user["email"], html.escape(db_user["display_name"] or "")
    try:
        _abo_mail(mail, "InkluDocs: Dein Konto wurde gelöscht", "Dein Konto wurde gelöscht",
                  [f"Hallo {name},",
                   "dein InkluDocs-Konto wurde soeben auf deinen Wunsch hin gelöscht. "
                   "Alle deine Projekte, Dokumente und Alt-Texte sind damit entfernt.",
                   "Falls du dich irgendwann wieder anmelden möchtest, kannst du jederzeit "
                   "ein neues Konto anlegen. Danke, dass du InkluDocs genutzt hast!"])
        _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs: Konto gelöscht ({mail})",
                  "Konto-Selbstlöschung",
                  [f"{html.escape(mail)} hat das eigene Konto gelöscht.",
                   f"Plan zum Zeitpunkt der Löschung: {db_user.get('plan') or 'free'}"
                   + (f", Team mit {len(mitglieder)} Mitgliedern aufgelöst." if mitglieder else ".")])
    except Exception:
        logger.exception("Abschieds-Mail fehlgeschlagen (%s)", mail)

    # 6. + 7. Daten weg, Sitzung beenden.
    user_upload_dir = os.path.join(UPLOAD_DIR, str(db_user["id"]))
    user_results_dir = os.path.join(RESULTS_DIR, str(db_user["id"]))
    for pfad in (user_upload_dir, user_results_dir):
        try:
            if os.path.exists(pfad):
                shutil.rmtree(pfad)
        except Exception:
            logger.exception("Dateien beim Loeschen nicht entfernt: %s", pfad)
    delete_user_data(db_user["id"])
    antwort = JSONResponse({"ok": True, "message": "Konto wurde gelöscht"})
    antwort.delete_cookie("token")
    return antwort


@app.post("/api/change-email")
async def change_email(request: Request, user: dict = Depends(get_current_user)):
    """Request email change – sends confirmation link to new address."""
    data = await request.json()
    new_email = data.get("new_email", "").strip().lower()

    if not new_email or "@" not in new_email:
        raise HTTPException(status_code=400, detail="Bitte eine gueltige E-Mail-Adresse eingeben")

    if new_email == user["email"]:
        raise HTTPException(status_code=400, detail="Das ist bereits deine aktuelle E-Mail-Adresse")

    existing = get_user_by_email(new_email)
    if existing:
        raise HTTPException(status_code=409, detail="Diese E-Mail-Adresse ist bereits vergeben")

    db_user = get_user_by_id(user["id"])
    token = create_email_change_token(user["id"], new_email)
    confirm_url = f"{BASE_URL}/api/confirm-email?token={token}"

    _subj, email_body = _email_change_mail(_mail_lang(db_user), db_user['display_name'], new_email, confirm_url)
    sent = send_email(new_email, _subj, email_body, bcc_admin=False)
    if not sent:
        raise HTTPException(status_code=500, detail="Bestaetigungsmail konnte nicht gesendet werden. Bitte spaeter erneut versuchen.")

    return {"ok": True, "message": f"Bestaetigungslink wurde an {new_email} gesendet. Bitte pruefe dein Postfach."}


@app.get("/api/confirm-email")
async def confirm_email(request: Request, token: str = ""):
    """Confirm email change via link from confirmation email."""
    lang = resolve_ui_language(request)
    _ = get_gettext(lang)
    if not token:
        raise HTTPException(status_code=400, detail="Token fehlt")

    result = confirm_email_change(token)
    if not result:
        neuer_link = _('Bitte fordere in den {einstellungen} einen neuen Link an.').format(
            einstellungen=f'<a href="/app">{_("Einstellungen")}</a>')
        return _auth_notice_page(lang,
            f'''<p style="margin:2rem 0;color:var(--error,#dc2626);font-weight:600;">{_("Dieser Bestätigungslink ist ungültig oder abgelaufen.")}</p>
<p>{neuer_link}</p>''',
            status_code=400)

    # Success – show confirmation page and auto-redirect
    db_user = get_user_by_id(result["user_id"])
    geaendert = _('Deine E-Mail-Adresse wurde erfolgreich auf {email} geändert.').format(email=result['new_email'])
    weiter = _('Du wirst in 3 Sekunden weitergeleitet.')
    return _auth_notice_page(lang,
        f'''<p style="margin:2rem 0;color:var(--success,#16a34a);font-weight:600;">{geaendert}</p>
<p>{weiter} <a href="/app">{_("Jetzt zur App")}</a></p>''',
        head_extra='<meta http-equiv="refresh" content="3;url=/app">')


@app.post("/api/change-displayname")
async def change_displayname(request: Request, user: dict = Depends(get_current_user)):
    data = await request.json()
    new_name = _normiere_display_name(data.get("display_name", ""))  # Befund 7: eine Zeile, max. 100

    if not new_name:
        raise HTTPException(status_code=400, detail="Bitte einen Namen eingeben")
    if len(new_name) > 100:
        raise HTTPException(status_code=400, detail="Name darf maximal 100 Zeichen lang sein")

    conn = get_db()
    conn.execute("UPDATE users SET display_name = ? WHERE id = ?", (new_name, user["id"]))
    conn.commit()
    conn.close()
    return {"ok": True, "message": "Name wurde geaendert", "display_name": new_name}


# ─── Password Reset ──────────────────────────────────────────

@app.post("/api/forgot-password")
async def forgot_password(request: Request):
    data = await request.json()
    email = data.get("email", "").strip().lower()
    user = get_user_by_email(email)

    # Always return success (don't reveal if email exists)
    if not user:
        return {"ok": True, "message": "Falls ein Konto existiert, wird ein Reset-Link angezeigt."}

    token = create_password_reset_token(user["id"])
    reset_url = f"{BASE_URL}/reset?token={token}"

    # Send reset link via email (never expose in API response)
    _subj, email_body = _reset_mail(_mail_lang(user), user['display_name'], reset_url)
    send_email(email, _subj, email_body, bcc_admin=False)

    return {
        "ok": True,
        "message": "Falls ein Konto mit dieser E-Mail existiert, wurde ein Reset-Link per E-Mail gesendet.",
    }


def _team_beitritt_nach_passwort(user_id: int) -> None:
    """Loest offene konto_neu-Einladungen ein, sobald die eingeladene Person
    ihr Passwort gesetzt hat (= ihre Registrierung; neues Abomodell 06.08.).

    Nur Einladungen mit konto_neu=1 (Konto entstand DURCH die Einladung) —
    Bestandskonten treten weiterhin ausschliesslich ueber den bewussten
    Annahme-Klick bei. Sitz-Check atomar wie im Annahme-Pfad; ist das Team
    inzwischen voll oder der Plan abgelaufen, bleibt die Person einfach auf
    Free (Konto und Projekte unberuehrt). Fehler werden geloggt, nie geworfen.
    """
    try:
        db_user = get_user_by_id(user_id)
        if not db_user:
            return
        conn = get_db()
        try:
            conn.isolation_level = None
            conn.execute("BEGIN IMMEDIATE")
            offene = conn.execute(
                "SELECT e.id, e.inhaber_id, u.plan, u.plan_gueltig_bis "
                "FROM team_einladungen e JOIN users u ON u.id = e.inhaber_id "
                "WHERE e.email = ? AND e.konto_neu = 1 AND e.eingeloest_am IS NULL "
                "AND e.gueltig_bis > datetime('now')",
                ((db_user.get("email") or "").lower(),)).fetchall()
            for e in offene:
                plan = billing.effektiver_plan(e)
                sitze_max = billing.sitze_fuer_plan(plan)
                if sitze_max is None:
                    continue
                belegt = int(conn.execute(
                    "SELECT COUNT(*) FROM team_mitgliedschaften WHERE inhaber_id = ?",
                    (e["inhaber_id"],)).fetchone()[0]) + 1
                if belegt >= sitze_max:
                    continue
                conn.execute("INSERT OR IGNORE INTO team_mitgliedschaften (inhaber_id, mitglied_id) "
                             "VALUES (?, ?)", (e["inhaber_id"], user_id))
                conn.execute("UPDATE users SET aktiver_topf = ? WHERE id = ? AND aktiver_topf IS NULL",
                             (e["inhaber_id"], user_id))
                conn.execute("UPDATE team_einladungen SET eingeloest_am = datetime('now') WHERE id = ?",
                             (e["id"],))
            conn.execute("COMMIT")
        except Exception:
            try:
                conn.execute("ROLLBACK")
            except Exception:
                pass
            raise
        finally:
            conn.close()
    except Exception:
        logger.exception("Team-Beitritt nach Passwort-Setzen fehlgeschlagen (user=%s)", user_id)


@app.post("/api/reset-password")
async def do_reset_password(request: Request):
    data = await request.json()
    token = data.get("token", "")
    new_password = data.get("new_password", "")

    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="Passwort muss mindestens 8 Zeichen lang sein")

    # user_id VOR dem Einloesen merken (reset_password verbraucht den Token) —
    # noetig fuer den automatischen Team-Beitritt eingeladener Neu-Konten.
    conn = get_db()
    try:
        _pr = conn.execute("SELECT user_id FROM password_resets WHERE token = ?",
                           (token,)).fetchone()
    finally:
        conn.close()

    if not reset_password(token, new_password):
        raise HTTPException(status_code=400, detail="Reset-Link ist ungueltig oder abgelaufen")

    if _pr:
        _team_beitritt_nach_passwort(int(_pr["user_id"]))

    return {"ok": True, "message": "Passwort wurde zurueckgesetzt. Sie koennen sich jetzt anmelden."}


# ─── Admin Routes ────────────────────────────────────────────

@app.get("/api/admin/users")
async def admin_list_users(user: dict = Depends(require_admin)):
    users = list_all_users()
    conn = get_db()
    try:
        for u in users:
            row = conn.execute("SELECT COUNT(*) as cnt FROM projects WHERE user_id = ?",
                               (u["id"],)).fetchone()
            u["project_count"] = row["cnt"]
            # Abo-Eckdaten gleich mitliefern (Steve 07.08.): Die Liste soll
            # zeigen, WAS jemand hat — sonst muss man fuer jede Kleinigkeit
            # erst die E-Mail abtippen und in einem anderen Formular nachsehen.
            u["plan"] = billing.effektiver_plan(u)
            u["plan_gueltig_bis"] = (u.get("plan_gueltig_bis") or "")[:10] or None
            u["plan_laufzeit_monate"] = u.get("plan_laufzeit_monate")
            u["plan_quelle"] = u.get("plan_quelle") or None
            av = u.get("auto_verlaengerung")
            u["auto_verlaengerung"] = bool(1 if av is None else av)
            u["team_name"] = u.get("team_name") or ""
            u["team_mitglieder"] = conn.execute(
                "SELECT COUNT(*) FROM team_mitgliedschaften WHERE inhaber_id = ?",
                (u["id"],)).fetchone()[0]
            # Passwort-Hash und Stripe-Kennungen gehoeren nicht in eine Liste,
            # die im Browser landet.
            for feld in ("password_hash", "stripe_customer_id", "stripe_subscription_id"):
                u.pop(feld, None)
    finally:
        conn.close()
    return {"users": users}


@app.get("/api/admin/users/{user_id}/report")
async def admin_kunden_report(user_id: int, monate: int = 12,
                              user: dict = Depends(require_admin)):
    """Verbrauchs- und Abo-Report zu einem Konto (Steve-Wunsch 07.08.2026).

    Beantwortet die Fragen, die vor einem Kundengespraech aufkommen: Was hat
    die Person gebucht, was verbraucht sie, wie entwickelt sich das, und was
    kostet uns das ungefaehr.

    Die Kostenangabe ist ausdruecklich eine SCHAETZUNG (siehe
    billing.KOSTEN_PRO_CREDIT_EUR) — echte Modellkosten fallen pro Aufruf an
    und sind einzelnen Kunden nicht sauber zuzuordnen. Lieber eine ehrlich
    beschriftete Groessenordnung als eine Zahl, die Genauigkeit vortaeuscht.
    """
    ziel = get_user_by_id(user_id)
    if not ziel:
        raise HTTPException(status_code=404, detail="Konto nicht gefunden")
    monate = max(1, min(int(monate or 12), 36))

    zustand = billing.pruefe_kontingent(user_id)
    # Hoertest-Befund 11.08.2026 (Steve): pruefe_kontingent rechnet mit dem
    # AKTIVEN Arbeits-Topf (_konto_fuer). Arbeitet die Person gerade aus einem
    # fremden Team-Topf, zeigte der Report "Plan: Single" und darunter das
    # Team-Kontingent (z. B. 100 statt 50) — ohne den Zusammenhang zu nennen.
    # Deshalb: fremden Topf explizit mitliefern, das Frontend sagt ihn an.
    topf_fremd = None
    topf_id = billing._konto_fuer(user_id)
    if topf_id != user_id:
        topf_inhaber = get_user_by_id(topf_id)
        if topf_inhaber:
            topf_fremd = {
                "inhaber": topf_inhaber["display_name"],
                "team_name": topf_inhaber.get("team_name") or "",
            }
    conn = get_db()
    try:
        # Verbrauch je Monat des EIGENEN Topfes (Steve 11.08.2026: das ist
        # die Zahl, die die Kosten DIESES Kontos treibt — bei Team-Inhabern
        # inklusive der Mitglieder). Was die Person in fremden Toepfen tut,
        # steht in der Aufstellung "Verbrauch nach Topf" darunter.
        verlauf = [dict(r) for r in conn.execute(
            "SELECT strftime('%Y-%m', created_at) AS monat, "
            "       SUM(CASE WHEN user_id = ? THEN credits ELSE 0 END) AS selbst, "
            "       SUM(credits) AS topf "
            "FROM usage_events WHERE COALESCE(konto_user_id, user_id) = ? "
            "  AND created_at >= date('now', ?) "
            "GROUP BY monat ORDER BY monat DESC",
            (user_id, user_id, f"-{monate} months")).fetchall()]
        gesamt = conn.execute(
            "SELECT COALESCE(SUM(credits), 0) FROM usage_events "
            "WHERE user_id = ? OR konto_user_id = ?", (user_id, user_id)).fetchone()[0]
        projekte = conn.execute("SELECT COUNT(*) FROM projects WHERE user_id = ?",
                                (user_id,)).fetchone()[0]
        bilder = conn.execute(
            "SELECT COUNT(*) FROM images WHERE project_id IN "
            "(SELECT id FROM projects WHERE user_id = ?)", (user_id,)).fetchone()[0]
        letzte_aktion = conn.execute(
            "SELECT MAX(created_at) FROM usage_events WHERE user_id = ?",
            (user_id,)).fetchone()[0]
        pakete = [dict(r) for r in conn.execute(
            "SELECT groesse, verbleibend, quelle, notiz, erstellt_am, verfaellt_am "
            "FROM quota_pakete WHERE user_id = ? ORDER BY erstellt_am DESC LIMIT 20",
            (user_id,)).fetchall()]
        mitglieder = [dict(r) for r in conn.execute(
            "SELECT u.display_name, u.email FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.mitglied_id WHERE m.inhaber_id = ? "
            "ORDER BY u.display_name", (user_id,)).fetchall()]
        teams = [dict(r) for r in conn.execute(
            "SELECT u.id AS inhaber_id, u.display_name AS inhaber, u.team_name "
            "FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.inhaber_id WHERE m.mitglied_id = ?",
            (user_id,)).fetchall()]
        api_aufrufe = conn.execute(
            "SELECT COUNT(*) FROM api_usage WHERE user_id = ?", (user_id,)).fetchone()[0]
        # Verbrauch der Person aufgeteilt nach dem belasteten Topf (Steve
        # 11.08.2026, gestaffelt): "abends privat, tagsueber im Team" muss
        # der Admin als getrennte Zahlen sehen — Monat, Gesamt und die
        # geschaetzten KI-Kosten je Topf.
        topf_rows = conn.execute(
            "SELECT COALESCE(konto_user_id, user_id) AS konto, "
            "       COALESCE(SUM(CASE WHEN strftime('%Y-%m', created_at) = "
            "                strftime('%Y-%m', 'now') THEN credits ELSE 0 END), 0) AS monat, "
            "       COALESCE(SUM(credits), 0) AS gesamt "
            "FROM usage_events WHERE user_id = ? "
            "GROUP BY konto", (user_id,)).fetchall()
        # Was DIESES Konto uns insgesamt kostet: alles, was seinen Topf
        # belastet hat (bei Inhabern inkl. Mitglieder) — Basis der Kostenzeile.
        eigener_topf_gesamt = conn.execute(
            "SELECT COALESCE(SUM(credits), 0) FROM usage_events "
            "WHERE COALESCE(konto_user_id, user_id) = ?", (user_id,)).fetchone()[0]
    finally:
        conn.close()

    satz = billing.KOSTEN_PRO_CREDIT_EUR
    for zeile in verlauf:
        zeile["kosten_geschaetzt"] = round((zeile["topf"] or 0) * satz, 2)
    plan = billing.effektiver_plan(ziel)
    monatspreis = (billing.preis_pro_monat(plan, ziel.get("plan_laufzeit_monate") or 0)
                   if plan in billing.PLAN_PREISE_EUR else 0.0)

    # Aufstellung "Verbrauch nach Topf": eigener Topf und JEDE Mitgliedschaft
    # erscheinen immer (auch mit 0), fruehere Toepfe nur bei Verbrauch.
    # Je Topf: Monat, Gesamt und geschaetzte KI-Kosten des SELBST-Anteils.
    topf_map = {r["konto"]: {"monat": r["monat"], "gesamt": r["gesamt"]}
                for r in topf_rows}

    def _topf_zeile(basis, daten):
        daten = daten or {"monat": 0, "gesamt": 0}
        basis["monat"] = daten["monat"]
        basis["gesamt"] = daten["gesamt"]
        basis["kosten_geschaetzt_eur"] = round(daten["gesamt"] * satz, 2)
        return basis

    verbrauch_nach_topf = [_topf_zeile({"eigen": True, "plan": plan},
                                       topf_map.pop(user_id, None))]
    for tm in teams:
        verbrauch_nach_topf.append(_topf_zeile({
            "eigen": False, "inhaber": tm["inhaber"],
            "team_name": tm.get("team_name") or "",
        }, topf_map.pop(tm["inhaber_id"], None)))
    for fremd_id, fremd_daten in topf_map.items():
        frueher = get_user_by_id(fremd_id)
        verbrauch_nach_topf.append(_topf_zeile({
            "eigen": False,
            "inhaber": frueher["display_name"] if frueher else "?",
            "team_name": (frueher.get("team_name") or "") if frueher else "",
        }, fremd_daten))
    selbst_gesamt = sum(r["gesamt"] for r in topf_rows)

    return {
        "ok": True,
        "konto": {
            "id": ziel["id"], "name": ziel["display_name"], "email": ziel["email"],
            "angelegt_am": (ziel.get("created_at") or "")[:10],
            "aktiv": bool(ziel.get("is_active", 1)),
            "ist_betreiber": bool(ziel.get("is_admin")),
            # Missbrauchsbremse (11.08.2026): individuell oder Standard.
            "api_tageslimit": ziel.get("api_tageslimit"),
            "api_tageslimit_effektiv": effektives_api_tageslimit(ziel),
        },
        "abo": {
            "plan": plan,
            "laufzeit_monate": ziel.get("plan_laufzeit_monate"),
            "gueltig_bis": (ziel.get("plan_gueltig_bis") or "")[:10] or None,
            "auto_verlaengerung": bool(1 if ziel.get("auto_verlaengerung") is None
                                       else ziel.get("auto_verlaengerung")),
            "quelle": ziel.get("plan_quelle") or None,
            "geplanter_plan": ziel.get("geplanter_plan") or None,
            "monatspreis_eur": monatspreis,
            "erloes_laufzeit_eur": round(monatspreis * (ziel.get("plan_laufzeit_monate") or 0), 2),
            "team_name": ziel.get("team_name") or "",
            "sitze_inklusive": billing.PLAN_SITZE.get(plan),
        },
        "guthaben": {
            "kontingent": zustand.get("kontingent"),
            "uebertrag": zustand.get("uebertrag", 0),
            "verfuegbar_monat": zustand.get("verfuegbar_monat"),
            "verbraucht_monat": zustand.get("verbraucht", 0),
            "rest": zustand.get("rest"),
            "pakete_rest": zustand.get("pakete_rest", 0),
            "pakete": pakete,
            "topf_fremd": topf_fremd,
            "verbrauch_nach_topf": verbrauch_nach_topf,
        },
        "nutzung": {
            "projekte": projekte, "bilder": bilder, "api_aufrufe": api_aufrufe,
            "credits_gesamt": gesamt, "selbst_gesamt": selbst_gesamt,
            "letzte_aktion": letzte_aktion,
            "verlauf": verlauf,
        },
        "kosten": {
            "satz_pro_credit_eur": satz,
            # Seit 11.08.2026: NUR der eigene Topf — sonst zaehlt Team-
            # Verbrauch im Report des Mitglieds UND des Inhabers doppelt.
            "gesamt_geschaetzt_eur": round(eigener_topf_gesamt * satz, 2),
            "hinweis": "Schätzung auf Basis eines Durchschnittssatzes je Credit — "
                       "keine kundengenaue Abrechnung der Modellkosten.",
        },
        "team": {"mitglieder": mitglieder, "mitglied_in": teams},
    }


@app.post("/api/admin/users/{user_id}/toggle-active")
async def admin_toggle_active(user_id: int, user: dict = Depends(require_full_admin)):
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User nicht gefunden")
    if target["id"] == user["id"]:
        raise HTTPException(status_code=400, detail="Sie koennen sich nicht selbst deaktivieren")
    new_status = 0 if target["is_active"] else 1
    update_user_active(user_id, new_status)
    return {"ok": True, "is_active": new_status}


@app.post("/api/admin/users/{user_id}/reset-password")
async def admin_reset_user_password(user_id: int, request: Request, user: dict = Depends(require_full_admin)):
    data = await request.json()
    new_password = data.get("new_password", "")
    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="Passwort muss mindestens 8 Zeichen lang sein")
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User nicht gefunden")
    admin_reset_password(user_id, new_password)
    return {"ok": True, "message": f"Passwort fuer {target['email']} wurde zurueckgesetzt"}


@app.post("/api/admin/users/create")
async def admin_create_user(request: Request, user: dict = Depends(require_full_admin)):
    """Admin: Create a new user account."""
    data = await request.json()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    display_name = _normiere_display_name(data.get("display_name", ""))  # Befund 7: eine Zeile, max. 100

    if not _email_plausibel(email):
        raise HTTPException(status_code=400, detail="Bitte eine gueltige E-Mail-Adresse eingeben")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Passwort muss mindestens 8 Zeichen lang sein")
    if not display_name:
        raise HTTPException(status_code=400, detail="Bitte einen Namen eingeben")

    existing = get_user_by_email(email)
    if existing:
        raise HTTPException(status_code=409, detail="Diese E-Mail-Adresse ist bereits registriert")

    try:
        user_id = create_user(email, password, display_name)
    except Exception:
        raise HTTPException(status_code=500, detail="Benutzer konnte nicht erstellt werden")

    # Send welcome email with credentials
    send_welcome = data.get("send_email", True)
    email_sent = False
    if send_welcome:
        # Steve 24.08.2026: gleicher Karten-Rahmen, Kontrastfarbe und
        # verallgemeinerter Claim wie bei allen anderen Konto-Mails.
        inhalt = (
            f'{_MAIL_P}Hallo {display_name},</p>'
            f'{_MAIL_P}dein Zugang zu InkluDocs wurde erstellt — deinem KI-Werkzeug '
            'für barrierefreie Dokumente und Bilder.</p>'
            '<h2 style="color:#1b2a4a;font-size:17px;margin:18px 0 10px;">Deine Zugangsdaten</h2>'
            f'{_MAIL_P}<strong>Login-Seite:</strong> <a href="{BASE_URL}">{BASE_URL}</a><br>'
            f'<strong>E-Mail:</strong> {email}<br>'
            f'<strong>Passwort:</strong> {password}</p>'
            '<p style="background-color:#fff3eb;padding:14px;border-left:3px solid #c75000;'
            'border-radius:0 4px 4px 0;margin:0 0 14px;line-height:1.55;">'
            'Bitte ändere dein Passwort nach dem ersten Login unter <strong>Einstellungen</strong>.</p>'
            '<h2 style="color:#1b2a4a;font-size:17px;margin:18px 0 10px;">So funktioniert InkluDocs</h2>'
            '<ol style="line-height:1.7;margin:0 0 4px;padding-left:22px;">'
            f'<li>Melde dich auf <a href="{BASE_URL}">{BASE_URL}</a> an.</li>'
            '<li>Lade ein PDF oder Bilder hoch oder gib eine Website-Adresse ein.</li>'
            '<li>Klicke auf „Alt-Texte generieren“.</li>'
            '<li>Bearbeite die Alt-Texte bei Bedarf und exportiere sie.</li></ol>')
        email_body = _auth_mail_rahmen("de", "Willkommen bei InkluDocs", inhalt)
        email_sent = send_email(email, f"Dein InkluDocs-Zugang", email_body)

    msg = f"Benutzer {display_name} ({email}) wurde erstellt"
    if email_sent:
        msg += " – Zugangsdaten per E-Mail gesendet"
    else:
        msg += " – Zugangsdaten bitte manuell weitergeben"

    return {"ok": True, "message": msg, "user_id": user_id, "email_sent": email_sent}


@app.delete("/api/admin/users/{user_id}")
async def admin_delete_user(user_id: int, user: dict = Depends(require_full_admin)):
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User nicht gefunden")
    if target["id"] == user["id"]:
        raise HTTPException(status_code=400, detail="Sie koennen sich nicht selbst loeschen")
    # Delete user files from disk
    user_upload_dir = os.path.join(UPLOAD_DIR, str(user_id))
    user_results_dir = os.path.join(RESULTS_DIR, str(user_id))
    if os.path.exists(user_upload_dir):
        shutil.rmtree(user_upload_dir)
    if os.path.exists(user_results_dir):
        shutil.rmtree(user_results_dir)
    # Delete from DB (DSGVO-konform: alle Daten werden geloescht)
    delete_user_data(user_id)
    return {"ok": True, "message": f"User {target['email']} und alle Daten wurden geloescht"}


# ─── Abo-Stufenverwaltung (Etappe 2, 31.07.2026) ─────────────
# Voll-Admins setzen den Plan von Hand — der Weg fuer Gruender-Tarif und
# Rechnungskunden, solange es keine Online-Zahlung gibt (Etappe 3).

def _setze_plan_kern(user_id: int, target: dict, plan: str, laufzeit,
                     gueltig_bis, auto_verlaengerung: int, quelle) -> tuple:
    """Der eigentliche Plan-Wechsel — EINE Mechanik fuer Admin-/Rechnungsweg
    UND Stripe-Webhook (06.08.2026 nachts). Rueckgabe:
    (gueltig_bis, geloeste_mitgliedschaften, ueberbelegte_sitze)."""
    conn = get_db()
    geloest = 0
    try:
        # plan_treue wird bei JEDEM Plan-Wechsel zurueckgesetzt (zentral,
        # Geldfluss-Review Befund 6, 24.08.2026); der invoice.paid-Webhook
        # setzt es fuer laufende Treue-Abos direkt danach wieder auf 1.
        if plan == "free":
            # Free hat weder Laufzeit noch Verlaengerung noch Quelle.
            conn.execute("UPDATE users SET plan = 'free', plan_gueltig_bis = NULL, "
                         "plan_laufzeit_monate = NULL, auto_verlaengerung = 1, "
                         "plan_quelle = NULL, plan_treue = 0 WHERE id = ?", (user_id,))
            gueltig_bis = None
        elif gueltig_bis is not None:
            conn.execute("UPDATE users SET plan = ?, plan_gueltig_bis = ?, "
                         "plan_laufzeit_monate = ?, auto_verlaengerung = ?, "
                         "plan_quelle = ?, plan_treue = 0 WHERE id = ?",
                         (plan, gueltig_bis, laufzeit, auto_verlaengerung, quelle, user_id))
        else:
            conn.execute("UPDATE users SET plan = ?, "
                         "plan_gueltig_bis = date('now', ?), "
                         "plan_laufzeit_monate = ?, auto_verlaengerung = ?, "
                         "plan_quelle = ?, plan_treue = 0 WHERE id = ?",
                         (plan, f"+{laufzeit} months", laufzeit, auto_verlaengerung, quelle, user_id))
            gueltig_bis = conn.execute("SELECT plan_gueltig_bis FROM users WHERE id = ?",
                                       (user_id,)).fetchone()[0]
        # Review-Befund 5 (31.07.2026): Wechsel WEG von einem Topf-Plan darf
        # kein Geister-Team hinterlassen — Mitgliedschaften loesen; wer gerade
        # in diesem Topf arbeitete, faellt auf den eigenen Kontext zurueck.
        if target.get("plan") in billing.PLAN_SITZE and plan not in billing.PLAN_SITZE:
            cur = conn.execute("DELETE FROM team_mitgliedschaften WHERE inhaber_id = ?",
                               (user_id,))
            geloest = cur.rowcount
            conn.execute("UPDATE users SET aktiver_topf = NULL WHERE aktiver_topf = ?",
                         (user_id,))
        # Review-Befund K6 (06.08.): Downgrade zwischen Sitz-Plaenen behaelt
        # die Mitglieder bewusst — Ueberbelegung wird aber gemeldet.
        ueberbelegt = 0
        neuer_deckel = billing.sitze_fuer_plan(plan)
        if neuer_deckel is not None:
            belegt = int(conn.execute(
                "SELECT COUNT(*) FROM team_mitgliedschaften WHERE inhaber_id = ?",
                (user_id,)).fetchone()[0]) + 1
            ueberbelegt = max(0, belegt - neuer_deckel)
        conn.commit()
    finally:
        conn.close()
    return gueltig_bis, geloest, ueberbelegt


@app.post("/api/admin/users/{user_id}/plan")
async def admin_set_plan(user_id: int, request: Request, user: dict = Depends(require_full_admin)):
    """Admin: Abo-Plan eines Kontos setzen — der Rechnungsweg (Actino/INKLUTEC).

    Neues Abomodell (06.08.2026): Bezahl-Plaene werden fest fuer 3/6/12
    Monate gebucht (laufzeit_monate); gueltig_bis errechnet sich daraus.
    Alternativ kann weiterhin ein exaktes gueltig_bis uebergeben werden.
    auto_verlaengerung (Default an) steuert, ob der Plan am Laufzeitende
    automatisch verlaengert wird (siehe _abo_tageslauf). Der Kunde bekommt
    eine Bestaetigungs-Mail und sieht den Plan sofort in seinem Abo-Bereich.
    """
    data = await request.json()
    plan = (data.get("plan") or "").strip().lower()
    if plan not in billing.PLAN_KONTINGENTE:
        raise HTTPException(status_code=400,
                            detail=f"Ungueltiger Plan. Erlaubt: {', '.join(billing.PLAN_KONTINGENTE)}")
    laufzeit = data.get("laufzeit_monate")
    if laufzeit is not None:
        try:
            laufzeit = int(laufzeit)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="laufzeit_monate muss eine Zahl sein")
        # Monatsabo (Laufzeit 1) gibt es NUR online ueber Stripe — der
        # Rechnungsweg bleibt bei festen Laufzeiten (Steve/Michael 19.08.2026).
        if laufzeit not in billing.PLAN_LAUFZEITEN_RECHNUNG:
            raise HTTPException(status_code=400,
                                detail=f"Laufzeit muss {', '.join(str(m) for m in billing.PLAN_LAUFZEITEN_RECHNUNG)} Monate sein (Monatsabo nur online)")
    gueltig_bis = data.get("gueltig_bis") or None
    if gueltig_bis is not None:
        try:
            datetime.fromisoformat(str(gueltig_bis))
            gueltig_bis = str(gueltig_bis)[:10]
        except ValueError:
            raise HTTPException(status_code=400,
                                detail="gueltig_bis muss ein ISO-Datum sein (JJJJ-MM-TT) oder null")
    _av_raw = data.get("auto_verlaengerung")
    auto_verlaengerung = 1 if _av_raw is None else (1 if _av_raw in (1, True, "1", "true") else 0)
    if plan != "free" and laufzeit is None and gueltig_bis is None:
        raise HTTPException(status_code=400,
                            detail="Bezahl-Plaene brauchen eine Laufzeit (3, 6 oder 12 Monate)")
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User nicht gefunden")
    gueltig_bis, geloest, ueberbelegt = _setze_plan_kern(
        user_id, target, plan, laufzeit, gueltig_bis, auto_verlaengerung,
        quelle="rechnung")
    # Review-Befund 9: Rechnungskunden duerfen nicht schlechter stehen als
    # Stripe-Kunden — steigt das Kontingent, startet es auch hier frisch.
    if (billing.PLAN_KONTINGENTE.get(plan, 0)
            > billing.PLAN_KONTINGENTE.get(target.get("plan"), 0)):
        billing.starte_kontingent_neu(user_id)
    # Bestaetigungs-Mail an den Kunden — gleicher Text wie spaeter beim
    # Online-Kauf: er sieht sofort, was gebucht wurde und bis wann es laeuft.
    if plan != "free":
        try:
            _sende_plan_bestaetigung(target, plan, gueltig_bis, laufzeit, auto_verlaengerung)
        except Exception:
            logger.exception("Plan-Bestaetigungs-Mail an %s fehlgeschlagen", target.get("email"))
    msg = f"Plan von {target['email']} ist jetzt '{plan}'"
    if gueltig_bis:
        msg += f" (gueltig bis {str(gueltig_bis)[:10]})"
    if geloest:
        msg += f" – {geloest} Team-Mitgliedschaften wurden gelöst"
    if ueberbelegt:
        msg += (f" – ACHTUNG: {ueberbelegt} Mitglieder über dem neuen Sitz-Deckel, "
                "bitte über die Abo-Seite des Kunden entfernen")
    return {"ok": True, "message": msg, "plan": plan,
            "plan_gueltig_bis": gueltig_bis, "team_mitglieder_geloest": geloest}


# ─── Administrator-Verwaltung (14.06.2026) ──────────────────
# Voll-Admins koennen bestehende Konten zu Admins machen und die
# Rechte-Stufe setzen ('full' = alle Rechte, 'view' = nur Einsicht).
# Sicherungen: niemand kann sich selbst die Admin-Rechte entziehen,
# und der letzte Voll-Admin kann weder entfernt noch herabgestuft
# werden (Schutz gegen Aussperren).

@app.get("/api/admin/admins")
async def admin_list_admins(user: dict = Depends(require_admin)):
    """Liste aller Administratoren inkl. Rechte-Stufe (auch fuer Nur-Einsicht lesbar)."""
    return {"admins": list_admins()}


@app.post("/api/admin/admins")
async def admin_add_admin(request: Request, user: dict = Depends(require_full_admin)):
    """Ein bestehendes Konto zum Admin machen (per E-Mail) und Stufe setzen."""
    data = await request.json()
    email = data.get("email", "").strip().lower()
    level = data.get("level", "full")
    if level not in ("full", "view"):
        raise HTTPException(status_code=400, detail="Ungueltige Rechte-Stufe")
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Bitte eine gueltige E-Mail-Adresse eingeben")
    target = get_user_by_email(email)
    if not target:
        raise HTTPException(status_code=404, detail="Kein Konto mit dieser E-Mail. Bitte die Person zuerst unter 'Neuen Benutzer anlegen' anlegen und danach zum Admin machen.")
    if target["is_admin"]:
        raise HTTPException(status_code=409, detail=f"{target['display_name']} ist bereits Administrator")
    set_user_admin(target["id"], 1, level)
    stufe = "Voll-Admin" if level == "full" else "Nur-Einsicht"
    return {"ok": True, "message": f"{target['display_name']} ist jetzt Administrator ({stufe})"}


@app.put("/api/admin/admins/{user_id}")
async def admin_set_level(user_id: int, request: Request, user: dict = Depends(require_full_admin)):
    """Rechte-Stufe eines Admins aendern (full <-> view)."""
    data = await request.json()
    level = data.get("level", "")
    if level not in ("full", "view"):
        raise HTTPException(status_code=400, detail="Ungueltige Rechte-Stufe")
    target = get_user_by_id(user_id)
    if not target or not target["is_admin"]:
        raise HTTPException(status_code=404, detail="Administrator nicht gefunden")
    if level == "view" and target["admin_level"] == "full" and count_full_admins() <= 1:
        raise HTTPException(status_code=400, detail="Der letzte Voll-Admin kann nicht auf 'Nur-Einsicht' gesetzt werden")
    set_user_admin(user_id, 1, level)
    stufe = "Voll-Admin" if level == "full" else "Nur-Einsicht"
    return {"ok": True, "message": f"{target['display_name']} ist jetzt {stufe}", "level": level}


@app.delete("/api/admin/admins/{user_id}")
async def admin_revoke_admin(user_id: int, user: dict = Depends(require_full_admin)):
    """Admin-Rechte entziehen (das Konto bleibt als normaler Nutzer bestehen)."""
    target = get_user_by_id(user_id)
    if not target or not target["is_admin"]:
        raise HTTPException(status_code=404, detail="Administrator nicht gefunden")
    if target["id"] == user["id"]:
        raise HTTPException(status_code=400, detail="Sie koennen sich die Admin-Rechte nicht selbst entziehen")
    if target["admin_level"] == "full" and count_full_admins() <= 1:
        raise HTTPException(status_code=400, detail="Der letzte Voll-Admin kann nicht entfernt werden")
    set_user_admin(user_id, 0)
    return {"ok": True, "message": f"Admin-Rechte von {target['display_name']} entzogen"}


# ─── Team-Verwaltung (neues Abomodell, 06.08.2026) ──────────────
# Team-/Enterprise-Topf: Mitglieder (team_mitgliedschaften) koennen aus dem
# Kontingent des zahlenden Inhabers verbrauchen, wenn ihr aktiver_topf darauf
# steht. Verwalten (einladen, entfernen, benennen) darf NUR der Inhaber —
# Mitglieder koennen sich ausdruecklich NICHT selbst entfernen (Steve 06.08.).

def _require_team_inhaber(user: dict) -> dict:
    """Liefert den frischen DB-Datensatz des Topf-Inhabers oder wirft 403.

    Frisch aus der DB (nicht aus dem Token), damit ein Plan-Wechsel sofort
    greift. Inhaber ist, wessen EIGENER effektiver Plan Sitze kennt
    (team/enterprise); ein abgelaufener Plan verliert damit auch die
    Verwaltungs-Rechte (Lazy-Rueckfall). Dass der Inhaber selbst irgendwo
    Mitglied ist, ist im neuen Modell erlaubt und hier egal.
    """
    db_user = get_user_by_id(user["id"])
    if not db_user or billing.effektiver_plan(db_user) not in billing.PLAN_SITZE:
        raise HTTPException(status_code=403,
                            detail="Nur fuer Inhaber eines Team- oder Enterprise-Abos verfuegbar")
    return db_user


def _team_sitze_belegt(inhaber_id: int) -> int:
    """Belegte Sitze = Mitgliedschaften + 1 (der Inhaber selbst).

    Eine Stelle fuer die Zaehlweise, weil der Sitz-Deckel ZWEIMAL geprueft
    wird: beim Einladen (freundliche Sofort-Ablehnung) und beim Einloesen
    der Einladung (verbindlich, siehe Review-Befund 1).
    """
    conn = get_db()
    try:
        row = conn.execute("SELECT COUNT(*) FROM team_mitgliedschaften WHERE inhaber_id = ?",
                           (inhaber_id,)).fetchone()
    finally:
        conn.close()
    return int(row[0]) + 1


@app.get("/api/team")
async def team_uebersicht(user: dict = Depends(get_current_user)):
    """Topf-Uebersicht fuer den Inhaber eines Team-/Enterprise-Abos.

    WICHTIG: Die Topf-Zahlen werden fuer das INHABER-KONTO gerechnet, nicht
    ueber pruefe_kontingent(inhaber) — der Inhaber koennte ja gerade selbst
    im Kontext eines fremden Teams arbeiten.
    """
    inhaber = _require_team_inhaber(user)
    plan = billing.effektiver_plan(inhaber)
    kontingent = billing.PLAN_KONTINGENTE[plan]
    uebertrag = billing._uebertrag(inhaber["id"], plan, kontingent)
    verbraucht = billing.monats_verbrauch(inhaber["id"])
    pakete = billing.pakete_rest(inhaber["id"])
    verfuegbar = kontingent + uebertrag
    conn = get_db()
    try:
        # Inhaber + Mitglieder in einer Abfrage; verbraucht_monat je Person =
        # was sie DIESEM Topf entnommen hat (konto_user_id = Inhaber) — so
        # sieht der Inhaber, wer wieviel zieht; Eigenverbrauch der Mitglieder
        # in anderen Kontexten geht ihn nichts an.
        rows = conn.execute(
            "SELECT u.id, u.display_name, u.email, "
            "COALESCE((SELECT SUM(e.credits) FROM usage_events e "
            "  WHERE e.user_id = u.id AND e.konto_user_id = ? "
            "  AND e.created_at >= (SELECT MAX(date('now','start of month'), "
            "      COALESCE(iu.kontingent_reset_am,'0000-01-01')) FROM users iu WHERE iu.id = ?)"
            "  ), 0) AS verbraucht_monat "
            "FROM users u WHERE u.id = ? OR u.id IN "
            "(SELECT mitglied_id FROM team_mitgliedschaften WHERE inhaber_id = ?) "
            "ORDER BY (u.id != ?), u.display_name COLLATE NOCASE",
            # Fuenf Platzhalter: Topf, Reset-Konto, Inhaber, Mitgliedschaft, Sortierung.
            (inhaber["id"],) * 5,
        ).fetchall()
    finally:
        conn.close()
    mitglieder = []
    for r in rows:
        m = dict(r)
        m["ist_inhaber"] = (r["id"] == inhaber["id"])
        mitglieder.append(m)
    return {
        "ok": True,
        "plan": plan,
        "team_name": inhaber.get("team_name") or "",
        "mitglieder": mitglieder,
        "kontingent": kontingent,
        "uebertrag": uebertrag,
        "verfuegbar_monat": verfuegbar,
        "verbraucht_gesamt": verbraucht,
        "rest": max(0, verfuegbar - verbraucht),
        "pakete_rest": pakete,
        "zeitraum_ende": billing._monatsende_iso(),
        "sitze_belegt": len(mitglieder),
        "sitze_inklusive": billing.sitze_fuer_plan(plan),
    }


@app.post("/api/team/name")
async def team_name_setzen(request: Request, user: dict = Depends(get_current_user)):
    """Inhaber benennt sein Team (Anzeige beim Mitglied und im Umschalter).

    Auch der Weg, ein Team ERSTMALS anzulegen: Ohne Namen gibt es kein Team,
    erst danach kann eingeladen werden (Steve 07.08.).
    """
    inhaber = _require_team_inhaber(user)
    data = await request.json()
    name = _normiere_display_name(data.get("team_name") or "")
    if not name:
        raise HTTPException(status_code=400, detail="Bitte einen Teamnamen eingeben")
    conn = get_db()
    try:
        conn.execute("UPDATE users SET team_name = ? WHERE id = ?", (name, inhaber["id"]))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "team_name": name}


@app.delete("/api/team")
async def team_aufloesen(user: dict = Depends(get_current_user)):
    """Loest das Team auf — SOFORT und unabhaengig vom Abo (Steve 07.08.).

    Wichtig: Das Abo bleibt unveraendert bestehen (der Inhaber behaelt Plan
    und Credits, nutzt sie dann allein) — „Team" und „Abo" sind zwei Dinge.
    Alle Mitgliedschaften fallen weg, wer im Team-Kontext arbeitete, faellt
    auf sein eigenes Guthaben zurueck, alle bekommen eine Mail. Der Teamname
    wird geleert, damit beim naechsten Mal wieder sauber „Team erstellen"
    steht.
    """
    inhaber = _require_team_inhaber(user)
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT u.id, u.email, u.display_name FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.mitglied_id WHERE m.inhaber_id = ?",
            (inhaber["id"],)).fetchall()
        mitglieder = [dict(r) for r in rows]
        conn.execute("DELETE FROM team_mitgliedschaften WHERE inhaber_id = ?", (inhaber["id"],))
        conn.execute("UPDATE users SET aktiver_topf = NULL WHERE aktiver_topf = ?", (inhaber["id"],))
        conn.execute("DELETE FROM team_einladungen WHERE inhaber_id = ? AND eingeloest_am IS NULL",
                     (inhaber["id"],))
        conn.execute("UPDATE users SET team_name = '' WHERE id = ?", (inhaber["id"],))
        conn.commit()
    finally:
        conn.close()
    team_label = html.escape(inhaber.get("team_name") or "") or f"das Team von {html.escape(inhaber['display_name'])}"
    for m in mitglieder:
        try:
            _abo_mail(m["email"], "InkluDocs: Euer gemeinsames Guthaben wurde aufgelöst",
                      "Team aufgelöst",
                      [f"Hallo {html.escape(m['display_name'] or '')},",
                       f"{html.escape(inhaber['display_name'])} hat {team_label} aufgelöst. "
                       "Ab sofort arbeitest du wieder mit deinem eigenen Guthaben "
                       "(oder dem kostenlosen Free-Plan).",
                       "Dein Konto und alle deine Projekte bleiben davon unberührt — es ändert "
                       "sich nur, aus welchem Guthaben neue Alt-Texte bezahlt werden."])
        except Exception:
            logger.exception("Aufloesungs-Mail an %s fehlgeschlagen", m["email"])
    return {"ok": True, "entfernt": len(mitglieder),
            "message": f"Team aufgelöst — {len(mitglieder)} Mitglieder wurden informiert. "
                       "Dein Abo läuft unverändert weiter."}


@app.post("/api/team/einladen")
async def team_einladen(request: Request, user: dict = Depends(get_current_user)):
    """Team-Inhaber laedt per E-Mail ein — IMMER nur als Einladung, nie mehr
    als stilles Umhaengen (Review-Befund 1, 31.07.2026):

    - Bestehendes Konto: Einladungs-Token (7 Tage gueltig) + Mail an den
      EINGELADENEN; Mitglied wird er erst, wenn er den Annahme-Link
      /team-einladung/{token} eingeloggt bestaetigt. Vorher konnte ein
      Inhaber fremde Bestandskonten per UPDATE kapern.
    - Unbekannte Adresse: Konto entsteht ERST durch die Einladung, darum darf
      die Mitgliedschaft dort direkt angelegt werden (niemandem wird etwas
      weggenommen); Mail mit Passwort-Reset-Link und Team-Hinweis.

    Antwort bewusst NEUTRAL (immer "Einladung versandt" — kein Name, kein
    neu_angelegt-Flag): sonst koennte ein Inhaber ueber die Antwort auslesen,
    welche E-Mail-Adressen ein InkluDocs-Konto haben (E-Mail-Enumeration).
    """
    inhaber = _require_team_inhaber(user)
    plan = billing.effektiver_plan(inhaber)
    sitze_max = billing.sitze_fuer_plan(plan)
    data = await request.json()
    email = (data.get("email") or "").strip().lower()
    if not _email_plausibel(email):
        raise HTTPException(status_code=400, detail="Bitte eine gueltige E-Mail-Adresse eingeben")
    if email == (inhaber.get("email") or "").lower():
        raise HTTPException(status_code=400, detail="Du kannst dich nicht selbst einladen")
    if email.rsplit("@", 1)[-1] in billing.WEGWERF_DOMAINS:
        raise HTTPException(status_code=400,
                            detail="Wegwerf-E-Mail-Adressen koennen nicht eingeladen werden")

    # Einladungs-Bremse pro Inhaber (Sicherheits-Review 06.08.): begrenzt
    # Mail-Versand an frei waehlbare Adressen. Nur Erfolge zaehlen.
    _jetzt = time.time()
    _einladung_attempts[inhaber["id"]] = [t for t in _einladung_attempts[inhaber["id"]]
                                          if _jetzt - t < EINLADUNG_WINDOW_SECONDS]
    if len(_einladung_attempts[inhaber["id"]]) >= MAX_EINLADUNGEN_PER_WINDOW:
        raise HTTPException(status_code=429,
                            detail="Zu viele Einladungen in kurzer Zeit. Bitte spaeter erneut versuchen")

    # Sitz-Deckel je Plan (Team 5 / Enterprise 25, inkl. Inhaber): belegte
    # Sitze PLUS offene Einladungen — sonst laedt man mehr ein, als beim
    # Einloesen Platz haben. Verbindlich geprueft wird beim EINLOESEN.
    conn = get_db()
    try:
        offene = int(conn.execute(
            "SELECT COUNT(*) FROM team_einladungen WHERE inhaber_id = ? "
            "AND eingeloest_am IS NULL AND gueltig_bis > datetime('now')",
            (inhaber["id"],)).fetchone()[0])
    finally:
        conn.close()
    if _team_sitze_belegt(inhaber["id"]) + offene >= sitze_max:
        raise HTTPException(
            status_code=400,
            detail=f"Alle {sitze_max} Plätze deines Abos sind belegt oder durch offene "
                   "Einladungen reserviert. Mehr Nutzer gibt es in der nächsthöheren Abo-Stufe.")

    neutrale_antwort = {"ok": True, "message": "Einladung versandt"}

    # Dedup (Sicherheits-Review 06.08.): pro (Inhaber, Adresse) hoechstens
    # EINE offene Einladung — verhindert Mail-Wiederholung und Token-Wildwuchs.
    # Antwort bleibt neutral.
    conn = get_db()
    try:
        schon_offen = conn.execute(
            "SELECT 1 FROM team_einladungen WHERE inhaber_id = ? AND email = ? "
            "AND eingeloest_am IS NULL AND gueltig_bis > datetime('now')",
            (inhaber["id"], email)).fetchone()
    finally:
        conn.close()
    if schon_offen:
        return neutrale_antwort
    # Befund 7: Anzeigename ist zwar an den Setz-Stellen normiert, im
    # Mail-HTML wird trotzdem escaped (Verteidigung in der Tiefe).
    inhaber_name = html.escape(inhaber["display_name"])
    team_label = html.escape(inhaber.get("team_name") or "") or f"das Team von {inhaber_name}"

    ziel = get_user_by_email(email)
    if ziel:
        # Neues Modell: ein eigener Bezahl-Plan oder andere Mitgliedschaften
        # sind KEIN Hindernis mehr — beides darf parallel bestehen, der
        # Kontext-Umschalter regelt den Verbrauch. Nur wer bereits Mitglied
        # DIESES Teams ist, bekommt keine sinnlose Mail (Antwort bleibt
        # neutral — keine E-Mail-Enumeration).
        conn = get_db()
        try:
            schon = conn.execute(
                "SELECT 1 FROM team_mitgliedschaften WHERE inhaber_id = ? AND mitglied_id = ?",
                (inhaber["id"], ziel["id"])).fetchone()
        finally:
            conn.close()
        if schon:
            return neutrale_antwort
        einladungs_token = secrets.token_urlsafe(32)
        conn = get_db()
        try:
            conn.execute(
                "INSERT INTO team_einladungen (token, inhaber_id, email, gueltig_bis) "
                "VALUES (?, ?, ?, datetime('now', '+7 days'))",
                (einladungs_token, inhaber["id"], email))
            conn.commit()
        finally:
            conn.close()
        annahme_url = f"{BASE_URL}/team-einladung/{einladungs_token}"
        # Steve 24.08.2026: neuer Karten-Rahmen + Kontrastfarbe wie alle Mails.
        inhalt = (
            f'{_MAIL_P}Hallo,</p>'
            f'{_MAIL_P}<strong>{inhaber_name}</strong> möchte dich bei InkluDocs in {team_label} '
            'aufnehmen. Ihr arbeitet dann aus einem gemeinsamen Credit-Kontingent; dein Konto, '
            'deine Projekte und ein eventuell vorhandenes eigenes Abo bleiben unberührt — du '
            'kannst jederzeit wählen, aus welchem Guthaben du arbeitest.</p>'
            f'{_MAIL_P}Wenn du einverstanden bist, bestätige hier:</p>'
            + _mail_knopf(annahme_url, "Team-Einladung annehmen")
            + f'{_MAIL_KLEIN}Oder kopiere diesen Link: {annahme_url}</p>'
            + f'{_MAIL_KLEIN}Der Link ist 7 Tage gültig und funktioniert nur, wenn du mit deinem '
            f'Konto ({html.escape(email)}) angemeldet bist. Wenn du die Einladung nicht annehmen '
            'möchtest, ignoriere diese Mail einfach — an deinem Konto ändert sich dann nichts.</p>')
        email_body = _auth_mail_rahmen("de", "Einladung in ein InkluDocs-Team", inhalt)
        # Betreff bewusst OHNE Nutzereingabe (Header-Injektionsflaeche, Befund 7).
        send_email(email, "Einladung in ein InkluDocs-Team", email_body, bcc_admin=False)
        _einladung_attempts[inhaber["id"]].append(_jetzt)
        return neutrale_antwort

    # Unbekannte Adresse: Konto anlegen mit Zufallspasswort, das niemand je
    # sieht. Die MITGLIEDSCHAFT entsteht bewusst NOCH NICHT (Sicherheits-
    # Review 06.08.: sonst verriete die Team-Liste sofort, ob eine Adresse
    # schon ein Konto hatte — E-Mail-Enumeration). Stattdessen wird eine
    # Einladung mit konto_neu=1 hinterlegt; sobald die Person ihr Passwort
    # setzt (= ihre Registrierung, Reset-Link aus der Mail), loest
    # do_reset_password die Einladung automatisch ein — Sitz-Check dort
    # atomar. Der Reset-Link ist 1h gueltig, danach hilft 'Passwort
    # vergessen' (die Einladung selbst gilt 7 Tage).
    zufallspasswort = secrets.token_urlsafe(16)
    display_name = _normiere_display_name(data.get("display_name") or "") or email.split("@")[0]
    try:
        neu_id = create_user(email, zufallspasswort, display_name)
    except Exception:
        # Bewusst KEIN Detail nach aussen (neutraler 500): auch der seltene
        # Anlage-Fehler darf nicht verraten, ob es die Adresse schon gab.
        raise HTTPException(status_code=500, detail="Einladung konnte nicht verarbeitet werden")
    einladungs_token = secrets.token_urlsafe(32)
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO team_einladungen (token, inhaber_id, email, gueltig_bis, konto_neu) "
            "VALUES (?, ?, ?, datetime('now', '+7 days'), 1)",
            (einladungs_token, inhaber["id"], email))
        conn.commit()
    finally:
        conn.close()

    reset_token = create_password_reset_token(neu_id)
    reset_url = f"{BASE_URL}/reset?token={reset_token}"
    sicherer_name = html.escape(display_name)
    # Steve 24.08.2026: neuer Rahmen + verallgemeinerter Claim (nicht nur Alt-Texte).
    inhalt = (
        f'{_MAIL_P}Hallo {sicherer_name},</p>'
        f'{_MAIL_P}<strong>{inhaber_name}</strong> möchte dich bei InkluDocs in {team_label} '
        'aufnehmen. InkluDocs ist dein KI-Werkzeug für barrierefreie Dokumente und Bilder — '
        'als Team-Mitglied arbeitet ihr aus einem gemeinsamen Credit-Kontingent.</p>'
        f'{_MAIL_P}Du hast noch kein Konto bei uns. Zum Loslegen registrierst du dich einfach, '
        'indem du einmal dein persönliches Passwort setzt — damit trittst du zugleich dem '
        'Team bei:</p>'
        + _mail_knopf(reset_url, "Passwort festlegen und starten")
        + f'{_MAIL_KLEIN}Oder kopiere diesen Link: {reset_url}</p>'
        + f'{_MAIL_KLEIN}Der Link ist 1 Stunde gültig. Danach kannst du auf '
        f'<a href="{BASE_URL}">{BASE_URL}</a> jederzeit die Funktion „Passwort vergessen“ '
        f'nutzen — dein Konto ist die Adresse {html.escape(email)}.</p>')
    email_body = _auth_mail_rahmen("de", "Einladung in ein InkluDocs-Team", inhalt)
    send_email(email, "Einladung in ein InkluDocs-Team", email_body, bcc_admin=False)
    _einladung_attempts[inhaber["id"]].append(_jetzt)
    return neutrale_antwort


@app.get("/team-einladung/{token}", response_class=HTMLResponse)
async def team_einladung_annehmen(token: str, request: Request):
    """Annahme-Seite fuer Team-Einladungen (Review-Befund 1, 31.07.2026).

    Die ZUSTIMMUNG des Eingeladenen: erst dieser eingeloggte Klick haengt
    sein Konto in den Team-Topf um. Aufbau wie die anderen
    Bestaetigungsseiten (_auth_notice_page): h1, Fliesstext, Rueckweg-Link —
    bewusst kein eigenes Template.
    """
    lang = resolve_ui_language(request)
    _ = get_gettext(lang)
    zum_dashboard = f'<p style="margin-top:1.5rem;"><a href="/dashboard">{_("Zur Startseite")}</a></p>'

    # Login-Pflicht mit derselben Redirect-Logik wie die geschuetzten Seiten
    # (_render_protected_template): ohne gueltiges Cookie zur Anmeldung —
    # nach dem Login fuehrt der Link aus der Mail erneut hierher.
    user = get_optional_user(request)
    if not user:
        return RedirectResponse("/")

    conn = get_db()
    try:
        einladung = conn.execute(
            "SELECT * FROM team_einladungen WHERE token = ? AND eingeloest_am IS NULL "
            "AND gueltig_bis > datetime('now')",
            (token,),
        ).fetchone()
    finally:
        conn.close()
    if not einladung:
        return _auth_notice_page(lang,
            f'<p style="color:#dc2626;margin:2rem 0;">{_("Dieser Einladungslink ist ungültig oder abgelaufen.")}</p>' + zum_dashboard,
            status_code=400)

    db_user = get_user_by_id(user["id"])
    if not db_user or (db_user.get("email") or "").lower() != einladung["email"].lower():
        # Falsches Konto eingeloggt: Einladung bleibt offen und unangetastet.
        return _auth_notice_page(lang,
            f'<p style="color:#dc2626;margin:2rem 0;">{_("Diese Einladung gehört zu einer anderen E-Mail-Adresse. Bitte melde dich mit dem eingeladenen Konto an.")}</p>' + zum_dashboard,
            status_code=403)

    # Zwischen Einladen und Annehmen kann sich alles geaendert haben — die
    # fachlichen Regeln hier VERBINDLICH erneut pruefen: der Inhaber muss
    # noch ein aktives Team-/Enterprise-Abo haben. Ein eigener Plan oder
    # andere Mitgliedschaften des Eingeladenen sind im neuen Modell KEIN
    # Hindernis — beides existiert parallel, der Kontext-Umschalter regelt
    # den Verbrauch.
    inhaber = get_user_by_id(einladung["inhaber_id"])
    inhaber_plan = billing.effektiver_plan(inhaber) if inhaber else "free"
    if not inhaber or inhaber_plan not in billing.PLAN_SITZE:
        return _auth_notice_page(lang,
            f'<p style="color:#dc2626;margin:2rem 0;">{_("Diese Einladung kann nicht mehr angenommen werden.")}</p>' + zum_dashboard,
            status_code=409)

    conn = get_db()
    try:
        # Sitz-Check und Beitritt atomar (BEGIN IMMEDIATE): zwei gleichzeitig
        # klickende Eingeladene koennen den Sitz-Deckel sonst ueberrennen.
        conn.isolation_level = None
        conn.execute("BEGIN IMMEDIATE")
        schon_mitglied = conn.execute(
            "SELECT 1 FROM team_mitgliedschaften WHERE inhaber_id = ? AND mitglied_id = ?",
            (einladung["inhaber_id"], db_user["id"])).fetchone() is not None
        belegt = int(conn.execute(
            "SELECT COUNT(*) FROM team_mitgliedschaften WHERE inhaber_id = ?",
            (einladung["inhaber_id"],)).fetchone()[0]) + 1
        if not schon_mitglied and belegt >= billing.sitze_fuer_plan(inhaber_plan):
            conn.execute("ROLLBACK")
            return _auth_notice_page(lang,
                f'<p style="color:#dc2626;margin:2rem 0;">{_("Alle Plätze in diesem Team sind bereits belegt.")}</p>' + zum_dashboard,
                status_code=409)
        if not schon_mitglied:
            conn.execute("INSERT OR IGNORE INTO team_mitgliedschaften (inhaber_id, mitglied_id) "
                         "VALUES (?, ?)", (einladung["inhaber_id"], db_user["id"]))
            # Der Beitritt schaltet den Arbeits-Kontext aufs neue Team — das
            # Erwartbare nach dem Klick auf 'annehmen'; zurueckwechseln geht
            # jederzeit auf der Abo-Seite.
            conn.execute("UPDATE users SET aktiver_topf = ? WHERE id = ?",
                         (einladung["inhaber_id"], db_user["id"]))
        # ALLE offenen Einladungen dieses Paars einloesen (Sicherheits-Review
        # 06.08.): sonst bliebe ein zweiter Token als Hintertuer offen, mit
        # dem ein spaeter entferntes Mitglied ohne neue Einladung zurueckkommt.
        conn.execute("UPDATE team_einladungen SET eingeloest_am = datetime('now') "
                     "WHERE inhaber_id = ? AND email = ? AND eingeloest_am IS NULL",
                     (einladung["inhaber_id"], einladung["email"]))
        conn.execute("COMMIT")
    finally:
        conn.close()

    inhaber_name = html.escape(inhaber["display_name"])
    team_label = html.escape(inhaber.get("team_name") or "")
    team_satz = (f'{_("Du bist jetzt Mitglied im Team")} „{team_label}“ ({inhaber_name}).'
                 if team_label else
                 f'{_("Du bist jetzt Mitglied im Team von")} {inhaber_name}.')
    return _auth_notice_page(lang,
        f'''<p style="color:#16a34a;font-size:1.2rem;margin:2rem 0;font-weight:600;">&#10003; {_("Team-Beitritt bestätigt!")}</p>
    <p>{team_satz} {_("Ihr arbeitet aus einem gemeinsamen Credit-Kontingent. Auf der Seite „Abo & Verbrauch“ wählst du jederzeit, aus welchem Guthaben du arbeitest.")}</p>''' + zum_dashboard,
        title_suffix=_("Team-Beitritt bestätigt"))


@app.delete("/api/team/mitglied/{mitglied_id}")
async def team_mitglied_entfernen(mitglied_id: int, user: dict = Depends(get_current_user)):
    """Mitglied aus dem eigenen Topf loesen — NUR der Inhaber darf das
    (Mitglieder koennen sich nicht selbst entfernen, Steve 06.08.2026).
    Das Konto bleibt bestehen und arbeitet weiter mit eigenem Abo/Free;
    bisherige Verbraeuche bleiben als Protokoll auf dem Topf-Konto. Der
    Entfernte bekommt eine kurze Info-Mail."""
    inhaber = _require_team_inhaber(user)
    ziel = get_user_by_id(mitglied_id)
    conn = get_db()
    try:
        mitglied = conn.execute(
            "SELECT 1 FROM team_mitgliedschaften WHERE inhaber_id = ? AND mitglied_id = ?",
            (inhaber["id"], mitglied_id)).fetchone()
        if not ziel or not mitglied:
            raise HTTPException(status_code=404, detail="Kein Mitglied deines Teams")
        conn.execute("DELETE FROM team_mitgliedschaften WHERE inhaber_id = ? AND mitglied_id = ?",
                     (inhaber["id"], mitglied_id))
        # Arbeitete die Person gerade in diesem Topf, faellt sie auf den
        # eigenen Kontext zurueck — sonst wuerde _konto_fuer zwar still
        # zurueckfallen, aber die Oberflaeche zeigte einen toten Kontext.
        conn.execute("UPDATE users SET aktiver_topf = NULL WHERE id = ? AND aktiver_topf = ?",
                     (mitglied_id, inhaber["id"]))
        # Offene Einladungen dieses Inhabers an die Person mitloeschen —
        # ein alter Token darf nach dem Entfernen keine Rueckkehr-Tuer sein.
        conn.execute("DELETE FROM team_einladungen WHERE inhaber_id = ? AND email = ? "
                     "AND eingeloest_am IS NULL", (inhaber["id"], ziel["email"]))
        conn.commit()
    finally:
        conn.close()
    try:
        team_label = html.escape(inhaber.get("team_name") or "") or f"dem Team von {html.escape(inhaber['display_name'])}"
        inhalt = (
            f"{_MAIL_P}Hallo {html.escape(ziel['display_name'])},</p>"
            f'{_MAIL_P}du wurdest aus {team_label} entfernt. Ab jetzt arbeitest du wieder mit '
            'deinem eigenen Tarif und dem Guthaben, das dazu gehört — hast du kein eigenes Abo, '
            'ist das der kostenlose Free-Tarif. Dein Konto und alle deine Projekte bleiben '
            'unverändert bestehen.</p>')
        info_body = _auth_mail_rahmen("de", "Änderung an deinem InkluDocs-Konto", inhalt)
        send_email(ziel["email"], "InkluDocs: Team-Mitgliedschaft beendet", info_body, bcc_admin=False)
    except Exception:
        logger.exception("Info-Mail an entferntes Mitglied %s fehlgeschlagen", mitglied_id)
    return {"ok": True, "message": f"{ziel['display_name']} wurde aus dem Team entfernt"}


# ─── Kontext-Umschalter + Kuendigung (neues Abomodell, 06.08.2026) ───

@app.post("/api/abo/kontext")
async def abo_kontext_setzen(request: Request, user: dict = Depends(get_current_user)):
    """Waehlt, aus welchem Credit-Topf das Konto arbeitet.

    topf: 'eigen' (eigenes Abo/Free) oder die Nutzer-ID eines Team-Inhabers,
    bei dem eine Mitgliedschaft besteht und dessen Plan aktiv ist. Gilt fuer
    JEDEN Verbrauch (Generierung, Export, Chatbot, API) — billing._konto_fuer
    liest genau diesen Wert und validiert ihn bei jeder Buchung erneut.
    """
    data = await request.json()
    topf = data.get("topf")
    if topf in ("eigen", None, "", 0):
        conn = get_db()
        try:
            conn.execute("UPDATE users SET aktiver_topf = NULL WHERE id = ?", (user["id"],))
            conn.commit()
        finally:
            conn.close()
        return {"ok": True, "aktiver_topf": "eigen"}
    try:
        topf = int(topf)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Ungueltiger Topf")
    conn = get_db()
    try:
        inhaber = conn.execute(
            "SELECT u.plan, u.plan_gueltig_bis FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.inhaber_id "
            "WHERE m.inhaber_id = ? AND m.mitglied_id = ?",
            (topf, user["id"])).fetchone()
        if inhaber is None or billing.effektiver_plan(inhaber) not in billing.PLAN_SITZE:
            raise HTTPException(status_code=404,
                                detail="Dieses Team steht dir nicht (mehr) zur Verfügung")
        conn.execute("UPDATE users SET aktiver_topf = ? WHERE id = ?", (topf, user["id"]))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "aktiver_topf": topf}


def _stripe_kuendigung_sync(db_user: dict, kuendigen: bool) -> None:
    """Kuendigungs-Status mit Stripe abgleichen (nur Stripe-verwaltete Abos).

    Wichtig: Bei Stripe-Abos MUSS der Abgleich gelingen — sonst stuende
    lokal 'gekuendigt', waehrend Stripe echtes Geld weiter abbucht.
    """
    sub = (db_user.get("stripe_subscription_id") or "").strip()
    if not sub or db_user.get("plan_quelle") != "stripe" or not stripe_zahlung.AKTIV:
        return
    try:
        # Befund 07.08.: Ein vorgemerkter Wechsel (Subscription-Schedule)
        # blockiert bei Stripe jede Kuendigungs-Aenderung. Beim Kuendigen
        # wird der Wechsel ohnehin hinfaellig — also erst freigeben.
        if kuendigen:
            stripe_zahlung.loese_schedule(sub)
        stripe_zahlung.kuendige_subscription(sub, kuendigen)
        # Kuendigungs-WIDERRUF (24.08.2026): Die Kuendigung hatte auch den
        # Treue-Anschluss (AGB Ziffer 7) geloest — bei fester Laufzeit
        # wieder aufsetzen, sonst verlaengert sich der reaktivierte Vertrag
        # still um die volle Laufzeit statt als Treue-Monatsabo.
        if not kuendigen and int(db_user.get("plan_laufzeit_monate") or 0) > 1:
            try:
                stripe_zahlung.plane_treue_anschluss(sub, db_user.get("plan"))
            except Exception:
                logger.exception("Treue-Anschluss nach Kuendigungs-Widerruf "
                                 "nicht angelegt (user=%s)", db_user["id"])
    except Exception:
        logger.exception("Stripe-Kuendigungs-Abgleich fehlgeschlagen (user=%s)", db_user["id"])
        raise HTTPException(status_code=502,
                            detail="Die Kündigung konnte bei unserem Zahlungsdienst gerade nicht "
                                   "hinterlegt werden. Bitte in ein paar Minuten erneut versuchen")


@app.post("/api/abo/kuendigen")
async def abo_kuendigen(user: dict = Depends(get_current_user)):
    """Eigenes Abo zum Laufzeitende kuendigen (auto_verlaengerung aus).

    Betrifft IMMER nur den eigenen Plan — ein Team, in dem man nur Mitglied
    ist, kann man nicht kuendigen (das macht dessen Inhaber). Der Plan
    laeuft bis zum Laufzeitende normal weiter und faellt dann auf Free.
    Stripe-Abos werden zusaetzlich bei Stripe zum Periodenende gekuendigt.
    """
    db_user = get_user_by_id(user["id"])
    if not db_user or billing.effektiver_plan(db_user) == "free":
        raise HTTPException(status_code=400, detail="Kein aktives Abo zum Kündigen")
    _stripe_kuendigung_sync(db_user, True)
    conn = get_db()
    try:
        # Eine Kuendigung hebt einen vorgemerkten Plan-Wechsel auf — sonst
        # wuerde am Stichtag ein neuer (bezahlpflichtiger) Plan starten,
        # obwohl der Kunde gerade gekuendigt hat.
        conn.execute("UPDATE users SET auto_verlaengerung = 0, geplanter_plan = NULL, "
                     "geplante_laufzeit = NULL, geplant_ab = NULL WHERE id = ?", (user["id"],))
        conn.commit()
    finally:
        conn.close()
    bis = (db_user.get("plan_gueltig_bis") or "")[:10] or None
    # Team-Mitglieder rechtzeitig informieren (Steve 07.08.): Wer ein
    # Team-/Enterprise-Abo kuendigt, beendet damit auch den gemeinsamen Topf.
    if bis and billing.effektiver_plan(db_user) in billing.PLAN_SITZE:
        try:
            _team_enddatum_mail(db_user, bis, "kuendigung")
        except Exception:
            logger.exception("Team-Enddatum-Mails nach Kuendigung fehlgeschlagen")
    return {"ok": True, "gueltig_bis": bis,
            "message": "Gekündigt — dein Abo läuft bis zum Laufzeitende weiter und endet dann"}


# Bremsen fuer den oeffentlichen Kuendigungsknopf. Zwei Ebenen, bewusst
# unterschiedlich streng (Selbst-Review 08.08.): Der Knopf ist gesetzlich
# vorgeschrieben und muss verfuegbar sein — eine zu enge Bremse waere selbst
# ein Rechtsverstoss. Hinter einer Firmen-IP koennen mehrere Personen
# nacheinander kuendigen wollen, darum ist das IP-Limit weit. Eng ist nur
# das Limit PRO ADRESSE: dieselbe Adresse mehrfach zu kuendigen bringt
# niemandem etwas, waere aber eine Mail-Kanone gegen diese Adresse.
_kuendigung_ip = defaultdict(list)
_kuendigung_mail = defaultdict(list)
MAX_KUENDIGUNGEN_IP = 20
MAX_KUENDIGUNGEN_MAIL = 3
KUENDIGUNG_WINDOW_SECONDS = 3600
# Selbst-Review 08.08.: Der Adress-Zaehler waechst mit JEDER angefragten
# Adresse — von aussen frei waehlbar. Ohne Aufraeumen koennte man den
# Prozess-Speicher damit langsam vollschreiben. Darum werden abgelaufene
# Eintraege regelmaessig entfernt, statt sie ewig mitzuschleppen.
MAX_BREMS_SCHLUESSEL = 5000


def _bremse_aufraeumen(speicher: dict, fenster: int) -> None:
    """Entfernt abgelaufene Zaehler aus einem Bremsen-Speicher."""
    jetzt = time.time()
    for schluessel in [k for k, v in speicher.items()
                       if not v or jetzt - v[-1] > fenster]:
        speicher.pop(schluessel, None)


@app.post("/api/kuendigung")
async def oeffentliche_kuendigung(request: Request):
    """Gesetzlicher Kuendigungsknopf nach § 312k BGB (08.08.2026).

    Muss OHNE Anmeldung funktionieren — genau das ist der Sinn der
    Vorschrift. Rechtlich zaehlt der Zeitpunkt des ZUGANGS, darum wird jede
    Erklaerung protokolliert, auch wenn sich kein Konto dazu finden laesst.

    Missbrauchsabwaegung: Wer eine fremde Adresse eintraegt, kann damit ein
    fremdes Abo zum Laufzeitende kuendigen. Der Schaden ist gering und
    umkehrbar — das Abo laeuft mit vollem Guthaben weiter, die betroffene
    Person wird sofort per Mail informiert und kann die Kuendigung im
    Abo-Bereich mit einem Klick widerrufen. Die Alternative (Bestaetigungs-
    link vor Wirksamkeit) wuerde die vom Gesetz geforderte unmittelbare
    Kuendigungsmoeglichkeit aushebeln. Deshalb: sofort wirksam, sofort
    informieren, jederzeit widerrufbar.

    Aus demselben Grund verraet die Antwort NICHT, ob es zu der Adresse ein
    Konto gibt (keine Konto-Ausforschung) — die Rueckmeldung ist immer
    gleich, die Einzelheiten stehen in der Mail an die Adresse selbst.
    """
    schluessel = absender.limit_schluessel(request)
    jetzt = time.time()
    _kuendigung_ip[schluessel] = [t for t in _kuendigung_ip[schluessel]
                                  if jetzt - t < KUENDIGUNG_WINDOW_SECONDS]
    if len(_kuendigung_ip[schluessel]) >= MAX_KUENDIGUNGEN_IP:
        raise HTTPException(status_code=429,
                            detail="Zu viele Anfragen von dieser Verbindung. Bitte später erneut "
                                   "versuchen oder direkt an kontakt@inklutec.de schreiben — "
                                   "eine Kündigung per E-Mail ist genauso wirksam.")
    data = await request.json()
    mail = (data.get("email") or "").strip()
    if not _email_plausibel(mail):
        raise HTTPException(status_code=400, detail="Bitte eine gültige E-Mail-Adresse angeben")
    mail_key = mail.lower()
    if len(_kuendigung_mail) > MAX_BREMS_SCHLUESSEL:
        _bremse_aufraeumen(_kuendigung_mail, KUENDIGUNG_WINDOW_SECONDS)
    _kuendigung_mail[mail_key] = [t for t in _kuendigung_mail[mail_key]
                                  if jetzt - t < KUENDIGUNG_WINDOW_SECONDS]
    if len(_kuendigung_mail[mail_key]) >= MAX_KUENDIGUNGEN_MAIL:
        raise HTTPException(status_code=429,
                            detail="Für diese Adresse liegt uns bereits eine Kündigung vor. "
                                   "Die Bestätigung ist unterwegs; bei Fragen schreib uns bitte "
                                   "an kontakt@inklutec.de.")
    art = (data.get("art") or "ordentlich").strip().lower()
    if art not in ("ordentlich", "ausserordentlich", "widerruf"):
        art = "ordentlich"
    name = (data.get("name") or "").strip()[:200]
    zusatz = (data.get("zusatz") or "").strip()[:2000]
    _kuendigung_ip[schluessel].append(jetzt)
    _kuendigung_mail[mail_key].append(jetzt)

    db_user = get_user_by_email(mail)
    hat_abo = bool(db_user and billing.effektiver_plan(db_user) != "free")
    wirksam_zum = (db_user.get("plan_gueltig_bis") or "")[:10] if hat_abo else None

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO kuendigungen (email, name, art, zusatz, user_id, vollzogen, "
            "wirksam_zum, absender) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (mail, name, art, zusatz, db_user["id"] if db_user else None,
             1 if hat_abo else 0, wirksam_zum, schluessel))
        conn.commit()
    finally:
        conn.close()

    if hat_abo:
        # Denselben Weg gehen wie der Knopf im Abo-Bereich — eine Mechanik.
        try:
            _stripe_kuendigung_sync(db_user, True)
        except Exception:
            logger.exception("Stripe-Kuendigung ueber den oeffentlichen Weg fehlgeschlagen (%s)", mail)
            # Geldfluss-Review Befund 2 (24.08.2026): Hier NUR zu loggen hiesse:
            # der Kunde bekommt "Kuendigung bestaetigt", Stripe bucht aber
            # weiter ab. Der Zugang der Erklaerung zaehlt rechtlich trotzdem —
            # also Vorgang als NICHT vollzogen markieren und den Betreiber
            # alarmieren, damit er die Stripe-Seite von Hand nachzieht.
            conn = get_db()
            try:
                conn.execute("UPDATE kuendigungen SET vollzogen = 0 WHERE id = "
                             "(SELECT MAX(id) FROM kuendigungen WHERE email = ?)", (mail,))
                conn.commit()
            finally:
                conn.close()
            try:
                _abo_mail(NOTIFICATION_EMAIL,
                          f"InkluDocs ALARM: Stripe-Kündigung fehlgeschlagen ({mail})",
                          "Stripe-Kündigung von Hand nachziehen",
                          [f"Die Erklärung ({html.escape(art)}) von {html.escape(mail)} ist "
                           "protokolliert und dem Kunden bestätigt, aber der Stripe-Abgleich "
                           "ist fehlgeschlagen — die Subscription läuft dort noch.",
                           "Bitte im Stripe-Dashboard zum Periodenende kündigen, sonst wird "
                           "weiter abgebucht. Details im Server-Log."])
            except Exception:
                logger.exception("Alarm-Mail zur fehlgeschlagenen Stripe-Kuendigung scheiterte")
        conn = get_db()
        try:
            conn.execute("UPDATE users SET auto_verlaengerung = 0, geplanter_plan = NULL, "
                         "geplante_laufzeit = NULL, geplant_ab = NULL WHERE id = ?",
                         (db_user["id"],))
            conn.commit()
        finally:
            conn.close()
        if wirksam_zum and billing.effektiver_plan(db_user) in billing.PLAN_SITZE:
            try:
                _team_enddatum_mail(db_user, wirksam_zum, "kuendigung")
            except Exception:
                logger.exception("Team-Enddatum-Mails nach oeffentlicher Kuendigung fehlgeschlagen")

    # Bestaetigung in Textform — Pflicht nach § 312k Abs. 4 BGB. Sie geht an
    # die angegebene Adresse; nur dort stehen die Einzelheiten.
    art_text = {"ordentlich": "ordentliche Kündigung zum Laufzeitende",
                "ausserordentlich": "außerordentliche Kündigung aus wichtigem Grund",
                "widerruf": "Widerruf innerhalb von 14 Tagen"}[art]
    saetze = [f"Hallo {html.escape(name) or 'zusammen'},",
              f"deine Erklärung ist bei uns eingegangen: {art_text}, "
              f"eingegangen am {_datum_lang(datetime.now().strftime('%Y-%m-%d'))} "
              f"um {datetime.now().strftime('%H:%M')} Uhr."]
    if hat_abo:
        saetze.append(f"Dein Abo endet damit am {_datum_lang(wirksam_zum)}. Bis dahin kannst du es mit "
                      "vollem Guthaben weiternutzen; danach läuft dein Konto im "
                      "kostenlosen Free-Tarif weiter.")
        saetze.append("Falls du das nicht wolltest: Im Bereich „Abo &amp; Verbrauch“ kannst du "
                      "die Kündigung bis zum Laufzeitende mit einem Klick widerrufen.")
    else:
        saetze.append("Zu dieser Adresse ist bei uns kein laufendes kostenpflichtiges Abo "
                      "hinterlegt. Wir haben deine Erklärung trotzdem festgehalten und sehen "
                      "sie uns an — falls etwas offen ist, melden wir uns.")
    if art == "widerruf":
        saetze.append("Zum Widerruf melden wir uns gesondert wegen der Rückabwicklung. "
                      "Einzelheiten stehen in unserer Widerrufsbelehrung.")
    try:
        _abo_mail(mail,
                  ("InkluDocs: Dein Widerruf ist eingegangen" if art == "widerruf"
                   else "InkluDocs: Deine Kündigung ist eingegangen"),
                  ("Widerruf bestätigt" if art == "widerruf" else "Kündigung bestätigt"),
                  saetze)
        _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs: Kündigung über die Kündigungsseite ({mail})",
                  "Kündigung eingegangen",
                  [f"Adresse: {html.escape(mail)}",
                   f"Name: {html.escape(name) or '(ohne)'}",
                   f"Art: {art_text}",
                   f"Konto gefunden: {'ja' if db_user else 'nein'}; "
                   f"laufendes Abo: {'ja, beendet zum ' + str(wirksam_zum) if hat_abo else 'nein'}",
                   f"Anmerkung: {html.escape(zusatz) or '(keine)'}"])
    except Exception:
        logger.exception("Kuendigungs-Bestaetigung konnte nicht versandt werden (%s)", mail)

    if art == "widerruf":
        return {"ok": True, "meldungen": [
            "Dein Widerruf ist eingegangen. Wir haben ihn mit Datum und Uhrzeit festgehalten.",
            "Eine Bestätigung in Textform geht gerade an die angegebene E-Mail-Adresse.",
            "Falls du innerhalb weniger Minuten keine E-Mail bekommst, schreib uns bitte an "
            "kontakt@inklutec.de.",
        ]}
    return {"ok": True, "meldungen": [
        "Deine Kündigung ist eingegangen. Wir haben sie mit Datum und Uhrzeit festgehalten.",
        "Eine Bestätigung in Textform geht gerade an die angegebene E-Mail-Adresse — dort "
        "steht auch, wann dein Vertrag genau endet.",
        "Falls du innerhalb weniger Minuten keine E-Mail bekommst, schreib uns bitte an "
        "kontakt@inklutec.de.",
    ]}


# Bremsen fuer das Kontaktformular (gleiche Bauart wie beim Kuendigungsknopf).
_kontakt_ip = defaultdict(list)
_kontakt_mail = defaultdict(list)
MAX_KONTAKT_IP = 10
MAX_KONTAKT_MAIL = 3
KONTAKT_WINDOW_SECONDS = 3600
SUPPORT_EMAIL = os.environ.get("SUPPORT_EMAIL", "support@inkludocs.de")


@app.post("/api/kontakt")
async def oeffentlicher_kontakt(request: Request):
    """Kontaktformular (24.08.2026, § 5 DDG zweiter Weg) — ohne Anmeldung.

    Die Nachricht geht ans Support-Postfach mit Antwort-an = Absender.
    BEWUSST keine automatische Mail an die eingetragene Adresse: sonst
    liesse sich das Formular als Mail-Kanone gegen fremde Adressen nutzen.
    Der Honigtopf (Feld 'firma') faengt Formular-Bots: gefuellt = still
    verworfen, damit der Bot keinen Unterschied sieht.
    """
    schluessel = absender.limit_schluessel(request)
    jetzt = time.time()
    _kontakt_ip[schluessel] = [t for t in _kontakt_ip[schluessel]
                               if jetzt - t < KONTAKT_WINDOW_SECONDS]
    if len(_kontakt_ip[schluessel]) >= MAX_KONTAKT_IP:
        raise HTTPException(status_code=429,
                            detail="Zu viele Anfragen von dieser Verbindung. Bitte später erneut "
                                   "versuchen oder direkt an support@inkludocs.de schreiben.")
    data = await request.json()
    if (data.get("firma") or "").strip():
        return {"ok": True}
    mail = (data.get("email") or "").strip()
    nachricht = (data.get("nachricht") or "").strip()[:5000]
    name = (data.get("name") or "").strip()[:200]
    # Betreff (freiwillig, 25.08.2026) — landet in der Betreffzeile der
    # Support-Mail, damit das Postfach sortierbar bleibt.
    betreff = (data.get("betreff") or "").strip()[:150]
    # Angemeldete Absender: Konto-Adresse NUR intern in die Support-Mail
    # (Zuordnung im Postfach). Das Formular wird bewusst nicht vorausgefuellt
    # (Steve 25.08.2026: die Antwort soll an die eingetragene Adresse gehen).
    # Steht so in der Datenschutzerklaerung, Abschnitt „Kontaktformular“.
    angemeldet = get_optional_user(request)
    if not _email_plausibel(mail):
        raise HTTPException(status_code=400, detail="Bitte eine gültige E-Mail-Adresse angeben")
    if not nachricht:
        raise HTTPException(status_code=400, detail="Bitte eine Nachricht eingeben")
    mail_key = mail.lower()
    if len(_kontakt_mail) > MAX_BREMS_SCHLUESSEL:
        _bremse_aufraeumen(_kontakt_mail, KONTAKT_WINDOW_SECONDS)
    _kontakt_mail[mail_key] = [t for t in _kontakt_mail[mail_key]
                               if jetzt - t < KONTAKT_WINDOW_SECONDS]
    if len(_kontakt_mail[mail_key]) >= MAX_KONTAKT_MAIL:
        raise HTTPException(status_code=429,
                            detail="Für diese Adresse liegen uns bereits Nachrichten vor. Bitte "
                                   "warte auf unsere Antwort oder schreib direkt an "
                                   "support@inkludocs.de.")
    _kontakt_ip[schluessel].append(jetzt)
    _kontakt_mail[mail_key].append(jetzt)
    konto_zeile = ""
    if angemeldet and angemeldet.get("email"):
        konto_zeile = (f'<p>Angemeldet als: {html.escape(angemeldet["email"])} '
                       f'(Konto-ID {int(angemeldet["id"])})</p>')
    body = (f'<!DOCTYPE html><html lang="de"><head><meta charset="utf-8"></head><body>'
            f'<h1>Kontaktformular</h1>'
            f'<p>Von: {html.escape(name) or "(ohne Namen)"} &lt;{html.escape(mail)}&gt;</p>'
            f'{konto_zeile}'
            f'<p>Betreff: {html.escape(betreff) or "(ohne Betreff)"}</p>'
            f'<p style="white-space:pre-wrap;">{html.escape(nachricht)}</p></body></html>')
    try:
        send_email(SUPPORT_EMAIL, f"InkluDocs Kontaktformular: {betreff or name or mail}",
                   body, bcc_admin=False, reply_to=mail)
    except Exception:
        logger.exception("Kontaktformular-Mail fehlgeschlagen (%s)", mail)
        raise HTTPException(status_code=502,
                            detail="Die Nachricht konnte gerade nicht zugestellt werden. Bitte "
                                   "später erneut versuchen oder direkt an support@inkludocs.de "
                                   "schreiben.")
    return {"ok": True}


@app.post("/api/abo/kuendigung-widerrufen")
async def abo_kuendigung_widerrufen(user: dict = Depends(get_current_user)):
    """Kuendigung zuruecknehmen, solange die Laufzeit noch laeuft."""
    db_user = get_user_by_id(user["id"])
    if not db_user or billing.effektiver_plan(db_user) == "free":
        raise HTTPException(status_code=400, detail="Kein aktives Abo")
    _stripe_kuendigung_sync(db_user, False)
    conn = get_db()
    try:
        conn.execute("UPDATE users SET auto_verlaengerung = 1 WHERE id = ?", (user["id"],))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "message": "Die Kündigung wurde zurückgenommen — dein Abo verlängert sich wieder automatisch"}


# ─── Stripe-Buchungsweg (06.08.2026 nachts, zuerst Testmodus) ───
# Checkout und Portal sind duenn: die Freischaltung passiert IMMER im
# Webhook (auch wenn der Kunde den Erfolgs-Redirect nie sieht).

def _team_enddatum_mail(inhaber: dict, bis: str, grund: str) -> None:
    """Informiert alle Team-Mitglieder, dass der gemeinsame Topf endet.

    grund: 'downgrade' (Inhaber wechselt auf einen Plan ohne Sitze) oder
    'kuendigung' (Abo laeuft aus). Bis zum Stichtag aendert sich fuer die
    Mitglieder NICHTS — sie sollen nur rechtzeitig Bescheid wissen.
    """
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT u.email, u.display_name FROM team_mitgliedschaften m "
            "JOIN users u ON u.id = m.mitglied_id WHERE m.inhaber_id = ?",
            (inhaber["id"],)).fetchall()
    finally:
        conn.close()
    team_label = html.escape(inhaber.get("team_name") or "") or f"das Team von {html.escape(inhaber['display_name'])}"
    for r in rows:
        try:
            _abo_mail(r["email"], "InkluDocs: Euer gemeinsames Guthaben endet bald",
                      "Änderung an eurem Team",
                      [f"Hallo {html.escape(r['display_name'] or '')},",
                       f"{html.escape(inhaber['display_name'])} hat das Abo für {team_label} geändert: "
                       f"Das gemeinsame Guthaben endet am {_datum_lang(bis)}.",
                       "Bis dahin ändert sich für dich nichts. Danach arbeitest du automatisch "
                       "wieder mit deinem eigenen Guthaben (oder dem kostenlosen Free-Plan) "
                       "weiter — dein Konto und alle deine Projekte bleiben unverändert."])
        except Exception:
            logger.exception("Team-Enddatum-Mail an %s fehlgeschlagen", r["email"])


@app.post("/api/abo/wechseln")
async def abo_wechseln(request: Request, user: dict = Depends(get_current_user)):
    """Plan wechseln aus einem laufenden Abo heraus (07.08.2026).

    UPGRADE (teurer): sofort wirksam, Stripe rechnet den Restwert des alten
    Plans an und fakturiert nur die Differenz.
    DOWNGRADE (guenstiger): wird vorgemerkt und wirkt zum Ende der bezahlten
    Laufzeit; bis dahin laeuft alles unveraendert weiter.
    Massstab ist der Monatspreis — nicht die Laufzeit.
    """
    data = await request.json()
    plan = (data.get("plan") or "").strip().lower()
    if plan not in billing.PLAN_PREISE_EUR:
        raise HTTPException(status_code=400, detail="Ungueltiger Plan")
    try:
        monate = int(data.get("laufzeit_monate"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="laufzeit_monate fehlt")
    if monate not in billing.PLAN_LAUFZEITEN:
        raise HTTPException(status_code=400, detail="Laufzeit muss 1 (Monatsabo), 3, 6 oder 12 Monate sein")

    db_user = get_user_by_id(user["id"])
    aktueller = billing.effektiver_plan(db_user)
    if aktueller == "free":
        raise HTTPException(status_code=400,
                            detail="Du hast noch kein Abo — bitte über „Abo buchen“ starten")
    aktuelle_laufzeit = db_user.get("plan_laufzeit_monate") or 0
    if aktueller == plan and aktuelle_laufzeit == monate:
        raise HTTPException(status_code=400, detail="Diesen Plan hast du bereits")
    # Geldfluss-Review Befund 5 (24.08.2026): "Gleicher Plan als Monatsabo"
    # wuerde die zugesagte Treuekondition (Effektivpreis der Laufzeit) still
    # durch das TEURERE regulaere Monatsabo ersetzen — dieselbe Leistung,
    # mehr Geld. Der Weiterlauf als Monatsabo passiert am Laufzeitende
    # ohnehin automatisch und guenstiger.
    if plan == aktueller and monate == 1 and int(aktuelle_laufzeit) > 1:
        raise HTTPException(
            status_code=400,
            detail="Das brauchst du nicht zu buchen: Nach dem Laufzeitende läuft dein "
                   "Abo automatisch als Monatsabo zur Treuekondition weiter — zum "
                   "Monatspreis deiner jetzigen Laufzeit, günstiger als das reguläre "
                   "Monatsabo. Kündigen kannst du das Anschluss-Abo jederzeit monatlich.")

    sub_id = (db_user.get("stripe_subscription_id") or "").strip()
    stripe_abo = bool(sub_id) and db_user.get("plan_quelle") == "stripe" and stripe_zahlung.AKTIV
    # Review-Befund 1 (KRITISCH): Ohne zahlenden Stripe-Vertrag darf hier
    # NIEMAND einen hoeheren Plan bekommen — sonst waere „Enterprise gratis"
    # nur ein API-Aufruf entfernt. Rechnungskunden laufen ueber Actino/
    # INKLUTEC, dort stellt der Admin um (mit Rechnung).
    if not stripe_abo:
        raise HTTPException(
            status_code=400,
            detail="Dein Abo läuft über den Rechnungsweg. Schreib uns kurz an "
                   "support@inkludocs.de — wir stellen den Plan für dich um.")
    # Review-Befund 8: Ein gekuendigtes Abo darf nicht durch einen Wechsel
    # stillschweigend reaktiviert werden.
    if not int(db_user.get("auto_verlaengerung") or 0):
        raise HTTPException(
            status_code=400,
            detail="Dein Abo ist gekündigt. Nimm zuerst die Kündigung zurück, "
                   "dann kannst du den Plan wechseln.")

    # Review-Befund 7: Der Preisvergleich allein reicht nicht — bei GLEICHEM
    # Plan entscheidet die Laufzeit. Laengere Bindung gilt sofort (mehr
    # Leistung fuer den Kunden, mehr Umsatz), kuerzere zum Laufzeitende.
    if plan == aktueller:
        ist_upgrade = monate > aktuelle_laufzeit
    else:
        ist_upgrade = billing.PLAN_PREISE_EUR[plan] > billing.PLAN_PREISE_EUR[aktueller]

    if ist_upgrade:
        # Review-Befund 1 (07.08. KRITISCH): Beim Upgrade startet eine NEUE
        # Laufzeit ab heute — sie darf das bezahlte Ende aber nie NACH VORN
        # ziehen. Beispiel: team/12 bis naechsten August, Wechsel auf
        # enterprise/3 haette 9 bezahlte Monate gekostet (und bei Stripe ein
        # Guthaben erzeugt, das bei Kuendigung verfaellt).
        alt_ende = (db_user.get("plan_gueltig_bis") or "")[:10]
        if alt_ende:
            conn = get_db()
            try:
                neu_ende = conn.execute("SELECT date('now', ?)",
                                        (f"+{monate} months",)).fetchone()[0]
            finally:
                conn.close()
            if neu_ende < alt_ende:
                raise HTTPException(
                    status_code=400,
                    detail=f"Dein jetziger Plan ist noch bis {alt_ende} bezahlt. Bitte wähle "
                           "eine Laufzeit, die mindestens bis dahin reicht — sonst würde "
                           "bezahlte Zeit verloren gehen.")
        # SOFORT: erst bei Stripe umstellen (dort haengt das Geld), dann lokal.
        # Review-Befund 5: Idempotency-Key gegen Doppelklick/Doppelbuchung.
        try:
            stripe_zahlung.wechsle_sofort(
                sub_id, plan, monate,
                # Review-Befund 5: KEINE Zeitkomponente — mit
                # billing_cycle_anchor='now' wuerde ein Wiederholungs-Aufruf
                # sonst eine zweite volle Periode berechnen. Der Schluessel
                # enthaelt das ALTE Laufzeitende: derselbe Wechsel ist
                # idempotent, ein spaeterer echter Wechsel (dann ist das Ende
                # ein anderes) bleibt moeglich.
                idempotency_key=f"wechsel-{sub_id}-{plan}-{monate}-"
                                f"{(db_user.get('plan_gueltig_bis') or 'x')[:10]}")
        except stripe_zahlung.ZahlungOffen as z:
            # Review-Befund 3: Stripe hat umgestellt, aber die Differenz ist
            # NICHT bezahlt -> lokal nichts freischalten, Kunde zur Rechnung
            # schicken. Der invoice.paid-Webhook holt den Plan nach, sobald
            # gezahlt wurde.
            logger.warning("Upgrade offen (unbezahlt) user=%s", user["id"])
            raise HTTPException(
                status_code=402,
                detail="Die Zahlung für den höheren Plan ist noch offen. Bitte schließe sie "
                       "ab — danach wird der Plan automatisch freigeschaltet."
                       + (f" Rechnung: {z.invoice_url}" if z.invoice_url else ""))
        except HTTPException:
            raise
        except Exception:
            logger.exception("Stripe-Upgrade fehlgeschlagen (user=%s)", user["id"])
            raise HTTPException(status_code=502,
                                detail="Der Wechsel konnte bei unserem Zahlungsdienst gerade "
                                       "nicht durchgeführt werden. Bitte später erneut versuchen")
        # Steve 07.08.2026: Beim Upgrade startet eine NEUE Laufzeit ab heute
        # (nicht das alte Ende weiterschieben). Bezahlte Restzeit verfaellt
        # dabei NICHT — Stripe rechnet sie als Gutschrift an, siehe
        # stripe_zahlung.wechsle_sofort.
        gueltig_bis, geloest, ueberbelegt = _setze_plan_kern(
            db_user["id"], db_user, plan, monate, None, 1, quelle="stripe")
        # ... und der Kunde bekommt SOFORT das volle neue Kontingent: der
        # Verbrauch des laufenden Zeitraums zaehlt ab hier neu. Die Funktion
        # bucht vorher offenen Ueberhang ab und gewaehrt den Neustart
        # hoechstens EINMAL pro Kalendermonat (Review-Befunde 2 und 8).
        billing.starte_kontingent_neu(db_user["id"])
        # Ein evtl. vorgemerkter Downgrade ist mit dem Upgrade hinfaellig.
        conn = get_db()
        try:
            conn.execute("UPDATE users SET geplanter_plan = NULL, geplante_laufzeit = NULL, "
                         "geplant_ab = NULL WHERE id = ?", (db_user["id"],))
            conn.commit()
        finally:
            conn.close()
        try:
            _sende_plan_bestaetigung(db_user, plan, gueltig_bis, monate, 1, quelle="stripe")
            _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs Plan-Upgrade: {db_user['email']}",
                      "Kunde hat hochgestuft",
                      [f"{html.escape(db_user['email'])} hat von "
                       f"{PLAN_ANZEIGENAMEN.get(aktueller, aktueller)} auf "
                       f"{PLAN_ANZEIGENAMEN.get(plan, plan)} hochgestuft "
                       f"({monate} Monate, gültig bis {str(gueltig_bis)[:10]}). "
                       "Die Differenz hat Stripe anteilig abgerechnet."])
        except Exception:
            logger.exception("Bestaetigungs-Mail nach Upgrade fehlgeschlagen")
        return {"ok": True, "sofort": True, "plan": plan, "gueltig_bis": str(gueltig_bis)[:10],
                "message": "Dein Plan wurde sofort umgestellt — die höheren Credits stehen ab jetzt bereit.",
                "team_geloest": geloest, "ueberbelegt": ueberbelegt}

    # DOWNGRADE: nur vormerken, wirksam zum Laufzeitende.
    bis = (db_user.get("plan_gueltig_bis") or "")[:10]
    if not bis:
        raise HTTPException(status_code=400,
                            detail="Für dieses Konto ist kein Laufzeitende hinterlegt — bitte an "
                                   "support@inkludocs.de wenden")
    if stripe_abo:
        try:
            stripe_zahlung.plane_wechsel_zum_periodenende(sub_id, plan, monate)
        except Exception:
            logger.exception("Stripe-Downgrade-Planung fehlgeschlagen (user=%s)", user["id"])
            raise HTTPException(status_code=502,
                                detail="Der Wechsel konnte bei unserem Zahlungsdienst gerade nicht "
                                       "vorgemerkt werden. Bitte später erneut versuchen")
    conn = get_db()
    try:
        # Review-Befund 2: Der Stichtag wird FEST gespeichert. plan_gueltig_bis
        # kann sich durch Stripe-Webhooks verschieben — der vorgemerkte
        # Wechsel muss trotzdem am ursprünglichen Datum greifen.
        conn.execute("UPDATE users SET geplanter_plan = ?, geplante_laufzeit = ?, "
                     "geplant_ab = ? WHERE id = ?", (plan, monate, bis, db_user["id"]))
        conn.commit()
    finally:
        conn.close()
    # Verliert das Konto zum Stichtag die Team-Faehigkeit, muessen die
    # Mitglieder das rechtzeitig erfahren (Steve 07.08.).
    if aktueller in billing.PLAN_SITZE and plan not in billing.PLAN_SITZE:
        try:
            _team_enddatum_mail(db_user, bis, "downgrade")
        except Exception:
            logger.exception("Team-Enddatum-Mails nach Downgrade fehlgeschlagen")
    try:
        _abo_mail(db_user["email"], "InkluDocs: Dein Plan-Wechsel ist vorgemerkt",
                  "Plan-Wechsel vorgemerkt",
                  [f"Hallo {html.escape(db_user['display_name'])},",
                   f"dein Wechsel auf den Plan {PLAN_ANZEIGENAMEN.get(plan, plan)} ist vorgemerkt. "
                   f"Bis zum {_datum_lang(bis)} läuft dein jetziger Plan unverändert weiter — du behältst also "
                   "alles, wofür du bezahlt hast.",
                   f"Am {_datum_lang(bis)} stellen wir dann automatisch um; ab da gilt der neue Plan mit "
                   f"{billing.PLAN_KONTINGENTE[plan]} Credits im Monat.",
                   "Du kannst den Wechsel bis dahin jederzeit auf der Seite „Abo & Verbrauch“ "
                   "wieder zurücknehmen."])
    except Exception:
        logger.exception("Downgrade-Bestaetigung fehlgeschlagen")
    return {"ok": True, "sofort": False, "plan": plan, "ab": bis,
            "message": f"Wechsel vorgemerkt — dein jetziger Plan läuft bis {bis}, danach gilt der neue."}


@app.post("/api/abo/wechsel-widerrufen")
async def abo_wechsel_widerrufen(user: dict = Depends(get_current_user)):
    """Nimmt einen vorgemerkten Downgrade wieder zurueck."""
    db_user = get_user_by_id(user["id"])
    if not db_user or not db_user.get("geplanter_plan"):
        raise HTTPException(status_code=400, detail="Es ist kein Wechsel vorgemerkt")
    # Review-Befund 11: Ist der Stichtag erreicht, kann der Tageslauf gerade
    # umstellen — dann waere „zurueckgenommen" gelogen.
    _ab = (db_user.get("geplant_ab") or "")[:10]
    if _ab and _ab < datetime.utcnow().strftime("%Y-%m-%d"):
        raise HTTPException(status_code=409,
                            detail="Der Wechsel wird gerade vollzogen und kann nicht mehr "
                                   "zurückgenommen werden. Du kannst den Plan danach neu wechseln.")
    sub_id = (db_user.get("stripe_subscription_id") or "").strip()
    if sub_id and db_user.get("plan_quelle") == "stripe" and stripe_zahlung.AKTIV:
        try:
            # Plan + Laufzeit mitgeben, damit der Treue-Anschluss des
            # LAUFENDEN Plans direkt wieder aufgesetzt wird (24.08.2026).
            stripe_zahlung.widerrufe_geplanten_wechsel(
                sub_id, db_user.get("plan"), db_user.get("plan_laufzeit_monate"))
        except Exception:
            logger.exception("Stripe-Wechsel-Widerruf fehlgeschlagen (user=%s)", user["id"])
            raise HTTPException(status_code=502,
                                detail="Der Widerruf konnte bei unserem Zahlungsdienst gerade nicht "
                                       "hinterlegt werden. Bitte später erneut versuchen")
    conn = get_db()
    try:
        conn.execute("UPDATE users SET geplanter_plan = NULL, geplante_laufzeit = NULL, "
                     "geplant_ab = NULL WHERE id = ?", (db_user["id"],))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "message": "Der vorgemerkte Wechsel wurde zurückgenommen — dein Plan bleibt bestehen."}


@app.post("/api/abo/checkout")
async def abo_checkout(request: Request, user: dict = Depends(get_current_user)):
    """Stripe-Checkout-Link fuer eine Abo-Buchung (Plan + Laufzeit)."""
    if not stripe_zahlung.AKTIV:
        raise HTTPException(status_code=503, detail="Online-Zahlung ist noch nicht freigeschaltet")
    data = await request.json()
    plan = (data.get("plan") or "").strip().lower()
    if plan not in billing.PLAN_PREISE_EUR:
        raise HTTPException(status_code=400, detail="Ungueltiger Plan")
    try:
        monate = int(data.get("laufzeit_monate"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="laufzeit_monate fehlt")
    if monate not in billing.PLAN_LAUFZEITEN:
        raise HTTPException(status_code=400, detail="Laufzeit muss 1 (Monatsabo), 3, 6 oder 12 Monate sein")
    # Verbraucherrecht (08.08.2026): Weil die Leistung sofort bereitsteht,
    # braucht es VOR der Buchung die ausdrueckliche Zustimmung zum Beginn
    # vor Ablauf der Widerrufsfrist — sonst koennte nach Wochen der Nutzung
    # noch der volle Betrag zurueckgefordert werden. Die Zustimmung wird
    # protokolliert, weil im Streitfall WIR sie beweisen muessen.
    if data.get("widerruf_zustimmung") is not True:
        raise HTTPException(
            status_code=400,
            detail="Bitte bestätige zuerst den sofortigen Leistungsbeginn und die "
                   "Kenntnisnahme der Widerrufsbelehrung.")
    db_user = get_user_by_id(user["id"])
    if billing.effektiver_plan(db_user) != "free":
        raise HTTPException(status_code=400,
                            detail="Du hast bereits ein aktives Abo. Plan-Wechsel bitte über "
                                   "support@inkludocs.de — Online-Wechsel folgt")
    try:
        kunde = stripe_zahlung.kunde_fuer(db_user)
        if kunde != db_user.get("stripe_customer_id"):
            conn = get_db()
            try:
                conn.execute("UPDATE users SET stripe_customer_id = ? WHERE id = ?",
                             (kunde, db_user["id"]))
                conn.commit()
            finally:
                conn.close()
        conn = get_db()
        try:
            cur = conn.execute(
                "INSERT INTO widerruf_zustimmungen (user_id, plan, laufzeit_monate, "
                "summe_cent, belehrung_fassung, absender) VALUES (?, ?, ?, ?, ?, ?)",
                (db_user["id"], plan, monate,
                 round(billing.preis_pro_monat(plan, monate) * monate * 100),
                 WIDERRUFSBELEHRUNG_FASSUNG, absender.limit_schluessel(request)))
            zustimmung_id = cur.lastrowid
            conn.commit()
        finally:
            conn.close()
        # Die Zustimmungs-Nummer reist als Metadatum durch Stripe zurueck —
        # so dokumentiert die Bestaetigungsmail genau DIE Zustimmung, die zu
        # dieser Zahlung gehoert, nicht die eines abgebrochenen Versuchs
        # (Geldfluss-Review Befund 7, 24.08.2026).
        url = stripe_zahlung.checkout_abo(db_user, plan, monate, BASE_URL, kunde,
                                          zustimmung_id=zustimmung_id)
    except HTTPException:
        raise
    except Exception:
        logger.exception("Stripe-Checkout fehlgeschlagen (user=%s)", user["id"])
        raise HTTPException(status_code=502, detail="Die Zahlungsseite konnte nicht erstellt werden")
    return {"ok": True, "url": url}


@app.post("/api/abo/paket-checkout")
async def abo_paket_checkout(request: Request, user: dict = Depends(get_current_user)):
    """Stripe-Checkout-Link fuer ein Zusatz-Credit-Paket (Einmalzahlung)."""
    if not stripe_zahlung.AKTIV:
        raise HTTPException(status_code=503, detail="Online-Zahlung ist noch nicht freigeschaltet")
    data = await request.json()
    try:
        groesse = int(data.get("groesse"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="groesse fehlt")
    if groesse not in billing.PAKET_PREISE:
        raise HTTPException(status_code=400, detail="Ungueltige Paketgroesse")
    db_user = get_user_by_id(user["id"])
    try:
        kunde = stripe_zahlung.kunde_fuer(db_user)
        if kunde != db_user.get("stripe_customer_id"):
            conn = get_db()
            try:
                conn.execute("UPDATE users SET stripe_customer_id = ? WHERE id = ?",
                             (kunde, db_user["id"]))
                conn.commit()
            finally:
                conn.close()
        url = stripe_zahlung.checkout_paket(db_user, groesse, BASE_URL, kunde)
    except Exception:
        logger.exception("Stripe-Paket-Checkout fehlgeschlagen (user=%s)", user["id"])
        raise HTTPException(status_code=502, detail="Die Zahlungsseite konnte nicht erstellt werden")
    return {"ok": True, "url": url}


@app.post("/api/abo/portal")
async def abo_portal(user: dict = Depends(get_current_user)):
    """Link ins Stripe-Kundenportal (Rechnungen, Zahlungsart, Kuendigung)."""
    if not stripe_zahlung.AKTIV:
        raise HTTPException(status_code=503, detail="Online-Zahlung ist noch nicht freigeschaltet")
    db_user = get_user_by_id(user["id"])
    kunde = (db_user.get("stripe_customer_id") or "").strip()
    if not kunde:
        raise HTTPException(status_code=400, detail="Für dieses Konto gibt es noch keine Online-Zahlungen")
    try:
        url = stripe_zahlung.portal_url(kunde, BASE_URL)
    except Exception:
        logger.exception("Stripe-Portal fehlgeschlagen (user=%s)", user["id"])
        raise HTTPException(status_code=502, detail="Das Zahlungs-Portal ist gerade nicht erreichbar")
    return {"ok": True, "url": url}


@app.post("/api/stripe/webhook")
async def stripe_webhook(request: Request):
    """Stripe meldet Zahlungs-Ereignisse — HIER passiert die Freischaltung.

    Signaturgeprueft (STRIPE_WEBHOOK_SECRET). Antwortet Stripe gegenueber
    immer schnell mit 200, sobald das Ereignis verarbeitet oder bewusst
    ignoriert wurde; nur Signatur-/Formatfehler geben 400.
    """
    payload = await request.body()
    signatur = request.headers.get("stripe-signature", "")
    try:
        event = stripe_zahlung.pruefe_webhook(payload, signatur)
    except Exception:
        logger.warning("Stripe-Webhook mit ungueltiger Signatur abgewiesen")
        raise HTTPException(status_code=400, detail="Ungueltige Signatur")

    typ = event["type"]
    obj = event["data"]["object"]
    try:
        if typ == "checkout.session.completed":
            meta = obj.get("metadata") or {}
            user_id = int(meta.get("idoc_user_id") or 0)
            target = get_user_by_id(user_id) if user_id else None
            if not target:
                logger.warning("Stripe-Webhook: unbekanntes Konto in %s", typ)
                return {"ok": True}
            if obj.get("mode") == "subscription" and meta.get("idoc_plan"):
                plan = meta["idoc_plan"]
                monate = int(meta.get("idoc_laufzeit") or 6)
                gueltig_bis, _g, _u = _setze_plan_kern(
                    user_id, target, plan, monate, None, 1, quelle="stripe")
                conn = get_db()
                try:
                    conn.execute("UPDATE users SET stripe_subscription_id = ?, "
                                 "stripe_customer_id = ?, plan_treue = 0 WHERE id = ?",
                                 (obj.get("subscription"), obj.get("customer"), user_id))
                    conn.commit()
                finally:
                    conn.close()
                # Treue-Anschluss (AGB Ziffer 7, 24.08.2026): Feste Laufzeiten
                # laufen danach als Monatsabo zur Treuekondition weiter — der
                # Zeitplan wird direkt bei der Buchung hinterlegt. Scheitert
                # das hier, heilt der invoice.paid-Webhook nach.
                if monate > 1:
                    try:
                        stripe_zahlung.plane_treue_anschluss(obj.get("subscription"), plan)
                    except Exception:
                        logger.exception("Treue-Anschluss nach Buchung nicht angelegt (user=%s)", user_id)
                try:
                    _sende_plan_bestaetigung(target, plan, gueltig_bis, monate, 1,
                                             quelle="stripe",
                                             zustimmung=_letzte_widerruf_zustimmung(
                                                 user_id, plan, monate,
                                                 zustimmung_id=meta.get("idoc_zustimmung")))
                except Exception:
                    logger.exception("Bestaetigungs-Mail nach Stripe-Buchung fehlgeschlagen")
                _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs Stripe-Buchung: {target['email']}",
                          "Neue Online-Buchung",
                          [f"{html.escape(target['email'])} hat {PLAN_ANZEIGENAMEN.get(plan, plan)} "
                           f"für {monate} Monate über Stripe gebucht (gültig bis {str(gueltig_bis)[:10]})."])
            elif obj.get("mode") == "payment" and meta.get("idoc_paket"):
                groesse = int(meta["idoc_paket"])
                # Steves Regel (24.08.2026): GEKAUFTE Pakete verfallen nie
                # und sind immer nutzbar — auch fuer Free-Konten. Datums-
                # Verfall gibt es nur noch fuer Kulanz-Geschenke (Admin-Weg).
                billing.schenke_credits(user_id, groesse, notiz="Stripe-Kauf",
                                        quelle="stripe", verfall_monate=None)
                _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs Stripe-Paketkauf: {target['email']}",
                          "Credit-Paket gekauft",
                          [f"{html.escape(target['email'])} hat {groesse} Zusatz-Credits über "
                           "Stripe gekauft."])
        elif typ == "invoice.paid":
            # DIE letzte Wahrheit fuer Stripe-Abos (Review-Befunde 2 und 4,
            # 07.08.2026): Jede bezahlte Abo-Rechnung setzt Plan, Laufzeit
            # UND Laufzeitende so, wie tatsaechlich bezahlt wurde — egal ob
            # Zyklus-Verlaengerung, sofortiges Upgrade (subscription_update)
            # oder der vollzogene Downgrade aus einem Schedule. Damit kommt
            # ein bezahlter Plan auch dann an, wenn der ausloesende Aufruf
            # unterwegs abgebrochen ist, und ein Downgrade wird genau dann
            # wirksam, wenn Stripe ihn abrechnet.
            grund = obj.get("billing_reason") or ""
            sub_id = obj.get("subscription")
            if sub_id and grund.startswith("subscription"):
                zeilen_liste = (obj.get("lines") or {}).get("data") or []
                periode_ende = max((z.get("period", {}).get("end") or 0)
                                   for z in zeilen_liste) if zeilen_liste else 0
                # Bezahlten Plan aus dem Preis-Lookup der Abo-Zeile lesen.
                neuer_plan, neue_laufzeit, ist_treue = None, None, False
                for z in zeilen_liste:
                    lk = ((z.get("price") or {}).get("lookup_key")
                          or (z.get("pricing") or {}).get("price_details", {}).get("lookup_key"))
                    p, m, tr = stripe_zahlung.plan_aus_lookup(lk)
                    if p:
                        neuer_plan, neue_laufzeit, ist_treue = p, m, tr
                conn = get_db()
                try:
                    row = conn.execute("SELECT id FROM users WHERE stripe_subscription_id = ?",
                                       (sub_id,)).fetchone()
                finally:
                    conn.close()
                if row:
                    ziel = get_user_by_id(row["id"])
                    if ziel and neuer_plan and neuer_plan in billing.PLAN_KONTINGENTE:
                        gueltig = (datetime.utcfromtimestamp(int(periode_ende)).strftime("%Y-%m-%d")
                                   if periode_ende else None)
                        alter_plan = ziel.get("plan")
                        alte_laufzeit = int(ziel.get("plan_laufzeit_monate") or 0)
                        # Geldfluss-Review Befund 2 (24.08.2026): Eine ZYKLUS-
                        # Abrechnung sagt nichts ueber eine Kuendigung — sie darf
                        # auto_verlaengerung=0 nicht wieder auf 1 heben (sonst
                        # lebt eine Kuendigung nach der letzten Zahlung wieder
                        # auf). Neubuchung/Upgrade setzen dagegen bewusst 1.
                        if grund == "subscription_cycle":
                            _av_alt = ziel.get("auto_verlaengerung")
                            auto_neu = 1 if _av_alt is None else int(_av_alt)
                        else:
                            auto_neu = 1
                        _setze_plan_kern(ziel["id"], ziel, neuer_plan, neue_laufzeit,
                                         gueltig, auto_neu, quelle="stripe")
                        conn = get_db()
                        try:
                            conn.execute("UPDATE users SET geplanter_plan = NULL, "
                                         "geplante_laufzeit = NULL, geplant_ab = NULL, "
                                         "plan_treue = ? WHERE id = ?",
                                         (1 if ist_treue else 0, ziel["id"]))
                            conn.commit()
                        finally:
                            conn.close()
                        # Review-Befund 6: Kontingent-Neustart auch beim reinen
                        # Laufzeit-Upgrade (single/3 -> single/12), aber NIE bei
                        # der bloßen Verlaengerung desselben Plans. Der TREUE-
                        # Uebergang (12 Monate -> Treue-Monatsabo) ist dieselbe
                        # Stufe in neuem Rhythmus — KEIN Neustart (24.08.2026).
                        if (billing.PLAN_KONTINGENTE.get(neuer_plan, 0)
                                > billing.PLAN_KONTINGENTE.get(alter_plan, 0)
                                or (alter_plan == neuer_plan and not ist_treue
                                    and neue_laufzeit != alte_laufzeit)):
                            billing.starte_kontingent_neu(ziel["id"])
                        # Selbstheilung Treue-Anschluss (24.08.2026): Jede
                        # bezahlte FESTE Laufzeit muss ihren Anschluss-Zeitplan
                        # haben — auch wenn der Checkout-Webhook ausfiel oder
                        # das Upgrade vor dem Anlegen abbrach. nur_wenn_frei
                        # laesst vorgemerkte Downgrades unangetastet.
                        if (not ist_treue and neue_laufzeit and int(neue_laufzeit) > 1
                                and stripe_zahlung.AKTIV):
                            try:
                                stripe_zahlung.plane_treue_anschluss(
                                    sub_id, neuer_plan, nur_wenn_frei=True)
                            except Exception:
                                logger.exception("Treue-Anschluss-Selbstheilung "
                                                 "fehlgeschlagen (%s)", sub_id)
                        if ist_treue and alte_laufzeit > 1 and _kv_einmal(
                                f"abo_mail_treue_{ziel['id']}_{gueltig}"):
                            # Erste Treue-Abrechnung: dem Kunden den Uebergang
                            # bestaetigen (AGB Ziffer 7).
                            try:
                                _preis = f"{billing.treue_preis_eur(neuer_plan):.2f}".replace(".", ",")
                                _abo_mail(ziel["email"],
                                          "InkluDocs: Dein Abo läuft jetzt als Monatsabo weiter",
                                          "Dein Abo läuft als Monatsabo weiter",
                                          [f"Hallo {html.escape(ziel.get('display_name') or '')},",
                                           f"deine feste Laufzeit ist abgelaufen. Wie in den Nutzungsbedingungen "
                                           f"zugesagt, läuft dein {PLAN_ANZEIGENAMEN.get(neuer_plan, neuer_plan)}-Abo "
                                           f"jetzt als Monatsabo zur Treuekondition weiter: {_preis} € pro Monat — "
                                           "dein bisheriger Monatspreis, jetzt ohne feste Laufzeit und monatlich kündbar.",
                                           "Du kannst jederzeit zum Ende des laufenden Monats kündigen oder erneut "
                                           "eine feste Laufzeit buchen — beides auf der Seite „Abo &amp; Verbrauch“."])
                            except Exception:
                                logger.exception("Treue-Uebergangs-Mail fehlgeschlagen")
                        if alter_plan != neuer_plan and _kv_einmal(
                                f"abo_mail_stripewechsel_{ziel['id']}_{gueltig}_{neuer_plan}"):
                            try:
                                _sende_plan_bestaetigung(ziel, neuer_plan, gueltig,
                                                         neue_laufzeit, 1, quelle="stripe")
                            except Exception:
                                logger.exception("Bestaetigung nach Stripe-Abrechnung fehlgeschlagen")
                    elif row and periode_ende:
                        # Kein erkennbarer Plan (Fremdpreis): wenigstens das
                        # Laufzeitende mitziehen.
                        conn = get_db()
                        try:
                            conn.execute("UPDATE users SET plan_gueltig_bis = ? "
                                         "WHERE stripe_subscription_id = ?",
                                         (datetime.utcfromtimestamp(int(periode_ende)).strftime("%Y-%m-%d"),
                                          sub_id))
                            conn.commit()
                        finally:
                            conn.close()
        elif typ in ("customer.subscription.updated", "customer.subscription.deleted"):
            # Kuendigung im Stripe-Portal (oder durch Stripe selbst) mit
            # unserem auto_verlaengerung-Schalter synchron halten.
            #
            # BEFUND 07.08.2026 (im Test gefunden): Stripe-Events koennen sich
            # UEBERHOLEN. Beim Kuendigen loesen wir erst den Wechsel-Schedule
            # (Event: cancel_at_period_end=false) und kuendigen dann (Event:
            # true) — trifft das erste Event zuletzt ein, wuerde es die
            # Kuendigung stillschweigend aufheben. Darum wird der Zustand
            # frisch bei Stripe abgefragt statt dem (evtl. veralteten)
            # Payload zu glauben.
            sub_id = obj.get("id")
            beendet = (typ == "customer.subscription.deleted")
            aktuell = obj
            if not beendet and sub_id and stripe_zahlung.AKTIV:
                try:
                    aktuell = stripe_zahlung.stripe.Subscription.retrieve(sub_id)
                except Exception:
                    logger.exception("Subscription-Nachfrage fehlgeschlagen (%s)", sub_id)
            if not beendet and aktuell.get("status") in ("canceled", "incomplete_expired"):
                beendet = True
            auslaufend = bool(aktuell.get("cancel_at_period_end")) or beendet
            conn = get_db()
            try:
                # Zustand VOR dem Update merken: Der Treue-Zeitplan unten darf
                # nur nach einer ECHTEN Reaktivierung (0 -> 1) neu entstehen.
                treue_row = conn.execute(
                    "SELECT plan, plan_laufzeit_monate, "
                    "COALESCE(auto_verlaengerung, 1) AS auto_verlaengerung "
                    "FROM users WHERE stripe_subscription_id = ?", (sub_id,)).fetchone()
                conn.execute("UPDATE users SET auto_verlaengerung = ? "
                             "WHERE stripe_subscription_id = ?",
                             (0 if auslaufend else 1, sub_id))
                # Review-Befund 3: Beendet Stripe die Subscription (etwa nach
                # gescheiterten Zahlungsversuchen), darf lokal keine bezahlte
                # Restlaufzeit stehen bleiben — sonst laeuft ein nie bezahlter
                # Plan monatelang weiter. Auf das echte Periodenende kappen.
                if beendet:
                    ende = (aktuell.get("current_period_end") or aktuell.get("ended_at")
                            or obj.get("current_period_end") or obj.get("ended_at"))
                    if ende:
                        conn.execute(
                            "UPDATE users SET plan_gueltig_bis = ? "
                            "WHERE stripe_subscription_id = ? AND "
                            "(plan_gueltig_bis IS NULL OR date(plan_gueltig_bis) > date(?))",
                            (datetime.utcfromtimestamp(int(ende)).strftime("%Y-%m-%d"), sub_id,
                             datetime.utcfromtimestamp(int(ende)).strftime("%Y-%m-%d")))
                conn.commit()
            finally:
                conn.close()
            # Reaktivierung AUCH ueber diesen Weg (Portal statt App,
            # 24.08.2026): Die Kuendigung hatte den Treue-Zeitplan geloest —
            # bei fester Laufzeit wieder aufsetzen, sonst haengt der Treue-
            # Anschluss (AGB Ziffer 7) davon ab, WO der Kunde reaktiviert.
            # nur_wenn_frei laesst einen vorhandenen Zeitplan unangetastet.
            # NUR bei echter Reaktivierung (vorher gekuendigt) — sonst wuerde
            # das Release-Event einer GERADE LAUFENDEN Kuendigung hier den
            # Zeitplan neu anlegen und den cancel-Aufruf blockieren
            # (Geldfluss-Review Befund 2, 24.08.2026).
            if (not auslaufend and treue_row
                    and int(treue_row["auto_verlaengerung"]) == 0
                    and int(treue_row["plan_laufzeit_monate"] or 0) > 1
                    and stripe_zahlung.AKTIV):
                try:
                    stripe_zahlung.plane_treue_anschluss(
                        sub_id, treue_row["plan"], nur_wenn_frei=True)
                except Exception:
                    logger.exception("Treue-Anschluss nach Portal-Reaktivierung "
                                     "nicht angelegt (%s)", sub_id)
    except Exception:
        # Verarbeitungsfehler loggen, aber 200 melden — Stripe wiederholt
        # sonst endlos; der Vorgang laesst sich im Log/Dashboard nachziehen.
        logger.exception("Stripe-Webhook-Verarbeitung fehlgeschlagen (%s)", typ)
    return {"ok": True}


# ─── Lizenzschluessel: Modell VERWORFEN 06.08.2026 (Steve+Michael) ─────
# Das Schluessel-Modell (03.–05.08.) wurde zugunsten der Abo-Stufen
# Single/Team/Enterprise aufgegeben. Alle Routen und die Oberflaeche sind
# entfernt; die DB-Tabelle lizenzschluessel bleibt als Beleg stehen.
# Bestands-Toepfe wurden per Migration auf plan='team' ueberfuehrt.


# ─── Abo-Tageslauf: Erinnerungen + Auto-Verlaengerung (06.08.2026) ─────
# Ohne Cron: /api/me stoesst den Lauf an (Hintergrund-Thread), system_kv
# stellt sicher, dass er hoechstens 1x pro Tag arbeitet und keine Mail
# doppelt rausgeht. Fachregeln (Steve 06.08.): Bezahl-Plaene laufen fest
# 3/6/12 Monate; ohne Kuendigung verlaengert sich der Plan automatisch um
# die gebuchte Laufzeit (die Rechnung folgt — Betreiber werden erinnert);
# mit Kuendigung endet er am Laufzeitende (Lazy-Rueckfall in billing).

PLAN_ANZEIGENAMEN = {"free": "Free", "single": "Single", "team": "Team",
                     "enterprise": "Enterprise"}

_MAIL_FUSS = ('<p style="margin:0 0 6px;">Bei Fragen wende dich an '
              '<a href="mailto:support@inkludocs.de" style="color:#1b2a4a;">support@inkludocs.de</a>.</p>'
              '<p style="color:#64748b;font-size:13px;margin:18px 0 0;">'
              'InkluDocs ist ein Produkt von INKLUTEC · kontakt@inklutec.de</p>')


def _abo_mail(empfaenger: str, betreff: str, ueberschrift: str, saetze: list) -> None:
    """Einheitlicher Rahmen fuer alle Abo-Mails (Bestaetigung/Erinnerung).

    Aufgeraeumtes, mail-client-sicheres Layout (Steve, 24.08.2026): alle
    Stile inline, ein Markenkopf, ruhige Abstaende, klare Fusszeile —
    wirkt auf ALLE Systemmails gleichzeitig.
    """
    absaetze = "".join(
        f'<p style="margin:0 0 14px;line-height:1.55;">{s}</p>' for s in saetze)
    body = (
        '<!DOCTYPE html><html lang="de"><head><meta charset="utf-8"></head>'
        '<body style="margin:0;padding:0;background-color:#f1f5f9;">'
        '<div style="max-width:600px;margin:0 auto;padding:24px 16px;'
        'font-family:Arial,Helvetica,sans-serif;color:#1e293b;font-size:16px;">'
        '<p style="margin:0 0 18px;font-size:22px;font-weight:bold;color:#1b2a4a;">'
        'Inklu<span style="font-weight:normal;">Docs</span></p>'
        '<div style="background-color:#ffffff;border-radius:8px;padding:28px 24px;'
        'border:1px solid #e2e8f0;">'
        f'<h1 style="color:#1b2a4a;font-size:20px;margin:0 0 18px;">{ueberschrift}</h1>'
        f'{absaetze}'
        '</div>'
        f'<div style="padding:18px 8px 0;font-size:14px;color:#475569;">{_MAIL_FUSS}</div>'
        '</div></body></html>')
    send_email(empfaenger, betreff, body, bcc_admin=False)


def _letzte_widerruf_zustimmung(user_id: int, plan: str, laufzeit,
                                zustimmung_id=None) -> dict:
    """Protokollierte Zustimmung zum sofortigen Leistungsbeginn fuer die
    Dokumentation in der Bestaetigungsmail nach § 312f BGB (24.08.2026).
    Bevorzugt die per Checkout-Metadatum mitgereiste Nummer (beweissicher
    genau DIESER Kauf); Rueckfall: juengster Eintrag zu Plan + Laufzeit.
    None, wenn keine vorliegt (Admin-/Rechnungsweg erhebt keine)."""
    conn = get_db()
    try:
        if zustimmung_id:
            row = conn.execute(
                "SELECT belehrung_fassung, created_at, summe_cent FROM widerruf_zustimmungen "
                "WHERE id = ? AND user_id = ?", (int(zustimmung_id), user_id)).fetchone()
            if row:
                return dict(row)
        row = conn.execute(
            "SELECT belehrung_fassung, created_at, summe_cent FROM widerruf_zustimmungen "
            "WHERE user_id = ? AND plan = ? AND laufzeit_monate = ? "
            "ORDER BY id DESC LIMIT 1", (user_id, plan, int(laufzeit or 0))).fetchone()
        return dict(row) if row else None
    except Exception:
        logger.exception("Widerruf-Zustimmung nicht lesbar (user=%s)", user_id)
        return None
    finally:
        conn.close()


_MONATSNAMEN = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli",
                "August", "September", "Oktober", "November", "Dezember")


def _datum_lang(iso: str) -> str:
    """'2026-08-24…' -> '24. August 2026' (fuer Mailtexte; Steve 24.08.:
    Datumsangaben ausgeschrieben statt als Technik-Kuerzel)."""
    try:
        j, m, t = str(iso)[:10].split("-")
        return f"{int(t)}. {_MONATSNAMEN[int(m) - 1]} {j}"
    except Exception:
        return str(iso)[:10]


def _sende_plan_bestaetigung(konto: dict, plan: str, gueltig_bis, laufzeit,
                             auto_verlaengerung: int, quelle: str = "rechnung",
                             zustimmung: dict = None) -> None:
    """Bestaetigungs-Mail nach einer Plan-Buchung (Admin-/Rechnungsweg UND
    Stripe-Weg). quelle steuert den Verlaengerungs-Text: Online-Vertraege
    (stripe) mit fester Laufzeit laufen nach AGB Ziffer 7 als Monatsabo zur
    Treuekondition weiter; Rechnungsvertraege verlaengern sich um dieselbe
    Laufzeit. zustimmung dokumentiert die beim Checkout erteilte Zustimmung
    zum sofortigen Leistungsbeginn (§ 312f BGB, 24.08.2026)."""
    name = html.escape(konto.get("display_name") or "")
    plan_name = PLAN_ANZEIGENAMEN.get(plan, plan)
    datum = str(gueltig_bis or "")[:10]
    laufzeit_zahl = int(laufzeit or 0)
    # Betrag nur beim Online-Kauf nennen (Steve 24.08.); Rechnungskunden
    # bekommen ihn ueber Michaels Rechnung.
    preis_zusatz = ""
    if quelle == "stripe" and laufzeit_zahl > 1:
        _cent = ((zustimmung or {}).get("summe_cent")
                 or round(billing.preis_pro_monat(plan, laufzeit_zahl) * laufzeit_zahl * 100))
        preis_zusatz = f" zum Gesamtpreis von {_cent / 100:.2f} €".replace(".", ",")
    elif quelle == "stripe" and laufzeit_zahl == 1:
        _monat = f"{billing.preis_pro_monat(plan, 1):.2f}".replace(".", ",")
        preis_zusatz = f" für {_monat} € pro Monat"
    saetze = [f"Hallo {name},",
              f"dein InkluDocs-Abo wurde umgestellt: Du hast jetzt den Plan <strong>{plan_name}</strong>"
              + (f" mit einer Laufzeit von {laufzeit} Monaten" if laufzeit_zahl > 1 else
                 (" als Monatsabo" if laufzeit_zahl == 1 else ""))
              + preis_zusatz
              + (f", gültig bis {_datum_lang(datum)}" if datum else "") + "."]
    if auto_verlaengerung:
        if quelle == "stripe" and laufzeit_zahl > 1:
            treue = f"{billing.treue_preis_eur(plan):.2f}".replace(".", ",")
            saetze.append(f"Nach dem Laufzeitende läuft das Abo automatisch als Monatsabo zur "
                          f"Treuekondition weiter: {treue} € pro Monat — dein Monatspreis "
                          "bleibt also gleich, nur ohne feste Laufzeit und monatlich kündbar. "
                          "Du kannst stattdessen auch erneut eine feste Laufzeit buchen. "
                          "Kündigen kannst du jederzeit zum Laufzeitende auf der Seite "
                          "„Abo &amp; Verbrauch“.")
        elif quelle == "stripe" and laufzeit_zahl == 1:
            saetze.append("Das Monatsabo verlängert sich um jeweils einen Monat und ist "
                          "jederzeit zum Ende des laufenden Monatszeitraums kündbar — auf der "
                          "Seite „Abo &amp; Verbrauch“.")
        else:
            saetze.append("Wenn du nicht kündigst, verlängert sich das Abo am Laufzeitende "
                          "automatisch um dieselbe Laufzeit. Kündigen kannst du jederzeit "
                          "zum Laufzeitende auf der Seite „Abo &amp; Verbrauch“.")
    else:
        saetze.append(f"Das Abo läuft am {_datum_lang(datum)} aus und wird nicht automatisch verlängert.")
    if zustimmung:
        # § 312f BGB: Die Bestaetigung muss die erteilte Zustimmung zum
        # Leistungsbeginn VOR Ablauf der Widerrufsfrist dokumentieren.
        # Formulierung nach Steves Hoertest 24.08.: freundlicher Absatz
        # statt Behoerdenton; nur das Buchungsdatum (der sekundengenaue
        # Zeitstempel bleibt im Protokoll), Fassung ausgeschrieben.
        wann = _datum_lang(str(zustimmung.get("created_at") or ""))
        fassung = _datum_lang(str(zustimmung.get("belehrung_fassung") or ""))
        saetze.append("Dein Widerrufsrecht: Als Verbraucher kannst du diesen Vertrag "
                      "innerhalb von 14 Tagen widerrufen — am einfachsten online über "
                      f'unsere <a href="{BASE_URL}/widerrufen">Widerrufsseite</a>. '
                      "Damit du InkluDocs sofort nutzen kannst, hast du bei der Buchung"
                      + (f" am {wann}" if wann else "")
                      + " zugestimmt, dass wir direkt mit der Leistung beginnen. "
                      "Widerrufst du innerhalb der Frist, erstatten wir den Betrag "
                      "abzüglich des bereits genutzten Anteils; ist die Leistung "
                      "vollständig erbracht, erlischt das Widerrufsrecht. Alle "
                      "Einzelheiten stehen in der "
                      f'<a href="{BASE_URL}/widerruf">Widerrufsbelehrung</a>'
                      + (f" (Fassung vom {fassung})" if fassung else "") + ".")
    saetze.append("Deinen aktuellen Stand siehst du jederzeit unter „Abo &amp; Verbrauch“ in der App.")
    _abo_mail(konto["email"], f"InkluDocs: Dein {plan_name}-Abo ist aktiv",
              "Dein InkluDocs-Abo", saetze)


def _kv_setzen(key: str, value: str) -> None:
    """Schluessel/Wert in system_kv ablegen (anlegen oder ueberschreiben)."""
    conn = get_db()
    try:
        conn.execute("INSERT INTO system_kv (key, value) VALUES (?, ?) "
                     "ON CONFLICT(key) DO UPDATE SET value = excluded.value, "
                     "updated_at = datetime('now')", (key, value))
        conn.commit()
    finally:
        conn.close()


def _kv_lesen(key: str):
    conn = get_db()
    try:
        row = conn.execute("SELECT value FROM system_kv WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else None
    finally:
        conn.close()


def _kv_einmal(key: str) -> bool:
    """True genau beim ERSTEN Aufruf fuer diesen Schluessel (Doppel-Schutz)."""
    conn = get_db()
    try:
        cur = conn.execute("INSERT OR IGNORE INTO system_kv (key, value) VALUES (?, '1')", (key,))
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


# Prozesslokaler Schnell-Merker (Review-Befund 06.08.): /api/me soll nicht
# bei jedem Aufruf einen Thread + SQLite-Schreibzugriff erzeugen. Erst wenn
# der Prozess den heutigen Lauf noch nicht gesehen hat, wird ueberhaupt
# gestartet; den Rest entscheidet der atomare system_kv-Claim.
_tageslauf_lock = threading.Lock()
_tageslauf_datum = None


def _abo_tageslauf_noetig() -> bool:
    heute = datetime.utcnow().strftime("%Y-%m-%d")
    return _tageslauf_datum != heute


def _abo_tageslauf() -> None:
    """Taeglicher Abo-Lauf (Erinnerungen, Auto-Verlaengerungen), max. 1x/Tag.

    Atomarer Tages-Claim ueber system_kv: von zwei parallelen Anstoessen
    arbeitet nur einer; der Prozess-Merker verhindert danach jeden weiteren
    Anlauf des Tages. Fehler je Konto werden geloggt und geschluckt.
    """
    global _tageslauf_datum
    try:
        heute = datetime.utcnow().strftime("%Y-%m-%d")
        with _tageslauf_lock:
            if _tageslauf_datum == heute:
                return
            conn = get_db()
            try:
                claimed = conn.execute(
                    "UPDATE system_kv SET value = ?, updated_at = datetime('now') "
                    "WHERE key = 'abo_tageslauf' AND value != ?", (heute, heute)).rowcount
                if not claimed:
                    claimed = conn.execute(
                        "INSERT OR IGNORE INTO system_kv (key, value) VALUES ('abo_tageslauf', ?)",
                        (heute,)).rowcount
                conn.commit()
                # Ob WIR den Claim haben oder ein anderer Prozess: fuer diesen
                # Prozess ist der Tag damit erledigt.
                _tageslauf_datum = heute
                if not claimed:
                    return
                # Untergrenze -62 Tage (Review-Befund K5): laengst abgelaufene
                # Bestandskonten werden weder taeglich durchgekaut noch beim
                # Erst-Deploy mit Abschieds-Mails geflutet; die Verlaengerungs-
                # Aufholung deckt trotzdem bis zu zwei Monate Server-Pause.
                faellige = conn.execute(
                    "SELECT id, email, display_name, plan, plan_gueltig_bis, plan_laufzeit_monate, "
                    "plan_quelle, stripe_subscription_id, geplanter_plan, geplante_laufzeit, geplant_ab, "
                    "COALESCE(auto_verlaengerung, 1) AS auto_verlaengerung "
                    "FROM users WHERE plan IN ('single', 'team', 'enterprise') "
                    "AND ((plan_gueltig_bis IS NOT NULL "
                    "      AND date(plan_gueltig_bis) <= date('now', '+14 days') "
                    "      AND date(plan_gueltig_bis) >= date('now', '-62 days')) "
                    "  OR (geplanter_plan IS NOT NULL AND geplant_ab IS NOT NULL "
                    "      AND date(geplant_ab) <= date('now')))").fetchall()
            finally:
                conn.close()
        for k in faellige:
            try:
                _abo_konto_pruefen(dict(k), heute)
            except Exception:
                logger.exception("Abo-Tageslauf: Konto %s fehlgeschlagen", k["id"])
        # Abschluss festhalten (10.08.2026): Nur so kann ein Waechter von aussen
        # sehen, ob der Lauf wirklich durchlief — ein blosses "angestossen"
        # meldet auch dann gruen, wenn er unterwegs abbricht.
        _kv_setzen("abo_tageslauf_fertig", datetime.utcnow().strftime("%Y-%m-%d %H:%M") + " UTC")
    except Exception:
        logger.exception("Abo-Tageslauf fehlgeschlagen")
        # Nach einem Fehlschlag den Tag WIEDER FREIGEBEN (10.08.2026): Sonst
        # gilt der Tag als erledigt, obwohl nichts passiert ist — weder ein
        # spaeterer Anstoss vom Cron noch der Weg ueber /api/me kaeme dann
        # nochmal zum Zug, und Erinnerungen fielen still aus. Die einzelnen
        # Mails sind ohnehin gegen Doppelversand gesichert (_kv_einmal), ein
        # zweiter Anlauf am selben Tag ist also gefahrlos.
        try:
            _tageslauf_datum = None
            _kv_setzen("abo_tageslauf", "fehlgeschlagen")
            _kv_setzen("abo_tageslauf_fehler", datetime.utcnow().strftime("%Y-%m-%d %H:%M") + " UTC")
        except Exception:
            pass


@app.post("/api/intern/tageslauf")
async def intern_tageslauf(request: Request):
    """Taeglicher Anstoss von aussen (10.08.2026) — fuer Cron/Uptime-Kuma.

    Bis hierher hing der Abo-Tageslauf allein daran, dass sich jemand anmeldet
    (/api/me). Meldet sich einen Tag lang niemand an, verschieben sich
    Erinnerungen, Verlaengerungen und vorgemerkte Wechsel um einen Tag — bei
    der zugesagten 14-Tage-Erinnerung ist das ein Versprechen, das wir dann
    nicht halten. Dieser Endpunkt stoesst denselben Lauf an. Sein Selbstschutz
    bleibt unveraendert (ein Lauf je Tag, atomar ueber system_kv), und der Weg
    ueber /api/me bleibt als zweites Standbein bestehen.

    Antwort: angestossen = ob DIESER Aufruf den Lauf gestartet hat,
    letzter_abschluss = wann der Lauf zuletzt vollstaendig durchlief.
    """
    token = (os.environ.get("TAGESLAUF_TOKEN") or "").strip()
    if not token:
        raise HTTPException(status_code=503, detail="Tageslauf-Anstoss ist nicht eingerichtet.")
    mitgegeben = (request.headers.get("X-Tageslauf-Token") or "").strip()
    if not secrets.compare_digest(mitgegeben, token):
        raise HTTPException(status_code=401, detail="Nicht berechtigt.")
    angestossen = _abo_tageslauf_noetig()
    if angestossen:
        # Im Hintergrund, aber mit kurzer Wartezeit: Der Lauf verschickt Mails
        # und darf den Waechter nicht in seinen Zeitablauf ziehen. 20 Sekunden
        # reichen im Normalfall, damit die Antwort den frischen Abschluss schon
        # enthaelt; dauert es laenger, laeuft er trotzdem zu Ende und der
        # naechste Aufruf sieht das Ergebnis.
        t = threading.Thread(target=_abo_tageslauf, daemon=True)
        t.start()
        t.join(timeout=20)
    # lauf_datum ist der Tag, an dem der Lauf zuletzt WIRKLICH dran war. Ohne
    # dieses Feld kann ein Waechter "heute schon erledigt" nicht von "seit
    # Tagen nichts gelaufen" unterscheiden — beide zeigten sonst nur
    # angestossen=false. Genau das ist beim ersten Test passiert (10.08.2026).
    return {"ok": True, "angestossen": angestossen,
            "lauf_datum": _kv_lesen("abo_tageslauf"),
            "letzter_abschluss": _kv_lesen("abo_tageslauf_fertig"),
            "letzter_fehler": _kv_lesen("abo_tageslauf_fehler")}


def _abo_konto_pruefen(k: dict, heute: str) -> None:
    """Ein faelliges Konto abarbeiten: erinnern, verlaengern oder verabschieden."""
    ablauf = str(k["plan_gueltig_bis"])[:10]
    plan_name = PLAN_ANZEIGENAMEN.get(k["plan"], k["plan"])
    name = html.escape(k.get("display_name") or "")
    auto = int(k.get("auto_verlaengerung") or 0)
    laufzeit = k.get("plan_laufzeit_monate")

    # Vorgemerkter DOWNGRADE (07.08.2026): Am Stichtag wird der neue Plan
    # gesetzt — VOR jeder Verlaengerungs-Logik, denn der neue Plan bringt
    # seine eigene Laufzeit mit. Bis hierher lief alles unveraendert weiter,
    # der Kunde hat also bekommen, wofuer er bezahlt hat.
    # Stichtag aus geplant_ab (Review-Befund 2): plan_gueltig_bis kann sich
    # durch Stripe-Webhooks verschoben haben — der Wechsel muss trotzdem am
    # ursprünglich vorgemerkten Datum greifen. Bei STRIPE-Abos vollzieht ihn
    # der invoice.paid-Webhook (dort ist die Zahlung belegt); hier laufen nur
    # Rechnungsweg-Konten durch.
    geplant = (k.get("geplanter_plan") or "").strip()
    geplant_ab = (k.get("geplant_ab") or "")[:10]
    if geplant and geplant_ab and geplant_ab <= heute and k.get("plan_quelle") != "stripe":
        ziel_laufzeit = int(k.get("geplante_laufzeit") or laufzeit or 6)
        db_user = get_user_by_id(k["id"])
        if db_user:
            # Bedingtes Loeschen ZUERST (Review-Befund 11): Wer die Vormerkung
            # abraeumt, fuehrt den Wechsel aus — ein paralleler Lauf oder ein
            # gleichzeitiger Widerruf kommt dann nicht mehr zum Zug.
            conn = get_db()
            try:
                cur = conn.execute("UPDATE users SET geplanter_plan = NULL, "
                                   "geplante_laufzeit = NULL, geplant_ab = NULL "
                                   "WHERE id = ? AND geplanter_plan = ?", (k["id"], geplant))
                conn.commit()
                zustaendig = cur.rowcount > 0
            finally:
                conn.close()
            if not zustaendig:
                return
            gueltig_bis, geloest, _u = _setze_plan_kern(
                k["id"], db_user, geplant, ziel_laufzeit, None, 1,
                quelle=k.get("plan_quelle") or "rechnung")
            # Review-Befund 7: Kontingent-Neustart NUR, wenn der Zielplan
            # mehr Credits hat. Beim Downgrade waere er ein Geschenk (voller
            # neuer Topf zusaetzlich zum schon verbrauchten alten).
            if (billing.PLAN_KONTINGENTE.get(geplant, 0)
                    > billing.PLAN_KONTINGENTE.get(k.get("plan"), 0)):
                billing.starte_kontingent_neu(k["id"])
            if _kv_einmal(f"abo_mail_gewechselt_{k['id']}_{ablauf}"):
                saetze = [f"Hallo {name},",
                          f"wie vorgemerkt haben wir dein Abo heute auf den Plan "
                          f"{PLAN_ANZEIGENAMEN.get(geplant, geplant)} umgestellt. Ab sofort stehen dir "
                          f"{billing.PLAN_KONTINGENTE[geplant]} Credits im Monat zur Verfügung, "
                          f"die Laufzeit läuft bis {_datum_lang(str(gueltig_bis)[:10])}."]
                if geloest:
                    saetze.append(f"Dein Team wurde damit aufgelöst — die {geloest} Mitglieder "
                                  "arbeiten ab jetzt wieder mit ihrem eigenen Guthaben und wurden "
                                  "darüber informiert.")
                _abo_mail(k["email"], "InkluDocs: Dein Plan wurde umgestellt",
                          "Dein neuer Plan ist aktiv", saetze)
                _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs Plan-Wechsel vollzogen: {k['email']}",
                          "Vorgemerkter Plan-Wechsel ausgeführt",
                          [f"{html.escape(k['email'])} wurde heute planmäßig von {plan_name} auf "
                           f"{PLAN_ANZEIGENAMEN.get(geplant, geplant)} umgestellt "
                           f"({ziel_laufzeit} Monate, gültig bis {str(gueltig_bis)[:10]}).",
                           "Stripe-Abos stellt der Zahlungsplan automatisch um; Rechnungskunden "
                           "brauchen eine Rechnung über den neuen Plan."])
        return

    if ablauf >= heute:
        # Monatsabo (19.08.2026): KEINE monatliche Erinnerungsmail — das
        # waere Spam. Es ist jederzeit kuendbar, Stripe fakturiert selbst.
        if int(laufzeit or 0) == 1:
            return
        # Noch nicht abgelaufen -> Erinnerung (einmal pro Periode).
        if not _kv_einmal(f"abo_mail_erinnerung_{k['id']}_{ablauf}"):
            return
        # Review-Befund K2: "verlaengert sich automatisch" nur versprechen,
        # wenn die Verlaengerung auch ausgefuehrt wird (auto UND Laufzeit) —
        # sonst ehrlich das Ende ankuendigen.
        if auto and laufzeit:
            ist_stripe = k.get("plan_quelle") == "stripe"
            if ist_stripe and int(laufzeit) > 1:
                # Online-Vertrag (AGB Ziffer 7, 24.08.2026): KEINE erneute
                # feste Laufzeit — Weiterlauf als Monatsabo zur Treuekondition.
                #
                # RECONCILIATION (Geldfluss-Review Befund 1, 24.08.2026):
                # Ohne haengenden Treue-Zeitplan wuerde Stripe am Stichtag
                # still die VOLLE Laufzeit erneut abbuchen — und dieser Lauf
                # hier ist die letzte Kontrolle davor. Fehlt der Zeitplan,
                # wird er nachgelegt; klappt auch das nicht (z. B. fehlende
                # Treue-Preise im Live-Konto), bekommt der Betreiber einen
                # ALARM statt der Beruhigungsmail.
                zeitplan_ok = True
                sub_id_k = (k.get("stripe_subscription_id") or "").strip()
                if stripe_zahlung.AKTIV and sub_id_k:
                    try:
                        _sub = stripe_zahlung.stripe.Subscription.retrieve(sub_id_k)
                        if not _sub.get("schedule") and not _sub.get("cancel_at_period_end"):
                            stripe_zahlung.plane_treue_anschluss(
                                sub_id_k, k["plan"], nur_wenn_frei=True)
                            _sub = stripe_zahlung.stripe.Subscription.retrieve(sub_id_k)
                        zeitplan_ok = bool(_sub.get("schedule")) or bool(
                            _sub.get("cancel_at_period_end"))
                    except Exception:
                        logger.exception("Treue-Zeitplan-Kontrolle fehlgeschlagen (user=%s)",
                                         k["id"])
                        zeitplan_ok = False
                treue = f"{billing.treue_preis_eur(k['plan']):.2f}".replace(".", ",")
                saetze = [f"Hallo {name},",
                          f"dein {plan_name}-Abo bei InkluDocs erreicht am {_datum_lang(ablauf)} "
                          "das Laufzeitende. Danach läuft es automatisch als Monatsabo zur "
                          f"Treuekondition weiter: {treue} € pro Monat — dein bisheriger "
                          "Monatspreis, jetzt ohne feste Laufzeit und monatlich kündbar.",
                          # Formulierung nach Steves Hoertest 24.08. (die alte Fassung
                          # versprach faelschlich einen "guenstigeren" Laufzeitpreis).
                          "Das Anschluss-Abo ist jederzeit zum Ende des laufenden Monats "
                          "kündbar — du musst also nichts tun. Wenn du möchtest, kannst du "
                          "stattdessen erneut eine feste Laufzeit buchen. Und falls du nicht "
                          f"verlängern willst: Kündige einfach vor dem {_datum_lang(ablauf)} "
                          "auf der Seite „Abo &amp; Verbrauch“, dann endet das Abo zum "
                          "Laufzeitende."]
                _abo_mail(k["email"], "InkluDocs: Dein Abo läuft bald als Monatsabo weiter",
                          "Dein Abo läuft bald als Monatsabo weiter", saetze)
                if zeitplan_ok:
                    betreiber_satz = ("Stripe-Abo mit fester Laufzeit: Der hinterlegte Zahlungsplan "
                                      "stellt zum Laufzeitende automatisch auf das Treue-Monatsabo um "
                                      "und fakturiert monatlich — hier ist nichts zu tun (nur zur Kenntnis).")
                    betreff = f"InkluDocs Abo-Anschluss: {k['email']}"
                    kopf = "Abo läuft bald als Treue-Monatsabo weiter"
                else:
                    betreiber_satz = ("ALARM: Am Abo haengt KEIN Treue-Zahlungsplan, und er liess "
                                      "sich nicht nachlegen. Ohne Eingriff bucht Stripe am Stichtag "
                                      "die VOLLE Laufzeit erneut ab — entgegen AGB Ziffer 7. Bitte "
                                      "im Stripe-Dashboard pruefen (Subscription "
                                      f"{html.escape(sub_id_k or 'unbekannt')}); Ursache koennen "
                                      "fehlende Treue-Preise sein (sichere_produkte() ausfuehren).")
                    betreff = f"InkluDocs ALARM Treue-Zeitplan fehlt: {k['email']}"
                    kopf = "Treue-Zeitplan fehlt — Eingriff noetig"
                _abo_mail(NOTIFICATION_EMAIL, betreff, kopf,
                          [f"Das {plan_name}-Abo von {html.escape(k['email'])} erreicht am {ablauf} das "
                           f"Laufzeitende und soll dann als Treue-Monatsabo ({treue} €/Monat) weiterlaufen.",
                           betreiber_satz])
                return
            saetze = [f"Hallo {name},",
                      f"dein {plan_name}-Abo bei InkluDocs erreicht am {_datum_lang(ablauf)} "
                      "das Laufzeitende und verlängert sich dann automatisch"
                      + (f" um {laufzeit} Monate" if laufzeit else "") + ".",
                      "Wenn du das nicht möchtest, kündige einfach vorher auf der Seite "
                      "„Abo &amp; Verbrauch“ — dann endet das Abo zum Laufzeitende."]
            _abo_mail(k["email"], "InkluDocs: Dein Abo verlängert sich bald",
                      "Dein Abo verlängert sich bald", saetze)
            # Betreiber-Hinweis: beim Rechnungsweg muss eine neue Rechnung
            # raus; Stripe-Abos bucht und fakturiert Stripe von selbst.
            if ist_stripe:
                betreiber_satz = ("Stripe-Abo: Die Verlängerung bucht und fakturiert Stripe "
                                  "automatisch — hier ist nichts zu tun (nur zur Kenntnis).")
            else:
                betreiber_satz = ("Rechnungskunde: Jetzt eine neue Rechnung stellen "
                                  "(Actino/INKLUTEC) — die App verlängert am Stichtag automatisch.")
            _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs Abo-Verlängerung: {k['email']}",
                      "Abo-Verlängerung steht an",
                      [f"Das {plan_name}-Abo von {html.escape(k['email'])} verlängert sich am {ablauf} "
                       f"automatisch{f' um {laufzeit} Monate' if laufzeit else ''}.",
                       betreiber_satz])
        else:
            saetze = [f"Hallo {name},",
                      f"dein {plan_name}-Abo bei InkluDocs endet am {_datum_lang(ablauf)}. "
                      "Bis dahin kannst du es ganz normal weiter nutzen.",
                      "Danach wechselt dein Konto automatisch auf den Free-Plan mit "
                      f"{billing.PLAN_KONTINGENTE['free']} kostenlosen Credits pro Monat — deine "
                      "Projekte und Daten bleiben vollständig erhalten. Gekaufte Zusatz-Credits "
                      "bleiben dir ebenfalls erhalten und verfallen nicht — du kannst sie auch im "
                      "Free-Plan aufbrauchen.",
                      "Du hast es dir anders überlegt? Dann nimm die Kündigung unter "
                      "„Abo &amp; Verbrauch“ einfach wieder zurück."]
            _abo_mail(k["email"], "InkluDocs: Dein Abo endet bald",
                      "Dein Abo endet bald", saetze)
        return

    # Laufzeitende ueberschritten. Stripe-Abos verlaengert AUSSCHLIESSLICH
    # der invoice.paid-Webhook (sonst doppelt); scheitert dort die Zahlung,
    # setzt customer.subscription.deleted auto_verlaengerung=0 und dieses
    # Konto laeuft beim naechsten Tageslauf in den Abschieds-Zweig.
    if k.get("plan_quelle") == "stripe" and auto:
        return
    if auto and laufzeit:
        # Auto-Verlaengerung: gueltig_bis um die Laufzeit verlaengern, bis es
        # in der Zukunft liegt (faengt auch lange Server-Pausen ab).
        conn = get_db()
        try:
            neues = ablauf
            while neues < heute:
                neues = conn.execute("SELECT date(?, ?)",
                                     (neues, f"+{int(laufzeit)} months")).fetchone()[0]
            conn.execute("UPDATE users SET plan_gueltig_bis = ? WHERE id = ? AND plan_gueltig_bis = ?",
                         (neues, k["id"], k["plan_gueltig_bis"]))
            conn.commit()
        finally:
            conn.close()
        if _kv_einmal(f"abo_mail_verlaengert_{k['id']}_{neues}"):
            _abo_mail(k["email"], "InkluDocs: Dein Abo wurde verlängert",
                      "Dein Abo wurde verlängert",
                      [f"Hallo {name},",
                       f"dein {plan_name}-Abo wurde wie vereinbart um {laufzeit} Monate verlängert "
                       f"und läuft jetzt bis {_datum_lang(neues)}.",
                       "Wenn du das nicht mehr möchtest, kündige jederzeit zum Laufzeitende "
                       "auf der Seite „Abo &amp; Verbrauch“."])
            _abo_mail(NOTIFICATION_EMAIL, f"InkluDocs Abo verlängert: {k['email']}",
                      "Abo wurde automatisch verlängert",
                      [f"Das {plan_name}-Abo von {html.escape(k['email'])} wurde automatisch bis {neues} "
                       "verlängert. Bitte Rechnung stellen (Rechnungskunden)."])
        return

    # Abgelaufen ohne Verlaengerung: der Lazy-Rueckfall in billing macht das
    # Konto bereits zu Free — hier nur die Abschieds-Mail (einmalig, und nur
    # bei FRISCHEM Ablauf: Bestandskonten, die laengst ausgelaufen sind,
    # sollen beim Erst-Deploy keine verspaeteten Mails bekommen — Befund K5).
    frisch_ab = (datetime.utcnow() - timedelta(days=14)).strftime("%Y-%m-%d")
    if ablauf >= frisch_ab and _kv_einmal(f"abo_mail_abgelaufen_{k['id']}_{ablauf}"):
        _abo_mail(k["email"], "InkluDocs: Dein Abo ist beendet",
                  "Dein Abo ist beendet",
                  [f"Hallo {name},",
                   f"dein {plan_name}-Abo ist am {_datum_lang(ablauf)} ausgelaufen. Dein Konto steht jetzt "
                   "auf dem Free-Plan — alle Projekte und Daten bleiben erhalten, und du hast "
                   f"weiterhin {billing.PLAN_KONTINGENTE['free']} kostenlose Credits pro Monat.",
                   "Gekaufte Zusatz-Credits bleiben dir erhalten und verfallen nicht — du kannst "
                   "sie auch im Free-Plan aufbrauchen. Und wenn du weitermachen möchtest: Buche "
                   "jederzeit wieder ein Abo."])


@app.post("/api/admin/users/{user_id}/pakete")
async def admin_paket_anlegen(user_id: int, request: Request,
                              user: dict = Depends(require_full_admin)):
    """Admin: Zusatz-Credit-Paket fuer ein Konto anlegen (Rechnungsweg).

    Punkt 4 (04.08.2026): Preise laut billing.PAKET_PREISE (100/500/1000);
    andere Groessen sind fuer Kulanz erlaubt. verfall='kuendigung' (Vorgabe,
    Michaels Regel: Credits verfallen erst mit der Kuendigung) oder eine
    Monats-Zahl fuer klassische Datums-Pakete.
    """
    data = await request.json()
    try:
        groesse = int(data.get("groesse") or 0)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="groesse muss eine Zahl sein")
    if not 1 <= groesse <= 100000:
        raise HTTPException(status_code=400, detail="groesse muss zwischen 1 und 100000 liegen")
    verfall = data.get("verfall", "kuendigung")
    if verfall == "kuendigung":
        verfall_monate = None
    else:
        try:
            verfall_monate = int(verfall)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400,
                                detail="verfall muss 'kuendigung' oder eine Monats-Zahl sein")
        if not 1 <= verfall_monate <= 120:
            raise HTTPException(status_code=400, detail="verfall (Monate) muss zwischen 1 und 120 liegen")
    quelle = data.get("quelle", "rechnung")
    if quelle not in ("rechnung", "admin"):
        raise HTTPException(status_code=400, detail="quelle muss 'rechnung' oder 'admin' sein")
    notiz = (data.get("notiz") or "").strip()[:500]
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User nicht gefunden")
    # Review-Befund K1 (06.08.): Das Paket gehoert IMMER genau dem Konto, das
    # der Admin angegeben hat — NICHT dem fluechtigen Arbeits-Kontext
    # (aktiver_topf ist jederzeit umschaltbar; die Gutschrift eines Kaeufers
    # darf nie beim fremden Team-Inhaber landen, nur weil der Kaeufer gerade
    # dort arbeitet). Team-Pakete legt der Admin direkt auf die E-Mail des
    # Team-Inhabers.
    konto_id = target["id"]
    # Review-Befund K4: Kuendigungs-Pakete (verfaellt_am NULL) sind nur bei
    # aktivem Bezahl-Plan nutzbar — fuer ein effektiv freies Konto waeren die
    # Credits unsichtbar. Kulanz/Free bekommt automatisch ein Datums-Paket.
    hinweis = ""
    if verfall_monate is None and billing.effektiver_plan(target) == "free":
        verfall_monate = 12
        hinweis = " (Konto ohne aktives Abo: Paket verfaellt in 12 Monaten statt bei Kuendigung)"
    paket_id = billing.schenke_credits(konto_id, groesse, notiz=notiz,
                                       quelle=quelle, verfall_monate=verfall_monate)
    preis = billing.PAKET_PREISE.get(groesse)
    return {"ok": True, "paket_id": paket_id, "konto_user_id": konto_id,
            "groesse": groesse, "preis_eur": preis,
            "verfall": "kuendigung" if verfall_monate is None else f"{verfall_monate} Monate",
            "message": f"Paket ({groesse} Credits) fuer {target['email']} angelegt{hinweis}"}


@app.post("/api/admin/users/{user_id}/api-limit")
async def admin_setze_api_limit(user_id: int, request: Request,
                                user: dict = Depends(require_full_admin)):
    """Admin: API-Tageslimit (Missbrauchsbremse) pro Konto setzen.

    Steve 11.08.2026: Das starre globale DAILY_IMAGE_LIMIT bremste den
    InkluEdit-Einzug aus (alle Schluessel unter einem Konto). limit =
    ganze Zahl 0..1000000 (0 sperrt die API-Generierung), null/leer =
    zurueck auf den globalen Standard. Das Abo-Guthaben gilt immer
    zusaetzlich.
    """
    data = await request.json()
    roh = data.get("limit", None)
    if roh in (None, ""):
        limit = None
    else:
        try:
            limit = int(roh)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400,
                                detail="limit muss eine ganze Zahl oder leer sein")
        if not 0 <= limit <= 1000000:
            raise HTTPException(status_code=400,
                                detail="limit muss zwischen 0 und 1000000 liegen")
    ziel = get_user_by_id(user_id)
    if not ziel:
        raise HTTPException(status_code=404, detail="User nicht gefunden")
    conn = get_db()
    try:
        conn.execute("UPDATE users SET api_tageslimit = ? WHERE id = ?",
                     (limit, user_id))
        conn.commit()
    finally:
        conn.close()
    if limit is None:
        text = f"Standard ({DAILY_IMAGE_LIMIT} Bilder pro Tag)"
    else:
        text = f"{limit} Bilder pro Tag"
    return {"ok": True, "api_tageslimit": limit,
            "message": f"API-Tageslimit fuer {ziel['email']}: {text}"}


# ─── API Key Management ─────────────────────────────────────

@app.get("/api/api-keys")
async def api_list_keys(user: dict = Depends(get_current_user)):
    keys = list_api_keys(user["id"])
    return {"api_keys": keys}


@app.post("/api/api-keys")
async def api_create_key(request: Request, user: dict = Depends(get_current_user)):
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Bitte einen Namen fuer den API-Key eingeben")
    key_id, raw_key = create_api_key(user["id"], name)
    return {
        "ok": True,
        "key_id": key_id,
        "api_key": raw_key,
        "hinweis": "Dieser API-Key wird nur einmal angezeigt. Bitte sicher speichern.",
    }


@app.put("/api/api-keys/{key_id}")
async def api_rename_key(key_id: int, request: Request, user: dict = Depends(get_current_user)):
    data = await request.json()
    new_name = data.get("name", "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Bitte einen Namen eingeben")
    if not rename_api_key(user["id"], key_id, new_name):
        raise HTTPException(status_code=404, detail="API-Key nicht gefunden")
    return {"ok": True}


@app.delete("/api/api-keys/{key_id}")
async def api_delete_key(key_id: int, user: dict = Depends(get_current_user)):
    if not delete_api_key(user["id"], key_id):
        raise HTTPException(status_code=404, detail="API-Key nicht gefunden")
    return {"ok": True}


# ─── Project Routes ──────────────────────────────────────────

@app.get("/api/projects")
async def list_projects(user: dict = Depends(get_current_user)):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM projects WHERE user_id = ? ORDER BY created_at DESC", (user["id"],)
    ).fetchall()
    conn.close()
    return {"projects": [dict(r) for r in rows]}


@app.post("/api/projects")
async def create_project(request: Request, user: dict = Depends(get_current_user)):
    """Legt ein leeres Projekt an (Name + Werkzeug). Die Datei(en) werden
    anschliessend im jeweiligen Werkzeug hochgeladen (siehe Upload-Endpunkte)."""
    data = await request.json()
    name = (data.get("name") or "").strip()
    tool = (data.get("tool") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Bitte einen Projektnamen eingeben")
    if len(name) > 120:
        raise HTTPException(status_code=400, detail="Projektname darf maximal 120 Zeichen lang sein")
    if not is_valid_tool_key(tool):
        raise HTTPException(status_code=400, detail="Unbekanntes oder nicht verfuegbares Werkzeug")
    # project_type ist der Datentyp der Quelle (Anzeige-/Verzweigungslogik im Frontend);
    # "docx" = Word-Werkzeug (26.08.2026), "pdfform" = Quickinfo-Werkzeug (27.08.2026).
    project_type = {"pdf": "pdf", "web": "url", "grafik": "images", "word": "docx", "formular": "pdfform"}.get(tool)
    if not project_type:
        raise HTTPException(status_code=400, detail="Fuer dieses Werkzeug ist das Anlegen noch nicht verfuegbar")
    conn = get_db()
    cursor = conn.execute(
        "INSERT INTO projects (user_id, name, filename, original_path, status, project_type, tool) VALUES (?, ?, ?, '', 'neu', ?, ?)",
        (user["id"], name, name, project_type, tool)
    )
    project_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return {"ok": True, "project_id": project_id, "tool": tool, "project_type": project_type}


@app.patch("/api/projects/{project_id}")
async def rename_project(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Benennt ein Projekt um (freier Anzeigename)."""
    data = await request.json()
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Bitte einen Projektnamen eingeben")
    if len(name) > 120:
        raise HTTPException(status_code=400, detail="Projektname darf maximal 120 Zeichen lang sein")
    conn = get_db()
    cur = conn.execute("UPDATE projects SET name = ? WHERE id = ? AND user_id = ?", (name, project_id, user["id"]))
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {"ok": True, "name": name}


@app.post("/api/projects/{project_id}/context-setting")
async def set_context_setting(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Setzt den Projekt-Schalter 'KI-Kontext verwenden' (19.06.2026).

    use_context=1 -> kommende Generierungen geben der KI den Dokument-Kontext
    (Seiten-/Kapiteltext); =0 -> die KI bekommt keinen Kontext. Wirkt nur auf
    NEU generierte Bilder; bestehende bleiben, bis man sie neu generiert.
    """
    data = await request.json()
    use_ctx = 1 if data.get("use_context") else 0
    conn = get_db()
    cur = conn.execute(
        "UPDATE projects SET use_context = ? WHERE id = ? AND user_id = ?",
        (use_ctx, project_id, user["id"])
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {"ok": True, "use_context": bool(use_ctx)}


ALT_TEXT_LANGUAGES = ("de", "en", "da", "fr", "es", "sv")


# ── Konto-Mails in Empfaengersprache (03.07.2026) ──────────────────────────
def _mail_lang(user_row) -> str:
    """Empfaengersprache einer Mail: users.language mit Deutsch-Fallback."""
    try:
        lang = user_row["language"] if user_row else None
    except (KeyError, IndexError, TypeError):
        lang = None
    return lang if lang in SUPPORTED_LANGUAGES else "de"


_MAIL_P = '<p style="margin:0 0 14px;line-height:1.55;">'
_MAIL_KLEIN = '<p style="color:#64748b;font-size:14px;margin:0 0 14px;">'


def _mail_knopf(url: str, label: str) -> str:
    """Handlungs-Link im Knopf-Look — Kontrastfarbe #c75000 (4,5:1 mit
    weisser Schrift, app-weiter Farbwert seit 08.08.2026)."""
    return (f'<p style="margin:0 0 14px;"><a href="{url}" '
            'style="display:inline-block;background-color:#c75000;color:#ffffff;'
            'padding:12px 24px;border-radius:6px;text-decoration:none;'
            f'font-weight:bold;">{label}</a></p>')


def _auth_mail_rahmen(lang: str, ueberschrift: str, inhalt_html: str) -> str:
    """Karten-Rahmen fuer Konto-Mails — gleiche Optik wie _abo_mail
    (Steve, 24.08.2026): Markenkopf, weisse Karte, Fusszeile mit
    support@inkludocs.de. Alle Stile inline (Mail-Clients)."""
    _ = get_gettext(lang)
    fragen = _('Bei Fragen wende dich gerne an {mail}.').format(
        mail='<a href="mailto:support@inkludocs.de" style="color:#1b2a4a;">support@inkludocs.de</a>')
    return (
        f'<!DOCTYPE html><html lang="{lang}"><head><meta charset="utf-8"></head>'
        '<body style="margin:0;padding:0;background-color:#f1f5f9;">'
        '<div style="max-width:600px;margin:0 auto;padding:24px 16px;'
        'font-family:Arial,Helvetica,sans-serif;color:#1e293b;font-size:16px;">'
        '<p style="margin:0 0 18px;font-size:22px;font-weight:bold;color:#1b2a4a;">'
        'Inklu<span style="font-weight:normal;">Docs</span></p>'
        '<div style="background-color:#ffffff;border-radius:8px;padding:28px 24px;'
        'border:1px solid #e2e8f0;">'
        f'<h1 style="color:#1b2a4a;font-size:20px;margin:0 0 18px;">{ueberschrift}</h1>'
        f'{inhalt_html}'
        '</div>'
        '<div style="padding:18px 8px 0;font-size:14px;color:#475569;">'
        f'<p style="margin:0 0 6px;">{fragen}</p>'
        '<p style="color:#64748b;font-size:13px;margin:18px 0 0;">'
        f"{_('InkluDocs ist ein Produkt von INKLUTEC – support@inkludocs.de')}</p>"
        '</div></div></body></html>')


def _verification_mail_full(lang: str, display_name: str, verify_url: str):
    """Willkommens-/Bestaetigungsmail bei der Registrierung.

    Steve 24.08.2026: durchgehend Du-Form (vorher siezte diese Mail als
    einzige), verallgemeinerter Claim (nicht nur Alt-Texte), kein
    Meta-Gerede ueber Knoepfe, neuer Karten-Rahmen."""
    _ = get_gettext(lang)
    subject = _('InkluDocs: E-Mail-Adresse bestätigen')
    anmelden = _('Melde dich auf {url} an.').format(url=f'<a href="{BASE_URL}">{BASE_URL}</a>')
    inhalt = (
        f'{_MAIL_P}{_("Hallo")} {display_name},</p>'
        f'{_MAIL_P}{_("danke für deine Registrierung bei InkluDocs – deinem KI-Werkzeug für barrierefreie Dokumente und Bilder.")}</p>'
        f'{_MAIL_P}{_("Bitte bestätige deine E-Mail-Adresse, um dein Konto zu aktivieren:")}</p>'
        + _mail_knopf(verify_url, _('E-Mail-Adresse bestätigen'))
        + f'{_MAIL_KLEIN}{_("Oder kopiere diesen Link:")} {verify_url}</p>'
        + f'{_MAIL_KLEIN}{_("Der Link ist 24 Stunden gültig.")}</p>'
        + f'<h2 style="color:#1b2a4a;font-size:17px;margin:18px 0 10px;">{_("So funktioniert InkluDocs")}</h2>'
        + f'{_MAIL_P}{_("Nachdem du deine E-Mail-Adresse bestätigt hast:")}</p>'
        + '<ol style="line-height:1.7;margin:0 0 4px;padding-left:22px;">'
        + f'<li>{anmelden}</li>'
        + f'<li>{_("Lade ein PDF oder Bilder hoch oder gib eine Website-Adresse ein.")}</li>'
        + f'<li>{_("Klicke auf „Alt-Texte generieren“.")}</li>'
        + f'<li>{_("Bearbeite die Alt-Texte bei Bedarf und exportiere sie.")}</li></ol>')
    return subject, _auth_mail_rahmen(lang, _('Willkommen bei InkluDocs'), inhalt)


def _verification_mail_short(lang: str, display_name: str, verify_url: str):
    """Neuer Bestaetigungslink (Resend)."""
    _ = get_gettext(lang)
    subject = _('InkluDocs: E-Mail-Adresse bestätigen')
    inhalt = (
        f'{_MAIL_P}{_("Hallo")} {display_name},</p>'
        f'{_MAIL_P}{_("hier ist dein neuer Bestätigungslink:")}</p>'
        + _mail_knopf(verify_url, _('E-Mail-Adresse bestätigen'))
        + f'{_MAIL_KLEIN}{_("Oder kopiere diesen Link:")} {verify_url}</p>'
        + f'{_MAIL_KLEIN}{_("Der Link ist 24 Stunden gültig.")}</p>')
    return subject, _auth_mail_rahmen(lang, _('E-Mail-Adresse bestätigen'), inhalt)


def _email_change_mail(lang: str, display_name: str, new_email: str, confirm_url: str):
    """Bestaetigung einer E-Mail-Adress-Aenderung."""
    _ = get_gettext(lang)
    subject = _('InkluDocs: E-Mail-Adresse bestätigen')
    aendern = _('du hast angefordert, deine InkluDocs E-Mail-Adresse auf {email} zu ändern.').format(
        email=f'<strong>{new_email}</strong>')
    inhalt = (
        f'{_MAIL_P}{_("Hallo")} {display_name},</p>'
        f'{_MAIL_P}{aendern}</p>'
        + _mail_knopf(confirm_url, _('E-Mail-Adresse bestätigen'))
        + f'{_MAIL_KLEIN}{_("Oder kopiere diesen Link:")} {confirm_url}</p>'
        + f'{_MAIL_KLEIN}{_("Der Link ist 1 Stunde gültig. Falls du diese Änderung nicht angefordert hast, ignoriere diese E-Mail.")}</p>')
    return subject, _auth_mail_rahmen(lang, _('E-Mail-Adresse bestätigen'), inhalt)


def _reset_mail(lang: str, display_name: str, reset_url: str):
    """Passwort-Reset-Mail."""
    _ = get_gettext(lang)
    subject = _('InkluDocs: Passwort zurücksetzen')
    inhalt = (
        f'{_MAIL_P}{_("Hallo")} {display_name},</p>'
        f'{_MAIL_P}{_("du hast eine Passwort-Zurücksetzung für dein InkluDocs-Konto angefordert.")}</p>'
        + _mail_knopf(reset_url, _('Passwort jetzt zurücksetzen'))
        + f'{_MAIL_KLEIN}{_("Oder kopiere diesen Link:")} {reset_url}</p>'
        + f'{_MAIL_KLEIN}{_("Der Link ist 1 Stunde gültig. Falls du diese Anfrage nicht gestellt hast, kannst du diese E-Mail ignorieren.")}</p>')
    return subject, _auth_mail_rahmen(lang, _('Passwort zurücksetzen'), inhalt)


def _auth_notice_page(lang: str, inner_html: str, status_code: int = 200,
                      title_suffix: str = "", head_extra: str = "") -> HTMLResponse:
    """Kleine Hinweis-Seite im Anmelde-Layout (Bestaetigungsseiten nach
    Mail-Klick, Registrierung geschlossen). Sprache aus dem Request."""
    title = "InkluDocs" + ((" – " + title_suffix) if title_suffix else "")
    return HTMLResponse(f"""<!DOCTYPE html><html lang="{lang}"><head><meta charset="utf-8"><title>{title}</title>
<link rel="stylesheet" href="/static/style.css">{head_extra}</head><body>
<div class="auth-wrapper"><div class="auth-container" role="main">
<h1><span class="brand">Inklu</span>Docs</h1>
{inner_html}
</div></div></body></html>""", status_code=status_code)


@app.post("/api/projects/{project_id}/language-setting")
async def set_language_setting(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Setzt die Ausgabesprache der Alt-Texte (03.07.2026, alirodocs).

    Lebende Projekt-Einstellung wie use_context: gilt fuer alles, was AB JETZT
    generiert wird (Sammellauf und Einzel-Neu-Generieren). Gemischte Sprachen
    im selben Projekt sind gewollt (Steve 03.07.). Unabhaengig von der
    UI-Sprache (users.language).
    """
    data = await request.json()
    lang = str(data.get("language", "de")).lower()
    if lang not in ALT_TEXT_LANGUAGES:
        raise HTTPException(status_code=400, detail="Unbekannte Sprache")
    conn = get_db()
    cur = conn.execute(
        "UPDATE projects SET alt_language = ? WHERE id = ? AND user_id = ?",
        (lang, project_id, user["id"])
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {"ok": True, "language": lang}


# ─── Eigene Prompts (Prompt-Verwaltung, 06.07.2026) ─────────────────────────
# Vom Nutzer gespeicherte Zusatz-Anweisungen fuer die Alt-Text-Generierung.
# Ein Projekt referenziert ueber projects.prompt_id genau einen aktiven Prompt
# (NULL = keiner). Der Text geht additiv als Suffix in die v4-Pipeline
# (_user_prompt_suffix im Orchestrator) — die Premium-Prompts bleiben unangetastet.

PROMPT_NAME_MAX = 100
PROMPT_CATEGORY_MAX = 60
PROMPT_DESCRIPTION_MAX = 300
PROMPT_TEXT_MAX = 4000


def _clean_prompt_payload(data: dict) -> dict:
    """Validiert die Prompt-Felder; wirft HTTPException bei Verstoessen."""
    name = str(data.get("name") or "").strip()
    category = str(data.get("category") or "").strip()
    description = str(data.get("description") or "").strip()
    prompt_text = str(data.get("prompt_text") or "").strip()
    if not name or not prompt_text:
        raise HTTPException(status_code=400, detail="Name und Prompt-Text sind Pflichtfelder")
    if len(name) > PROMPT_NAME_MAX:
        raise HTTPException(status_code=400, detail=f"Name darf maximal {PROMPT_NAME_MAX} Zeichen lang sein")
    if len(category) > PROMPT_CATEGORY_MAX:
        raise HTTPException(status_code=400, detail=f"Kategorie darf maximal {PROMPT_CATEGORY_MAX} Zeichen lang sein")
    if len(description) > PROMPT_DESCRIPTION_MAX:
        raise HTTPException(status_code=400, detail=f"Beschreibung darf maximal {PROMPT_DESCRIPTION_MAX} Zeichen lang sein")
    if len(prompt_text) > PROMPT_TEXT_MAX:
        raise HTTPException(status_code=400, detail=f"Prompt-Text darf maximal {PROMPT_TEXT_MAX} Zeichen lang sein")
    return {"name": name, "category": category, "description": description, "prompt_text": prompt_text}


@app.get("/api/prompts")
async def list_prompts(user: dict = Depends(get_current_user)):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, name, description, category, prompt_text, created_at, updated_at "
        "FROM user_prompts WHERE user_id = ? ORDER BY LOWER(category), LOWER(name)",
        (user["id"],)
    ).fetchall()
    conn.close()
    return {"prompts": [dict(r) for r in rows]}


@app.post("/api/prompts")
async def create_prompt(request: Request, user: dict = Depends(get_current_user)):
    fields = _clean_prompt_payload(await request.json())
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO user_prompts (user_id, name, description, category, prompt_text) "
        "VALUES (?, ?, ?, ?, ?)",
        (user["id"], fields["name"], fields["description"], fields["category"], fields["prompt_text"])
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return {"ok": True, "id": new_id}


@app.patch("/api/prompts/{prompt_id}")
async def update_prompt(prompt_id: int, request: Request, user: dict = Depends(get_current_user)):
    fields = _clean_prompt_payload(await request.json())
    conn = get_db()
    cur = conn.execute(
        "UPDATE user_prompts SET name = ?, description = ?, category = ?, prompt_text = ?, "
        "updated_at = datetime('now') WHERE id = ? AND user_id = ?",
        (fields["name"], fields["description"], fields["category"], fields["prompt_text"], prompt_id, user["id"])
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Prompt nicht gefunden")
    return {"ok": True}


@app.delete("/api/prompts/{prompt_id}")
async def delete_prompt(prompt_id: int, user: dict = Depends(get_current_user)):
    conn = get_db()
    # Projekte, die diesen Prompt nutzen, fallen auf "kein eigener Prompt" zurueck.
    conn.execute(
        "UPDATE projects SET prompt_id = NULL WHERE prompt_id = ? AND user_id = ?",
        (prompt_id, user["id"])
    )
    cur = conn.execute(
        "DELETE FROM user_prompts WHERE id = ? AND user_id = ?",
        (prompt_id, user["id"])
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Prompt nicht gefunden")
    return {"ok": True}


@app.post("/api/projects/{project_id}/prompt-setting")
async def set_prompt_setting(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Setzt den aktiven eigenen Prompt eines Projekts (06.07.2026).

    Lebende Projekt-Einstellung wie die Ausgabesprache: gilt fuer alles, was
    AB JETZT generiert wird (Sammellauf, Einzel-Neu-Generieren, InkluAgent).
    prompt_id null/leer = kein eigener Prompt (Standardverhalten).
    """
    data = await request.json()
    raw = data.get("prompt_id")
    prompt_id = None
    if raw not in (None, "", 0, "0"):
        try:
            prompt_id = int(raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Ungueltige prompt_id")
    conn = get_db()
    if prompt_id is not None:
        owned = conn.execute(
            "SELECT 1 FROM user_prompts WHERE id = ? AND user_id = ?",
            (prompt_id, user["id"])
        ).fetchone()
        if not owned:
            conn.close()
            raise HTTPException(status_code=404, detail="Prompt nicht gefunden")
    cur = conn.execute(
        "UPDATE projects SET prompt_id = ? WHERE id = ? AND user_id = ?",
        (prompt_id, project_id, user["id"])
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {"ok": True, "prompt_id": prompt_id}


@app.get("/api/tools")
async def list_tools(user: dict = Depends(get_current_user)):
    """Liefert die im Dashboard angebotenen Werkzeuge (Quelle: tools.py)."""
    return {
        "tools": [
            {
                "key": t.key,
                "name": t.name,
                "description": t.description,
                "route": t.route,
                "status": t.status.value,
                "status_label": t.status_label,
                "is_available": t.is_available,
            }
            for t in TOOLS
        ]
    }


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), project_id: int = Form(None), user: dict = Depends(get_current_user)):
    """Upload a PDF or image file(s). Accepts PDF, JPG, JPEG, PNG, GIF, SVG, WEBP."""
    filename = file.filename or "unknown"
    ext = os.path.splitext(filename)[1].lower()

    is_pdf = ext == ".pdf"
    is_docx = ext == ".docx"       # Word-Werkzeug (26.08.2026)
    is_image = ext in IMAGE_EXTENSIONS

    # Word-Altformat und Makro-Dateien bewusst mit eigener Meldung abweisen —
    # .doc kann das Werkzeug nicht lesen (Binaerformat), .docm enthaelt Makros,
    # die wir weder pruefen noch weitergeben wollen.
    if ext == ".doc":
        raise HTTPException(status_code=400, detail="Das alte Word-Format .doc wird nicht unterstützt. Bitte in Word über „Speichern unter“ als .docx speichern.")
    if ext in (".docm", ".dotm", ".dotx"):
        raise HTTPException(status_code=400, detail="Word-Vorlagen und Dateien mit Makros (.docm, .dotm, .dotx) werden nicht verarbeitet. Bitte als normale .docx-Datei speichern.")
    if not is_pdf and not is_docx and not is_image:
        raise HTTPException(
            status_code=400,
            detail="Nur PDF-, Word- und Bilddateien erlaubt (PDF, DOCX, JPG, PNG, GIF, SVG, WebP, HEIC, BMP, TIFF)"
        )

    # Read and check file size
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Datei zu gross. Maximum: {MAX_UPLOAD_SIZE // (1024*1024)} MB"
        )

    # SVG sofort nach PNG wandeln (PIL kann kein SVG). Gespeichert und
    # verarbeitet wird nur das PNG; der Original-Dateiname (foo.svg) bleibt
    # als Anzeige-Name des Projekts/Bilds erhalten.
    if ext == ".svg":
        try:
            content = _svg_to_png_bytes(content)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"SVG konnte nicht verarbeitet werden: {e}")
        ext = ".png"

    # Create user directory
    user_dir = os.path.join(UPLOAD_DIR, str(user["id"]))
    os.makedirs(user_dir, exist_ok=True)

    # Save file with sanitized name
    # Endung aus `ext` (nicht aus dem Original-Namen): nach SVG-Konvertierung
    # traegt die gespeicherte Datei .png — sonst laege PNG-Inhalt unter .svg.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(filename))[0]
    safe_name = f"{timestamp}_{base_name}{ext}"
    file_path = os.path.join(user_dir, safe_name)
    with open(file_path, "wb") as f:
        f.write(content)

    if project_id is not None and not is_pdf and formular_api.ist_formular_projekt(project_id, user["id"]):
        # Formular-Projekte nehmen nur PDF an — abweisen, bevor eine Waise auf der Platte bleibt.
        try:
            os.unlink(file_path)
        except OSError:
            pass
        raise HTTPException(status_code=400, detail="In ein Formular-Projekt können nur PDF-Formulare hochgeladen werden.")
    if is_pdf and project_id is not None and formular_api.ist_formular_projekt(project_id, user["id"]):
        # QUICKINFO-WERKZEUG (27.08.2026): PDF in einem Formular-Projekt geht den
        # Formularweg (Felder statt Bilder). Vorpruefung im Request, damit der
        # Nutzer eine klare Meldung bekommt (kein Formular, Passwort, zu gross).
        try:
            validiere_formular(file_path)
        except FormularFehler as e:
            try:
                os.unlink(file_path)
            except OSError:
                pass
            raise HTTPException(status_code=400, detail=str(e))
        return await formular_api.handle_upload(file_path, filename, user, project_id)
    if is_pdf:
        return await _handle_pdf_upload(file_path, filename, user, project_id)
    elif is_docx:
        # Vorpruefung im Request (Zip, Grenzen, echtes DOCX ohne Makros), damit der
        # Nutzer eine klare Meldung bekommt statt eines still verschwundenen Dokuments.
        try:
            validiere_docx(file_path)
        except DocxFehler as e:
            try:
                os.unlink(file_path)
            except OSError:
                pass
            raise HTTPException(status_code=400, detail=str(e))
        return await _handle_pdf_upload(file_path, filename, user, project_id, art="docx")
    else:
        return await _handle_image_upload(file_path, filename, user, content, ext, project_id)


async def _handle_pdf_upload(file_path: str, filename: str, user: dict, project_id: int = None,
                             art: str = "pdf") -> dict:
    """Legt ein neues PDF-Projekt an ODER haengt eine weitere PDF an ein
    bestehendes PDF-Projekt an.

    art="docx" (Word-Werkzeug, 26.08.2026): identischer Ablauf fuer Word-
    Dateien — Werkzeug "word", project_type "docx", extraction_method "docx",
    Extraktor docx_processor. Ein Dokument = eine Word-Datei, Multi-Datei
    wie bei PDF. Bewusst KEINE Kopie dieser Funktion, damit Aenderungen am
    Upload-Ablauf beide Formate treffen.

    Multi-Datei Phase 1 (08.06.2026): Ein PDF-Projekt kann mehrere PDFs
    enthalten. Jede hochgeladene PDF wird zu einem Dokument
    (Tabelle documents). Bilder bekommen die document_id und ihre
    page_number bleibt die PDF-interne Seitennummer — Kollisionen zwischen
    Dokumenten loest die document_id auf. Sammel-Generierung verarbeitet
    weiterhin nur status='pending' (siehe _process_project), ueberschreibt
    also bereits fertige Alt-Texte der vorherigen Dokumente NICHT.
    """
    conn = get_db()
    is_append = False  # zeigt an, ob wir an ein bestehendes PDF-Projekt anhaengen
    if project_id is not None:
        proj = conn.execute(
            "SELECT id, tool, total_images FROM projects WHERE id = ? AND user_id = ?",
            (project_id, user["id"])
        ).fetchone()
        if not proj:
            conn.close()
            raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
        erwartet_tool = "word" if art == "docx" else "pdf"
        if proj["tool"] != erwartet_tool:
            conn.close()
            raise HTTPException(status_code=400, detail=("Dieses Projekt ist kein Word-Projekt" if art == "docx" else "Dieses Projekt ist kein PDF-Projekt"))
        # Multi-Datei: Anhaengen ist ab jetzt erlaubt — kein 409 mehr bei total_images > 0.
        is_append = bool(proj["total_images"] and proj["total_images"] > 0)
        # status auf extracting setzen; filename/original_path nur beim ERSTEN
        # PDF setzen, damit die "Hauptdatei" des Projekts (z.B. fuer den Projektkopf)
        # erhalten bleibt. Backfill setzt das ohnehin auf das erste Dokument.
        if is_append:
            conn.execute("UPDATE projects SET status = 'extracting' WHERE id = ?", (project_id,))
        else:
            conn.execute(
                "UPDATE projects SET filename = ?, original_path = ?, status = 'extracting' WHERE id = ?",
                (filename, file_path, project_id)
            )
        conn.commit()
    else:
        cursor = conn.execute(
            "INSERT INTO projects (user_id, filename, original_path, status, project_type, tool) VALUES (?, ?, ?, 'extracting', ?, ?)",
            (user["id"], filename, file_path, "docx" if art == "docx" else "pdf", "word" if art == "docx" else "pdf")
        )
        project_id = cursor.lastrowid
        conn.commit()

    # Naechsten doc_index ermitteln (1-basiert, fortlaufend pro Projekt).
    doc_index = (conn.execute(
        "SELECT COALESCE(MAX(doc_index), 0) FROM documents WHERE project_id = ?",
        (project_id,)
    ).fetchone()[0] or 0) + 1

    # Dokument-Eintrag anlegen (extraction_method wird unten nachgetragen).
    doc_cursor = conn.execute(
        """INSERT INTO documents
           (project_id, doc_index, original_filename, original_path, extraction_method, total_images)
           VALUES (?, ?, ?, ?, ?, 0)""",
        (project_id, doc_index, filename, file_path, "docx" if art == "docx" else "fitz")
    )
    document_id = doc_cursor.lastrowid
    conn.commit()

    conn.close()

    # ASYNCHRON SEIT 04.07.2026 (Julia-Befund strukturell geloest): Die
    # Extraktion (fitz/PDFix, Seitenansichten, Seitentext) lief frueher
    # synchron in diesem Request — bei grossen PDFs laenger als jedes
    # Proxy-Timeout, der Browser sah "Netzwerkfehler", obwohl der Server
    # fertig verarbeitete. Jetzt antwortet der Endpunkt sofort mit Status
    # 'extracting'; die Extraktion laeuft als Hintergrund-Task (Muster
    # _process_project), und das Frontend pollt GET /api/projects/{id},
    # bis der Status wechselt. total_images/added_images stehen deshalb
    # NICHT mehr in dieser Antwort — sie entstehen erst nach der Extraktion.
    asyncio.create_task(_extract_document(
        project_id, document_id, doc_index, file_path, user["id"], is_append, art=art))

    return {
        "ok": True,
        "project_id": project_id,
        "document_id": document_id,
        "doc_index": doc_index,
        "filename": filename,
        "project_type": "docx" if art == "docx" else "pdf",
        "appended": is_append,
        "status": "extracting",
    }


async def _extract_document(project_id: int, document_id: int, doc_index: int,
                            file_path: str, user_id: int, is_append: bool,
                            art: str = "pdf"):
    """Hintergrund-Extraktion einer hochgeladenen PDF (ausgelagert 04.07.2026).

    Der frueher synchrone Teil von _handle_pdf_upload, unveraendert in der
    Logik: Bilder extrahieren, images-Zeilen anlegen, Zaehler und Status
    fortschreiben. fitz/PDFix blockieren -> Executor-Thread, damit der
    Event-Loop (und damit die ganze App) waehrenddessen bedienbar bleibt.

    Fehlerpfad: documents-Zeile entfernen (das Verschwinden ist das Signal
    fuer das pollende Frontend) und den Projektstatus zuruecksetzen —
    'error' beim ersten Dokument, 'extracted' beim Anhaengen (die alten
    Dokumente sind unberuehrt; frueher blieb das Projekt hier faelschlich
    auf 'extracting' haengen, weil nur der HTTP-Fehler die Info trug).
    """
    img_dir = os.path.join(RESULTS_DIR, str(user_id), str(project_id), f"doc{doc_index}")
    os.makedirs(img_dir, exist_ok=True)
    loop = asyncio.get_event_loop()
    try:
        hinweise: dict = {}
        if art == "docx":
            # 27.08.2026: uebersprungene Elemente + Warnungen kommen als Hinweise mit
            images, hinweise = await loop.run_in_executor(
                None, extract_docx, file_path, img_dir, project_id)
        else:
            images = await loop.run_in_executor(
                None, extract_images_from_pdf, file_path, img_dir, project_id)
    except Exception as e:
        print(f"[upload] Extraktion fehlgeschlagen (Projekt {project_id}, Dokument {document_id}): {e}")
        conn = get_db()
        conn.execute("DELETE FROM documents WHERE id = ?", (document_id,))
        if is_append:
            conn.execute("UPDATE projects SET status = 'extracted' WHERE id = ?", (project_id,))
        else:
            conn.execute("UPDATE projects SET status = 'error' WHERE id = ?", (project_id,))
        conn.commit()
        conn.close()
        return

    conn = get_db()
    # Naechsten image_index fortlaufen lassen, damit alte Projekte weiter
    # konsistent durchnummeriert bleiben. Das pro-Dokument-Nummer-Feld leistet
    # die Anzeige im Frontend (dort einfach Position innerhalb des Dokuments).
    next_idx = conn.execute(
        "SELECT COALESCE(MAX(image_index), 0) FROM images WHERE project_id = ?", (project_id,)
    ).fetchone()[0] or 0

    for img in images:
        bbox = img.get("bbox")
        bbox_x0 = bbox[0] if bbox else None
        bbox_y0 = bbox[1] if bbox else None
        bbox_x1 = bbox[2] if bbox else None
        bbox_y1 = bbox[3] if bbox else None
        is_vector = 1 if img.get("is_vector") else 0
        next_idx += 1

        conn.execute(
            """INSERT INTO images (project_id, document_id, page_number, image_index, image_path, context_text,
               width, height, xref, bbox_x0, bbox_y0, bbox_x1, bbox_y1, is_vector,
               original_alt, page_view_path, page_text, docx_anker)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (project_id, document_id, img["page_number"], next_idx, img["image_path"],
             img["context_text"], img["width"], img["height"], img.get("xref"),
             bbox_x0, bbox_y0, bbox_x1, bbox_y1, is_vector,
             # Word: ein vom Autor gesetztes Dekorativ-Kennzeichen wird als
             # original_alt "dekorativ" uebernommen — so bleibt es beim Export
             # erhalten, auch wenn der Nutzer das Bild nie generieren laesst.
             ("dekorativ" if img.get("decorative_hint") else img.get("original_alt", "")),
             img.get("page_view_path", ""), img.get("page_text", ""),
             img.get("docx_anker", ""))
        )

    # PDFIX-INTEGRATION (24.04.2026): Extraktionsweg merken (fitz|pdfix)
    if art == "docx":
        # 27.08.2026 (Steve: "wenn moeglich dieselben Seiten wie im Dokument"):
        # "docx-seiten", wenn das Dokument Word-Seitenmarken traegt (page_number =
        # Seite wie zuletzt in Word gezeigt), sonst "docx" (page_number = Abschnitt).
        extraction_method = "docx-seiten" if images and images[0].get("docx_einheit") == "seite" else "docx"
    else:
        extraction_method = "pdfix" if images and any(i.get("source") == "pdfix" for i in images) else "fitz"
    conn.execute(
        "UPDATE documents SET extraction_method = ?, total_images = ?, hinweise = ? WHERE id = ?",
        (extraction_method, len(images),
         json.dumps(hinweise, ensure_ascii=False) if hinweise and (hinweise.get("uebersprungen") or hinweise.get("warnungen")) else "",
         document_id)
    )
    # Projekt-Summe = Summe ueber alle Dokumente. extraction_method des Projekts
    # bleibt der des ersten Dokuments (Anzeige im Kopf) — bei Mischfaellen sind
    # die Methoden je Dokument im UI sichtbar.
    total_imgs = conn.execute(
        "SELECT COUNT(*) FROM images WHERE project_id = ?", (project_id,)
    ).fetchone()[0]
    if is_append:
        conn.execute(
            "UPDATE projects SET status = 'extracted', total_images = ? WHERE id = ?",
            (total_imgs, project_id)
        )
    else:
        conn.execute(
            "UPDATE projects SET status = 'extracted', total_images = ?, extraction_method = ? WHERE id = ?",
            (total_imgs, extraction_method, project_id)
        )
    conn.commit()
    conn.close()


async def _handle_image_upload(file_path: str, filename: str, user: dict, content: bytes, ext: str, project_id: int = None) -> dict:
    """Process a direct image upload."""
    from PIL import Image as PILImage
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
    except ImportError:
        pass

    conn = get_db()
    if project_id is not None:
        proj = conn.execute(
            "SELECT id, tool FROM projects WHERE id = ? AND user_id = ?",
            (project_id, user["id"])
        ).fetchone()
        if not proj:
            conn.close()
            raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
        if proj["tool"] != "grafik":
            conn.close()
            raise HTTPException(status_code=400, detail="Dieses Projekt ist kein Grafik-Projekt")
    else:
        cursor = conn.execute(
            "INSERT INTO projects (user_id, filename, original_path, status, project_type, tool) VALUES (?, ?, ?, 'extracted', 'images', 'grafik')",
            (user["id"], filename, file_path)
        )
        project_id = cursor.lastrowid
        conn.commit()
    # Fortlaufende Bild-Nummer (auch beim Anhaengen weiterer Bilder)
    next_idx = conn.execute("SELECT COALESCE(MAX(image_index), 0) FROM images WHERE project_id = ?", (project_id,)).fetchone()[0] + 1

    # Create project image directory
    img_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project_id))
    os.makedirs(img_dir, exist_ok=True)

    # Convert HEIC/HEIF/BMP/TIFF to JPEG for model compatibility
    needs_conversion = ext in {".heic", ".heif", ".bmp", ".tiff", ".tif"}
    if needs_conversion:
        img_path = os.path.join(img_dir, f"img_{next_idx}.jpg")
        try:
            with PILImage.open(file_path) as img:
                rgb_img = img.convert("RGB")
                rgb_img.save(img_path, format="JPEG", quality=90)
        except Exception as e:
            conn.execute("UPDATE projects SET status = 'error' WHERE id = ?", (project_id,))
            conn.commit()
            conn.close()
            raise HTTPException(status_code=400, detail=f"Bildformat konnte nicht konvertiert werden: {str(e)}")
    else:
        img_path = os.path.join(img_dir, f"img_{next_idx}{ext}")
        shutil.copy2(file_path, img_path)

    # Get image dimensions
    width, height = 0, 0
    try:
        with PILImage.open(img_path) as img:
            width, height = img.size
    except Exception:
        pass

    # Multi-Datei Phase 2 (14.08.2026): Original-Dateiname pro Bild speichern —
    # die Grafik-Ansicht zeigt "Bild 1: urlaubsfoto.jpg" (Michael-Wunsch 10.06.).
    conn.execute(
        """INSERT INTO images (project_id, page_number, image_index, image_path, context_text,
           width, height, xref, original_filename)
           VALUES (?, 1, ?, ?, '', ?, ?, 0, ?)""",
        (project_id, next_idx, img_path, width, height, (filename or "")[:200])
    )
    total = conn.execute("SELECT COUNT(*) FROM images WHERE project_id = ?", (project_id,)).fetchone()[0]
    conn.execute(
        "UPDATE projects SET total_images = ?, status = 'extracted' WHERE id = ?",
        (total, project_id)
    )
    conn.commit()
    conn.close()

    return {
        "ok": True,
        "project_id": project_id,
        "filename": filename,
        "total_images": total,
        "project_type": "images",
    }


# ─────────────────────────────────────────────────────────────────────────────
# DEMO-MODUS — oeffentliche Kostprobe ohne Anmeldung (nur aktiv bei DEMO_MODE=on)
# ─────────────────────────────────────────────────────────────────────────────
# Diese Endpunkte sind funktional nur in der Demo-Instanz vorhanden (eigener
# Container + eigene Wegwerf-DB, siehe DEMO.md). Bei DEMO_MODE=off liefert der
# Gate-Check 404 — in Production/Staging sind sie damit unsichtbar/inert.
# Sitzungs- und Aufraeum-Logik steckt im Modul demo.py; hier nur die HTTP-Schicht,
# die bewusst die BESTEHENDEN Bausteine (_handle_image_upload, _process_project)
# wiederverwendet — keine eigene Generierungslogik. Limits + Bot-Leitplanke: Schritt 3.

def _require_demo():
    """Schaltet die Demo-Endpunkte ab, wenn die Instanz nicht im Demo-Modus laeuft."""
    if not demo_mod.demo_enabled():
        raise HTTPException(status_code=404, detail="Not found")


def _demo_user_ctx(user_id: int) -> dict:
    """Minimaler Nutzer-Kontext fuer die wiederverwendeten Helfer (brauchen nur id + is_admin)."""
    return {"id": user_id, "is_admin": False}


@app.post("/api/demo/generate")
async def demo_generate(request: Request, file: UploadFile = File(...),
                        language: Optional[str] = Form(None)):
    """Anonymer Einzelbild-Upload fuer die Demo. Legt bei Bedarf eine fluechtige
    Demo-Sitzung an (signiertes Cookie), speichert das Bild, startet die normale
    Generierungs-Pipeline im Hintergrund und gibt die Projekt-ID zurueck. Das
    Ergebnis wird ueber GET /api/demo/result abgeholt (Polling). Limits: Schritt 3."""
    _require_demo()
    # Sichtbare Fehlertexte in der Sprache des Besuchers (die Demo ist
    # oeffentlich und mehrsprachig). ACHTUNG: _()-Strings in main.py werden
    # vom Bau-Zeit-Check NICHT erfasst — .po-Eintraege manuell pflegen.
    _ = get_gettext(detect_language(request))
    filename = file.filename or "upload"
    ext = os.path.splitext(filename)[1].lower()
    if ext not in IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=_('Bitte ein Bild hochladen (JPG, PNG, GIF, WebP …). PDFs gibt es im vollen Werkzeug nach Anmeldung.'),
        )
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail=_('Datei zu groß. Maximum: {mb} MB').format(mb=MAX_UPLOAD_SIZE // (1024*1024)))

    # SVG sofort nach PNG wandeln (PIL kann kein SVG; gleiche Behandlung wie
    # /api/upload). Fehlertext: bewusst die bestehende, bereits uebersetzte
    # Format-Meldung wiederverwenden — kein neuer msgid (siehe Hinweis oben).
    if ext == ".svg":
        try:
            content = _svg_to_png_bytes(content)
        except Exception:
            raise HTTPException(
                status_code=400,
                detail=_('Bitte ein Bild hochladen (JPG, PNG, GIF, WebP …). PDFs gibt es im vollen Werkzeug nach Anmeldung.'),
            )
        ext = ".png"

    # Limit pruefen (server-seitig, VOR jeder teuren Generierung — nie vom Bot)
    chk = demo_mod.check_generation(request)
    if not chk["allowed"]:
        if chk["global_reached"]:
            raise HTTPException(status_code=429, detail=_('Die Demo ist heute stark gefragt — das Tageskontingent ist erreicht. Bitte morgen wieder probieren oder gleich kostenlos registrieren.'))
        raise HTTPException(status_code=429, detail=_('Deine {n} kostenlosen Analysen für heute sind aufgebraucht. Für unbegrenzte Bilder, ganze PDFs und den fertig getaggten Export: kostenlos registrieren.').format(n=demo_mod.daily_image_limit()))

    # Sitzung aus Cookie ableiten oder neu anlegen
    uid = demo_mod.resolve_session_user_id(request.cookies.get(demo_mod.DEMO_COOKIE_NAME))
    set_cookie = uid is None
    if uid is None:
        uid = demo_mod.create_demo_user(create_user)

    # Datei speichern (eigener Sitzungs-Ordner) — wie im normalen Upload-Pfad
    user_dir = os.path.join(UPLOAD_DIR, str(uid))
    os.makedirs(user_dir, exist_ok=True)
    # Endung aus `ext`: nach SVG-Konvertierung traegt die Datei .png (wie /api/upload)
    safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.splitext(os.path.basename(filename))[0]}{ext}"
    file_path = os.path.join(user_dir, safe_name)
    with open(file_path, "wb") as f:
        f.write(content)

    # BESTEHENDEN Bild-Upload-Handler wiederverwenden (legt Grafik-Projekt + Bild an)
    result = await _handle_image_upload(file_path, filename, _demo_user_ctx(uid), content, ext, None)
    project_id = result["project_id"]

    # Generierung wie im normalen Pfad starten (Hintergrund) — Frontend pollt /result
    conn = get_db()
    # Ausgabesprache der Alt-Texte (Sprachmenue der Demo, gleiches Prinzip wie
    # die Projekt-Einstellung im Werkzeug). Ungueltige oder fehlende Werte
    # fallen still auf den Projekt-Default de zurueck — das Menue sendet nur
    # gueltige Codes, alles andere waere Handarbeit gegen die API.
    alt_lang = (language or "").lower()
    if alt_lang in ALT_TEXT_LANGUAGES and alt_lang != "de":
        conn.execute("UPDATE projects SET alt_language = ? WHERE id = ?", (alt_lang, project_id))
    conn.execute("UPDATE projects SET status = 'processing' WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    asyncio.create_task(_process_project(project_id, uid))
    demo_mod.record_generation(request)  # Zaehler erst nach dem Start hochziehen

    # Secure-Flag korrekt hinter dem Proxy: https in Produktion (Nginx setzt
    # X-Forwarded-Proto), http beim lokalen Dunkel-Tunnel — sonst hielte die
    # Sitzung im Browser ueber den Tunnel nicht.
    is_https = (request.headers.get("x-forwarded-proto", "").lower() == "https"
                or request.url.scheme == "https")
    resp = JSONResponse({"ok": True, "project_id": project_id})
    if set_cookie:
        resp.set_cookie(
            demo_mod.DEMO_COOKIE_NAME, demo_mod.cookie_for(uid),
            max_age=demo_mod.session_ttl_seconds(), httponly=True, samesite="lax", secure=is_https,
        )
    return resp


@app.get("/api/demo/result")
async def demo_result(request: Request):
    """Status + Ergebnis der aktuellen Demo-Sitzung (Ziel des Frontend-Pollings)."""
    _require_demo()
    uid = demo_mod.resolve_session_user_id(request.cookies.get(demo_mod.DEMO_COOKIE_NAME))
    if uid is None:
        raise HTTPException(status_code=404, detail="Keine aktive Demo-Sitzung.")
    demo_mod.touch_session(uid)  # gleitendes TTL: Aktivitaet markieren
    conn = get_db()
    proj = conn.execute(
        "SELECT * FROM projects WHERE user_id = ? ORDER BY id DESC LIMIT 1", (uid,)
    ).fetchone()
    if not proj:
        conn.close()
        return {"ok": True, "status": "leer", "image": None, "limits": demo_mod.remaining_for(request)}
    img = conn.execute(
        "SELECT * FROM images WHERE project_id = ? ORDER BY id DESC LIMIT 1", (proj["id"],)
    ).fetchone()
    conn.close()
    image = None
    if img:
        image = {
            "id": img["id"],
            "status": img["status"],
            "alt_text": _ausgabe_alt_text(img) or "",
            "langbeschreibung": img["langbeschreibung"] or "",
            "image_type": img["image_type"] or "",
            "konfidenz": img["konfidenz"] or "",
            "width": img["width"],
            "height": img["height"],
        }
    return {"ok": True, "status": proj["status"], "project_id": proj["id"], "image": image,
            "limits": demo_mod.remaining_for(request)}


@app.delete("/api/demo/session")
async def demo_delete_session(request: Request):
    """Manuelles „Loeschen" durch den Besucher: aktuelle Inhalte sofort weg,
    die Sitzung selbst bleibt bestehen (er kann erneut testen)."""
    _require_demo()
    uid = demo_mod.resolve_session_user_id(request.cookies.get(demo_mod.DEMO_COOKIE_NAME))
    if uid is not None:
        demo_mod.delete_session_content(uid, UPLOAD_DIR, RESULTS_DIR)
    return {"ok": True}


@app.post("/api/demo/chat")
async def demo_chat(request: Request):
    """Anonymer Chat fuer die Demo — derselbe InkluAgent, arbeitet auf dem Bild der
    aktuellen Sitzung. Zwei Demo-Besonderheiten: ein server-seitiges Tageslimit
    (vor dem Modell-Aufruf geprueft) und die thematische Leitplanke, die NUR hier
    an den bestehenden System-Prompt angehaengt wird (siehe demo.DEMO_GUARDRAIL)."""
    _require_demo()
    uid = demo_mod.resolve_session_user_id(request.cookies.get(demo_mod.DEMO_COOKIE_NAME))
    if uid is None:
        raise HTTPException(status_code=404, detail="Keine aktive Demo-Sitzung.")
    data = await request.json()
    message = (data.get("message") or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="Nachricht darf nicht leer sein")
    if len(message) > 2000:
        raise HTTPException(status_code=400, detail="Nachricht zu lang (max. 2000 Zeichen)")

    chk = demo_mod.check_chat(request)
    if not chk["allowed"]:
        # Sichtbarer Text in Besuchersprache (.po-Eintrag manuell, s. demo_generate).
        _ = get_gettext(detect_language(request))
        raise HTTPException(status_code=429, detail=_('Du hast die {n} kostenlosen Chat-Nachrichten für heute genutzt. Für unbegrenztes Verfeinern: kostenlos registrieren.').format(n=demo_mod.daily_chat_limit()))

    conn = get_db()
    proj = conn.execute(
        "SELECT id FROM projects WHERE user_id = ? ORDER BY id DESC LIMIT 1", (uid,)
    ).fetchone()
    conn.close()
    if not proj:
        raise HTTPException(status_code=400, detail="Bitte zuerst ein Bild hochladen, dann hilft der Assistent.")
    project_id = proj["id"]
    demo_mod.touch_session(uid)  # gleitendes TTL

    from inkluagent import storage
    from inkluagent.chat_engine import process_message
    storage.append_message(project_id, "user", message)
    # WICHTIG: system_suffix haengt die Demo-Leitplanke an — der Standard-Prompt
    # des Agenten bleibt unveraendert (Default None im normalen, eingeloggten Pfad).
    result = await asyncio.get_event_loop().run_in_executor(
        None, lambda: process_message(project_id, message, uid, system_suffix=demo_mod.DEMO_GUARDRAIL)
    )
    storage.append_message(
        project_id, "assistant", result["reply"],
        image_refs=result.get("image_refs"), intent=result.get("intent"),
    )
    demo_mod.record_chat(request)
    return {
        "reply": result["reply"],
        "image_refs": result.get("image_refs"),
        "limits": demo_mod.remaining_for(request),
    }


@app.get("/api/demo/limits")
async def demo_limits(request: Request):
    """Aktuelle Rest-Kontingente des Besuchers — auch ohne Sitzung abrufbar
    (fuer die Anzeige 'noch X von 3' beim Laden der Seite)."""
    _require_demo()
    return {"ok": True, "limits": demo_mod.remaining_for(request)}


@app.get("/api/demo/image")
async def demo_image(request: Request):
    """Liefert die Bilddatei der aktuellen Demo-Sitzung (anonym, sitzungs-bezogen)
    — fuer die Wiederherstellung der Vorschau beim Neuladen der Seite."""
    _require_demo()
    uid = demo_mod.resolve_session_user_id(request.cookies.get(demo_mod.DEMO_COOKIE_NAME))
    if uid is None:
        raise HTTPException(status_code=404, detail="Keine aktive Demo-Sitzung.")
    conn = get_db()
    row = conn.execute(
        "SELECT i.image_path FROM images i JOIN projects p ON i.project_id = p.id "
        "WHERE p.user_id = ? ORDER BY i.id DESC LIMIT 1", (uid,)
    ).fetchone()
    conn.close()
    if not row or not os.path.exists(row["image_path"]):
        raise HTTPException(status_code=404, detail="Kein Bild vorhanden.")
    return FileResponse(row["image_path"])


@app.post("/api/demo/save")
async def demo_save(request: Request):
    """Manuelle Aenderungen an Alt-Text/Langbeschreibung speichern (damit der
    Chatbot sie sieht und ein Neuladen sie zeigt). Arbeitet auf dem Bild der Sitzung."""
    _require_demo()
    uid = demo_mod.resolve_session_user_id(request.cookies.get(demo_mod.DEMO_COOKIE_NAME))
    if uid is None:
        raise HTTPException(status_code=404, detail="Keine aktive Demo-Sitzung.")
    data = await request.json()
    alt = (data.get("alt_text") or "").strip()
    lang = (data.get("langbeschreibung") or "").strip()
    conn = get_db()
    row = conn.execute(
        "SELECT i.id FROM images i JOIN projects p ON i.project_id = p.id "
        "WHERE p.user_id = ? ORDER BY i.id DESC LIMIT 1", (uid,)
    ).fetchone()
    if row:
        conn.execute("UPDATE images SET alt_text_edited = ?, langbeschreibung = ? WHERE id = ?",
                     (alt, lang, row["id"]))
        # gleitendes TTL: Aktivitaet markieren (gleiche Verbindung, kein zweiter Connect)
        conn.execute("UPDATE projects SET updated_at = datetime('now') WHERE user_id = ?", (uid,))
        conn.commit()
    conn.close()
    return {"ok": True}


# ─── URL Scanner ─────────────────────────────────────────────

_LINK_PREFIX_PATTERNS = [
    re.compile(r"^(read\s+more|read|more|continue\s+reading)\s*[:\-–—]?\s*", re.IGNORECASE),
    re.compile(r"^(weiterlesen|weiter\s*lesen|mehr\s+lesen|mehr|weiter)\s*[:\-–—]?\s*", re.IGNORECASE),
    re.compile(r"^(lese\s+mehr|lesen)\s*[:\-–—]?\s*", re.IGNORECASE),
]
_LINK_GENERIC_LABELS = {"", "read", "more", "mehr", "weiter", "weiterlesen", "lesen"}


def _clean_link_label(label: str) -> str:
    """Strip generic 'Read more'-prefixes that WordPress screen-reader-text spans
    leak into link text (e.g. '<span class=\"screen-reader-text\">Read</span>:
    DIY Adventskalender' -> 'Read: DIY Adventskalender'). Returns empty string
    if the remainder is purely generic."""
    if not label:
        return ""
    cleaned = label.strip()
    for pat in _LINK_PREFIX_PATTERNS:
        cleaned = pat.sub("", cleaned).strip()
    if cleaned.lower() in _LINK_GENERIC_LABELS or len(cleaned) < 3:
        return ""
    return cleaned


def _web_page_text(soup) -> str:
    """Lesbarer Textinhalt einer gescannten Webseite fuer die
    "Seitentext anzeigen"-Klappe (Multi-Datei Phase 2, 14.08.2026).

    Gleiche Rolle wie images.page_text bei PDF-Seiten, nur pro Webseite
    (documents.page_text). Bewusst schlicht: Skript-/Stil-/Navigations-
    Gerippe raus, dann die sichtbaren Textelemente in Dokumentreihenfolge.
    Deckel 15.000 Zeichen — die Klappe soll vorlesbar bleiben, kein
    Datenarchiv sein. Liest NUR, veraendert die Soup nicht (sie wird vom
    Aufrufer weiterbenutzt)."""
    skip = {"script", "style", "noscript", "template", "svg", "nav"}
    body = soup.body or soup
    parts = []
    seen = set()
    for el in body.find_all(["h1", "h2", "h3", "h4", "h5", "h6", "p", "li",
                             "figcaption", "blockquote", "td", "th"]):
        if el.find_parent(lambda t: t.name in skip):
            continue
        txt = el.get_text(" ", strip=True)
        # Dedupe exakt gleicher Texte: <li><p>…</p></li> liefert den Absatz
        # sonst doppelt (einmal als li-, einmal als p-Treffer).
        if not txt or txt in seen:
            continue
        seen.add(txt)
        parts.append(txt)
        if sum(len(p) for p in parts) > 15000:
            break
    return "\n".join(parts)[:15000]


# --- SSRF-Schutz fuer den URL-Scanner (18.08.2026, Fable 5) ------------------
# Der Scanner ruft beliebige, vom Nutzer angegebene Adressen ab (Seite + alle
# Bilder). Ohne Schutz kann ein angemeldeter Nutzer den Server interne Ziele
# abrufen lassen (z.B. die Cloud-Metadaten-Adresse 169.254.169.254 -> AWS-
# Zugangsdaten, oder localhost/private Netze). _pruefe_url_sicher() loest den
# Host auf und blockiert interne Adressbereiche; _safe_fetch() folgt Redirects
# MANUELL und prueft JEDEN Hop (follow_redirects=True wuerde die Vorpruefung
# aushebeln) und begrenzt die Download-Groesse.
# Bekannte Restgrenze: DNS-Rebinding (Aufloesung aendert sich zwischen Pruefung
# und Abruf) ist nicht abgedeckt; die praktischen Angriffe (Metadaten,
# localhost, private Bereiche) sind geschlossen.
def _pruefe_url_sicher(url: str) -> None:
    import ipaddress, socket
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(status_code=400, detail="Nur HTTP/HTTPS-URLs werden unterstuetzt")
    host = parsed.hostname
    if not host:
        raise HTTPException(status_code=400, detail="Diese Adresse ist nicht erlaubt")
    # Testgeschirr (19.08.2026): verify_multidatei2 serviert Pruefseiten von
    # 127.0.0.1. Loopback ist NUR erlaubt, wenn der Betreiber es ausdruecklich
    # per Umgebungsvariable freigibt — auf Staging gesetzt, auf Prod/Demo NIE.
    if os.environ.get("SCAN_ERLAUBE_LOOPBACK") == "1" and host in ("127.0.0.1", "localhost"):
        return
    try:
        infos = socket.getaddrinfo(host, None)
    except Exception:
        raise HTTPException(status_code=400, detail="Diese Adresse ist nicht erlaubt")
    for info in infos:
        ip_str = info[4][0]
        try:
            addr = ipaddress.ip_address(ip_str)
        except ValueError:
            raise HTTPException(status_code=400, detail="Diese Adresse ist nicht erlaubt")
        mapped = getattr(addr, "ipv4_mapped", None)
        if mapped is not None:
            addr = mapped
        if (addr.is_private or addr.is_loopback or addr.is_link_local
                or addr.is_reserved or addr.is_multicast or addr.is_unspecified):
            raise HTTPException(status_code=400, detail="Diese Adresse ist nicht erlaubt")


async def _safe_fetch(client, url, *, max_redirects: int = 5, max_bytes: int | None = None):
    """SSRF-sicherer GET: prueft jeden Redirect-Hop, begrenzt die Groesse."""
    current = url
    for _ in range(max_redirects + 1):
        _pruefe_url_sicher(current)
        resp = await client.get(current, follow_redirects=False)
        if resp.is_redirect and resp.headers.get("location"):
            current = urljoin(current, resp.headers["location"])
            continue
        if max_bytes is not None and len(resp.content) > max_bytes:
            raise HTTPException(status_code=413, detail="Datei zu gross")
        return resp
    raise HTTPException(status_code=502, detail="Zu viele Weiterleitungen")


@app.post("/api/scan-url")
async def scan_url(request: Request, user: dict = Depends(get_current_user)):
    """Scan a URL for images and create a project."""
    data = await request.json()
    url = data.get("url", "").strip()
    project_id = data.get("project_id")  # gesetzt = an bestehendes Web-Projekt anhaengen (sonst neues Projekt)

    if not url:
        raise HTTPException(status_code=400, detail="Bitte eine URL eingeben")

    # Schema automatisch ergaenzen: erlaubt Eingaben wie "www.inklutec.de" oder "inklutec.de".
    # Wer bewusst http:// tippt, behaelt es; nur wenn gar kein Schema vorhanden ist, https:// voranstellen.
    if not (url.lower().startswith("http://") or url.lower().startswith("https://")):
        url = "https://" + url

    # Validate URL
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(status_code=400, detail="Nur HTTP/HTTPS-URLs werden unterstuetzt")

    try:
        _seiten_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36"
        }
        async with httpx.AsyncClient(timeout=30.0, headers=_seiten_headers) as client:
            response = await _safe_fetch(client, url, max_bytes=15 * 1024 * 1024)
            response.raise_for_status()
    except HTTPException:
        raise
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Zeitueberschreitung beim Laden der Seite")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=502, detail=f"Seite nicht erreichbar: HTTP {e.response.status_code}")
    except Exception:
        raise HTTPException(status_code=502, detail="Seite konnte nicht geladen werden")

    # Parse HTML
    soup = BeautifulSoup(response.text, "html.parser")
    img_tags = soup.find_all("img")

    if not img_tags:
        raise HTTPException(status_code=404, detail="Keine Bilder auf dieser Seite gefunden")

    # Extract page profile for better context (Seitenprofil)
    from context_engine import extract_page_profile
    page_profile = extract_page_profile(soup)

    # Projekt anlegen ODER an ein bestehendes Web-Projekt anhaengen.
    # Neuer Flow (seit 04.06.2026): Im Dashboard wird zuerst ein leeres "web"-Projekt
    # angelegt; dieses Feld scannt dann eine Seite in genau dieses Projekt, und es koennen
    # nacheinander weitere Seiten angehaengt werden. Ohne project_id (Alt-Flow / direkte API)
    # wird wie bisher ein neues Projekt erzeugt -> voll rueckwaertskompatibel.
    conn = get_db()
    page_title = soup.title.string.strip() if soup.title and soup.title.string else parsed.netloc
    if project_id is not None:
        proj = conn.execute(
            "SELECT id, tool, source_url FROM projects WHERE id = ? AND user_id = ?",
            (project_id, user["id"])
        ).fetchone()
        if not proj:
            conn.close()
            raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
        if proj["tool"] != "web":
            conn.close()
            raise HTTPException(status_code=400, detail="Dieses Projekt ist kein Webseiten-Projekt")
        # Erste gescannte Adresse als Projekt-Quelle merken, falls noch leer (frisch angelegt).
        if not proj["source_url"]:
            conn.execute("UPDATE projects SET source_url = ? WHERE id = ?", (url, project_id))
        conn.execute("UPDATE projects SET status = 'extracting' WHERE id = ?", (project_id,))
        conn.commit()
    else:
        cursor = conn.execute(
            "INSERT INTO projects (user_id, filename, original_path, status, project_type, tool, source_url) VALUES (?, ?, '', 'extracting', 'url', 'web', ?)",
            (user["id"], f"Website: {page_title[:80]}", url)
        )
        project_id = cursor.lastrowid
        conn.commit()

    # Multi-Datei Phase 2 (14.08.2026): Jede gescannte Adresse wird ein eigenes
    # "Webseite N"-Dokument — gleiches Muster wie die PDF-Dokumente aus Phase 1,
    # inklusive Umbenennen/Loeschen ueber die bestehenden documents-Endpunkte.
    # Die Zeile entsteht VOR dem Bild-Download, damit die Bilder direkt mit
    # document_id eingefuegt werden koennen; liefert die Seite am Ende nichts,
    # wird sie wieder entfernt (gleiches Aufraeum-Muster wie _extract_document).
    next_doc_index = conn.execute(
        "SELECT COALESCE(MAX(doc_index), 0) FROM documents WHERE project_id = ?",
        (project_id,)
    ).fetchone()[0] + 1
    doc_cursor = conn.execute(
        """INSERT INTO documents
           (project_id, doc_index, original_filename, original_path,
            extraction_method, total_images, source_url, page_text)
           VALUES (?, ?, ?, '', 'web', 0, ?, ?)""",
        (project_id, next_doc_index, page_title[:200], url, _web_page_text(soup))
    )
    document_id = doc_cursor.lastrowid
    conn.commit()

    # Fortlaufende Bildnummer: beim Anhaengen weiterer Seiten ab dem bisherigen Maximum
    # weiterzaehlen. Das haelt image_index eindeutig UND verhindert Dateinamens-Kollisionen
    # (web_<idx>...), weil schon vergebene Indizes nicht erneut benutzt werden.
    base_idx = conn.execute(
        "SELECT COALESCE(MAX(image_index), 0) FROM images WHERE project_id = ?",
        (project_id,)
    ).fetchone()[0]

    img_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project_id))
    os.makedirs(img_dir, exist_ok=True)

    downloaded = 0
    # Browser-like headers for image downloads: many WordPress sites block requests
    # without Referer header (hotlink protection / bot detection). Fix for 403 errors
    # on sites like dc-tischlermeister.de. Reported by Stephan Raithel, 25.03.2026.
    img_download_headers = {
        "Referer": url,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
        "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    }
    async with httpx.AsyncClient(timeout=15.0, follow_redirects=True, headers=img_download_headers) as client:
        for raw_idx, img_tag in enumerate(img_tags, 1):
            # Beim Anhaengen weiterer Seiten ab dem bisherigen Maximum weiterzaehlen (siehe base_idx).
            idx = base_idx + raw_idx
            # Support lazy-loaded images: check data-src, data-lazy-src first
            src = img_tag.get("data-src") or img_tag.get("data-lazy-src") or img_tag.get("src", "")

            # v2.2.3: Always check srcset for higher-resolution version
            srcset = img_tag.get("srcset") or img_tag.get("data-srcset") or ""
            if srcset:
                # Parse srcset entries and pick the largest by width descriptor
                best_url = ""
                best_width = 0
                for entry in srcset.split(","):
                    parts = entry.strip().split()
                    if len(parts) >= 1:
                        candidate_url = parts[0]
                        w = 0
                        if len(parts) >= 2 and parts[1].endswith("w"):
                            try:
                                w = int(parts[1][:-1])
                            except ValueError:
                                w = 0
                        elif len(parts) >= 2 and parts[1].endswith("x"):
                            try:
                                w = int(float(parts[1][:-1]) * 1000)  # treat 2x as 2000w
                            except ValueError:
                                w = 0
                        if w > best_width:
                            best_width = w
                            best_url = candidate_url
                # Use srcset version if it's larger than the src
                if best_url and best_width > 200:
                    src = best_url

            if not src or src.startswith("data:"):
                continue
            if not src:
                continue

            # Resolve relative URLs
            img_url = urljoin(url, src)

            # Get original alt text
            original_alt = img_tag.get("alt", "")

            # Detect intentionally hidden/decorative images from HTML attributes
            # NOTE: alt="" is NOT treated as hidden – many sites leave alt empty
            # instead of writing proper alt texts. The KI will classify these.
            is_hidden = (
                img_tag.get("aria-hidden") == "true"
                or img_tag.get("role") in ("presentation", "none")
            )
            has_empty_alt = img_tag.has_attr("alt") and img_tag["alt"] == ""

            # Hard bypass: only aria-hidden and role=presentation skip KI
            if is_hidden:
                conn.execute(
                    """INSERT INTO images (project_id, document_id, page_number, image_index, image_path, context_text,
                       width, height, xref, original_alt, status, image_type, alt_text)
                       VALUES (?, ?, 1, ?, '', '', 0, 0, 0, ?, 'done', 'dekorativ', '')""",
                    (project_id, document_id, idx, original_alt)
                )
                downloaded += 1
                continue

            # Get context: parent text, figcaption, title attribute
            context_parts = []
            if has_empty_alt:
                context_parts.append("[HTML-Hinweis] Bild hat alt=\"\" im Quellcode – prüfe ob wirklich dekorativ oder ob Alt-Text fehlt")
            if img_tag.get("title"):
                context_parts.append(f"[title] {img_tag['title']}")
            # Original alt-text is stored for display but NOT sent to the model
            # to prevent the "discussion bug" (model commenting on existing alt-text)
            # figcaption
            parent_figure = img_tag.find_parent("figure")
            if parent_figure:
                figcaption = parent_figure.find("figcaption")
                if figcaption:
                    context_parts.append(f"[Bildunterschrift] {figcaption.get_text(strip=True)}")
            # Check if image is a link (for linked image alt-text)
            parent_link = img_tag.find_parent("a")
            link_display_text = ""
            if parent_link:
                link_href = parent_link.get("href", "")
                # v2.2.3: Prefer human-readable link text over raw URL
                # v3.7: Sanitize generic "Read more"-prefixes from screen-reader-text spans
                link_label = _clean_link_label(parent_link.get("aria-label", "") or parent_link.get("title", ""))
                # Also check the visible text content of the link (excluding the image alt)
                link_text = parent_link.get_text(strip=True)
                if link_text and link_text.lower() not in {"", "bild", "grafik", "image", "img"}:
                    if not link_label:
                        link_label = _clean_link_label(link_text)[:150]
                if link_label:
                    context_parts.append(f"[Link-Beschriftung] {link_label}")
                    link_display_text = link_label
                if link_href:
                    # Only add raw URL if no readable label available
                    if not link_label:
                        context_parts.append(f"[Link-Ziel] {link_href}")
                        link_display_text = link_href
                    else:
                        # Still store href but label takes priority in context
                        context_parts.append(f"[Link-URL] {link_href}")

            # Improved context search: go beyond direct parent
            # 1. Nearest heading before the image
            prev_heading = img_tag.find_previous(["h1", "h2", "h3", "h4"])
            if prev_heading:
                context_parts.append(f"[Ueberschrift] {prev_heading.get_text(strip=True)[:150]}")

            # 2. Parent article/section (WordPress wrappers)
            content_parent = img_tag.find_parent(["article", "section"]) or img_tag.find_parent("div", class_=True)
            if content_parent and content_parent.name not in ("html", "body", "head"):
                parent_text = content_parent.get_text(strip=True)[:300]
                if parent_text and parent_text != original_alt:
                    context_parts.append(f"[Umgebungstext] {parent_text}")

            # 3. Text after the image (WordPress captions often below)
            next_sib = img_tag.find_next_sibling(["p", "span", "figcaption", "div"])
            if next_sib:
                next_text = next_sib.get_text(strip=True)[:150]
                if next_text:
                    context_parts.append(f"[Text danach] {next_text}")

            # Image-specific context FIRST, page profile as supplement
            all_context = []
            all_context.extend(context_parts)  # Bildspezifisch ZUERST
            if page_profile:
                all_context.append(page_profile)  # Seitenprofil als Ergaenzung
            context_text = "\n".join(all_context) if all_context else ""

            # Download image
            try:
                img_response = await _safe_fetch(client, img_url, max_bytes=20 * 1024 * 1024)
                img_response.raise_for_status()
                img_content = img_response.content

                # Determine extension from content type or URL
                content_type = img_response.headers.get("content-type", "")
                ext = _ext_from_content_type(content_type) or _ext_from_url(img_url)
                if not ext:
                    continue  # Skip unknown formats

                # SVG nach PNG wandeln — zentrale Hilfe _svg_to_png_bytes
                # (weisser Hintergrund, gleicher Weg wie die Upload-Endpunkte).
                # Bis 17.07.2026 war dieser Zweig unerreichbar: _ext_from_content_type
                # und _ext_from_url kannten SVG nicht, `if not ext: continue` oben
                # uebersprang SVGs still — jetzt liefern beide ".svg".
                if ext == ".svg" or "svg" in content_type:
                    try:
                        png_path = os.path.join(img_dir, f"web_{idx}.png")
                        with open(png_path, "wb") as f:
                            f.write(_svg_to_png_bytes(img_content))
                        img_path = png_path
                        ext = ".png"
                        img_filename = f"web_{idx}.png"
                    except Exception as e:
                        print(f"SVG-Konvertierung fehlgeschlagen fuer {img_url}: {e}")
                        continue
                else:
                    img_filename = f"web_{idx}{ext}"
                    img_path = os.path.join(img_dir, img_filename)
                    with open(img_path, "wb") as f:
                        f.write(img_content)

                # Get dimensions
                width, height = 0, 0
                try:
                    from PIL import Image as PILImage
                    with PILImage.open(img_path) as pimg:
                        width, height = pimg.size
                except Exception:
                    pass

                # Skip tiny images (likely tracking pixels or icons)
                if width > 0 and height > 0 and (width < 20 or height < 20):
                    os.remove(img_path)
                    continue

                # v2.2.3: If image is a small thumbnail, try to find a larger version
                if width > 0 and height > 0 and width < 250 and height < 250:
                    larger_urls = _try_larger_image_url(img_url)
                    for candidate_url in larger_urls:
                        try:
                            resolved_url = urljoin(url, candidate_url)
                            larger_response = await _safe_fetch(client, resolved_url, max_bytes=20 * 1024 * 1024)
                            if larger_response.status_code == 200:
                                larger_ct = larger_response.headers.get("content-type", "")
                                if "image" in larger_ct:
                                    larger_ext = _ext_from_content_type(larger_ct) or ext
                                    larger_path = os.path.join(img_dir, f"web_{idx}_large{larger_ext}")
                                    with open(larger_path, "wb") as lf:
                                        lf.write(larger_response.content)
                                    try:
                                        with PILImage.open(larger_path) as lpimg:
                                            lw, lh = lpimg.size
                                        if lw > width and lh > height:
                                            # Larger version found! Replace the small one
                                            os.remove(img_path)
                                            new_filename = f"web_{idx}{larger_ext}"
                                            new_path = os.path.join(img_dir, new_filename)
                                            os.rename(larger_path, new_path)
                                            img_path = new_path
                                            img_filename = new_filename
                                            print(f"v2.2.3 Upscale: {width}x{height} -> {lw}x{lh} fuer {img_url[-50:]}")
                                            width, height = lw, lh
                                            break
                                        else:
                                            os.remove(larger_path)
                                    except Exception:
                                        if os.path.exists(larger_path):
                                            os.remove(larger_path)
                        except Exception:
                            pass

                conn.execute(
                    """INSERT INTO images (project_id, document_id, page_number, image_index, image_path, context_text,
                       width, height, xref, original_alt)
                       VALUES (?, ?, 1, ?, ?, ?, ?, ?, 0, ?)""",
                    (project_id, document_id, idx, img_path, context_text, width, height, original_alt)
                )
                downloaded += 1

            except Exception as e:
                print(f"Fehler beim Download von {img_url}: {e}")
                continue

    # Kumulative Gesamtzahl (beim Anhaengen schon vorhandene Bilder mitzaehlen).
    total = conn.execute("SELECT COUNT(*) FROM images WHERE project_id = ?", (project_id,)).fetchone()[0]
    if downloaded == 0:
        # Leeres "Webseite N"-Dokument wieder entfernen — es haengen keine
        # Bilder dran (Phase 2; sonst bliebe ein leerer Block in der Ansicht).
        conn.execute("DELETE FROM documents WHERE id = ?", (document_id,))
        # Beim Anhaengen ein bestehendes Projekt NICHT zerstoeren: hat es schon Bilder,
        # nur freundlich melden und Status zurueck auf 'extracted'. Nur ein wirklich leeres
        # Projekt (Erst-Scan ohne Treffer) wird als Fehler markiert.
        if total > 0:
            conn.execute("UPDATE projects SET status = 'extracted' WHERE id = ?", (project_id,))
            conn.commit()
            conn.close()
            raise HTTPException(status_code=404, detail="Auf dieser Seite wurden keine neuen Bilder gefunden")
        conn.execute("UPDATE projects SET status = 'error' WHERE id = ?", (project_id,))
        conn.commit()
        conn.close()
        raise HTTPException(status_code=404, detail="Keine Bilder konnten heruntergeladen werden")

    conn.execute(
        "UPDATE documents SET total_images = ? WHERE id = ?",
        (downloaded, document_id)
    )
    conn.execute(
        "UPDATE projects SET status = 'extracted', total_images = ? WHERE id = ?",
        (total, project_id)
    )
    conn.commit()
    conn.close()

    return {
        "ok": True,
        "project_id": project_id,
        "document_id": document_id,
        "added": downloaded,
        "total_images": total,
        "source_url": url,
        "project_type": "url",
    }



# v2.2.3: Try to find larger versions of thumbnail images via URL rewriting
def _try_larger_image_url(img_url: str) -> list[str]:
    """Generate candidate URLs for larger versions of a thumbnail image.
    Returns a list of alternative URLs to try (most promising first)."""
    candidates = []

    # Pattern 1: BLE/Government CMS - ?__blob=thumbnail -> normal/wide
    if "__blob=thumbnail" in img_url:
        candidates.append(img_url.replace("__blob=thumbnail", "__blob=normal"))
        candidates.append(img_url.replace("__blob=thumbnail", "__blob=wide"))

    # Pattern 2: WordPress - remove size suffix (-150x150, -300x200, etc.)
    import re
    wp_match = re.search(r'-(\d{2,4})x(\d{2,4})\.(\w+)$', img_url)
    if wp_match:
        w, h = int(wp_match.group(1)), int(wp_match.group(2))
        if w <= 300 or h <= 300:
            original = re.sub(r'-\d{2,4}x\d{2,4}\.(\w+)$', r'.\1', img_url)
            candidates.append(original)

    # Pattern 3: Common thumbnail patterns
    replacements = [
        ("_thumb.", "_large."),
        ("_small.", "_large."),
        ("_thumbnail.", "."),
        ("/thumb/", "/"),
        ("/thumbnails/", "/images/"),
        ("/small/", "/large/"),
        ("_s.", "_l."),
        ("-thumb.", "."),
        ("?w=150", "?w=800"),
        ("?width=150", "?width=800"),
        ("&w=150", "&w=800"),
    ]
    for old, new in replacements:
        if old in img_url:
            candidates.append(img_url.replace(old, new))

    # Pattern 4: Shopify CDN
    shopify_match = re.search(r'_(\d+x\d+)\.', img_url)
    if shopify_match and "shopify" in img_url.lower():
        candidates.append(re.sub(r'_\d+x\d+\.', '.', img_url))

    return candidates


def _ext_from_content_type(ct: str) -> str:
    """Map content type to file extension."""
    ct = ct.lower().split(";")[0].strip()
    mapping = {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/gif": ".gif",
        "image/webp": ".webp",
        "image/avif": ".avif",
        "image/heic": ".heic",
        "image/heif": ".heif",
        # SVG wird vom Aufrufer sofort nach PNG konvertiert (_svg_to_png_bytes) —
        # PIL selbst kann kein SVG (frueher deshalb hier ausgeschlossen).
        "image/svg+xml": ".svg",
    }
    return mapping.get(ct, "")


def _ext_from_url(url: str) -> str:
    """Extract file extension from URL path."""
    path = urlparse(url).path.lower()
    for ext in IMAGE_EXTENSIONS:
        if path.endswith(ext):
            return ext
    return ""


def pdf_langbeschreibung_enabled():
    """Schalter (ENV PDF_LANGBESCHREIBUNG): ob bei PDF-Projekten Langbeschreibungen
    erzeugt + angezeigt werden. Standard AUS (Karbe-Wunsch 03.06.2026: bei PDF nicht
    gebraucht, spart den zweiten KI-Pass = Kosten). BILD-Projekte sind NIE betroffen.

    Umschalten OHNE Rebuild: PDF_LANGBESCHREIBUNG=on in .env.staging/.env.prod setzen,
    dann ./compose-staging.sh (bzw. ./compose-prod.sh) up -d  -> Container-Neustart.

    DAUERHAFT ENTFERNEN/AENDERN: diese Funktion + ihre Aufrufstellen (get_project +
    _process_project) sowie die Eintraege in den compose-/env-Dateien anpassen.
    Frontend-Gegenstueck: get_project liefert show_langbeschreibung -> renderImages(showLang)."""
    return os.getenv("PDF_LANGBESCHREIBUNG", "off").strip().lower() in ("1", "true", "on", "yes")


@app.get("/api/projects/{project_id}")
async def get_project(project_id: int, user: dict = Depends(get_current_user)):
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")

    # Multi-Datei Phase 1 (08.06.2026): Bilder werden weiter geordnet nach
    # Seite/image_index zurueckgegeben — das Frontend gruppiert sie pro Dokument.
    # Reihenfolge: doc_index, page_number, image_index (stabil pro Dokument).
    images = conn.execute(
        """SELECT i.*, (SELECT body FROM messages m WHERE m.image_id = i.id AND m.msg_type = 'review_note' LIMIT 1) AS review_note FROM images i
           LEFT JOIN documents d ON d.id = i.document_id
           WHERE i.project_id = ?
           ORDER BY COALESCE(d.doc_index, 0), i.page_number, i.image_index""",
        (project_id,)
    ).fetchall()
    documents = conn.execute(
        """SELECT id, doc_index, original_filename, display_name, extraction_method,
                  total_images, created_at, source_url, page_text, hinweise
           FROM documents WHERE project_id = ? ORDER BY doc_index""",
        (project_id,)
    ).fetchall()
    # Rollen-Workflow Etappe 2 (13.07.2026): Pruefstatus PRO ROLLE auch fuer den
    # Besitzer — Datenbasis der Badges „Lektorat: X / Herausgeber: Y" und der
    # Filterleiste. Zusaetzlich die aktiven Gast-Rollen des Projekts, damit das
    # Frontend weiss, welche Rollen-Badges/Filter es anbieten soll.
    review_rows = conn.execute(
        """SELECT r.image_id, r.role, r.status, r.reviewed_at FROM image_reviews r
           JOIN images i ON i.id = r.image_id WHERE i.project_id = ?""",
        (project_id,)
    ).fetchall()
    share_roles = [row["role"] for row in conn.execute(
        """SELECT DISTINCT COALESCE(NULLIF(role, ''), 'kunde') AS role FROM shares
           WHERE project_id = ? AND status IN ('active', 'completed') ORDER BY role""",
        (project_id,)
    ).fetchall()]
    # Rollen-Workflow Etappe 4 (15.07.2026): Nachrichten-Verlauf pro Bild.
    thread_rows = conn.execute(
        """SELECT m.image_id, COALESCE(m.sender_role, '') AS sender_role,
                  COALESCE(m.sender_name, '') AS sender_name, m.body, m.created_at
           FROM messages m JOIN images i ON i.id = m.image_id
           WHERE i.project_id = ? AND m.msg_type = 'chat'
           ORDER BY m.created_at, m.id""",
        (project_id,)
    ).fetchall()
    conn.close()

    thread_by_image = {}
    for m in thread_rows:
        thread_by_image.setdefault(m["image_id"], []).append({
            "sender_role": m["sender_role"], "sender_name": m["sender_name"],
            "body": m["body"], "created_at": m["created_at"]})
    reviews_by_image = {}
    for r in review_rows:
        reviews_by_image.setdefault(r["image_id"], {})[r["role"]] = {
            "status": r["status"], "reviewed_at": r["reviewed_at"]}
    image_dicts = []
    for img in images:
        d = dict(img)
        d["reviews"] = reviews_by_image.get(d["id"], {})
        d["thread"] = thread_by_image.get(d["id"], [])
        image_dicts.append(d)

    # Schalter PDF_LANGBESCHREIBUNG (Standard aus): Langbeschreibung im Frontend nur zeigen,
    # wenn kein PDF-Projekt ODER der Schalter an ist (siehe pdf_langbeschreibung_enabled).
    proj_dict = dict(project)
    _is_pdf = (proj_dict.get("tool") == "pdf" or proj_dict.get("project_type") == "pdf")
    # Review-Status nur zeigen, wenn das Projekt ueberhaupt zur Pruefung freigegeben
    # wurde (Steve 20.06.) -> Solo-Arbeit ohne Einladung bleibt frei von Pruef-Badges.
    # share_roles traegt dieselbe Information pro Rolle; der bool bleibt fuer
    # Bestands-Codepfade erhalten.
    return {
        "project": proj_dict,
        "images": image_dicts,
        "documents": [dict(d) for d in documents],
        "show_langbeschreibung": (not _is_pdf) or pdf_langbeschreibung_enabled(),
        "in_review": bool(share_roles),
        "share_roles": share_roles,
    }


@app.patch("/api/projects/{project_id}/documents/{document_id}")
async def rename_document(project_id: int, document_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Multi-Datei (08.06.2026): Anzeigename eines Dokuments setzen (Nice-to-have).

    Lasst der Nutzer das Feld leer, faellt die Anzeige im Frontend wieder auf
    den original_filename zurueck (display_name wird auf NULL gesetzt).
    """
    data = await request.json()
    name = (data.get("display_name") or "").strip()
    if len(name) > 200:
        raise HTTPException(status_code=400, detail="Anzeigename darf maximal 200 Zeichen lang sein")
    conn = get_db()
    proj = conn.execute(
        "SELECT id FROM projects WHERE id = ? AND user_id = ?",
        (project_id, user["id"])
    ).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    cur = conn.execute(
        "UPDATE documents SET display_name = ? WHERE id = ? AND project_id = ?",
        (name or None, document_id, project_id)
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Dokument nicht gefunden")
    return {"ok": True, "display_name": name or None}


@app.delete("/api/projects/{project_id}/documents/{document_id}")
async def delete_document(project_id: int, document_id: int, user: dict = Depends(get_current_user)):
    """Multi-Datei (10.06.2026): Ein einzelnes Dokument (eine hochgeladene PDF)
    samt seiner Bilder aus einem Projekt entfernen.

    Hintergrund (Michael Karbe, 09.06.2026): Da im Werkzeug stets die KOMPLETTE
    PDF angezeigt wird, loeschen wir bewusst nur auf Dokument-Ebene — einzelne
    Seiten/Bilder einer PDF bleiben unantastbar. (Grafik-/Web-Projekte werden in
    einem Folgeschritt auf Bild-Ebene loeschbar.)

    WICHTIG zur Nummerierung: doc_index wird NICHT neu vergeben. Er ist der
    stabile Ordnerschluessel auf der Platte (results/<user>/<projekt>/doc<N>),
    und neue Uploads leiten ihren Ordner aus MAX(doc_index)+1 ab. Wuerden wir
    umnummerieren, koennten spaetere Uploads in fremde Ordner schreiben. Die fuer
    den Nutzer sichtbare Nummer „Dokument 1, 2, 3" ist reine Anzeige-Position und
    wird im Frontend aus der sortierten Reihenfolge berechnet.

    Mandantensicher: sowohl Projekt- als auch Dokument-Zugriff sind auf den
    eingeloggten Nutzer eingegrenzt (kein IDOR).
    """
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")

    doc = conn.execute(
        "SELECT * FROM documents WHERE id = ? AND project_id = ?", (document_id, project_id)
    ).fetchone()
    if not doc:
        conn.close()
        raise HTTPException(status_code=404, detail="Dokument nicht gefunden")
    doc = dict(doc)

    # Dateien entfernen: der eigene Bild-/Vorschau-Ordner doc<N> und die
    # hochgeladene Quell-PDF. Best-effort — fehlende Dateien sind kein Fehler.
    doc_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project_id), f"doc{doc['doc_index']}")
    if os.path.isdir(doc_dir):
        shutil.rmtree(doc_dir, ignore_errors=True)
    src_pdf = doc.get("original_path") or ""
    if src_pdf and os.path.exists(src_pdf):
        try:
            os.remove(src_pdf)
        except OSError:
            pass

    # Multi-Datei Phase 2 (14.08.2026): Web-Dokumente haben KEINEN doc<N>-
    # Ordner — ihre Bilddateien liegen flach im Projektordner (web_<idx>.*).
    # Deshalb zusaetzlich die Dateien der betroffenen Bilder einzeln entfernen.
    # Best-effort und nur unterhalb von RESULTS_DIR (Pfade stammen aus unserer
    # DB; der Guard schuetzt vor Altlasten mit kaputten Pfaden). Fuer PDF-
    # Dokumente ist das ein No-op-Doppel zur doc-Ordner-Loeschung oben.
    _results_root = os.path.realpath(RESULTS_DIR) + os.sep
    img_rows = conn.execute(
        "SELECT id, image_path, page_view_path FROM images WHERE document_id = ? AND project_id = ?",
        (document_id, project_id)
    ).fetchall()
    for ir in img_rows:
        for p in (ir["image_path"], ir["page_view_path"]):
            if p and os.path.isfile(p) and os.path.realpath(p).startswith(_results_root):
                try:
                    os.remove(p)
                except OSError:
                    pass

    # DB-Zeilen entfernen (Bilder zuerst, dann das Dokument). Pruefstatus und
    # Nachrichten der geloeschten Bilder raeumen wir mit ab (Phase 2) — vorher
    # blieben sie als harmlose Waisen liegen (alle Lese-Wege JOINen auf images).
    # Quickinfo-Werkzeug (27.08.2026): Bilddateien der Formularfelder dieses
    # Dokuments (Ausschnitte, Seitenansichten) mit loeschen — nur unterhalb
    # von RESULTS_DIR (Realpath-Schutz wie bei den Bildern).
    for fr in conn.execute("SELECT ausschnitt_path, page_view_path FROM formularfelder WHERE document_id = ? AND project_id = ?",
                           (document_id, project_id)).fetchall():
        for p in (fr["ausschnitt_path"], fr["page_view_path"]):
            if p and os.path.isfile(p) and os.path.realpath(p).startswith(_results_root):
                try:
                    os.remove(p)
                except OSError:
                    pass
    _img_ids = [ir["id"] for ir in img_rows]
    if _img_ids:
        _ph = ",".join("?" * len(_img_ids))
        conn.execute(f"DELETE FROM image_reviews WHERE image_id IN ({_ph})", _img_ids)
        conn.execute(f"DELETE FROM messages WHERE image_id IN ({_ph})", _img_ids)
    conn.execute("DELETE FROM images WHERE document_id = ? AND project_id = ?", (document_id, project_id))
    # Quickinfo-Werkzeug (27.08.2026): Formularfelder des Dokuments (+ Gast-Urteile, 28.08.).
    conn.execute("DELETE FROM feld_reviews WHERE feld_id IN "
                 "(SELECT id FROM formularfelder WHERE document_id = ? AND project_id = ?)", (document_id, project_id))
    conn.execute("DELETE FROM formularfelder WHERE document_id = ? AND project_id = ?", (document_id, project_id))
    conn.execute("DELETE FROM documents WHERE id = ? AND project_id = ?", (document_id, project_id))

    # Projekt-Zaehler aus dem Ist-Stand neu berechnen. processed_images zaehlt
    # weiterhin nur fertige Bilder (status='done'), damit der Fortschritt stimmt.
    remaining_images = conn.execute(
        "SELECT COUNT(*) FROM images WHERE project_id = ?", (project_id,)
    ).fetchone()[0]
    processed = conn.execute(
        "SELECT COUNT(*) FROM images WHERE project_id = ? AND status = 'done'", (project_id,)
    ).fetchone()[0]
    remaining_docs = conn.execute(
        "SELECT COUNT(*) FROM documents WHERE project_id = ?", (project_id,)
    ).fetchone()[0]
    conn.execute(
        "UPDATE projects SET total_images = ?, processed_images = ?, updated_at = datetime('now') WHERE id = ?",
        (remaining_images, processed, project_id)
    )
    conn.commit()
    conn.close()
    return {
        "ok": True,
        "remaining_images": remaining_images,
        "remaining_documents": remaining_docs,
    }


def _require_non_pdf_project(conn, project_id: int, user_id: int) -> dict:
    """Gemeinsamer Wachposten der Einzelbild-Endpunkte (Phase 2, 14.08.2026):
    Projekt muss dem eingeloggten Nutzer gehoeren (kein IDOR) und darf KEIN
    PDF-Projekt sein — bei PDFs ist das Dokument die kleinste loesch- und
    benennbare Einheit (Michael-Vorgabe 09.06.2026: die komplette PDF wird
    angezeigt, einzelne Seiten/Bilder bleiben unantastbar)."""
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user_id)
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    project = dict(project)
    if project.get("tool") == "pdf" or project.get("project_type") == "pdf":
        conn.close()
        raise HTTPException(status_code=400,
                            detail="Bei PDF-Projekten werden ganze Dokumente bearbeitet, keine einzelnen Bilder")
    return project


@app.patch("/api/projects/{project_id}/images/{image_id}")
async def rename_image(project_id: int, image_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Multi-Datei Phase 2 (14.08.2026): Anzeigename EINES Bildes setzen —
    nur fuer Grafik- und Web-Projekte. Leeres Feld setzt auf NULL zurueck,
    die Anzeige faellt dann auf den Original-Dateinamen bzw. "Bild N"
    zurueck (gleiche Logik wie rename_document)."""
    data = await request.json()
    name = (data.get("display_name") or "").strip()
    if len(name) > 200:
        raise HTTPException(status_code=400, detail="Anzeigename darf maximal 200 Zeichen lang sein")
    conn = get_db()
    _require_non_pdf_project(conn, project_id, user["id"])
    cur = conn.execute(
        "UPDATE images SET display_name = ? WHERE id = ? AND project_id = ?",
        (name or None, image_id, project_id)
    )
    conn.commit()
    affected = cur.rowcount
    conn.close()
    if not affected:
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    return {"ok": True, "display_name": name or None}


@app.delete("/api/projects/{project_id}/images/{image_id}")
async def delete_image(project_id: int, image_id: int, user: dict = Depends(get_current_user)):
    """Multi-Datei Phase 2 (14.08.2026): EIN Bild aus einem Grafik- oder
    Web-Projekt entfernen. Bei Grafik ist das Einzelbild die atomare Einheit
    (vereinbart 10.06.2026), bei Web dient es dem Feinschnitt innerhalb einer
    Webseite (z.B. mitgeladene Deko-Grafiken aussortieren). PDF bleibt
    gesperrt — dort loescht man auf Dokument-Ebene (delete_document).

    Mandantensicher wie delete_document: Projekt- UND Bild-Zugriff sind auf
    den eingeloggten Nutzer eingegrenzt. Wird das letzte Bild einer Webseite
    entfernt, verschwindet auch deren documents-Zeile — sonst bliebe ein
    leerer "Webseite N"-Block stehen (und der Phantom-Cleanup beim naechsten
    Start wuerde ihn ohnehin wegputzen)."""
    conn = get_db()
    _require_non_pdf_project(conn, project_id, user["id"])
    img = conn.execute(
        "SELECT * FROM images WHERE id = ? AND project_id = ?", (image_id, project_id)
    ).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    img = dict(img)

    # Dateien best-effort entfernen, nur unterhalb von RESULTS_DIR (Pfade
    # stammen aus unserer DB; Guard wie in delete_document).
    _results_root = os.path.realpath(RESULTS_DIR) + os.sep
    for p in (img.get("image_path"), img.get("page_view_path")):
        if p and os.path.isfile(p) and os.path.realpath(p).startswith(_results_root):
            try:
                os.remove(p)
            except OSError:
                pass

    conn.execute("DELETE FROM image_reviews WHERE image_id = ?", (image_id,))
    conn.execute("DELETE FROM messages WHERE image_id = ?", (image_id,))
    conn.execute("DELETE FROM images WHERE id = ? AND project_id = ?", (image_id, project_id))

    removed_document = None
    doc_id = img.get("document_id")
    if doc_id:
        left_in_doc = conn.execute(
            "SELECT COUNT(*) FROM images WHERE document_id = ?", (doc_id,)
        ).fetchone()[0]
        if left_in_doc == 0:
            conn.execute("DELETE FROM documents WHERE id = ? AND project_id = ?", (doc_id, project_id))
            removed_document = doc_id
        else:
            conn.execute("UPDATE documents SET total_images = ? WHERE id = ?", (left_in_doc, doc_id))

    # Projekt-Zaehler aus dem Ist-Stand neu berechnen (wie delete_document).
    remaining_images = conn.execute(
        "SELECT COUNT(*) FROM images WHERE project_id = ?", (project_id,)
    ).fetchone()[0]
    processed = conn.execute(
        "SELECT COUNT(*) FROM images WHERE project_id = ? AND status = 'done'", (project_id,)
    ).fetchone()[0]
    conn.execute(
        "UPDATE projects SET total_images = ?, processed_images = ?, updated_at = datetime('now') WHERE id = ?",
        (remaining_images, processed, project_id)
    )
    conn.commit()
    conn.close()
    return {
        "ok": True,
        "remaining_images": remaining_images,
        "removed_document": removed_document,
    }


@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: int, user: dict = Depends(get_current_user)):
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")

    # Delete files
    project_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project_id))
    if os.path.exists(project_dir):
        shutil.rmtree(project_dir)
    if project["original_path"] and os.path.exists(project["original_path"]):
        os.remove(project["original_path"])

    conn.execute("DELETE FROM images WHERE project_id = ?", (project_id,))
    # Quickinfo-Werkzeug (27.08.2026): Formularfelder mit aufraeumen (+ Gast-Urteile, 28.08.).
    conn.execute("DELETE FROM feld_reviews WHERE feld_id IN (SELECT id FROM formularfelder WHERE project_id = ?)", (project_id,))
    conn.execute("DELETE FROM formularfelder WHERE project_id = ?", (project_id,))
    # Multi-Datei (08.06.2026): Dokument-Zeilen mit aufraeumen.
    conn.execute("DELETE FROM documents WHERE project_id = ?", (project_id,))
    conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/images/{image_id}/file")
async def get_image_file(image_id: int, user: dict = Depends(get_current_user)):
    conn = get_db()
    img = conn.execute(
        """SELECT i.* FROM images i
           JOIN projects p ON i.project_id = p.id
           WHERE i.id = ? AND p.user_id = ?""",
        (image_id, user["id"])
    ).fetchone()
    conn.close()
    if not img or not os.path.exists(img["image_path"]):
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    return FileResponse(img["image_path"])


@app.get("/api/images/{image_id}/page-view")
async def get_image_page_view(image_id: int, user: dict = Depends(get_current_user)):
    """Liefert die Seitenansicht-PNG (ganze PDF-Seite) zur visuellen Kontext-Anzeige."""
    conn = get_db()
    img = conn.execute(
        """SELECT i.page_view_path FROM images i
           JOIN projects p ON i.project_id = p.id
           WHERE i.id = ? AND p.user_id = ?""",
        (image_id, user["id"])
    ).fetchone()
    conn.close()
    if not img or not img["page_view_path"] or not os.path.exists(img["page_view_path"]):
        raise HTTPException(status_code=404, detail="Seitenansicht nicht gefunden")
    return FileResponse(img["page_view_path"])


@app.post("/api/projects/{project_id}/generate")
async def generate_alt_texts(project_id: int, user: dict = Depends(get_current_user)):
    # Check daily limit (admins are exempt)
    if not user.get("is_admin"):
        _tageslimit = effektives_api_tageslimit(user)
        daily_used = get_daily_image_count(user["id"])
        if daily_used >= _tageslimit:
            raise HTTPException(status_code=429, detail=f"Tageslimit erreicht ({_tageslimit} Bilder pro Tag). Das Limit wird um Mitternacht (UTC) zurueckgesetzt.")

    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")

    if project["status"] == "processing":
        conn.close()
        raise HTTPException(status_code=409, detail="Verarbeitung laeuft bereits")

    # Async-Upload (04.07.2026): Waehrend der Hintergrund-Extraktion einer PDF
    # wuerde die Sammel-Generierung nur die bis dahin eingefuegten Bilder
    # erwischen — deshalb sauber ablehnen, das Frontend pollt ohnehin.
    if project["status"] == "extracting":
        conn.close()
        raise HTTPException(status_code=409, detail="Die PDF wird noch verarbeitet. Bitte warte, bis die Bilder extrahiert sind.")

    conn.execute("UPDATE projects SET status = 'processing' WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()

    asyncio.create_task(_process_project(project_id, user["id"]))
    return {"ok": True, "message": "Alt-Text-Generierung gestartet"}


async def _process_project(project_id: int, user_id: int):
    # v2.2.3: Clear duplicate cache for this project
    clear_project_cache()
    conn = get_db()
    images = conn.execute(
        "SELECT * FROM images WHERE project_id = ? AND status = 'pending' ORDER BY page_number, image_index",
        (project_id,)
    ).fetchall()

    # processed_images zaehlt nur echte Erfolge (Steve 08.06.2026 — siehe
    # auch regenerate_image). Fehler werden separat ueber den Bild-Status
    # 'error' sichtbar, blaehen aber nicht den "X von Y Bildern generiert"-
    # Zaehler im UI-Kopf auf.
    processed = conn.execute(
        "SELECT COUNT(*) FROM images WHERE project_id = ? AND status = 'done'",
        (project_id,)
    ).fetchone()[0]
    # PDF-Schalter (pdf_langbeschreibung_enabled): bei PDF + Schalter aus den zweiten KI-Pass
    # (Langbeschreibung) ueberspringen = Kostenersparnis (Karbe 03.06.2026). Bild-Projekte unberuehrt.
    _proj = conn.execute("SELECT tool, project_type, use_context, alt_language FROM projects WHERE id = ?", (project_id,)).fetchone()
    # Word-Werkzeug (26.08.2026): .docx kennt nur descr/title, keine Langbeschreibung -> wie PDF.
    skip_langbeschreibung = bool(_proj and (_proj["tool"] in ("pdf", "word") or _proj["project_type"] in ("pdf", "docx"))) and not pdf_langbeschreibung_enabled()
    # KI-Kontext-Schalter (19.06.2026): bei "aus" bekommt die KI keinen Dokument-
    # Kontext (context=""). Default an (Spalte use_context Default 1). Der Vermerk
    # context_mode wird pro Bild gespeichert (nur Anzeige, NIE exportiert).
    use_context = True if (_proj is None or _proj["use_context"] is None) else bool(_proj["use_context"])
    ctx_mode = "mit" if use_context else "ohne"
    # Ausgabesprache (03.07.2026): Projekt-Einstellung, Default Deutsch.
    alt_lang = (_proj["alt_language"] if (_proj and _proj["alt_language"]) else "de")
    if alt_lang not in ALT_TEXT_LANGUAGES:
        alt_lang = "de"
    # Eigener Prompt (06.07.2026): aktiver gespeicherter Prompt des Projekt-
    # Besitzers — geht additiv als user_prompt in die Pipeline. Geloeschter
    # oder fremder Prompt liefert keine Zeile -> leer = Standardverhalten.
    user_prompt = ""
    _up = conn.execute(
        "SELECT up.prompt_text FROM user_prompts up "
        "JOIN projects p ON p.prompt_id = up.id AND p.user_id = up.user_id "
        "WHERE p.id = ?",
        (project_id,)
    ).fetchone()
    if _up and _up["prompt_text"]:
        user_prompt = _up["prompt_text"]

    for img in images:
        # Abo-Etappe-1: Kontingent VOR JEDEM Bild pruefen, nicht nur am Lauf-Start —
        # sonst rutscht eine 500-Bilder-PDF bei Reststand 1 komplett durch.
        # Bei ABO_ENFORCEMENT=off (Zaehl-Phase) ist erlaubt immer True.
        _kontingent = billing.pruefe_kontingent(user_id)
        if not _kontingent["erlaubt"]:
            # Sauber beenden: restliche Bilder bleiben 'pending' und laufen nach
            # Aufstockung ueber einen erneuten Sammellauf weiter.
            print(f"Sammellauf Projekt {project_id}: Monatskontingent erschoepft nach {processed} Bildern — Rest bleibt pending")
            break
        conn.execute("UPDATE images SET status = 'processing' WHERE id = ?", (img["id"],))
        conn.commit()

        try:
            # v2.2: Pass width, height, original_alt to pipeline
            img_width = img["width"] if img["width"] else 0
            img_height = img["height"] if img["height"] else 0
            img_original_alt = img["original_alt"] if img["original_alt"] else ""

            # KI-Kontext-Schalter: bei "aus" leeren Kontext an die KI geben.
            effective_context = img["context_text"] if use_context else ""

            # First pass: general prompt for type detection + alt-text
            result = await asyncio.get_event_loop().run_in_executor(
                None, generate_alt_text, img["image_path"], effective_context, None,
                img_width, img_height, img_original_alt, False, GENERATION_TEMPERATURE,  # force_regenerate=False; leicht lebendigere Standard-Texte
                alt_lang, "", user_prompt
            )

            # Second pass: if complex type detected, re-generate with specialized prompt
            # to get langbeschreibung automatically
            from context_engine import is_complex_type
            detected_type = result.get("bildtyp", "")
            langbeschreibung = result.get("langbeschreibung", "")

            if skip_langbeschreibung:
                # PDF + Schalter aus: zweiten Pass ueberspringen, keine Langbeschreibung speichern.
                langbeschreibung = ""
            elif is_complex_type(detected_type) and not langbeschreibung:
                specialized_result = await asyncio.get_event_loop().run_in_executor(
                    None, generate_alt_text, img["image_path"], effective_context, detected_type,
                    img_width, img_height, img_original_alt, False, GENERATION_TEMPERATURE,
                    alt_lang, "", user_prompt
                )
                if specialized_result.get("langbeschreibung"):
                    langbeschreibung = specialized_result["langbeschreibung"]
                # Use the specialized alt-text if it's better (has content)
                if specialized_result.get("alt_text") and len(specialized_result["alt_text"]) > 10:
                    result["alt_text"] = specialized_result["alt_text"]
                    result["konfidenz"] = specialized_result.get("konfidenz", result.get("konfidenz", "mittel"))
            conn.execute(
                """UPDATE images SET alt_text = ?, image_type = ?, konfidenz = ?, langbeschreibung = ?,
                   needs_review = ?, pipeline_steps = ?, validation_result = ?, context_mode = ?, gen_language = ?, status = 'done' WHERE id = ?""",
                (_append_link_reference(result["alt_text"], effective_context or "", alt_lang), result["bildtyp"], result.get("konfidenz", "mittel"), langbeschreibung,
                 1 if result.get("needs_review") else 0,
                 result.get("pipeline_steps", ""),
                 result.get("validation_result", ""),
                 ctx_mode,
                 alt_lang,
                 img["id"])
            )
            # Nur bei Erfolg den Zaehler hochziehen — Fehler werden separat
            # als status='error' an der Bild-Karte sichtbar.
            processed += 1
            # Bild-Update sofort festschreiben: die Verbuchung nutzt eine EIGENE
            # DB-Verbindung und wuerde sonst am offenen Schreib-Lock scheitern.
            conn.commit()
            # Abo-Etappe-1: nur echte Generierungen verbuchen — Fehllaeufe (except-Pfad)
            # und Cache-Treffer (from_cache, siehe pdf_processor) kosten keine Credits.
            if not result.get("from_cache"):
                billing.verbuche(user_id, "sammellauf", image_id=img["id"])
        except Exception as e:
            import traceback
            print(f"Fehler bei Bild {img['id']} ({img['image_path']}): {e}")
            print(traceback.format_exc())
            # 17.08.2026: alt_text NICHT mehr ueberschreiben. Das Feld wird
            # ausgeliefert (PDF /Alt, JSON, CSV, XLSX, Public API, Demo) und
            # wurde Screenreader-Nutzern als Bildbeschreibung vorgelesen —
            # im Versuch belegt, 23 Faelle auf Produktion. Die Ursache steht
            # im Log (print + traceback direkt darueber), der Zustand im
            # Status. Gleiche Bauweise wie Zeile 5749 (Steve, 08.06.2026):
            # "Bild ehrlich als Fehler markieren".
            conn.execute("UPDATE images SET status = 'error' WHERE id = ?",
                         (img["id"],))

        conn.execute(
            "UPDATE projects SET processed_images = ? WHERE id = ?",
            (processed, project_id)
        )
        conn.commit()

    conn.execute("UPDATE projects SET status = 'done', updated_at = datetime('now') WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()


@app.get("/api/projects/{project_id}/status")
async def get_project_status(project_id: int, user: dict = Depends(get_current_user)):
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    conn.close()
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    return {
        "status": project["status"],
        "total_images": project["total_images"],
        "processed_images": project["processed_images"],
    }


@app.post("/api/images/{image_id}/alt-text")
async def update_alt_text(image_id: int, request: Request, user: dict = Depends(get_current_user)):
    data = await request.json()
    conn = get_db()
    img = conn.execute(
        """SELECT i.id FROM images i
           JOIN projects p ON i.project_id = p.id
           WHERE i.id = ? AND p.user_id = ?""",
        (image_id, user["id"])
    ).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")

    conn.execute(
        "UPDATE images SET alt_text_edited = ? WHERE id = ?",
        (data.get("alt_text", ""), image_id)
    )
    # Also update langbeschreibung if provided
    if "langbeschreibung" in data:
        conn.execute(
            "UPDATE images SET langbeschreibung = ? WHERE id = ?",
            (data.get("langbeschreibung", ""), image_id)
        )
    conn.commit()
    conn.close()
    return {"ok": True}


def _display_alt_text(img):
    """Frontend-Fallback in Python: User-Edit > KI-Output > Original aus Quelle.
    Symmetrisch zur Render-Logik in app.html (textarea-value)."""
    return img["alt_text_edited"] or img["alt_text"] or img["original_alt"]


# 17.08.2026: Zweite Verteidigungslinie an der Systemgrenze.
# Befund: 23 Bilder auf Produktion trugen eine technische Fehlermeldung im
# Feld alt_text; im Versuch belegt, dass sie als /Alt im exportierten PDF
# landet und Screenreader-Nutzern als Bildbeschreibung vorgelesen wird.
# Die Ursache ist mit dem Fix an der Schreibstelle behoben (nur ein aktiver
# Weg, vollstaendige Inventur 17.08.). Dieser Filter deckt zwei Dinge ab:
# die Altdaten, die noch in der Datenbank stehen, und kuenftige Regressionen
# an einer Stelle, die niemand mehr im Blick hat.
# Treffgenauigkeit an ALLEN 2378 Prod-Alt-Texten gemessen: 23 Treffer,
# ausnahmslos echte Fehlermeldungen, kein einziger Fehlgriff. Erhalten
# bleiben z.B. "Navigationsfehler von Studienteilnehmern" und
# "Fehler bei der Analyse der Bodenproben" — der Doppelpunkt bzw. der volle
# Wortlaut ist Teil des Musters.
_FEHLER_PRAEFIXE = (
    "fehler bei der analyse:",
    "pipeline-fehler:",
    "bedrock invoke_model fehlgeschlagen",
    "mistral-aufruf fehlgeschlagen",
    "[modell-antwort konnte nicht verarbeitet werden",
    # 18.08.2026 (Fable 5): zusaetzliche technische Praefixe fuer rohe interne
    # wie externe Meldungen (AWS/boto3/HTTP/Python). An ALLEN 3585 echten
    # Prod-Texten geprueft: 0 Falschtreffer. Nur am Textanfang -> sicher.
    "an error occurred",
    "throttlingexception",
    "server error '",
    "traceback (most recent call",
    "botocore",
    "httperror",
    "connection error",
    "connecterror",
    "readtimeout",
    "500 internal server error",
    "502 bad gateway",
    "503 service",
)


def ist_pipeline_fehlertext(text) -> bool:
    """True, wenn der Text eine technische Fehlermeldung unserer Pipeline ist.

    Bewusst streng: nur unsere eigenen Praefixe und nur am Textanfang.
    """
    if not text:
        return False
    anfang = str(text).strip().lower()
    return any(anfang.startswith(p) for p in _FEHLER_PRAEFIXE)


def _ausgabe_alt_text(img):
    """Alt-Text fuer alles, was das System VERLAESST: Datei-Exporte, Public
    API, oeffentliche Demo.

    Unterschied zu _display_alt_text: technische Fehlermeldungen werden zu
    None. In der angemeldeten Oberflaeche bleibt die Meldung sichtbar — dort
    ist sie ein Hinweis, kein Inhalt. Nach aussen geht sie nie.
    """
    text = _display_alt_text(img)
    if ist_pipeline_fehlertext(text):
        return None
    return text


def _ohne_fehlertext(text):
    """Fuer Ausgabefelder OHNE img-Kontext (langbeschreibung, API-Strings):
    gibt '' zurueck, wenn der Text eine technische Meldung ist, sonst den Text.
    Zusammen mit _ausgabe_alt_text deckt das JEDEN Ausgang ab -> nie eine
    Meldung nach aussen, in KEINEM Textfeld."""
    return "" if ist_pipeline_fehlertext(text) else (text or "")


def _csv_safe(text):
    """CSV/XLSX-Formel-Injection entschaerfen: Zellen, die mit = + - @ (oder
    einem Steuerzeichen) beginnen, bekommen ein vorangestelltes Apostroph.
    Excel/Calc zeigen den Text unveraendert, fuehren ihn aber nicht als Formel aus."""
    s = "" if text is None else str(text)
    if s[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


@app.post("/api/images/{image_id}/feedback")
async def submit_feedback(image_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Submit feedback (good/bad) for a generated alt-text."""
    data = await request.json()
    feedback = data.get("feedback", "")
    if feedback not in ("good", "bad"):
        raise HTTPException(status_code=400, detail="Feedback muss 'good' oder 'bad' sein")

    conn = get_db()
    img = conn.execute(
        """SELECT i.*, p.filename as project_name, p.project_type, p.source_url
           FROM images i
           JOIN projects p ON i.project_id = p.id
           WHERE i.id = ? AND p.user_id = ?""",
        (image_id, user["id"])
    ).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")

    conn.execute("UPDATE images SET feedback = ? WHERE id = ?", (feedback, image_id))
    conn.commit()
    conn.close()

    # Send email notification for ALL feedback (positive + negative)
    alt_text = _display_alt_text(img)
    is_good = feedback == "good"
    color = "#16a34a" if is_good else "#dc2626"
    label = "positiv" if is_good else "negativ"
    emoji = "👍" if is_good else "👎"

    # Build source info based on project type
    project_type = img["project_type"] if "project_type" in img.keys() else "pdf"
    if project_type == "url":
        source_info = f'<p><strong>Quelle:</strong> <a href="{img["source_url"]}">{img["source_url"]}</a></p>'
    elif project_type == "pdf":
        source_info = f'<p><strong>Quelle:</strong> PDF "{img["project_name"]}", Seite {img["page_number"]}, Bild {img["image_index"]}</p>'
    else:
        source_info = f'<p><strong>Quelle:</strong> Einzelbild "{img["project_name"]}"</p>'

    langtext = img["langbeschreibung"] if img["langbeschreibung"] else ""
    lang_section = ""
    if langtext:
        lang_section = f'<p><strong>Langbeschreibung:</strong></p><blockquote style="border-left:3px solid #666;padding-left:1rem;color:#333;">{langtext}</blockquote>'

    email_body = f"""<!DOCTYPE html>
<html lang="de"><head><meta charset="utf-8"></head><body style="font-family:sans-serif;color:#1e293b;max-width:600px;">
<h2 style="color:{color};">{emoji} Alt-Text {label} bewertet</h2>
<p><strong>Benutzer:</strong> {user['email']}</p>
<p><strong>Projekt:</strong> {img['project_name']}</p>
{source_info}
<p><strong>Bildtyp:</strong> {img['image_type']} | <strong>Konfidenz:</strong> {img['konfidenz'] if img['konfidenz'] else 'mittel'}</p>
<p><strong>Bildgröße:</strong> {img['width']}x{img['height']}px</p>
<p><strong>Alt-Text:</strong></p>
<blockquote style="border-left:3px solid {color};padding-left:1rem;color:#333;">{alt_text}</blockquote>
{lang_section}
<p style="color:#64748b;font-size:0.85rem;margin-top:2rem;">InkluDocs Beta-Feedback | Das bewertete Bild ist als Anhang beigefügt.</p>
</body></html>"""

    send_email(
        NOTIFICATION_EMAIL,
        f"InkluDocs: Alt-Text {label} bewertet",
        email_body,
        bcc_admin=False,
        attachment_path=img["image_path"]
    )

    return {"ok": True}


# ─── Regenerate Single Image ────────────────────────────────

@app.post("/api/projects/{project_id}/regenerate/{image_id}")
async def regenerate_image(project_id: int, image_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Regenerate alt-text for a single image with optional specialized prompt."""
    data = await request.json()
    # Abo-Etappe-1: Einzel-Neu-Generieren ist eine echte neue KI-Anfrage und
    # unterliegt dem Monatskontingent (greift erst bei ABO_ENFORCEMENT=on).
    _kontingent = billing.pruefe_kontingent(user["id"])
    if not _kontingent["erlaubt"]:
        raise HTTPException(status_code=429, detail="Monatskontingent erreicht. Bitte Credits nachbuchen oder den Plan wechseln (Abo & Verbrauch).")
    image_type = data.get("image_type")  # Optional: foto, diagramm, karte, etc.
    want_long_desc = data.get("long_description", False)

    conn = get_db()
    img = conn.execute(
        """SELECT i.*, p.use_context AS use_context, p.alt_language AS alt_language, p.prompt_id AS prompt_id FROM images i
           JOIN projects p ON i.project_id = p.id
           WHERE i.id = ? AND i.project_id = ? AND p.user_id = ?""",
        (image_id, project_id, user["id"])
    ).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")

    conn.execute("UPDATE images SET status = 'processing' WHERE id = ?", (image_id,))
    conn.commit()

    try:
        # If long description requested and no explicit type, use the detected type
        effective_type = image_type
        if want_long_desc and not effective_type:
            effective_type = img["image_type"] if img["image_type"] != "unknown" else None

        # v2.2: Pass dimensions and original alt
        regen_width = img["width"] if img["width"] else 0
        regen_height = img["height"] if img["height"] else 0
        regen_original_alt = img["original_alt"] if img["original_alt"] else ""

        # KI-Kontext-Schalter (Projekt-Einstellung): bei "aus" keinen Kontext an die KI.
        _use_context = True if img["use_context"] is None else bool(img["use_context"])
        regen_context = img["context_text"] if _use_context else ""
        regen_ctx_mode = "mit" if _use_context else "ohne"
        # Ausgabesprache: aktuelle Projekt-Einstellung gilt auch fuer Einzel-Neu-Generieren.
        regen_lang = img["alt_language"] if img["alt_language"] in ALT_TEXT_LANGUAGES else "de"
        # Gezielte Variation (05.07.2026): der Text, den der Nutzer gerade sieht
        # (manuell editierte Fassung hat Vorrang), geht als Abgrenzungs-Vorlage mit.
        regen_previous = img["alt_text_edited"] or img["alt_text"] or ""
        # Eigener Prompt: aktive Projekt-Einstellung gilt auch beim Einzel-Neu-Generieren.
        regen_user_prompt = ""
        if img["prompt_id"]:
            _up = conn.execute(
                "SELECT prompt_text FROM user_prompts WHERE id = ? AND user_id = ?",
                (img["prompt_id"], user["id"])
            ).fetchone()
            if _up and _up["prompt_text"]:
                regen_user_prompt = _up["prompt_text"]

        # T6 (03.05.2026): Cache evictieren BEVOR die Pipeline laeuft. Ein expliziter
        # Re-Generate-Klick soll alle Cache-Varianten dieses Bildes loeschen (auch andere
        # image_type_override-Varianten), damit kein Mischzustand entsteht. force_regenerate
        # springt zusaetzlich am Cache-Lookup vorbei in generate_alt_text.
        from pdf_processor import _get_image_hash
        from cache import evict_by_content_hash
        try:
            content_hash = _get_image_hash(img["image_path"])
            evicted = evict_by_content_hash(content_hash)
            if evicted:
                print(f"T6 regenerate_image: {evicted} cache-Eintrag(e) fuer {img['image_path']} evictiert")
        except FileNotFoundError:
            pass  # Bild physikalisch weg — Pipeline wird das melden

        result = await asyncio.get_event_loop().run_in_executor(
            None, generate_alt_text, img["image_path"], regen_context, effective_type,
            regen_width, regen_height, regen_original_alt, True, REGENERATE_TEMPERATURE,  # force_regenerate=True; Variation beim Einzel-Neu-Generieren
            regen_lang, regen_previous, regen_user_prompt
        )

        langbeschreibung = result.get("langbeschreibung", "")
        conn.execute(
            """UPDATE images SET alt_text = ?, image_type = ?, konfidenz = ?,
               langbeschreibung = ?, alt_text_edited = NULL,
               needs_review = ?, pipeline_steps = ?, validation_result = ?, context_mode = ?, gen_language = ?, status = 'done' WHERE id = ?""",
            (_append_link_reference(result["alt_text"], regen_context or "", regen_lang), result["bildtyp"], result.get("konfidenz", "mittel"),
             langbeschreibung,
             1 if result.get("needs_review") else 0,
             result.get("pipeline_steps", ""),
             result.get("validation_result", ""),
             regen_ctx_mode,
             regen_lang,
             image_id)
        )
        # processed_images zaehlt nur ECHTE Erfolge (Steve 08.06.2026):
        # damit "X von Y Bildern generiert" im UI-Kopf nicht mit Fehlern
        # aufgeblaeht wird. Fehler werden separat ueber die Bild-Karten
        # sichtbar (status='error').
        processed_count = conn.execute(
            "SELECT COUNT(*) FROM images WHERE project_id = ? AND status = 'done'",
            (project_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE projects SET processed_images = ? WHERE id = ?",
            (processed_count, project_id)
        )
        conn.commit()
        conn.close()

        # Abo-Etappe-1: erfolgreiches Einzel-Neu-Generieren = 1 Credit
        # (laeuft mit force_regenerate, also nie aus dem Cache).
        billing.verbuche(user["id"], "einzeln", image_id=image_id)

        return {
            "ok": True,
            "alt_text": result["alt_text"],
            "bildtyp": result["bildtyp"],
            "konfidenz": result.get("konfidenz", "mittel"),
            "langbeschreibung": langbeschreibung,
        }
    except Exception as e:
        # Fehlerpfad: Bild ehrlich als Fehler markieren (Steve 08.06.2026).
        # Vorher wurde status='done' gesetzt — das hat die Bild-Karte als
        # erfolgreich generiert ausgewiesen, obwohl gar nichts produziert
        # wurde. Jetzt 'error' + Pflege von processed_images (nur Erfolge),
        # damit der UI-Zaehler korrekt bleibt.
        conn.execute("UPDATE images SET status = 'error' WHERE id = ?", (image_id,))
        processed_count = conn.execute(
            "SELECT COUNT(*) FROM images WHERE project_id = ? AND status = 'done'",
            (project_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE projects SET processed_images = ? WHERE id = ?",
            (processed_count, project_id)
        )
        conn.commit()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Fehler bei der Neugenerierung: {str(e)}")


# ─── Export Routes ───────────────────────────────────────────

# ─── Multi-Datei Export-Helpers (08.06.2026) ─────────────────
#
# Phase 1 dreht jedes "ein PDF pro Projekt" in "mehrere PDFs pro Projekt".
# Damit der bestehende Export-Code nicht doppelt geschrieben werden muss,
# kapseln wir die Datei-Erzeugung pro Dokument in kleinen Helfern und rufen
# diese aus den oeffentlichen Endpunkten heraus. Auswahl-Logik:
#   - kein document_id Parameter   -> alle Dokumente, gepackt als ZIP
#   - ein document_id Parameter    -> nur dieses Dokument, direkter Download
# Der ZIP-Name kann optional ueber den 'filename'-Body-Parameter gesetzt
# werden (ohne Erweiterung), genauso wie bei Einzelexports.

import zipfile


def _safe_filename_component(name: str, fallback: str = "datei") -> str:
    """Macht aus einem beliebigen Anzeigenamen einen Datei-System-tauglichen
    Bestandteil (keine Slashes, keine Steuerzeichen, max. 120 Zeichen).
    Wir bleiben bewusst konservativ, damit ZIP-Innennamen auch unter Windows
    sauber bleiben."""
    if not name:
        return fallback
    s = re.sub(r"[\\/\x00-\x1f<>:\"|?*]", "_", str(name)).strip(" .")
    return s[:120] or fallback


def _content_disposition(dateiname: str) -> str:
    """Content-Disposition nach RFC 6266: ASCII-Fallback + filename* in UTF-8.
    Kopfzeilen sind latin-1 — ein Projektname mit Zeichen ausserhalb davon
    (typografische Anfuehrungszeichen, Gedankenstrich) liess die Antwort sonst
    mit 500 abbrechen (Review-Befund 27.08.2026, gleiches Muster wie im
    Quickinfo-Werkzeug)."""
    ascii_name = dateiname.encode("ascii", "ignore").decode("ascii").replace('"', "") or "download"
    return "attachment; filename=\"%s\"; filename*=UTF-8''%s" % (ascii_name, urllib.parse.quote(dateiname))


def _warnings_header(warnungen) -> str:
    """X-Export-Warnings als reines ASCII (JSON-Escapes) — HTTP-Kopfzeilen vertragen
    kein UTF-8; vorher brach die Antwort bei Umlauten/Anfuehrungszeichen in einer
    Warnung mit 500 ab, NACHDEM die Credits verbucht waren (Review 27.08.2026)."""
    return json.dumps(list(warnungen), ensure_ascii=True)


async def _read_export_options(request: Optional[Request]) -> tuple[Optional[int], Optional[str]]:
    """Liest optionale Export-Parameter (document_id + filename) aus dem Request-Body.
    Beide Felder sind optional; fehlt der Body komplett, faellt alles auf None zurueck."""
    if request is None:
        return None, None
    try:
        body = await request.json()
    except Exception:
        return None, None
    doc_id = body.get("document_id") if isinstance(body, dict) else None
    if doc_id is not None:
        try:
            doc_id = int(doc_id)
        except (TypeError, ValueError):
            doc_id = None
    name = (body.get("filename") if isinstance(body, dict) else None) or None
    if name:
        name = str(name).strip()
        name = re.sub(r"\.[A-Za-z0-9]{1,5}$", "", name)
        name = _safe_filename_component(name)
    return doc_id, name


def _doc_label(doc: dict) -> str:
    """Anzeigename eines Dokuments fuer Dateinamen (display_name -> original_filename ohne .pdf)."""
    raw = (doc.get("display_name") or "").strip() or (doc.get("original_filename") or "").strip()
    raw = re.sub(r"\.(pdf|docx)$", "", raw, flags=re.IGNORECASE)
    return _safe_filename_component(raw, f"dokument_{doc.get('doc_index', 1)}")


def _load_pdf_export_units(project: dict, user_id: int, document_id: Optional[int]) -> list[dict]:
    """Lieferst Liste aus dicts mit doc-Metadaten + alt-Texten + image-Metadaten
    fuer jedes Dokument, das exportiert werden soll. Phase 1 verarbeitet nur
    PDF-Werkzeuge — andere Tools liefern eine einzelne synthetische Einheit
    (alles im Projekt zusammen)."""
    conn = get_db()
    docs = conn.execute(
        """SELECT * FROM documents WHERE project_id = ? ORDER BY doc_index""",
        (project["id"],)
    ).fetchall()
    if document_id is not None:
        docs = [d for d in docs if d["id"] == document_id]
        if not docs:
            conn.close()
            raise HTTPException(status_code=404, detail="Dokument nicht gefunden")

    units = []
    for d in docs:
        imgs = conn.execute(
            "SELECT * FROM images WHERE document_id = ? ORDER BY page_number, image_index",
            (d["id"],)
        ).fetchall()
        units.append({"doc": dict(d), "images": [dict(i) for i in imgs]})
    conn.close()

    # Fallback fuer den Uebergang: ein PDF-Projekt OHNE documents-Eintraege
    # (z.B. Race-Bedingung beim Erst-Migrate) wird wie ein einzelnes Dokument
    # behandelt, das auf project.filename/original_path zeigt.
    if not units:
        conn = get_db()
        imgs = conn.execute(
            "SELECT * FROM images WHERE project_id = ? ORDER BY page_number, image_index",
            (project["id"],)
        ).fetchall()
        conn.close()
        units.append({
            "doc": {
                "id": None,
                "doc_index": 1,
                "original_filename": project.get("filename") or "dokument.pdf",
                "display_name": None,
                "original_path": project.get("original_path") or "",
                "extraction_method": project.get("extraction_method") or "fitz",
            },
            "images": [dict(i) for i in imgs],
        })
    return units


def _exportable_alt_text(img) -> Optional[str]:
    """Alt-Text eines Bildes fuer den PDF-Export, oder None wenn das Bild
    NICHT exportiert werden soll.

    Regel (12.06.2026): Nur Bilder mit echtem Text (oder der expliziten
    Markierung "dekorativ") kommen in den Export. Unbearbeitete Bilder
    (leerer Text, z.B. status='pending') werden uebersprungen — sie bleiben
    in der PDF exakt wie im Original, inklusive eines evtl. vorhandenen
    Original-Alt-Textes. Vorher schrieb der Export leere /Alt-Eintraege in
    die Datei (Befund 12.06.2026: Export eines teilbearbeiteten Projekts
    erzeugte 6 leere Alt-Texte, die Pruefwerkzeuge als Fehler werten)."""
    alt_text = _ausgabe_alt_text(img)
    if alt_text is None:
        return None
    if alt_text == "dekorativ" or alt_text.strip():
        return alt_text
    return None


def _build_pdf_for_document(unit: dict, output_dir: str,
                            custom_title: Optional[str] = None) -> tuple[str, dict]:
    """Erzeugt die exportierte PDF fuer EIN Dokument (alle Alt-Texte
    eingebettet). Gibt (output_path, header_info) zurueck. Header_info
    enthaelt die gleichen Metriken wie der Single-Export davor.
    custom_title: vom Nutzer beim Export vergebener Name — wird als
    PDF-Dokumenttitel uebernommen (siehe finalize_export_pdf)."""
    doc = unit["doc"]
    images = unit["images"]
    extraction_method = doc.get("extraction_method") or "fitz"

    alt_texts = {}
    alt_texts_by_lfnr = {}
    image_metadata = []

    for img in images:
        alt_text = _exportable_alt_text(img)
        if alt_text is not None and img.get("xref"):
            alt_texts[img["xref"]] = alt_text
        if alt_text is not None and img.get("image_index"):
            alt_texts_by_lfnr[int(img["image_index"])] = alt_text
        bbox = None
        if img.get("bbox_x0") is not None:
            bbox = (img["bbox_x0"], img["bbox_y0"], img["bbox_x1"], img["bbox_y1"])
        image_metadata.append({
            "xref": img.get("xref"),
            "page_number": img.get("page_number"),
            "is_vector": bool(img.get("is_vector")) if img.get("is_vector") is not None else False,
            "bbox": bbox,
            "alt_text": alt_text,
            "image_path": img.get("image_path"),
        })

    os.makedirs(output_dir, exist_ok=True)
    base = doc.get("original_filename") or f"dokument_{doc.get('doc_index', 1)}.pdf"
    output_path = os.path.join(output_dir, f"inkludocs_{_safe_filename_component(base, base)}")
    info: dict = {}

    if extraction_method == "pdfix":
        try:
            import pdfix_roundtrip as _pdfix
            count = _pdfix.import_alt_texts_pdfix(
                doc["original_path"], output_path,
                alt_texts_by_lfnr, work_dir=output_dir)
            info["method"] = "pdfix"
            info["tagged"] = count
            info["total"] = len(alt_texts_by_lfnr)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"PDFix-Export fehlgeschlagen: {str(e)}")
    else:
        try:
            from pdf_export import write_alt_texts_to_pdf
            result = write_alt_texts_to_pdf(doc["original_path"], output_path, alt_texts, image_metadata)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Export fehlgeschlagen: {str(e)}")
        if isinstance(result, dict):
            info["method"] = "fitz"
            info["tagged"] = result.get("tagged_count", 0)
            info["total"] = len(alt_texts)
            warnings = result.get("warnings") or []
            if warnings:
                info["warnings"] = warnings

    # Gemeinsamer Abschluss fuer BEIDE Pfade (12.06.2026): Dokumentsprache,
    # Dokumenttitel (WCAG 3.1.1 / 2.4.2) und verwaiste /Alt-Altlasten der
    # Quell-PDF entfernen. Titel-Prioritaet: Export-Name des Nutzers >
    # Umbenennung in InkluDocs (display_name) > vorhandener Titel der
    # Quell-PDF > Dateiname ohne Endung. Der Dokument-Inhalt bleibt unberuehrt.
    explicit_title = (custom_title or "").strip() or (doc.get("display_name") or "").strip() or None
    fallback_title = re.sub(r"\.pdf$", "", (doc.get("original_filename") or "").strip(),
                            flags=re.IGNORECASE) or None
    try:
        from pdf_export import finalize_export_pdf
        info["a11y"] = finalize_export_pdf(output_path, title=explicit_title,
                                           fallback_title=fallback_title)
    except Exception as e:
        # Der Abschluss-Schritt darf den Export nicht scheitern lassen:
        # Die Alt-Texte sind zu diesem Zeitpunkt bereits korrekt gesetzt,
        # Sprache/Titel/Aufraeumen sind Zusatznutzen. Fehler aber loggen.
        print(f"WARNUNG: finalize_export_pdf fehlgeschlagen fuer {output_path}: {e}")

    return output_path, info


@app.post("/api/projects/{project_id}/export")
async def export_pdf(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """PDF-Export. Multi-Datei (08.06.2026):
    - ohne document_id im Body: alle Dokumente als ZIP
    - mit document_id: nur dieses Dokument, direkter Download
    """
    document_id, custom_name = await _read_export_options(request)
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    project_type = project["project_type"] if "project_type" in project.keys() else "pdf"
    if project_type != "pdf":
        conn.close()
        raise HTTPException(status_code=400, detail="PDF-Export ist nur fuer PDF-Projekte verfuegbar")
    project = dict(project)
    conn.close()

    units = _load_pdf_export_units(project, user["id"], document_id)
    output_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project["id"]), "_export")
    os.makedirs(output_dir, exist_ok=True)

    # Punkt 4 (04.08.2026, Michaels Modell): PDF-Export kostet 5 Credits —
    # pro Export-VORGANG, nur fuer Bezahl-Konten. Free bleibt bis zur
    # Wasserzeichen-Runde (Punkt 3) unverbucht: dort bekommt Free seinen
    # Wasserzeichen-Export, erst dann waere ein Zaehlen ehrlich.
    _abo_export = billing.pruefe_kontingent(user["id"])
    _export_kostenpflichtig = _abo_export.get("plan") != "free"
    if _export_kostenpflichtig and not _abo_export.get("erlaubt", True):
        raise HTTPException(status_code=429,
                            detail="Credit-Kontingent erschoepft. Bitte Credits nachbuchen oder bis zum Monatswechsel warten")

    if document_id is not None or len(units) == 1:
        # Einzelne Datei zurueckgeben (direkter Download, kein ZIP).
        unit = units[0]
        output_path, info = _build_pdf_for_document(unit, output_dir,
                                                    custom_title=custom_name)
        headers = {}
        if info.get("method"):
            headers["X-Export-Method"] = str(info["method"])
        if "tagged" in info:
            headers["X-Export-Tagged"] = str(info["tagged"])
        if "total" in info:
            headers["X-Export-Total"] = str(info["total"])
        if info.get("warnings"):
            headers["X-Export-Warnings"] = _warnings_header(info["warnings"])
        download_base = custom_name or _doc_label(unit["doc"])
        # Nur ERFOLGREICHE Exporte verbuchen: Antwort zuerst bauen, dann verbuchen
        # (Review 27.08.2026 — vorher konnte die Antwort nach der Verbuchung scheitern).
        response = FileResponse(
            output_path,
            filename=f"inkludocs_{download_base}.pdf",
            media_type="application/pdf",
            headers=headers,
        )
        if _export_kostenpflichtig:
            billing.verbuche(user["id"], "export", aktion="pdf_export")
        return response

    # Alle Dokumente -> ZIP.
    zip_base = custom_name or _safe_filename_component(project.get("name") or project.get("filename") or "projekt")
    zip_path = os.path.join(output_dir, f"{zip_base}_alle_pdfs.zip")
    aggregated_warnings: list[str] = []
    total_tagged = 0
    total_images = 0
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Nummerierung nach Anzeige-Position (1..N), passend zur „Dokument N"-
        # Anzeige im Frontend — NICHT nach doc_index, der nach Loeschungen Luecken
        # haben kann.
        for pos, unit in enumerate(units, start=1):
            out_path, info = _build_pdf_for_document(unit, output_dir)
            inner = f"{pos:02d}_{_doc_label(unit['doc'])}.pdf"
            zf.write(out_path, arcname=inner)
            total_tagged += int(info.get("tagged", 0) or 0)
            total_images += int(info.get("total", 0) or 0)
            for w in info.get("warnings", []) or []:
                aggregated_warnings.append(f"[{inner}] {w}")
    headers = {
        "X-Export-Tagged": str(total_tagged),
        "X-Export-Total": str(total_images),
    }
    if aggregated_warnings:
        headers["X-Export-Warnings"] = _warnings_header(aggregated_warnings)
    # Ein Export-Vorgang = 5 Credits, auch fuer das Alle-Dokumente-ZIP
    # (siehe billing.AKTIONS_PREISE); nur nach erfolgreichem Bau der Antwort.
    response = FileResponse(
        zip_path,
        filename=f"{zip_base}_alle_pdfs.zip",
        media_type="application/zip",
        headers=headers,
    )
    if _export_kostenpflichtig:
        billing.verbuche(user["id"], "export", aktion="pdf_export")
    return response


# ─── WORD-EXPORT (26.08.2026): Word-Datei mit Alt-Texten zurueckgeben ────────
#
# Spiegelbild des PDF-Exports: ein Dokument -> direkter Download der .docx,
# mehrere -> ZIP. Die Alt-Texte gehen ueber _exportable_alt_text (gleiche
# Regeln: leer = Bild unberuehrt lassen, "dekorativ" = Kennzeichen setzen).
# Der Anker images.docx_anker fuehrt zum Bildelement. Die Originaldatei wird
# NIE veraendert; docx_export schreibt ein neues Zip in RESULTS_DIR/.../_export.

_uploads_root = os.path.realpath(UPLOAD_DIR) + os.sep


def _build_docx_for_document(unit: dict, output_dir: str, custom_title: Optional[str] = None) -> tuple[str, dict]:
    doc = unit["doc"]
    src = doc.get("original_path") or ""
    # Nur Dateien aus dem Upload-Verzeichnis lesen (Schutz gegen manipulierte Pfade).
    if not src or not os.path.realpath(src).startswith(_uploads_root) or not os.path.isfile(src):
        raise HTTPException(status_code=404, detail="Die Originaldatei dieses Dokuments ist nicht mehr vorhanden")
    alt_texts: dict[str, str] = {}
    for img in unit["images"]:
        anker = (img.get("docx_anker") or "").strip()
        if not anker:
            continue
        text = _exportable_alt_text(img)
        if text is None:
            continue
        alt_texts[anker] = text
    base = custom_title or _doc_label(doc)
    out_path = os.path.join(output_dir, "inkludocs_" + _safe_filename_component(base) + ".docx")
    try:
        erg = docx_export.write_alt_texts_to_docx(src, out_path, alt_texts)
    except DocxFehler as e:
        raise HTTPException(status_code=400, detail=str(e))
    return out_path, {
        "method": "docx",
        "tagged": erg.geschrieben,
        "total": len(unit["images"]),
        "warnings": erg.warnungen,
    }


DOCX_MEDIA = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


@app.post("/api/projects/{project_id}/export/docx")
async def export_docx(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Word-Export: ohne document_id alle Dokumente als ZIP, mit document_id
    nur dieses Dokument als direkter Download. Kostet wie der PDF-Export
    (billing.AKTIONS_PREISE["docx_export"]), nur fuer Bezahl-Konten."""
    document_id, custom_name = await _read_export_options(request)
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    project = dict(project)
    conn.close()
    if project.get("project_type") != "docx":
        raise HTTPException(status_code=400, detail="Word-Export ist nur fuer Word-Projekte verfuegbar")

    units = _load_pdf_export_units(project, user["id"], document_id)
    output_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project["id"]), "_export")
    os.makedirs(output_dir, exist_ok=True)

    _abo_export = billing.pruefe_kontingent(user["id"])
    _export_kostenpflichtig = _abo_export.get("plan") != "free"
    if _export_kostenpflichtig and not _abo_export.get("erlaubt", True):
        raise HTTPException(status_code=429,
                            detail="Credit-Kontingent erschoepft. Bitte Credits nachbuchen oder bis zum Monatswechsel warten")

    if document_id is not None or len(units) == 1:
        unit = units[0]
        output_path, info = _build_docx_for_document(unit, output_dir, custom_title=custom_name)
        headers = {"X-Export-Method": "docx", "X-Export-Tagged": str(info["tagged"]),
                   "X-Export-Total": str(info["total"])}
        if info.get("warnings"):
            headers["X-Export-Warnings"] = _warnings_header(info["warnings"])
        download_base = custom_name or _doc_label(unit["doc"])
        response = FileResponse(output_path, filename=f"inkludocs_{download_base}.docx",
                                media_type=DOCX_MEDIA, headers=headers)
        if _export_kostenpflichtig:
            billing.verbuche(user["id"], "export", aktion="docx_export")
        return response

    zip_base = custom_name or _safe_filename_component(project.get("name") or project.get("filename") or "projekt")
    zip_path = os.path.join(output_dir, f"{zip_base}_alle_word.zip")
    aggregated_warnings: list[str] = []
    total_tagged = 0
    total_images = 0
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pos, unit in enumerate(units, start=1):
            out_path, info = _build_docx_for_document(unit, output_dir)
            inner = f"{pos:02d}_{_doc_label(unit['doc'])}.docx"
            zf.write(out_path, arcname=inner)
            total_tagged += int(info.get("tagged", 0) or 0)
            total_images += int(info.get("total", 0) or 0)
            for w in info.get("warnings", []) or []:
                aggregated_warnings.append(f"[{inner}] {w}")
    headers = {"X-Export-Tagged": str(total_tagged), "X-Export-Total": str(total_images)}
    if aggregated_warnings:
        headers["X-Export-Warnings"] = _warnings_header(aggregated_warnings)
    response = FileResponse(zip_path, filename=f"{zip_base}_alle_word.zip",
                            media_type="application/zip", headers=headers)
    if _export_kostenpflichtig:
        billing.verbuche(user["id"], "export", aktion="docx_export")
    return response


# ─── QUICKINFO-WERKZEUG (27.08.2026): eigener Router, eigene Tabellen ────────
# Alle Formular-Endpunkte liegen in formular_api.py; hier werden nur die
# gemeinsamen Regeln (Login, DB, Verzeichnisse, Credits, Export-Helfer)
# uebergeben, damit es keinen Ringimport gibt und die Regeln an EINER Stelle
# bleiben.
app.include_router(formular_api.build_router(formular_api.Deps(
    get_current_user=get_current_user,
    get_db=get_db,
    upload_dir=UPLOAD_DIR,
    results_dir=RESULTS_DIR,
    billing=billing,
    read_export_options=_read_export_options,
    safe_filename_component=_safe_filename_component,
    doc_label=_doc_label,
    csv_safe=_csv_safe,
    # Gast-Ansicht Formulare (28.08.2026): derselbe Wachposten wie bei den Bild-Freigaben.
    # (Lambda, weil _require_guest weiter unten in dieser Datei definiert ist und
    # zum Zeitpunkt des include_router noch nicht existiert.)
    require_guest=lambda request, token: _require_guest(request, token),
    guest_session=get_guest_session,
)))


# ─── Multi-Datei Export: JSON / CSV / XLSX (08.06.2026) ────────
#
# Diese drei Formate koennen INHALTLICH wahlweise nur ein Dokument
# oder alle gemeinsam darstellen. Wir gehen den simplen, fuer Nutzer
# erwartbaren Weg:
#   - document_id im Body  -> nur dieses Dokument, eine Datei
#   - sonst                -> alle Dokumente als ZIP (eine Datei pro Doc)
# Web- und Grafik-Projekte (kein documents-Eintrag) bleiben ein einzelnes
# logisches Dokument und liefern unveraendert eine flache Einzeldatei.

def _load_export_units_for_table(project: dict, document_id: Optional[int]) -> list[dict]:
    """Wie _load_pdf_export_units, aber NICHT auf project_type='pdf' beschraenkt.
    Web/Grafik-Projekte haben kein documents-Eintrag -> Fallback auf ein
    einziges synthetisches Dokument (alle Bilder zusammen)."""
    conn = get_db()
    docs = conn.execute(
        "SELECT * FROM documents WHERE project_id = ? ORDER BY doc_index",
        (project["id"],)
    ).fetchall()
    if docs:
        if document_id is not None:
            docs = [d for d in docs if d["id"] == document_id]
            if not docs:
                conn.close()
                raise HTTPException(status_code=404, detail="Dokument nicht gefunden")
        units = []
        for d in docs:
            imgs = conn.execute(
                "SELECT * FROM images WHERE document_id = ? ORDER BY page_number, image_index",
                (d["id"],)
            ).fetchall()
            units.append({"doc": dict(d), "images": [dict(i) for i in imgs]})
        conn.close()
        return units
    # Kein Dokument-Eintrag (Web/Grafik oder noch nicht migriert).
    imgs = conn.execute(
        "SELECT * FROM images WHERE project_id = ? ORDER BY page_number, image_index",
        (project["id"],)
    ).fetchall()
    conn.close()
    return [{
        "doc": {
            "id": None,
            "doc_index": 1,
            "original_filename": project.get("filename") or "export",
            "display_name": None,
        },
        "images": [dict(i) for i in imgs],
    }]


def _build_json_bytes(unit: dict, project_name: str) -> bytes:
    export_data = {
        "projekt": project_name,
        "dokument": _doc_label(unit["doc"]),
        "bilder": [],
    }
    for img in unit["images"]:
        alt_text = _ausgabe_alt_text(img)
        entry = {"alt_text": alt_text or ""}
        lang = _ohne_fehlertext(img.get("langbeschreibung"))
        if lang:
            entry["langbeschreibung"] = lang
        export_data["bilder"].append(entry)
    return json.dumps(export_data, ensure_ascii=False, indent=2).encode("utf-8")


def _build_csv_bytes(unit: dict) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Alt-Text", "Langbeschreibung"])
    for img in unit["images"]:
        alt_text = _ausgabe_alt_text(img)
        writer.writerow([_csv_safe(alt_text or ""), _csv_safe(_ohne_fehlertext(img.get("langbeschreibung")))])
    return output.getvalue().encode("utf-8-sig")


def _build_xlsx_bytes(unit: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Alt-Texte"
    ws["A1"] = "Bild"
    ws["B1"] = "Alt-Text"
    ws["C1"] = "Langbeschreibung"
    for cell in [ws["A1"], ws["B1"], ws["C1"]]:
        cell.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60

    for i, img in enumerate(unit["images"]):
        row = i + 2
        alt_text = _ausgabe_alt_text(img)
        langbeschreibung = _ohne_fehlertext(img.get("langbeschreibung"))
        img_path = img.get("image_path") or ""
        img_filename = os.path.basename(img_path) if img_path else "unbekannt"
        ws[f"A{row}"] = img_filename
        ws[f"A{row}"].alignment = Alignment(vertical="top")
        if img_path and os.path.exists(img_path):
            try:
                export_img_path = img_path
                if img_path.lower().endswith((".webp", ".avif", ".heic", ".heif")):
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(img_path)
                    if pil_img.mode not in ("RGB", "L"):
                        pil_img = pil_img.convert("RGB")
                    png_path = img_path.rsplit(".", 1)[0] + "_export.png"
                    pil_img.save(png_path, format="PNG")
                    export_img_path = png_path
                xl_img = XlImage(export_img_path)
                max_w = 150
                max_h = 120
                ratio = min(max_w / xl_img.width, max_h / xl_img.height, 1.0)
                xl_img.width = int(xl_img.width * ratio)
                xl_img.height = int(xl_img.height * ratio)
                ws.row_dimensions[row].height = max(xl_img.height * 0.75, 60)
                ws.add_image(xl_img, f"A{row}")
            except Exception:
                pass
        ws[f"B{row}"] = _csv_safe(alt_text or "")
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"C{row}"] = _csv_safe(langbeschreibung)
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def _table_export_dispatch(project_id: int, request: Request, user: dict, fmt: str):
    """Gemeinsamer Code fuer JSON/CSV/XLSX-Export."""
    document_id, custom_name = await _read_export_options(request)
    conn = get_db()
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    project = dict(project)
    conn.close()

    units = _load_export_units_for_table(project, document_id)
    project_name = project.get("name") or project.get("filename") or "Projekt"

    if fmt == "json":
        media = "application/json"
        suffix = ".json"
        single = lambda u: _build_json_bytes(u, project_name)
    elif fmt == "csv":
        media = "text/csv"
        suffix = ".csv"
        single = lambda u: _build_csv_bytes(u)
    elif fmt == "xlsx":
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        suffix = ".xlsx"
        single = lambda u: _build_xlsx_bytes(u)
    else:
        raise HTTPException(status_code=400, detail="Unbekanntes Format")

    if document_id is not None or len(units) == 1:
        unit = units[0]
        data = single(unit)
        base = custom_name or _doc_label(unit["doc"])
        return StreamingResponse(
            io.BytesIO(data),
            media_type=media,
            headers={"Content-Disposition": _content_disposition(f"inkludocs_{base}{suffix}")}
        )

    # Mehrere Dokumente -> ZIP
    zip_base = custom_name or _safe_filename_component(project_name)
    output_dir = os.path.join(RESULTS_DIR, str(user["id"]), str(project["id"]), "_export")
    os.makedirs(output_dir, exist_ok=True)
    zip_path = os.path.join(output_dir, f"{zip_base}_alle_{fmt}.zip")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Nummerierung nach Anzeige-Position (1..N), siehe PDF-Export oben.
        for pos, unit in enumerate(units, start=1):
            inner = f"{pos:02d}_{_doc_label(unit['doc'])}{suffix}"
            zf.writestr(inner, single(unit))
    return FileResponse(
        zip_path,
        filename=f"{zip_base}_alle_{fmt}.zip",
        media_type="application/zip",
    )


@app.post("/api/projects/{project_id}/export/json")
async def export_json(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Multi-Datei: ohne document_id alle Dokumente als ZIP, sonst Einzeldatei."""
    return await _table_export_dispatch(project_id, request, user, "json")


@app.post("/api/projects/{project_id}/export/csv")
async def export_csv(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Multi-Datei: ohne document_id alle Dokumente als ZIP, sonst Einzeldatei."""
    return await _table_export_dispatch(project_id, request, user, "csv")


@app.post("/api/projects/{project_id}/export/xlsx")
async def export_xlsx(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Multi-Datei: ohne document_id alle Dokumente als ZIP, sonst Einzeldatei."""
    return await _table_export_dispatch(project_id, request, user, "xlsx")


@app.post("/api/projects/{project_id}/export/summary")
async def export_summary(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Ehrlichkeits-Zusammenfassung VOR dem Export (18.08.2026, Fable 5).
    Zaehlt, wie viele Bilder eine echte Beschreibung bekommen und wie viele
    uebersprungen werden (mit Grund: Fehler oder noch nicht generiert). So
    sieht der Kunde vor dem Herunterladen, ob sein Dokument Luecken hat.
    Es wird NUR gezaehlt — nichts davon wandert in den Export oder den Inhalt.
    Zaehlweise identisch zum Export: _ausgabe_alt_text (Fehlertext-gefiltert)."""
    document_id, _name = await _read_export_options(request)
    conn = get_db()
    projekt = conn.execute(
        "SELECT id FROM projects WHERE id = ? AND user_id = ?", (project_id, user["id"])
    ).fetchone()
    if not projekt:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    if document_id is not None:
        rows = conn.execute(
            "SELECT * FROM images WHERE project_id = ? AND document_id = ?",
            (project_id, document_id),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM images WHERE project_id = ?", (project_id,)
        ).fetchall()
    conn.close()

    total = beschrieben = dekorativ = fehler = offen = 0
    for r in rows:
        img = dict(r)
        total += 1
        ausgabe = _ausgabe_alt_text(img)
        if ausgabe == "dekorativ":
            dekorativ += 1
        elif ausgabe and ausgabe.strip():
            beschrieben += 1
        elif img.get("status") == "error":
            fehler += 1
        else:
            offen += 1
    return JSONResponse(content={
        "total": total,
        "beschrieben": beschrieben,
        "dekorativ": dekorativ,
        "uebersprungen": fehler + offen,
        "fehler": fehler,
        "offen": offen,
    })


# ─── Public API ──────────────────────────────────────────────

@app.post("/api/v1/alt-text")
async def api_generate_alt_text(request: Request):
    """Public API endpoint for alt-text generation. Requires X-API-Key header.
    Accepts multipart/form-data (file upload) or application/json (base64 image)."""
    import base64 as b64module

    api_user = get_api_user(request)
    api_key_id = api_user["api_key_id"]

    # Daily limit check — Missbrauchsbremse, pro Konto einstellbar
    # (11.08.2026). get_api_user liefert nur die Schluessel-Kurzform,
    # deshalb hier die Konto-Zeile fuer das individuelle Limit nachladen.
    tageslimit = effektives_api_tageslimit(get_user_by_id(api_user["id"]) or {})
    daily_used = get_daily_api_count(api_user["id"])
    daily_remaining = max(0, tageslimit - daily_used)
    if daily_used >= tageslimit:
        raise HTTPException(
            status_code=429,
            detail=f"Tageslimit erreicht ({tageslimit} Bilder pro Tag). Das Limit wird um Mitternacht (UTC) zurueckgesetzt.",
            headers={"X-RateLimit-Limit": str(tageslimit), "X-RateLimit-Remaining": "0"},
        )

    # Rate limiting
    rate_info = check_api_rate_limit(api_key_id)

    # Abo-Etappe-1: API zieht aus DEMSELBEN Credit-Topf wie die App
    # (greift erst bei ABO_ENFORCEMENT=on; maschinenlesbares Fehlerformat folgt in Etappe 4).
    _kontingent = billing.pruefe_kontingent(api_user["id"])
    if not _kontingent["erlaubt"]:
        raise HTTPException(status_code=429, detail="Monatskontingent erreicht. Bitte Credits nachbuchen oder den Plan wechseln.")

    content_type = request.headers.get("content-type", "")
    context_text = ""
    language = "de"
    image_type = None
    content = None
    ext = ".jpg"

    if "application/json" in content_type:
        # JSON mode: base64-encoded image
        try:
            data = await request.json()
        except Exception:
            raise HTTPException(status_code=400, detail="Ungueltiges JSON")

        image_b64 = data.get("image_base64", "")
        if not image_b64:
            log_api_usage(api_key_id, api_user["id"], success=False,
                          error_message="Kein Bild (Base64)")
            raise HTTPException(status_code=400, detail="Feld 'image_base64' fehlt oder ist leer")

        context_text = data.get("context", "")
        language = data.get("language", "de")
        image_type = data.get("image_type")

        # Strip data URI prefix if present (e.g. "data:image/png;base64,...")
        if "," in image_b64 and image_b64.startswith("data:"):
            mime_part = image_b64.split(",")[0]  # "data:image/png;base64"
            image_b64 = image_b64.split(",", 1)[1]
            # Detect extension from MIME
            if "svg" in mime_part: ext = ".svg"  # wird unten nach PNG konvertiert
            elif "png" in mime_part: ext = ".png"
            elif "gif" in mime_part: ext = ".gif"
            elif "webp" in mime_part: ext = ".webp"
            else: ext = ".jpg"

        try:
            content = b64module.b64decode(image_b64)
        except Exception:
            log_api_usage(api_key_id, api_user["id"], success=False,
                          error_message="Base64 ungueltig")
            raise HTTPException(status_code=400, detail="Base64-Daten konnten nicht dekodiert werden")

    else:
        # Multipart mode: file upload
        form = await request.form()
        file = form.get("file") or form.get("image")
        if not file:
            log_api_usage(api_key_id, api_user["id"], success=False,
                          error_message="Kein Bild mitgeschickt")
            raise HTTPException(status_code=400, detail="Bitte eine Bilddatei als 'file' oder 'image' hochladen")

        context_text = form.get("context", "")
        language = form.get("language", "de")
        image_type = form.get("image_type")

        filename = file.filename or "image.jpg"
        ext = os.path.splitext(filename)[1].lower()
        if ext not in IMAGE_EXTENSIONS:
            log_api_usage(api_key_id, api_user["id"], success=False,
                          error_message=f"Ungueltiges Format: {ext}")
            raise HTTPException(
                status_code=400,
                detail="Nur Bilddateien erlaubt (JPG, PNG, GIF, WebP, BMP, TIFF, HEIC, SVG)"
            )
        content = await file.read()

    # SVG sofort nach PNG wandeln (PIL kann kein SVG; gleiche Behandlung wie /api/upload)
    if ext == ".svg":
        try:
            content = _svg_to_png_bytes(content)
        except Exception as e:
            log_api_usage(api_key_id, api_user["id"], success=False,
                          error_message=f"SVG-Konvertierung fehlgeschlagen: {e}")
            raise HTTPException(status_code=400, detail="SVG konnte nicht verarbeitet werden")
        ext = ".png"

    image_size = len(content)
    if image_size > 10 * 1024 * 1024:  # 10 MB limit for API
        log_api_usage(api_key_id, api_user["id"], image_size_bytes=image_size,
                      success=False, error_message="Bild zu gross")
        raise HTTPException(status_code=413, detail="Bild zu gross. Maximum: 10 MB")

    # Save temporarily
    tmp_dir = os.path.join(UPLOAD_DIR, "api_tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_path = os.path.join(tmp_dir, f"{secrets.token_hex(8)}{ext}")
    with open(tmp_path, "wb") as f:
        f.write(content)

    start_time = time.time()
    try:
        # v2.2: Get image dimensions for thumbnail guard
        try:
            from PIL import Image as PILImage
            with PILImage.open(tmp_path) as _pil_img:
                api_img_width, api_img_height = _pil_img.size
        except Exception:
            api_img_width, api_img_height = 0, 0

        if str(language).lower() not in ALT_TEXT_LANGUAGES:
            language = "de"
        result = await asyncio.get_event_loop().run_in_executor(
            None, generate_alt_text_for_image, tmp_path, context_text, image_type,
            api_img_width, api_img_height, "", str(language).lower()
        )
        processing_time_ms = int((time.time() - start_time) * 1000)
        model_used = result.get("model_used", "")

        # Log successful usage
        log_api_usage(api_key_id, api_user["id"],
                      processing_time_ms=processing_time_ms,
                      model_used=model_used,
                      image_size_bytes=image_size, success=True)

        # Abo-Etappe-1: API-Generierung = 1 Credit aus dem Konto-Topf
        # (Cache-Treffer kosten nichts, siehe pdf_processor from_cache).
        if not result.get("from_cache"):
            billing.verbuche(api_user["id"], "api")

        # result_id generieren und Ergebnis speichern
        result_id = secrets.token_urlsafe(16)
        create_api_result(
            result_id=result_id,
            user_id=api_user["id"],
            api_key_id=api_key_id,
            alt_text=_ohne_fehlertext(result.get("alt_text")),
            langbeschreibung=_ohne_fehlertext(result.get("langbeschreibung")),
            bildtyp=result.get("bildtyp", "unbekannt"),
            konfidenz=result.get("konfidenz", "mittel"),
            model_used=model_used,
            processing_time_ms=processing_time_ms,
            language=language,
            context_text=context_text,
        )

        response_data = {
            "result_id": result_id,
            # 17.08.2026: keine technischen Fehlermeldungen an API-Kunden
            "alt_text": ("" if ist_pipeline_fehlertext(result.get("alt_text"))
                         else result.get("alt_text", "")),
            "langbeschreibung": _ohne_fehlertext(result.get("langbeschreibung")),
            "bildtyp": result.get("bildtyp", "unbekannt"),
            "konfidenz": result.get("konfidenz", "mittel"),
            "processing_time_ms": processing_time_ms,
        }
        return JSONResponse(
            content=response_data,
            headers={
                "X-RateLimit-Remaining-Minute": str(rate_info["minute_remaining"]),
                "X-RateLimit-Remaining-Day": str(rate_info["day_remaining"]),
                "X-DailyLimit-Limit": str(tageslimit),
                "X-DailyLimit-Used": str(daily_used + 1),
                "X-DailyLimit-Remaining": str(max(0, daily_remaining - 1)),
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        processing_time_ms = int((time.time() - start_time) * 1000)
        log_api_usage(api_key_id, api_user["id"],
                      processing_time_ms=processing_time_ms,
                      image_size_bytes=image_size, success=False,
                      error_message=str(e)[:500])
        raise HTTPException(status_code=500, detail="Interner Fehler bei der Alt-Text-Generierung")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


@app.get("/api/v1/alt-text/{result_id}")
async def api_get_alt_text(result_id: str, request: Request):
    """Ruft ein gespeichertes API-Ergebnis ab. Erfordert X-API-Key Header."""
    api_user = get_api_user(request)

    result = get_api_result(result_id, api_user["id"])
    if not result:
        raise HTTPException(status_code=404, detail="Ergebnis nicht gefunden")

    return JSONResponse(content={
        "result_id": result["id"],
        "alt_text": _ohne_fehlertext(result["alt_text"]),
        "langbeschreibung": _ohne_fehlertext(result["langbeschreibung"]),
        "bildtyp": result["bildtyp"],
        "konfidenz": result["konfidenz"],
        "created_at": result["created_at"],
        "updated_at": result["updated_at"],
    })


@app.patch("/api/v1/alt-text/{result_id}")
async def api_update_alt_text(result_id: str, request: Request):
    """Aktualisiert den Alt-Text und/oder die Langbeschreibung eines API-Ergebnisses.
    Erfordert X-API-Key Header. Nur der Besitzer kann sein Ergebnis aendern."""
    api_user = get_api_user(request)

    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ungueltiges JSON")

    alt_text = data.get("alt_text")
    langbeschreibung = data.get("langbeschreibung")

    if alt_text is None and langbeschreibung is None:
        raise HTTPException(
            status_code=400,
            detail="Mindestens 'alt_text' oder 'langbeschreibung' muss angegeben werden"
        )

    if alt_text is not None and not isinstance(alt_text, str):
        raise HTTPException(status_code=400, detail="'alt_text' muss ein String sein")
    if langbeschreibung is not None and not isinstance(langbeschreibung, str):
        raise HTTPException(status_code=400, detail="'langbeschreibung' muss ein String sein")

    success = update_api_result(
        result_id=result_id,
        user_id=api_user["id"],
        alt_text=alt_text,
        langbeschreibung=langbeschreibung,
    )

    if not success:
        raise HTTPException(status_code=404, detail="Ergebnis nicht gefunden")

    updated = get_api_result(result_id, api_user["id"])
    return JSONResponse(content={
        "result_id": updated["id"],
        "alt_text": updated["alt_text"],
        "langbeschreibung": updated["langbeschreibung"],
        "bildtyp": updated["bildtyp"],
        "updated_at": updated["updated_at"],
    })


@app.get("/api/api-usage-stats")
async def api_usage_stats(user: dict = Depends(get_current_user)):
    """Get API usage statistics for the current user."""
    stats = get_api_usage_stats(user["id"])
    return stats


# ─── API Documentation ──────────────────────────────────────

@app.get("/api/v1/docs", response_class=HTMLResponse)
async def api_docs():
    """Accessible API documentation page (WCAG 2.2 AA)."""
    base = BASE_URL.rstrip("/")
    return """<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>InkluDocs API – Dokumentation</title>
<style>
:root { --primary: #1b2a4a; --accent: #e87722; --bg: #f8f9fa; --text: #1e293b; --muted: #64748b; --border: #e2e8f0; --code-bg: #1e293b; --code-text: #e2e8f0; }
*, *::before, *::after { box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; color: var(--text); background: var(--bg); margin: 0; padding: 0; line-height: 1.6; }
a { color: var(--accent); }
a:focus-visible { outline: 3px solid var(--accent); outline-offset: 2px; border-radius: 2px; }
.container { max-width: 800px; margin: 0 auto; padding: 2rem 1.5rem; }
h1 { color: var(--primary); font-size: 1.8rem; margin-bottom: 0.5rem; }
h1 span { color: var(--accent); }
h2 { color: var(--primary); font-size: 1.3rem; margin-top: 2.5rem; padding-bottom: 0.3rem; border-bottom: 2px solid var(--accent); }
h3 { color: var(--primary); font-size: 1.1rem; margin-top: 1.5rem; }
.badge { display: inline-block; padding: 0.2rem 0.6rem; border-radius: 4px; font-size: 0.85rem; font-weight: 600; }
.badge-post { background: #16a34a; color: white; }
.badge-get { background: #2563eb; color: white; }
pre { background: var(--code-bg); color: var(--code-text); padding: 1.2rem; border-radius: 8px; overflow-x: auto; font-size: 0.9rem; line-height: 1.5; }
code { font-family: 'SF Mono', Consolas, 'Liberation Mono', Menlo, monospace; }
p code, li code { background: #e2e8f0; color: var(--primary); padding: 0.15rem 0.4rem; border-radius: 3px; font-size: 0.9em; }
table { width: 100%; border-collapse: collapse; margin: 1rem 0; }
th, td { text-align: left; padding: 0.6rem 0.8rem; border-bottom: 1px solid var(--border); }
th { background: var(--primary); color: white; font-weight: 600; }
tr:nth-child(even) { background: rgba(0,0,0,0.02); }
.note { padding: 1rem; background: #fff7ed; border-left: 4px solid var(--accent); border-radius: 0 4px 4px 0; margin: 1rem 0; }
footer { margin-top: 3rem; padding-top: 1rem; border-top: 1px solid var(--border); color: var(--muted); font-size: 0.85rem; }
.sr-only { position: absolute; width: 1px; height: 1px; padding: 0; margin: -1px; overflow: hidden; clip: rect(0,0,0,0); white-space: nowrap; border: 0; }
@media (prefers-color-scheme: dark) {
    :root { --bg: #0f172a; --text: #e2e8f0; --muted: #94a3b8; --border: #334155; --code-bg: #1e293b; }
    p code, li code { background: #334155; color: #e2e8f0; }
    tr:nth-child(even) { background: rgba(255,255,255,0.03); }
    .note { background: #1c1917; }
}
</style>
</head>
<body>
<main class="container" role="main">
<h1><span>Inklu</span>Docs API</h1>
<p style="color:var(--muted);">Version 1.0 &ndash; Alt-Text-Generierung fuer Bilder</p>

<nav aria-label="Inhaltsverzeichnis">
<h2 id="nav">Inhalt</h2>
<ul>
<li><a href="#auth">Authentifizierung</a></li>
<li><a href="#endpoint">Endpoint</a></li>
<li><a href="#request">Request</a></li>
<li><a href="#response">Response</a></li>
<li><a href="#get-result">Ergebnis abrufen (GET)</a></li>
<li><a href="#patch-result">Ergebnis bearbeiten (PATCH)</a></li>
<li><a href="#errors">Fehler-Codes</a></li>
<li><a href="#ratelimit">Rate-Limits</a></li>
<li><a href="#examples">Beispiele</a></li>
</ul>
</nav>

<h2 id="auth">Authentifizierung</h2>
<p>Alle API-Anfragen erfordern einen gueltigen API-Schluessel im <code>X-API-Key</code> Header.</p>
<p>API-Schluessel kannst du in der <a href="/app">InkluDocs-App</a> unter <strong>Einstellungen</strong> erstellen.</p>
<pre><code>X-API-Key: idocs_deinSchluesselHier</code></pre>

<h2 id="endpoint">Endpoint</h2>
<p><span class="badge badge-post">POST</span> <code>/api/v1/alt-text</code></p>
<p>Generiert einen barrierefreien Alt-Text fuer ein hochgeladenes Bild.</p>

<h2 id="request">Request</h2>
<p>Der Endpoint akzeptiert zwei Formate:</p>

<h3>Option A: Datei-Upload (Multipart)</h3>
<p>Content-Type: <code>multipart/form-data</code></p>
<table>
<caption class="sr-only">Request-Parameter Multipart</caption>
<thead><tr><th scope="col">Parameter</th><th scope="col">Typ</th><th scope="col">Pflicht</th><th scope="col">Beschreibung</th></tr></thead>
<tbody>
<tr><td><code>file</code></td><td>Datei</td><td>Ja</td><td>Bilddatei (JPG, PNG, GIF, WebP, BMP, TIFF, HEIC). Max. 10 MB.</td></tr>
<tr><td><code>context</code></td><td>Text</td><td>Nein</td><td>Umgebungstext fuer bessere Beschreibung (z.B. Bildunterschrift, Seitentitel).</td></tr>
<tr><td><code>language</code></td><td>Text</td><td>Nein</td><td>Sprache des Alt-Texts. Standard: <code>de</code></td></tr>
<tr><td><code>image_type</code></td><td>Text</td><td>Nein</td><td>Hinweis auf Bildtyp: <code>foto</code>, <code>diagramm</code>, <code>logo</code>, <code>icon</code>, <code>karte</code>, <code>screenshot</code>, <code>infografik</code>, <code>tabelle</code></td></tr>
</tbody>
</table>

<h3>Option B: Base64 (JSON)</h3>
<p>Content-Type: <code>application/json</code></p>
<table>
<caption class="sr-only">Request-Parameter JSON</caption>
<thead><tr><th scope="col">Parameter</th><th scope="col">Typ</th><th scope="col">Pflicht</th><th scope="col">Beschreibung</th></tr></thead>
<tbody>
<tr><td><code>image_base64</code></td><td>String</td><td>Ja</td><td>Bild als Base64-String. Data-URI-Prefix optional (z.B. <code>data:image/png;base64,...</code>).</td></tr>
<tr><td><code>context</code></td><td>String</td><td>Nein</td><td>Umgebungstext fuer bessere Beschreibung.</td></tr>
<tr><td><code>language</code></td><td>String</td><td>Nein</td><td>Sprache des Alt-Texts. Standard: <code>de</code></td></tr>
<tr><td><code>image_type</code></td><td>String</td><td>Nein</td><td>Hinweis auf Bildtyp.</td></tr>
</tbody>
</table>

<h2 id="response">Response</h2>
<p>Content-Type: <code>application/json</code></p>
<pre><code>{
  "result_id": "abc123xyz",
  "alt_text": "Beschreibung des Bildes",
  "langbeschreibung": "Ausfuehrliche Beschreibung...",
  "bildtyp": "foto",
  "konfidenz": "hoch",
  "processing_time_ms": 1234
}</code></pre>

<table>
<caption class="sr-only">Response-Felder</caption>
<thead><tr><th scope="col">Feld</th><th scope="col">Beschreibung</th></tr></thead>
<tbody>
<tr><td><code>result_id</code></td><td>Eindeutige ID dieses Ergebnisses – fuer GET und PATCH</td></tr>
<tr><td><code>alt_text</code></td><td>Der generierte Alt-Text (kurz, fuer das alt-Attribut)</td></tr>
<tr><td><code>langbeschreibung</code></td><td>Ausfuehrliche Beschreibung (fuer aria-describedby oder Langtext)</td></tr>
<tr><td><code>bildtyp</code></td><td>Erkannter Bildtyp (foto, diagramm, logo, etc.)</td></tr>
<tr><td><code>konfidenz</code></td><td>Vertrauen in die Erkennung: hoch, mittel, niedrig</td></tr>
<tr><td><code>processing_time_ms</code></td><td>Verarbeitungszeit in Millisekunden</td></tr>
</tbody>
</table>

<h2 id="get-result">Ergebnis abrufen</h2>
<p><span class="badge badge-get">GET</span> <code>/api/v1/alt-text/{result_id}</code></p>
<p>Ruft ein gespeichertes Ergebnis anhand seiner <code>result_id</code> ab. Erfordert denselben <code>X-API-Key</code> wie beim Erstellen.</p>
<pre><code>curl %%BASE_URL%%/api/v1/alt-text/abc123xyz \\
  -H "X-API-Key: idocs_deinSchluessel"</code></pre>
<p>Response-Felder: <code>result_id</code>, <code>alt_text</code>, <code>langbeschreibung</code>, <code>bildtyp</code>, <code>konfidenz</code>, <code>created_at</code>, <code>updated_at</code></p>

<h2 id="patch-result">Ergebnis bearbeiten</h2>
<p><span class="badge" style="background:#7c3aed;color:white;">PATCH</span> <code>/api/v1/alt-text/{result_id}</code></p>
<p>Aendert den Alt-Text und/oder die Langbeschreibung eines gespeicherten Ergebnisses. Mindestens eines der Felder muss angegeben werden.</p>
<pre><code>curl -X PATCH %%BASE_URL%%/api/v1/alt-text/abc123xyz \\
  -H "X-API-Key: idocs_deinSchluessel" \\
  -H "Content-Type: application/json" \\
  -d '{"alt_text": "Mein korrigierter Alt-Text"}'</code></pre>
<table>
<caption class="sr-only">PATCH Request-Felder</caption>
<thead><tr><th scope="col">Feld</th><th scope="col">Typ</th><th scope="col">Beschreibung</th></tr></thead>
<tbody>
<tr><td><code>alt_text</code></td><td>String</td><td>Neuer Alt-Text (optional, aber mindestens eines der beiden Felder)</td></tr>
<tr><td><code>langbeschreibung</code></td><td>String</td><td>Neue Langbeschreibung (optional)</td></tr>
</tbody>
</table>

<h2 id="errors">Fehler-Codes</h2>
<table>
<caption class="sr-only">HTTP-Fehler-Codes</caption>
<thead><tr><th scope="col">Code</th><th scope="col">Bedeutung</th></tr></thead>
<tbody>
<tr><td><code>400</code></td><td>Kein Bild mitgeschickt, ungueltiges Format oder fehlendes JSON-Feld</td></tr>
<tr><td><code>401</code></td><td>Ungueltiger oder fehlender API-Key</td></tr>
<tr><td><code>404</code></td><td>Ergebnis nicht gefunden (falsche result_id oder falscher API-Key)</td></tr>
<tr><td><code>413</code></td><td>Bild zu gross (max. 10 MB)</td></tr>
<tr><td><code>429</code></td><td>Rate-Limit ueberschritten</td></tr>
<tr><td><code>500</code></td><td>Interner Serverfehler</td></tr>
</tbody>
</table>

<h2 id="ratelimit">Rate-Limits</h2>
<p>Pro API-Schluessel gelten folgende Limits:</p>
<ul>
<li><strong>100 Bilder pro Tag</strong> (Alt-Text-Generierung, Reset um Mitternacht UTC)</li>
<li><strong>60 Anfragen pro Minute</strong></li>
</ul>
<p>Die verbleibenden Anfragen werden in Response-Headern mitgeteilt:</p>
<pre><code>X-RateLimit-Remaining-Minute: 58
X-RateLimit-Remaining-Day: 997</code></pre>
<p>Bei Ueberschreitung erhaeltst du HTTP <code>429</code> mit einem <code>Retry-After</code> Header.</p>

<h2 id="examples">Beispiele</h2>

<h3>Einfacher Aufruf mit curl</h3>
<pre><code>curl -X POST %%BASE_URL%%/api/v1/alt-text \\
  -H "X-API-Key: idocs_deinSchluessel" \\
  -F "file=@foto.jpg"</code></pre>

<h3>Mit Kontext fuer bessere Ergebnisse</h3>
<pre><code>curl -X POST %%BASE_URL%%/api/v1/alt-text \\
  -H "X-API-Key: idocs_deinSchluessel" \\
  -F "file=@diagramm.png" \\
  -F "context=Jahresbericht 2025, Kapitel Umsatzentwicklung" \\
  -F "image_type=diagramm"</code></pre>

<h3>Python-Beispiel</h3>
<pre><code>import requests

response = requests.post(
    "%%BASE_URL%%/api/v1/alt-text",
    headers={"X-API-Key": "idocs_deinSchluessel"},
    files={"file": open("bild.jpg", "rb")},
    data={"context": "Produktseite eines Online-Shops"},
)

data = response.json()
print(data["alt_text"])</code></pre>

<h3>JavaScript/Node.js-Beispiel</h3>
<pre><code>const form = new FormData();
form.append('file', fs.createReadStream('bild.jpg'));
form.append('context', 'Blog-Artikel ueber Barrierefreiheit');

const res = await fetch('%%BASE_URL%%/api/v1/alt-text', {
    method: 'POST',
    headers: { 'X-API-Key': 'idocs_deinSchluessel' },
    body: form,
});

const data = await res.json();
console.log(data.alt_text);</code></pre>

<h3>Base64-Beispiel (JSON)</h3>
<pre><code>curl -X POST %%BASE_URL%%/api/v1/alt-text \\
  -H "X-API-Key: idocs_deinSchluessel" \\
  -H "Content-Type: application/json" \\
  -d '{"image_base64": "data:image/jpeg;base64,/9j/4AAQ...", "context": "Startseite"}'</code></pre>

<div class="note" role="note">
<p><strong>Hinweis:</strong> Die API generiert Alt-Texte mit der gleichen KI-Pipeline wie die Web-App. Fuer beste Ergebnisse sende moeglichst viel Kontext im <code>context</code>-Feld mit (z.B. Seitentitel, umgebender Text, Bildunterschrift).</p>
</div>

<footer>
<p>InkluDocs API v1.0 &ndash; <a href="mailto:kontakt@inklutec.de">kontakt@inklutec.de</a> &ndash; <a href="/">Zurueck zu InkluDocs</a></p>
</footer>
</main>
</body>
</html>""".replace("%%BASE_URL%%", base)


# ─── InkluAgent (Chatbot pro Projekt) ────────────────────────

INKLUAGENT_ENABLED = os.environ.get("INKLUAGENT_ENABLED", "false").lower() in ("true", "1", "yes")


def _require_inkluagent():
    if not INKLUAGENT_ENABLED:
        raise HTTPException(status_code=404, detail="Nicht aktiviert")


def _require_project_owned(project_id: int, user_id: int):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM projects WHERE id = ? AND user_id = ?",
            (project_id, user_id),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")


@app.get("/api/projects/{project_id}/chat/history")
async def chat_get_history(project_id: int, user: dict = Depends(get_current_user)):
    _require_inkluagent()
    _require_project_owned(project_id, user["id"])
    from inkluagent import storage
    return {"messages": storage.get_history(project_id)}


@app.post("/api/projects/{project_id}/chat")
async def chat_send_message(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    _require_inkluagent()
    _require_project_owned(project_id, user["id"])
    data = await request.json()
    message = (data.get("message") or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="Nachricht darf nicht leer sein")
    if len(message) > 5000:
        raise HTTPException(status_code=400, detail="Nachricht zu lang (max. 5000 Zeichen)")

    from inkluagent import storage
    from inkluagent.chat_engine import process_message

    storage.append_message(project_id, "user", message)
    loop = asyncio.get_running_loop()

    def _antwort(result: dict) -> dict:
        storage.append_message(
            project_id, "assistant", result["reply"],
            image_refs=result.get("image_refs"),
            intent=result.get("intent"),
            werkzeuge=result.get("werkzeuge"),
        )
        return {
            "reply": result["reply"],
            "intent": result.get("intent"),
            "image_refs": result.get("image_refs"),
            "actions": result.get("actions", []),
            "werkzeuge": result.get("werkzeuge", []),
        }

    # Werkzeug-Transparenz (Steve 28.08.2026): mit Accept: application/x-ndjson streamt der
    # Endpunkt je Werkzeugaufruf eine Zeile {"type":"tool","name":...} und am Ende
    # {"type":"reply", ...} — die Oberflaeche zeigt live „Ruft gerade auf: …“. Ohne den
    # Header bleibt die bisherige JSON-Antwort (API-Kunden, Demo, Tests).
    if "application/x-ndjson" not in (request.headers.get("accept") or ""):
        result = await loop.run_in_executor(None, process_message, project_id, message, user["id"])
        return _antwort(result)

    queue: asyncio.Queue = asyncio.Queue()

    def on_tool(name, args):
        loop.call_soon_threadsafe(queue.put_nowait, {"type": "tool", "name": name})

    async def lauf():
        try:
            r = await loop.run_in_executor(
                None, lambda: process_message(project_id, message, user["id"], on_tool=on_tool))
        except Exception as e:
            logger.exception("Chat-Stream: process_message crashte")
            r = {"reply": "Entschuldigung, dabei ist ein Fehler aufgetreten. Bitte erneut versuchen.",
                 "intent": "error", "image_refs": None, "actions": [], "werkzeuge": []}
        await queue.put({"type": "done", "result": r})

    async def strom():
        aufgabe = asyncio.ensure_future(lauf())
        try:
            while True:
                ev = await queue.get()
                if ev["type"] == "done":
                    out = _antwort(ev["result"])
                    out["type"] = "reply"
                    yield json.dumps(out, ensure_ascii=False) + "\n"
                    break
                yield json.dumps(ev, ensure_ascii=False) + "\n"
        finally:
            if not aufgabe.done():
                await aufgabe

    return StreamingResponse(strom(), media_type="application/x-ndjson",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.delete("/api/projects/{project_id}/chat")
async def chat_clear_history(project_id: int, user: dict = Depends(get_current_user)):
    _require_inkluagent()
    _require_project_owned(project_id, user["id"])
    from inkluagent import storage
    deleted = storage.clear_history(project_id)
    return {"deleted": deleted}


# ─── News / Neuigkeiten ─────────────────────────────────────


# === LINK-REFERENCE-POSTPROCESSOR (13.05.2026) ============================
# WCAG-relevant: Verlinkte Bilder bekommen "(verweist auf: ...)" an den
# Alt-Text gehaengt, deterministisch nach der LLM-Pipeline. Quelle ist die
# vom Web-Scraper extrahierte [Link-Beschriftung] im context_text.

_LINK_REF_GENERIC = {
    "mehr info", "mehr informationen", "weiterlesen", "hier klicken",
    "klicken", "link", "details", "mehr", "weiter", "mehr erfahren",
    "lesen sie mehr", "mehr lesen", "weiterlesen...",
    # Qualitaetsrunde 21.08.2026: Beschriftungen, die nur die Verlinkung
    # selbst beschreiben, sind kein Ziel — Zusatz waere reine Doppelung.
    "startseite", "zur startseite", "home", "homepage", "start",
}

# Zusatz darf den Alt-Text nicht sprengen: Schema-Cap der Generierung ist
# 400, die App erlaubt manuell bis 500 — mit Label-Deckel 80 bleibt die
# Summe immer unter 500 (Befund budni-Test 21.08.2026: 452 Zeichen).
_LINK_REF_LABEL_MAX = 80

# Schlagwort-Fix 21.08.2026 (Steve): Der Zusatz erscheint in der
# Ausgabesprache des Projekts, nicht mehr fest deutsch.
_LINK_REF_PHRASES = {
    "de": "verweist auf", "en": "links to", "fr": "renvoie à",
    "es": "enlaza a", "da": "henviser til", "sv": "länkar till",
}

def _append_link_reference(alt_text: str, context_text: str, language: str = "de") -> str:
    """Haengt "(verweist auf: <Beschriftung>)" an alt_text, wenn das Bild
    verlinkt ist und eine sinnvolle Beschriftung im Kontext steht.

    Quellen im context_text (erste Treffer gewinnt):
    1. [Link-Beschriftung] <text>
    2. [title] verweist auf: <text>  (Scraper-vorformatiert)

    Skip-Regeln: leerer alt_text, kein Kontext, alt_text enthaelt bereits
    "verweist auf", Beschriftung generisch (Mehr Info, Startseite etc.),
    <3 Zeichen, im Alt-Text bereits enthalten (Doppelung) oder eine
    Meta-Beschriftung, die nur die Verlinkung beschreibt ("Logo verlinkt
    auf die Startseite"). Lange Beschriftungen werden auf
    _LINK_REF_LABEL_MAX Zeichen gekappt.
    """
    if not alt_text or not context_text:
        return alt_text
    phrase = _LINK_REF_PHRASES.get((language or "de").lower(), "verweist auf")
    alt_low_gesamt = alt_text.lower()
    if "verweist auf" in alt_low_gesamt or phrase in alt_low_gesamt:
        return alt_text

    def _brauchbares_label(label: str) -> bool:
        low = label.lower()
        if len(label) < 3 or low in _LINK_REF_GENERIC:
            return False
        # Doppelung: steht die Beschriftung (oder ihr Kern) schon im Alt-Text,
        # traegt der Zusatz nichts bei ("Logo budni — Link zur Startseite").
        alt_low = alt_text.lower()
        if low in alt_low:
            return False
        # Wort-Abgleich faengt Umformulierungen ("InkluTec - Zur Startseite"
        # vs. "Logo InkluTec — Link zur Startseite"): Wenn JEDES Wort der
        # Beschriftung mit >=3 Zeichen schon im Alt-Text steht, ist der
        # Zusatz redundant.
        woerter = [w for w in re.findall(r"\w+", low) if len(w) >= 3]
        if woerter and all(w in alt_low for w in woerter):
            return False
        # Meta-Beschriftung beschreibt die Verlinkung statt des Ziels.
        if "verlinkt" in low or low.startswith("logo "):
            return False
        return True

    def _mit_label(label: str) -> str:
        if len(label) > _LINK_REF_LABEL_MAX:
            label = label[:_LINK_REF_LABEL_MAX].rstrip() + "…"
        return f"{alt_text.rstrip()} ({phrase}: {label})"

    m = re.search(r"\[Link-Beschriftung\]\s*(.+?)(?:\n|$)", context_text)
    if m:
        label = m.group(1).strip()
        if label and _brauchbares_label(label):
            return _mit_label(label)
    m = re.search(r"\[title\]\s*verweist auf:\s*(.+?)(?:\n|$)", context_text)
    if m:
        ref = m.group(1).strip()
        if ref and _brauchbares_label(ref):
            return _mit_label(ref)
    return alt_text


NEUIGKEITEN = [
    # Neueste zuerst. Nur Eintraege, die fuer Nutzer relevant + auf Production live sind.
    {"datum": "17.07.2026", "text": "Großes Update der Bildbeschreibungen: Alle 16 Bildkategorien — von Foto über Diagramm und Tabelle bis Strukturformel — liefern jetzt spürbar präzisere Texte: Kernaussage zuerst, Zahlen wortgetreu. Ein automatischer Redakteur prüft kritische Bilder zusätzlich gegen das Bild und korrigiert offensichtliche Fehler."},
    {"datum": "17.07.2026", "text": "Prüf-Workflow vereinfacht: Jedes Bild trägt genau einen Status — Neu, In Bearbeitung, Lektorat Freigabe/Änderung, Herausgeber Freigabe/Änderung — und genau danach kannst du filtern. Beim Abschluss einer Prüfung genügt eine Anmerkung."},
    {"datum": "17.07.2026", "text": "Neu: SVG-Dateien (Vektorgrafiken) können jetzt direkt hochgeladen werden — sie werden automatisch umgewandelt und durchlaufen die normale Generierung."},
    {"datum": "17.07.2026", "text": "Neue Sprache: Die Oberfläche und die Alt-Text-Ausgabe gibt es jetzt auch auf Schwedisch."},
    {"datum": "09.07.2026", "text": "Neu: Eigene Prompts. Du speicherst eigene Anweisungen für die KI – zum Beispiel ‚verwende einfache Sprache‘ – und wählst pro Projekt aus, welcher Prompt bei der Generierung zusätzlich gelten soll. Auch beim einzelnen Neu-Generieren."},
    {"datum": "09.07.2026", "text": "PDF-Auslesung verbessert: Texte werden bei Spalten- und Seitenumbrüchen nicht mehr doppelt erfasst, und Tabelleninhalte werden vollständiger übernommen. Die KI bekommt dadurch sauberere Zusammenhänge für ihre Beschreibungen."},
    {"datum": "09.07.2026", "text": "Bildbeschreibungen nochmals verbessert: kompaktere Alt-Texte, wichtige Personen im Bild werden zuverlässiger erwähnt, und eine automatische Gegenprüfung kontrolliert kritische Angaben."},
    {"datum": "05.07.2026", "text": "InkluDocs in deiner Sprache: Die Oberfläche gibt es jetzt auf Deutsch, Englisch, Französisch, Spanisch und Dänisch. Deine Sprache stellst du in den Einstellungen um."},
    {"datum": "05.07.2026", "text": "Alt-Texte in mehreren Sprachen: Je Projekt wählst du die Ausgabesprache der generierten Texte – jederzeit umschaltbar. Das Anhören nutzt automatisch die passende Stimme."},
    {"datum": "05.07.2026", "text": "Upload verbessert: Das Hochladen antwortet sofort, die Verarbeitung läuft im Hintergrund weiter – auch bei großen PDFs."},
    {"datum": "05.07.2026", "text": "Foto-Beschreibungen verbessert: Essen, Landschaften und Architektur werden präziser beschrieben, bekannte Wahrzeichen benannt."},
    {"datum": "25.06.2026", "text": "Passwort anzeigen: Bei Anmeldung, Registrierung und Passwort-Änderung kannst du dein Passwort jetzt per Häkchen sichtbar machen."},
    {"datum": "22.06.2026", "text": "Gastzugang: Du kannst ein Projekt mit einer Person ohne Konto teilen – sie kann deine Alt-Texte ansehen, kommentieren und freigeben."},
    {"datum": "19.06.2026", "text": "Seitentext übersichtlicher dargestellt; Alt-Texte und Beschreibungen lassen sich jetzt anhören."},
    {"datum": "10.06.2026", "text": "Mehrere PDFs in einem Projekt sammeln, einzeln umbenennen und löschen."},
    {"datum": "04.06.2026", "text": "Neu: Alt-Texte für ganze Webseiten. Adresse eingeben, die Bilder der Seite werden geladen – auch mehrere Seiten nacheinander."},
    {"datum": "02.06.2026", "text": "Neues Dashboard mit fester Seitenleiste: eine übersichtliche Startseite, von der aus du alles erreichst."},
    {"datum": "02.06.2026", "text": "Getrennte Werkzeuge für PDFs, Webseiten und Grafiken – du legst ein Projekt an und wählst das passende Werkzeug."},
    {"datum": "02.06.2026", "text": "PDF-Bearbeitung verbessert, mit Seitenvorschau für mehr Überblick."},
    {"datum": "02.06.2026", "text": "Verbesserte Barrierefreiheit: klarere Struktur, bessere Bedienung mit Screenreader und Tastatur."},
    {"datum": "13.05.2026", "text": "Chat-Assistent neu: Jedes Projekt hat jetzt einen eigenen Chatbot, der Bilder einsieht, Alt-Texte bewertet, Vorschläge macht und nach Bestätigung direkt in den Text übernimmt. Aktuell auf Deutsch."},
    {"datum": "13.05.2026", "text": "Mehrsprachigkeit in Vorbereitung: Englisch, Französisch und Spanisch werden in den kommenden Wochen nach und nach ausgerollt."},
    {"datum": "14.04.2026", "text": "Neue dreistufige Prüfpipeline aktiv: Klassifikation, Generierung und automatische Qualitätsprüfung gegen Halluzinationen. Laufende Auswertung zur weiteren Verbesserung."},
    {"datum": "07.04.2026", "text": "Alt-Text-Qualität verbessert: Produktbilder, Diagramme und verlinkte Bilder werden besser erkannt"},
    {"datum": "07.04.2026", "text": "Neue Bildformate: AVIF und HEIC werden jetzt unterstützt"},
    {"datum": "06.04.2026", "text": "Tageslimit: 100 Bilder pro Tag – Anzeige im Dashboard"},
    {"datum": "06.04.2026", "text": "Registrierung offen: Konto erstellen mit E-Mail-Bestätigung"},
    {"datum": "06.04.2026", "text": "InkluDocs unterstützen: Freiwillige Beiträge per PayPal möglich"},
]

def _news_date_key(d):
    """\"DD.MM.YYYY\" -> sortierbarer Schluessel \"YYYYMMDD\"."""
    import re as _re
    mm = _re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", d or "")
    return (mm.group(3) + mm.group(2) + mm.group(1)) if mm else (d or "")

def _newest_news_key():
    keys = [_news_date_key(n["datum"]) for n in NEUIGKEITEN]
    return max(keys) if keys else ""

@app.get("/api/news")
async def get_news(user: dict = Depends(get_current_user)):
    """Changelog-Eintraege + kontobezogener Lese-Stand (news_seen_until, Schluessel YYYYMMDD)."""
    conn = get_db()
    row = conn.execute("SELECT news_seen_until FROM users WHERE id = ?", (user["id"],)).fetchone()
    conn.close()
    seen = row["news_seen_until"] if row and row["news_seen_until"] else ""
    return {"news": NEUIGKEITEN, "seen_until": seen}

@app.post("/api/news/seen")
async def mark_news_seen(user: dict = Depends(get_current_user)):
    """Markiert alle aktuellen Neuigkeiten als gelesen (news_seen_until = neuester Schluessel)."""
    key = _newest_news_key()
    conn = get_db()
    conn.execute("UPDATE users SET news_seen_until = ? WHERE id = ?", (key, user["id"]))
    conn.commit()
    conn.close()
    return {"ok": True, "seen_until": key}

# ─── Frontend Routes ─────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    # Demo-Instanz: die Wurzel IST die oeffentliche Demo-Seite (statt Login/Dashboard).
    # Seit 04.07.2026 als Jinja2-Template fuenfsprachig; Sprache per Browser-
    # Erkennung bzw. lang-Cookie — bewusst ohne Umschalter (Steve 03.07.2026).
    if demo_mod.demo_enabled():
        return templates.TemplateResponse(
            "demo.html", template_context(request, detect_language(request)))
    lang = detect_language(request)
    registration_enabled = os.getenv("REGISTRATION_ENABLED", "true").lower() not in ("false", "0", "no")
    is_staging = "staging" in BASE_URL
    return templates.TemplateResponse(
        "index.html",
        template_context(
            request, lang,
            registration_enabled=registration_enabled,
            is_staging=is_staging,
            # Konto-Selbstloeschung (08.08.2026): Nach dem Loeschen landet man
            # hier — mit einer Bestaetigung statt eines wortlosen Rauswurfs.
            geloescht=request.query_params.get("geloescht") == "1",
        ),
    )


# 15.06.2026: Rechtsseiten der Demo IM Demo-Rahmen (Sidebar + Footer wie das
# Werkzeug). Bewusst eigene, NICHT login-geschuetzte Routen — die -app-Varianten
# des Werkzeugs setzen ein Login voraus und leiten anonyme Besucher zur Wurzel um.
# Der Rechtstext selbst kommt per fetch aus den oeffentlichen Routen (Single Source).
@app.get("/demo-impressum", response_class=HTMLResponse)
async def demo_impressum_page(request: Request):
    if demo_mod.demo_enabled():
        return templates.TemplateResponse(
            "demo-impressum.html", template_context(request, detect_language(request)))
    return RedirectResponse("/")


@app.get("/demo-datenschutz", response_class=HTMLResponse)
async def demo_datenschutz_page(request: Request):
    if demo_mod.demo_enabled():
        return templates.TemplateResponse(
            "demo-datenschutz.html", template_context(request, detect_language(request)))
    return RedirectResponse("/")


@app.get("/demo-nutzungsbedingungen", response_class=HTMLResponse)
async def demo_nutzungsbedingungen_page(request: Request):
    if demo_mod.demo_enabled():
        return templates.TemplateResponse(
            "demo-nutzungsbedingungen.html", template_context(request, detect_language(request)))
    return RedirectResponse("/")


@app.get("/set-language/{lang}")
async def set_language(lang: str, request: Request):
    """Wechsle UI-Sprache via Session-Cookie (spaeter auch in DB fuer User)."""
    referer = request.headers.get("referer", "/")
    # Zielort immer zurueck zum Referer, aber nur same-origin (Sicherheit).
    # Host-Vergleich statt BASE_URL-Praefix: die App laeuft unter mehreren
    # Domains (inkludocs.de, www.inkludocs.de, inkludocs.inklutec.de) —
    # der alte Praefix-Check passte nur auf EINE davon; auf allen anderen
    # landete der Sprachwechsel faelschlich auf "/" (= Anmeldemaske).
    # Weitergeleitet wird nur der Pfad, nie eine fremde absolute URL.
    from urllib.parse import urlsplit
    redirect_to = "/"
    if referer.startswith("/") and not referer.startswith("//"):
        redirect_to = referer
    else:
        try:
            parts = urlsplit(referer)
            req_host = (request.url.hostname or "").lower()
            if parts.hostname and req_host and parts.hostname.lower() == req_host:
                redirect_to = parts.path or "/"
                if parts.query:
                    redirect_to += "?" + parts.query
        except ValueError:
            pass
    response = RedirectResponse(redirect_to, status_code=303)
    if lang in SUPPORTED_LANGUAGES:
        # 1 Jahr gueltig, samesite=lax, nicht HttpOnly (Client koennte lesen).
        # Der Cookie deckt Gaeste und die aktuelle Sitzung ab.
        response.set_cookie("lang", lang, max_age=365*24*60*60, samesite="lax")
        # Eingeloggte Nutzer: Wahl zusaetzlich dauerhaft am Konto speichern,
        # damit sie ueber Geraete und die Cookie-Lebensdauer hinaus gilt
        # (Stufe 1 der Sprach-Erkennung in i18n.detect_language).
        user = get_optional_user(request)
        if user:
            set_user_language(user["id"], lang)
    return response

@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    if os.getenv("REGISTRATION_ENABLED", "true").lower() in ("false", "0", "no"):
        _lang = resolve_ui_language(request)
        _ = get_gettext(_lang)
        _beta = _('InkluDocs befindet sich in der geschlossenen Beta-Phase. Wenn du einen Testzugang erhalten möchtest, schreib bitte eine E-Mail an {mail}.').format(
            mail='<strong>kontakt@inklutec.de</strong>')
        return _auth_notice_page(_lang,
            f'''<p style="margin:2rem 0;font-size:1.1rem;">{_("Die Registrierung ist derzeit geschlossen.")}</p>
        <p>{_beta}</p>
        <p style="margin-top:2rem;"><a href="/">{_("Zurück zur Anmeldung")}</a></p>''',
            title_suffix=_("Registrierung geschlossen"))
    # Seit 03.07.2026 Jinja2-Template (i18n); Staging-Titel/-Brand kommen aus dem Template.
    lang = resolve_ui_language(request)
    return templates.TemplateResponse(
        "register.html",
        template_context(request, lang, is_staging=("staging" in BASE_URL)),
    )

@app.get("/forgot", response_class=HTMLResponse)
async def forgot_page(request: Request):
    lang = resolve_ui_language(request)
    return templates.TemplateResponse(
        "forgot.html",
        template_context(request, lang, is_staging=("staging" in BASE_URL)),
    )

@app.get("/reset", response_class=HTMLResponse)
async def reset_page(request: Request):
    lang = resolve_ui_language(request)
    return templates.TemplateResponse(
        "reset.html",
        template_context(request, lang, is_staging=("staging" in BASE_URL)),
    )

def _serve_protected_page(request: Request, filename: str):
    """Liefert eine login-geschuetzte HTML-Seite aus dem frontend-Verzeichnis.
    Leitet zu / um, wenn kein gueltiges Login-Cookie vorliegt."""
    token = request.cookies.get("token")
    if not token:
        return RedirectResponse("/")
    try:
        jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        return RedirectResponse("/")
    html = open(f"/app/frontend/{filename}").read()
    if "staging" in BASE_URL:
        html = html.replace("<title>InkluDocs</title>", "<title>InkluDocs (Testumgebung)</title>")
        html = html.replace('<span class="brand">Inklu</span>Docs', '<span class="brand">Inklu</span>Docs <span style="font-size:0.6em;color:#c75000;font-weight:normal;">(Testumgebung)</span>')
    return html


def resolve_ui_language(request: Request) -> str:
    """Ermittelt die anzuzeigende UI-Sprache fuer diesen Request.

    Eingeloggte Nutzer: Praeferenz aus users.language (Stufe 1 der Erkennung).
    Sonst faellt detect_language auf Cookie -> Browser-Header -> Deutsch zurueck.
    """
    user = get_optional_user(request)
    user_lang = None
    if user:
        db_user = get_user_by_id(user["id"])
        if db_user:
            user_lang = db_user.get("language")
    return detect_language(request, user_language=user_lang)


def _render_protected_template(request: Request, template_name: str, **extra):
    """Wie _serve_protected_page, aber rendert ein Jinja2-Template mit
    Sprach-Aufloesung. Fuer bereits auf i18n migrierte, eingeloggte Seiten."""
    token = request.cookies.get("token")
    if not token:
        return RedirectResponse("/")
    try:
        jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        return RedirectResponse("/")
    lang = resolve_ui_language(request)
    return templates.TemplateResponse(
        template_name,
        template_context(request, lang, is_staging=("staging" in BASE_URL), **extra),
    )


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(request: Request):
    return _render_protected_template(request, "dashboard.html")


@app.get("/projekte", response_class=HTMLResponse)
async def projects_page(request: Request):
    return _render_protected_template(request, "projekte.html")


@app.get("/projekt-neu", response_class=HTMLResponse)
async def new_project_page(request: Request):
    return _render_protected_template(request, "projekt-neu.html")


@app.get("/einstellungen", response_class=HTMLResponse)
async def settings_page(request: Request):
    return _render_protected_template(request, "einstellungen.html")


@app.get("/geteilte-projekte", response_class=HTMLResponse)
async def shared_projects_page(request: Request):
    return _render_protected_template(request, "geteilte_projekte.html")


@app.get("/abo", response_class=HTMLResponse)
async def abo_page(request: Request):
    # Abo & Verbrauch (Etappe 2, 31.07.2026): eigene Unterseite; das
    # Template abo.html liefert der Frontend-Teil dieser Etappe.
    return _render_protected_template(request, "abo.html")


@app.get("/team", response_class=HTMLResponse)
async def team_page(request: Request):
    """Team-Verwaltung als eigene Seite (08.08.2026, Steve).

    Vorher stand die ganze Verwaltung mitten im Abo-Bereich und machte die
    Seite lang — auch fuer alle, die kein Team fuehren wollen. Bewusst OHNE
    Eintrag in der Seitenleiste: Der Weg fuehrt ueber „Abo & Verbrauch",
    sonst stuende der Punkt auch bei Konten ohne Team-Tarif herum.
    Wer hier ohne passenden Tarif landet, bekommt eine Erklaerung statt
    leerer Formulare — die Endpunkte pruefen die Rechte ohnehin selbst.
    """
    return _render_protected_template(request, "team.html")


@app.get("/konto", response_class=HTMLResponse)
async def konto_page(request: Request):
    # E-Mail & Passwort (31.07.2026): Die Einstellungen sind eine schlanke
    # Uebersicht; Anmeldedaten aendert man auf dieser Unterseite.
    return _render_protected_template(request, "konto.html")


@app.get("/api-schluessel", response_class=HTMLResponse)
async def api_schluessel_page(request: Request):
    # API-Schluessel-Verwaltung (31.07.2026): aus den Einstellungen
    # ausgelagert, damit die Uebersicht kurz und hoerbar bleibt.
    return _render_protected_template(request, "api_schluessel.html")


@app.get("/prompts", response_class=HTMLResponse)
async def prompts_page(request: Request):
    return _render_protected_template(request, "prompts.html")


@app.get("/stammdaten", response_class=HTMLResponse)
async def stammdaten_page(request: Request):
    """QUICKINFO-WERKZEUG (27.08.2026): Stammdaten-Bibliothek des Kontos, wie „Meine Prompts“."""
    return _render_protected_template(request, "stammdaten.html")


@app.get("/benutzer", response_class=HTMLResponse)
async def users_page(request: Request):
    # api_limit_standard: aktueller Standard der Missbrauchsbremse — auf
    # Staging per ENV 5000, auf Prod 100. Die Seite zeigt ihn im Limit-Feld.
    return _render_protected_template(request, "benutzer.html",
                                      api_limit_standard=DAILY_IMAGE_LIMIT)


@app.get("/benutzer/report/{user_id}", response_class=HTMLResponse)
async def user_report_page(user_id: int, request: Request):
    # Eigene Report-Seite je Kunde (11.08.2026, Steve). Die Rechtepruefung
    # macht der Daten-Endpunkt /api/admin/users/<id>/report — die Seite
    # selbst zeigt ohne Rechte nur die Kein-Zugriff-Meldung.
    return _render_protected_template(request, "benutzer_report.html")


@app.get("/datensicherheit", response_class=HTMLResponse)
async def datensicherheit_page(request: Request):
    # Steve/Karbe 08.06.2026: eigene In-App-Sicht der Datenschutzerklaerung,
    # damit die Sidebar beim Aufruf sichtbar bleibt. Der Inhalt wird im
    # Frontend per fetch aus /datenschutz (Single Source) geladen.
    return _serve_protected_page(request, "datensicherheit.html")


@app.get("/preise", response_class=HTMLResponse)
async def preise_page(request: Request):
    """Oeffentliche Tarif-/Preisseite (Umbau-Punkt 7, 04.08.2026).

    Alle Zahlen kommen aus billing.py — eine Preisaenderung ist dort EINE
    Konfigurationszeile, die Seite zieht automatisch mit. Deutsche
    Komma-Formatierung mit zwei Nachkommastellen (9,95 / 87,50).
    Noch nirgends verlinkt — Verlinkung (Login-Fusszeile) kommt bewusst
    erst mit dem Livegang der Online-Zahlung.
    """
    def _eur(betrag: float) -> str:
        return f"{betrag:.2f}".replace(".", ",")

    ctx = template_context(request, detect_language(request),
                           is_staging="staging" in BASE_URL)
    ctx.update({
        "free_credits": billing.PLAN_KONTINGENTE["free"],
        "single_preis": _eur(billing.PLAN_PREISE_EUR["single"]),
        "single_credits": billing.PLAN_KONTINGENTE["single"],
        "team_preis": _eur(billing.PLAN_PREISE_EUR["team"]),
        "team_credits": billing.PLAN_KONTINGENTE["team"],
        "team_sitze": billing.PLAN_SITZE["team"],
        "enterprise_preis": _eur(billing.PLAN_PREISE_EUR["enterprise"]),
        "enterprise_credits": billing.PLAN_KONTINGENTE["enterprise"],
        "enterprise_sitze": billing.PLAN_SITZE["enterprise"],
        "single_preis_monat": _eur(billing.PLAN_PREISE_MONATLICH_EUR["single"]),
        "team_preis_monat": _eur(billing.PLAN_PREISE_MONATLICH_EUR["team"]),
        "enterprise_preis_monat": _eur(billing.PLAN_PREISE_MONATLICH_EUR["enterprise"]),
        "laufzeiten": "3, 6 oder 12",
        "pakete": [(n, _eur(p)) for n, p in sorted(billing.PAKET_PREISE.items())],
    })
    return templates.TemplateResponse("preise.html", ctx)


@app.get("/impressum-app", response_class=HTMLResponse)
async def impressum_app_page(request: Request):
    # Steve 08.06.2026: In-App-Sicht des Impressums (Schwester zu /datensicherheit).
    # Wird vom Footer der eingeloggten App-Bereiche statt /impressum verlinkt, damit
    # Dashboard/Sidebar beim Klick sichtbar bleiben. Inhalt kommt per fetch aus
    # /impressum (Single Source — keine Doppelpflege).
    return _serve_protected_page(request, "impressum-app.html")


@app.get("/widerruf-app", response_class=HTMLResponse)
async def widerruf_app_page(request: Request):
    """In-App-Sicht der Widerrufsbelehrung (08.08.2026), Inhalt aus /widerruf.

    Die Kuendigungsseite bekommt bewusst KEINE App-Sicht: § 312k BGB
    verlangt einen Weg, der ohne Anmeldung erreichbar ist — der Link aus der
    Fusszeile fuehrt deshalb direkt auf /kuendigen.
    """
    return _serve_protected_page(request, "widerruf-app.html")


@app.get("/nutzungsbedingungen-app", response_class=HTMLResponse)
async def nutzungsbedingungen_app_page(request: Request):
    # Steve 08.06.2026: In-App-Sicht der Nutzungsbedingungen. Gleiches Muster
    # wie /datensicherheit und /impressum-app. Inhalt aus /nutzungsbedingungen
    # per fetch geladen.
    return _serve_protected_page(request, "nutzungsbedingungen-app.html")


@app.get("/app", response_class=HTMLResponse)
async def app_page(request: Request):
    # Seit 03.07.2026 Jinja2-Template (i18n): Sprach-Aufloesung, window.I18N,
    # Staging-Titel und Cache-Busting kommen aus base_app.html.
    # max_upload_mb: Upload-Limit fuer Hinweis/Fehlertext im Upload-Bereich
    # (eine Quelle, Julia-Feedback 02.07.2026).
    return _render_protected_template(request, "app.html",
                                      max_upload_mb=MAX_UPLOAD_SIZE // (1024 * 1024))


# ============================================================
#  Gastzugang / Projekt-Freigabe — Routen (19.06.2026, Etappe 2a)
#  Alle additiv. Kein bestehender (Login-)Pfad wird veraendert.
# ============================================================
@app.get("/freigabe/{token}", response_class=HTMLResponse)
async def freigabe_page(token: str, request: Request):
    """Gast-Eingang: liefert dieselbe app.html im GAST-MODUS (kein Login noetig).
    Ungueltiger/abgelaufener/widerrufener Token -> neutraler Hinweis (kein Datenleck).
    Der Datenzugriff ist zusaetzlich durch das E-Mail-Gate geschuetzt."""
    if not sharing.is_valid_share(token):
        # Viersprachig (03.07.2026): Gaeste haben kein Konto — Sprache kommt aus
        # Cookie/Accept-Language. Echte Umlaute (vorher ue-Schreibweise sichtbar).
        lang = resolve_ui_language(request)
        _ = get_gettext(lang)
        return HTMLResponse(
            f"<!doctype html><html lang='{lang}'><head><meta charset='utf-8'>"
            "<title>InkluDocs</title></head>"
            "<body style='font-family:sans-serif;max-width:40rem;margin:3rem auto;padding:0 1rem'>"
            f"<h1>{_('Link ungültig oder abgelaufen')}</h1>"
            f"<p>{_('Diese Freigabe ist nicht (mehr) gültig. Bitte wenden Sie sich an die Person, die Ihnen den Link geschickt hat.')}</p></body></html>",
            status_code=404,
        )
    # Seit 03.07.2026 dasselbe Jinja2-Template wie /app; GAST-MODUS-Flags kommen
    # als Template-Variablen (head_extra-Block) statt per String-Injektion.
    # Sprache fuer Gaeste: Cookie -> Accept-Language -> Deutsch (resolve_ui_language).
    lang = resolve_ui_language(request)
    # Gast-Ansicht Formulare (28.08.2026): das Werkzeug des Projekts geht als
    # window.GUEST_TOOL mit, damit das E-Mail-Gate schon vor dem Laden weiss,
    # ob es „Alt-Texte" oder „Quickinfos" ansagt.
    share_row = sharing.get_share(token)
    conn = get_db()
    tool_row = conn.execute("SELECT tool FROM projects WHERE id = ?", (share_row["project_id"],)).fetchone() if share_row else None
    conn.close()
    return templates.TemplateResponse(
        "app.html",
        template_context(
            request, lang,
            is_staging=("staging" in BASE_URL),
            guest_mode=True, share_token=token,
            guest_tool=((tool_row["tool"] if tool_row else "") or ""),
        ),
    )


@app.post("/api/freigabe/{token}/confirm")
async def freigabe_confirm(token: str, request: Request):
    """E-Mail-Gate: prueft die eingegebene E-Mail gegen die Einladung. Bei Erfolg
    wird ein guest_token-Cookie gesetzt (an genau diese Freigabe gebunden)."""
    data = await request.json()
    email = (data.get("email") or "").strip()
    if not sharing.confirm_guest_email(token, email):
        raise HTTPException(status_code=403, detail="E-Mail stimmt nicht mit der Einladung ueberein.")
    resp = JSONResponse({"ok": True})
    resp.set_cookie("guest_token", create_guest_token(token, email.lower()),
                    httponly=True, secure=True, samesite="strict", max_age=TOKEN_EXPIRE_HOURS * 3600)
    return resp


def _require_guest(request, token):
    """Gemeinsamer Wachposten: gueltige Gast-Sitzung + nutzbare Freigabe -> share,
    sonst HTTPException(401)."""
    session = get_guest_session(request, token)
    share = sharing.get_share(token) if session else None
    if not session or not share or share["status"] not in ("active", "completed"):
        raise HTTPException(status_code=401, detail="Bitte zuerst die E-Mail bestaetigen.")
    return share


@app.get("/api/freigabe/{token}")
async def freigabe_data(token: str, request: Request):
    """Projektdaten fuer den Gast — nur mit gueltiger Gast-Sitzung, nur das eine
    Projekt dieser Freigabe. Spiegelt get_project, aber token- statt user-begrenzt."""
    share = _require_guest(request, token)
    pid = share["project_id"]
    conn = get_db()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (pid,)).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden.")
    if project["tool"] == "formular":
        # Gast-Ansicht Formulare (28.08.2026): die Felder holt formular.js ueber
        # /api/freigabe/{token}/felder; hier nur der Projektkopf OHNE Serverpfade,
        # damit app.html die Weiche auf die Formular-Ansicht nehmen kann.
        conn.close()
        proj_dict = dict(project)
        aussen = {k: proj_dict.get(k) for k in ("id", "name", "filename", "status", "tool", "project_type",
                                                 "alt_language", "created_at", "updated_at")}
        return {"project": aussen, "images": [], "documents": [], "guest": True,
                "role": (dict(share).get("role") or "kunde"), "formular": True}
    images = conn.execute(
        """SELECT i.*, (SELECT body FROM messages m WHERE m.image_id = i.id AND m.msg_type = 'review_note' LIMIT 1) AS review_note FROM images i
           LEFT JOIN documents d ON d.id = i.document_id
           WHERE i.project_id = ?
           ORDER BY COALESCE(d.doc_index, 0), i.page_number, i.image_index""",
        (pid,)
    ).fetchall()
    documents = conn.execute(
        """SELECT id, doc_index, original_filename, display_name, extraction_method,
                  total_images, created_at, source_url, page_text, hinweise
           FROM documents WHERE project_id = ? ORDER BY doc_index""",
        (pid,)
    ).fetchall()
    # Rollen-Workflow (10.07.2026): Pruefstatus pro Rolle an jedes Bild haengen —
    # der Gast sieht seinen eigenen Stand, der Kunde zusaetzlich die Lektorat-
    # Vorgeschichte. Wie alles hier strikt auf DIESES Projekt begrenzt.
    review_rows = conn.execute(
        """SELECT r.image_id, r.role, r.status, r.reviewed_at FROM image_reviews r
           JOIN images i ON i.id = r.image_id WHERE i.project_id = ?""",
        (pid,)
    ).fetchall()
    # Rollen-Workflow Etappe 4 (15.07.2026): Nachrichten-Verlauf pro Bild —
    # Gaeste sehen denselben gemeinsamen Verlauf wie der Besitzer.
    thread_rows = conn.execute(
        """SELECT m.image_id, COALESCE(m.sender_role, '') AS sender_role,
                  COALESCE(m.sender_name, '') AS sender_name, m.body, m.created_at
           FROM messages m JOIN images i ON i.id = m.image_id
           WHERE i.project_id = ? AND m.msg_type = 'chat'
           ORDER BY m.created_at, m.id""",
        (pid,)
    ).fetchall()
    conn.close()
    thread_by_image = {}
    for m in thread_rows:
        thread_by_image.setdefault(m["image_id"], []).append({
            "sender_role": m["sender_role"], "sender_name": m["sender_name"],
            "body": m["body"], "created_at": m["created_at"]})
    reviews_by_image = {}
    for r in review_rows:
        reviews_by_image.setdefault(r["image_id"], {})[r["role"]] = {
            "status": r["status"], "reviewed_at": r["reviewed_at"]}
    image_dicts = []
    for i in images:
        d = dict(i)
        d["reviews"] = reviews_by_image.get(d["id"], {})
        d["thread"] = thread_by_image.get(d["id"], [])
        image_dicts.append(d)
    proj_dict = dict(project)
    _is_pdf = (proj_dict.get("tool") == "pdf" or proj_dict.get("project_type") == "pdf")
    return {
        "project": proj_dict,
        "images": image_dicts,
        "documents": [dict(d) for d in documents],
        "show_langbeschreibung": (not _is_pdf) or pdf_langbeschreibung_enabled(),
        "guest": True,
        "role": (dict(share).get("role") or "kunde"),
    }


@app.get("/api/freigabe/{token}/images/{image_id}/file")
async def freigabe_image_file(token: str, image_id: int, request: Request):
    """Bild-Datei fuer den Gast — gast-begrenzt + nur Bilder DES freigegebenen Projekts."""
    share = _require_guest(request, token)
    conn = get_db()
    img = conn.execute("SELECT * FROM images WHERE id = ? AND project_id = ?",
                       (image_id, share["project_id"])).fetchone()
    conn.close()
    if not img or not os.path.exists(img["image_path"]):
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    return FileResponse(img["image_path"])


@app.get("/api/freigabe/{token}/images/{image_id}/page-view")
async def freigabe_image_page_view(token: str, image_id: int, request: Request):
    """Seitenansicht-PNG fuer den Gast — gast-begrenzt + projektgebunden."""
    share = _require_guest(request, token)
    conn = get_db()
    img = conn.execute("SELECT page_view_path FROM images WHERE id = ? AND project_id = ?",
                       (image_id, share["project_id"])).fetchone()
    conn.close()
    if not img or not img["page_view_path"] or not os.path.exists(img["page_view_path"]):
        raise HTTPException(status_code=404, detail="Seitenansicht nicht gefunden")
    return FileResponse(img["page_view_path"])


@app.post("/api/freigabe/{token}/images/{image_id}/alttext")
async def freigabe_save_alttext(token: str, image_id: int, request: Request):
    """Gast bearbeitet den Alt-Text haendisch (lesen + editieren erlaubt, KEINE KI).
    Speichert in alt_text_edited — nur fuer Bilder DES freigegebenen Projekts.
    Spiegelt update_alt_text (Besitzer), aber token-/gast-begrenzt."""
    share = _require_guest(request, token)
    data = await request.json()
    conn = get_db()
    img = conn.execute("SELECT id FROM images WHERE id = ? AND project_id = ?",
                       (image_id, share["project_id"])).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    conn.execute("UPDATE images SET alt_text_edited = ? WHERE id = ?",
                 (data.get("alt_text", ""), image_id))
    if "langbeschreibung" in data:
        conn.execute("UPDATE images SET langbeschreibung = ? WHERE id = ?",
                     (data.get("langbeschreibung", ""), image_id))
    # Rollen-Workflow (10.07.2026): Haendisches Editieren ohne gesetzten Status
    # schaltet das Bild fuer DIESE Rolle automatisch auf 'in_bearbeitung' —
    # ein gesetztes Urteil (freigegeben/ruecksprache/...) wird NIE ueberschrieben.
    role = (dict(share).get("role") or "kunde")
    cur = conn.execute("SELECT status FROM image_reviews WHERE image_id = ? AND role = ?",
                       (image_id, role)).fetchone()
    auto_status = None
    if not cur or (cur["status"] or "offen") == "offen":
        conn.execute(
            "INSERT INTO image_reviews (image_id, role, status, reviewed_at) VALUES (?, ?, 'in_bearbeitung', datetime('now')) "
            "ON CONFLICT(image_id, role) DO UPDATE SET status = 'in_bearbeitung', reviewed_at = datetime('now')",
            (image_id, role))
        conn.execute("UPDATE images SET review_status = 'in_bearbeitung', reviewed_at = datetime('now') WHERE id = ?",
                     (image_id,))
        auto_status = "in_bearbeitung"
    # Wieder-Einstieg (Steve 15.07.2026): arbeitet ein Gast mit bereits
    # abgeschlossener Pruefung weiter (Status setzen / Text aendern), springt
    # seine Freigabe automatisch auf 'in Pruefung' zurueck. Nachrichten und
    # reines Lesen loesen das bewusst NICHT aus.
    conn.execute("UPDATE shares SET status = 'active' WHERE token = ? AND status = 'completed'", (token,))
    conn.commit()
    conn.close()
    return {"ok": True, "auto_status": auto_status}


@app.post("/api/freigabe/{token}/images/{image_id}/review")
async def freigabe_set_review(token: str, image_id: int, request: Request):
    """Gast setzt den Pruefstatus eines Bildes (offen/freigegeben/zu_ueberarbeiten)
    und optional einen Kommentar. Genau EINE review_note pro Bild (in messages);
    leerer Kommentar loescht sie. Alles gast-/projektbegrenzt."""
    share = _require_guest(request, token)
    session = get_guest_session(request, token)
    data = await request.json()
    role = (dict(share).get("role") or "kunde")
    status = data.get("status", "offen")
    if status not in ("offen", "in_bearbeitung", "freigegeben", "zu_ueberarbeiten", "ruecksprache"):
        raise HTTPException(status_code=400, detail="Ungueltiger Status.")
    if status == "ruecksprache" and role != "lektorat":
        raise HTTPException(status_code=403, detail="Ruecksprache kann nur das Lektorat setzen.")
    comment = (data.get("comment") or "").strip()
    conn = get_db()
    img = conn.execute("SELECT id FROM images WHERE id = ? AND project_id = ?",
                       (image_id, share["project_id"])).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    conn.execute(
        "INSERT INTO image_reviews (image_id, role, status, reviewed_at) VALUES (?, ?, ?, datetime('now')) "
        "ON CONFLICT(image_id, role) DO UPDATE SET status = excluded.status, reviewed_at = excluded.reviewed_at",
        (image_id, role, status))
    # Spiegel fuer Besitzer-Badge/Zaehler: haelt das jeweils LETZTE Gast-Urteil.
    conn.execute("UPDATE images SET review_status = ?, reviewed_at = datetime('now') WHERE id = ?",
                 (status, image_id))
    # Genau eine review_note pro Bild: alte ersetzen / bei leerem Kommentar entfernen.
    conn.execute("DELETE FROM messages WHERE image_id = ? AND msg_type = 'review_note'", (image_id,))
    if comment:
        conn.execute(
            "INSERT INTO messages (project_id, image_id, sender_guest, recipient_user_id, body, msg_type) "
            "VALUES (?, ?, ?, ?, ?, 'review_note')",
            (share["project_id"], image_id, (session or {}).get("guest", ""), share["created_by"], comment),
        )
    # Wieder-Einstieg (Steve 15.07.2026): arbeitet ein Gast mit bereits
    # abgeschlossener Pruefung weiter (Status setzen / Text aendern), springt
    # seine Freigabe automatisch auf 'in Pruefung' zurueck. Nachrichten und
    # reines Lesen loesen das bewusst NICHT aus.
    conn.execute("UPDATE shares SET status = 'active' WHERE token = ? AND status = 'completed'", (token,))
    conn.commit()
    conn.close()
    return {"ok": True, "review_status": status, "comment": comment, "role": role}


def _mail_escape(t):
    """Minimaler HTML-Schutz fuer Gast-Eingaben in Benachrichtigungs-Mails."""
    return (t or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ─── Rollen-Workflow Etappe 4 (15.07.2026): Nachrichten-Verlauf pro Bild ───
# Modell "Kommentare unter einem Dokument" (Spec Abschnitt 4): EIN gemeinsamer
# Verlauf pro Bild am Projekt, keine Adressierung. Besitzer sieht immer alles,
# Gaeste ueber ihren Link; jede Nachricht traegt die Absender-Rolle. Neue
# Nachrichten loesen kurze Mails an alle anderen Beteiligten aus.
MSG_MAX_LEN = 2000


def _msg_role_label(_, role):
    if role == "lektorat":
        return _('Lektorat')
    if role == "kunde":
        return _('Herausgeber')
    return _('Ersteller')


def _image_mail_label(conn, _, image_id):
    """Bild-Bezeichnung fuer Mails — gleiche Anzeige-Nummerierung wie im
    Frontend (pro Dokument fortlaufend ab 1; Web/Grafik behaelt image_index)."""
    img = conn.execute(
        "SELECT id, image_index, page_number, document_id FROM images WHERE id = ?",
        (image_id,)).fetchone()
    if not img:
        return ""
    nr = img["image_index"]
    doc_name = ""
    if img["document_id"]:
        rows = conn.execute(
            "SELECT id FROM images WHERE document_id = ? ORDER BY page_number, image_index",
            (img["document_id"],)).fetchall()
        nr = next((i + 1 for i, r in enumerate(rows) if r["id"] == image_id), nr)
        doc = conn.execute(
            "SELECT COALESCE(NULLIF(display_name, ''), original_filename) AS n FROM documents WHERE id = ?",
            (img["document_id"],)).fetchone()
        doc_name = (doc["n"] if doc else "") or ""
    base = (_('Bild {n}, Seite {p}').format(n=nr, p=img["page_number"])
            if img["page_number"] and img["page_number"] > 0 else _('Bild {n}').format(n=nr))
    return _('Dokument „{name}", {bild}').format(name=doc_name, bild=base) if doc_name else base


def _notify_message_recipients(project_row, image_id, sender_role, sender_name, body,
                               exclude_share_token=None, skip_owner=False):
    """Kurze Benachrichtigung "Neue Nachricht im Projekt X" an Besitzer + alle
    Gaeste mit nutzbarer Freigabe — ausser an den Absender selbst. Mail-Fehler
    blockieren das Speichern nicht (send_email faengt selbst)."""
    pid = project_row["id"]
    pname = project_row["name"] or project_row["filename"] or ("Projekt " + str(pid))
    from database import get_user_by_id
    owner = get_user_by_id(project_row["user_id"])
    conn = get_db()
    shares_rows = conn.execute(
        "SELECT token, guest_email FROM shares WHERE project_id = ? AND status IN ('active', 'completed')",
        (pid,)).fetchall()
    _ = get_gettext(_mail_lang(owner))
    label = _image_mail_label(conn, _, image_id)
    conn.close()
    sender_label = _msg_role_label(_, sender_role) + ((" (" + sender_name + ")") if sender_name else "")
    intro = _('{absender} hat eine neue Nachricht zu {bild} im Projekt „{projekt}" geschrieben:').format(
        absender=_mail_escape(sender_label), bild=(label or _('einem Bild')), projekt=_mail_escape(pname))
    text_html = ("<p>" + intro + "</p><p>" + _mail_escape(body).replace(chr(10), "<br>") + "</p>")
    subject = _('InkluDocs: Neue Nachricht im Projekt „{projekt}"').format(projekt=pname)
    if not skip_owner and owner and owner.get("email"):
        link = BASE_URL + "/app?projekt=" + str(pid)
        send_email(owner["email"], subject,
                   text_html + "<p>" + _('Im Werkzeug ansehen:') + " <a href='" + link + "'>" + link + "</a></p>",
                   bcc_admin=False)
    for sh in shares_rows:
        if exclude_share_token and sh["token"] == exclude_share_token:
            continue
        if not sh["guest_email"]:
            continue
        glink = BASE_URL + "/freigabe/" + sh["token"]
        send_email(sh["guest_email"], subject,
                   text_html + "<p>" + _('Im Werkzeug ansehen:') + " <a href='" + glink + "'>" + glink + "</a></p>",
                   bcc_admin=False)


def _insert_chat_message(conn, project_id, image_id, sender_user_id, sender_guest,
                         sender_role, sender_name, body):
    conn.execute(
        "INSERT INTO messages (project_id, image_id, sender_user_id, sender_guest, body, msg_type, sender_role, sender_name) "
        "VALUES (?, ?, ?, ?, ?, 'chat', ?, ?)",
        (project_id, image_id, sender_user_id, sender_guest, body, sender_role, sender_name))


@app.post("/api/images/{image_id}/messages")
async def post_image_message_owner(image_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Besitzer schreibt eine Nachricht zum Bild; alle Gaeste werden per Mail informiert."""
    data = await request.json()
    body = (data.get("text") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Leere Nachricht.")
    if len(body) > MSG_MAX_LEN:
        raise HTTPException(status_code=400, detail="Nachricht zu lang (max. 2000 Zeichen).")
    conn = get_db()
    project = conn.execute(
        """SELECT p.* FROM projects p JOIN images i ON i.project_id = p.id
           WHERE i.id = ? AND p.user_id = ?""",
        (image_id, user["id"])).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    sender_name = (user.get("display_name") or "").strip()
    _insert_chat_message(conn, project["id"], image_id, user["id"], None, "besitzer", sender_name, body)
    conn.commit()
    conn.close()
    _notify_message_recipients(project, image_id, "besitzer", sender_name, body, skip_owner=True)
    return {"ok": True}


@app.post("/api/freigabe/{token}/images/{image_id}/messages")
async def post_image_message_guest(token: str, image_id: int, request: Request):
    """Gast (Lektorat/Herausgeber) schreibt eine Nachricht zum Bild; Besitzer und
    die jeweils anderen Gaeste werden per Mail informiert. Alles gast-/projektbegrenzt."""
    share = _require_guest(request, token)
    data = await request.json()
    body = (data.get("text") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Leere Nachricht.")
    if len(body) > MSG_MAX_LEN:
        raise HTTPException(status_code=400, detail="Nachricht zu lang (max. 2000 Zeichen).")
    role = (dict(share).get("role") or "kunde")
    sender_name = (dict(share).get("guest_name") or "").strip()
    conn = get_db()
    img = conn.execute("SELECT id FROM images WHERE id = ? AND project_id = ?",
                       (image_id, share["project_id"])).fetchone()
    if not img:
        conn.close()
        raise HTTPException(status_code=404, detail="Bild nicht gefunden")
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (share["project_id"],)).fetchone()
    _insert_chat_message(conn, share["project_id"], image_id, None, share["guest_email"],
                         role, sender_name, body)
    conn.commit()
    conn.close()
    _notify_message_recipients(project, image_id, role, sender_name, body, exclude_share_token=token)
    return {"ok": True}


# ─── Etappe 4: Besitzer erstellt/verwaltet Freigaben ───
@app.post("/api/projects/{project_id}/share")
async def create_project_share(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    """Besitzer legt eine Freigabe (Gastzugang) an. Bei notify=true (Standard) verschickt der
    Server die Einladung serverseitig: From-Anzeigename "<Besitzer> ueber InkluDocs", Reply-To
    = Besitzer (Antworten gehen an ihn, NICHT an InkluTec). Bei notify=false wird NICHT versendet
    (Besitzer holt sich nur den Link und mailt selbst). Optionale persoenliche Nachricht ersetzt
    nur den Einleitungssatz; Link + Bestaetigungs-Hinweis bleiben fest."""
    import html
    data = await request.json()
    guest_email = (data.get("guest_email") or "").strip()
    guest_name = (data.get("guest_name") or "").strip()
    custom_message = (data.get("message") or "").strip()
    notify = data.get("notify", True)
    role = (data.get("role") or "kunde").strip().lower()
    if role not in sharing.VALID_ROLES:
        raise HTTPException(status_code=400, detail="Ungueltige Rolle.")
    if not guest_email or "@" not in guest_email:
        raise HTTPException(status_code=400, detail="Bitte eine gueltige E-Mail-Adresse angeben.")
    conn = get_db()
    proj = conn.execute("SELECT id, name, filename, tool FROM projects WHERE id = ? AND user_id = ?",
                        (project_id, user["id"])).fetchone()
    conn.close()
    if not proj:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    # Gast-Ansicht Formulare (28.08.2026): Formular-Projekte laden zum Pruefen der
    # Quickinfos ein, nicht der Alt-Texte — gleicher Ablauf, anderer Wortlaut.
    ist_formular = (proj["tool"] == "formular")
    token = sharing.create_share(project_id, guest_email, user["id"], guest_name=guest_name, role=role)
    url = BASE_URL + "/freigabe/" + token
    sent = False
    if notify:
        owner = get_user_by_id(user["id"]) or {}
        # Sprache des Einladenden als beste Annaeherung an die Gast-Sprache
        # (03.07.2026): wer auf Englisch arbeitet, laedt englischsprachig ein.
        _ = get_gettext(_mail_lang(owner))
        owner_name = (owner.get("display_name") or "").strip() or _('Ein InkluDocs-Nutzer')
        owner_email = owner.get("email") or user["email"]
        proj_name = (proj["name"] or proj["filename"] or ("Projekt " + str(project_id)))
        greet = _('Hallo {name},').format(name=html.escape(guest_name)) if guest_name else _('Hallo,')
        if custom_message:
            intro = html.escape(custom_message).replace("\n", "<br>")
        elif ist_formular:
            intro = _('{name} möchte Sie bitten, die Quickinfos für das Formular „{projekt}“ zu prüfen.').format(
                name=html.escape(owner_name), projekt=html.escape(proj_name))
        else:
            intro = _('{name} möchte Sie bitten, die Alt-Texte für das Projekt „{projekt}“ zu prüfen.').format(
                name=html.escape(owner_name), projekt=html.escape(proj_name))
        body = (
            "<p>" + greet + "</p>"
            "<p>" + intro + "</p>"
            "<p>" + _('Bitte öffnen Sie diesen Link und bestätigen Sie Ihre E-Mail-Adresse:') + "</p>"
            "<p><a href='" + url + "'>" + url + "</a></p>"
            "<p>" + _('Eingeladen von:') + " " + html.escape(owner_email) + "</p>"
            "<p>" + _('Vielen Dank!') + "</p>"
        )
        subject = (_('Bitte Quickinfos prüfen: {projekt}') if ist_formular
                   else _('Bitte Alt-Texte prüfen: {projekt}')).format(projekt=proj_name)
        sent = send_email(guest_email, subject, body, bcc_admin=False,
                          reply_to=owner_email, from_name=(owner_name + " über InkluDocs"))
    return {"ok": True, "token": token, "url": url, "guest_email": guest_email, "sent": bool(sent), "role": role}


@app.get("/api/projects/{project_id}/shares")
async def list_project_shares(project_id: int, user: dict = Depends(get_current_user)):
    conn = get_db()
    proj = conn.execute("SELECT id FROM projects WHERE id = ? AND user_id = ?",
                        (project_id, user["id"])).fetchone()
    conn.close()
    if not proj:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")
    shares = sharing.list_shares_for_project(project_id, user["id"])
    for sh in shares:
        sh["url"] = BASE_URL + "/freigabe/" + sh["token"]
    return {"shares": shares}


@app.post("/api/projects/{project_id}/shares/revoke")
async def revoke_project_share(project_id: int, request: Request, user: dict = Depends(get_current_user)):
    data = await request.json()
    if not sharing.revoke_share(data.get("token", ""), user["id"]):
        raise HTTPException(status_code=404, detail="Freigabe nicht gefunden")
    return {"ok": True}


# ─── Etappe 5: Gast schliesst die Pruefung ab -> Besitzer wird benachrichtigt ───
@app.post("/api/freigabe/{token}/complete")
async def freigabe_complete(token: str, request: Request):
    """Gast schliesst die Pruefung ab: Freigabe -> 'completed', Zusammenfassung +
    optionale Nachricht per E-Mail an den Besitzer. Bleibt danach weiter lesbar
    (Gast kann jederzeit zurueckkommen)."""
    share = _require_guest(request, token)
    session = get_guest_session(request, token)
    data = await request.json()
    message = (data.get("message") or "").strip()
    pid = share["project_id"]
    conn = get_db()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (pid,)).fetchone()
    role = (dict(share).get("role") or "kunde")
    # Zaehler aus der Pro-Rolle-Tabelle DIESER Rolle — der Lektor schliesst
    # seine eigene Pruefung ab, nicht die des Kunden (und umgekehrt).
    ist_formular = bool(project and project["tool"] == "formular")
    if ist_formular:
        # Gast-Ansicht Formulare (28.08.2026): Zaehler und Anmerkungen aus den
        # Feld-Urteilen statt aus den Bildern; keine Ruecksprache-Liste (Stufe 1).
        imgs = conn.execute(
            """SELECT COALESCE(r.status, 'offen') AS review_status FROM formularfelder f
               LEFT JOIN feld_reviews r ON r.feld_id = f.id AND r.role = ?
               WHERE f.project_id = ?""",
            (role, pid)
        ).fetchall()
        notes = conn.execute(
            "SELECT review_note AS body FROM formularfelder WHERE project_id = ? AND COALESCE(review_note, '') <> ''", (pid,)
        ).fetchall()
    else:
        imgs = conn.execute(
            """SELECT COALESCE(r.status, 'offen') AS review_status FROM images i
               LEFT JOIN image_reviews r ON r.image_id = i.id AND r.role = ?
               WHERE i.project_id = ?""",
            (role, pid)
        ).fetchall()
        notes = conn.execute(
            "SELECT body FROM messages WHERE project_id = ? AND msg_type = 'review_note'", (pid,)
        ).fetchall()
    # Etappe 3 (15.07.2026): Schliesst das LEKTORAT ab, bekommt der Besitzer die
    # Ruecksprache-Bilder als eigene Liste in der Mail (Dokument, Bild, Seite,
    # Anmerkung). Anzeige-Nummer wie im Frontend: pro Dokument fortlaufend ab 1
    # (Web/Grafik ohne Dokument behaelt image_index).
    rueck_items = []
    if role == "lektorat" and not ist_formular:
        all_rows = conn.execute(
            """SELECT i.id, i.image_index, i.page_number, i.document_id,
                      COALESCE(d.doc_index, 0) AS doc_index,
                      COALESCE(NULLIF(d.display_name, ''), d.original_filename) AS doc_name,
                      COALESCE(r.status, '') AS lek_status,
                      (SELECT body FROM messages m WHERE m.image_id = i.id AND m.msg_type = 'review_note' LIMIT 1) AS note
               FROM images i
               LEFT JOIN documents d ON d.id = i.document_id
               LEFT JOIN image_reviews r ON r.image_id = i.id AND r.role = 'lektorat'
               WHERE i.project_id = ?
               ORDER BY doc_index, i.page_number, i.image_index""",
            (pid,)
        ).fetchall()
        doc_counters = {}
        for row in all_rows:
            dk = row["document_id"] or 0
            if dk == 0:
                nr = row["image_index"]
            else:
                doc_counters[dk] = doc_counters.get(dk, 0) + 1
                nr = doc_counters[dk]
            if row["lek_status"] == "ruecksprache":
                rueck_items.append({"nr": nr, "page": row["page_number"],
                                    "doc": row["doc_name"], "note": row["note"]})
    conn.close()
    freigegeben = sum(1 for i in imgs if i["review_status"] == "freigegeben")
    zu_ueber = sum(1 for i in imgs if i["review_status"] == "zu_ueberarbeiten")
    ruecksprache = sum(1 for i in imgs if i["review_status"] == "ruecksprache")
    in_bearb = sum(1 for i in imgs if i["review_status"] == "in_bearbeitung")
    offen = sum(1 for i in imgs if (i["review_status"] or "offen") == "offen")
    sharing.mark_completed(token)
    from database import get_user_by_id
    owner = get_user_by_id(share["created_by"])
    pname = (project["name"] if project and project["name"] else (project["filename"] if project else "Projekt"))
    guest = (session or {}).get("guest", share["guest_email"])
    link = BASE_URL + "/freigabe/" + token
    notes_html = "".join("<li>" + _mail_escape(n["body"]) + "</li>" for n in notes)
    _ = get_gettext(_mail_lang(owner))
    # Etappe 3: Ruecksprache-Block fuer die Besitzer-Mail (nur Lektorat, nur wenn vorhanden).
    rueck_html = ""
    if rueck_items:
        rueck_lis = []
        for it in rueck_items:
            label = (_('Bild {n}, Seite {p}').format(n=it["nr"], p=it["page"])
                     if it["page"] and it["page"] > 0 else _('Bild {n}').format(n=it["nr"]))
            if it["doc"]:
                label = _('Dokument „{name}", {bild}').format(name=it["doc"], bild=label)
            li = "<li>" + _mail_escape(label)
            if it["note"]:
                li += "<br><em>" + _mail_escape(_('Anmerkung:') + " " + it["note"]) + "</em>"
            rueck_lis.append(li + "</li>")
        rueck_heading = (_('1 Bild zur Rücksprache markiert:') if len(rueck_items) == 1
                         else _('{n} Bilder zur Rücksprache markiert:').format(n=len(rueck_items)))
        rueck_html = ("<p><strong>" + rueck_heading + "</strong></p><ul>"
                      + "".join(rueck_lis) + "</ul>")
    body = (
        "<p>" + _('{gast} hat die Prüfung für das Projekt „{projekt}“ abgeschlossen.').format(
            gast=_mail_escape(guest), projekt=_mail_escape(pname)) + "</p>"
        + "<p><strong>" + _('Ergebnis:') + "</strong> "
        + ((_('Quickinfos: ') if ist_formular else "")
           + _('{f} freigegeben, {z} zu überarbeiten, {r} Rücksprache, {b} in Bearbeitung, {o} offen.').format(
            f=freigegeben, z=zu_ueber, r=ruecksprache, b=in_bearb, o=offen)) + "</p>"
        + "<p><strong>" + _('Rolle:') + "</strong> " + (_('Lektorat') if role == "lektorat" else _('Herausgeber')) + "</p>"
        + rueck_html
        + (("<p><strong>" + _('Nachricht:') + "</strong><br>" + _mail_escape(message).replace(chr(10), "<br>") + "</p>") if message else "")
        + (("<p><strong>" + _('Anmerkungen:') + "</strong></p><ul>" + notes_html + "</ul>") if notes_html else "")
        + "<p>" + _('Im Werkzeug ansehen:') + " <a href='" + link + "'>" + link + "</a></p>"
    )
    if owner and owner.get("email"):
        send_email(owner["email"], _('InkluDocs: Prüfung abgeschlossen ({projekt})').format(projekt=pname), body, bcc_admin=False)
    return {"ok": True, "freigegeben": freigegeben, "zu_ueberarbeiten": zu_ueber,
            "ruecksprache": ruecksprache, "in_bearbeitung": in_bearb, "offen": offen, "role": role}


@app.get("/api/review-overview")
async def review_overview(user: dict = Depends(get_current_user)):
    """Projekte des Besitzers, die zur Pruefung freigegeben sind, mit Review-Zaehlern
    fuers Dashboard-Panel (offen / zu ueberarbeiten / freigegeben)."""
    conn = get_db()
    rows = conn.execute(
        """SELECT p.id AS id, p.name AS name, p.filename AS filename,
                  SUM(CASE WHEN COALESCE(i.review_status,'offen')='offen' THEN 1 ELSE 0 END) AS offen,
                  SUM(CASE WHEN i.review_status='zu_ueberarbeiten' THEN 1 ELSE 0 END) AS zu_ueberarbeiten,
                  SUM(CASE WHEN i.review_status='freigegeben' THEN 1 ELSE 0 END) AS freigegeben,
                  SUM(CASE WHEN EXISTS (SELECT 1 FROM image_reviews r
                        WHERE r.image_id = i.id AND r.role = 'lektorat' AND r.status = 'ruecksprache')
                      THEN 1 ELSE 0 END) AS ruecksprache,
                  (SELECT COUNT(*) FROM messages m
                    WHERE m.project_id = p.id AND m.msg_type = 'chat') AS nachrichten,
                  COUNT(i.id) AS total,
                  (SELECT GROUP_CONCAT(DISTINCT COALESCE(NULLIF(g.guest_name,''), g.guest_email)) FROM shares g
                    WHERE g.project_id = p.id AND g.status IN ('active','completed')) AS guests
           FROM projects p
           JOIN images i ON i.project_id = p.id
           WHERE p.user_id = ?
             AND EXISTS (SELECT 1 FROM shares s WHERE s.project_id = p.id AND s.status IN ('active','completed'))
           GROUP BY p.id
           ORDER BY p.updated_at DESC""",
        (user["id"],)
    ).fetchall()
    # Gast-Ansicht Formulare (28.08.2026): Formular-Projekte haben keine Bilder —
    # dieselben Zaehler aus formularfelder/feld_reviews, dieselbe Form fuers Frontend
    # (Dashboard-Knopf + /geteilte-projekte), damit Formulare dort nicht unsichtbar sind.
    form_rows = conn.execute(
        """SELECT p.id AS id, p.name AS name, p.filename AS filename,
                  SUM(CASE WHEN COALESCE(f.review_status,'offen')='offen' THEN 1 ELSE 0 END) AS offen,
                  SUM(CASE WHEN f.review_status='zu_ueberarbeiten' THEN 1 ELSE 0 END) AS zu_ueberarbeiten,
                  SUM(CASE WHEN f.review_status='freigegeben' THEN 1 ELSE 0 END) AS freigegeben,
                  SUM(CASE WHEN EXISTS (SELECT 1 FROM feld_reviews r
                        WHERE r.feld_id = f.id AND r.role = 'lektorat' AND r.status = 'ruecksprache')
                      THEN 1 ELSE 0 END) AS ruecksprache,
                  0 AS nachrichten,
                  COUNT(f.id) AS total,
                  (SELECT GROUP_CONCAT(DISTINCT COALESCE(NULLIF(g.guest_name,''), g.guest_email)) FROM shares g
                    WHERE g.project_id = p.id AND g.status IN ('active','completed')) AS guests
           FROM projects p
           JOIN formularfelder f ON f.project_id = p.id
           WHERE p.user_id = ? AND p.tool = 'formular'
             AND EXISTS (SELECT 1 FROM shares s WHERE s.project_id = p.id AND s.status IN ('active','completed'))
           GROUP BY p.id
           ORDER BY p.updated_at DESC""",
        (user["id"],)
    ).fetchall()
    rows = list(rows) + list(form_rows)
    # Geteilte-Projekte-Kasten (15.07.2026): Lese-Stand + Rollen-Status + Neu-Zaehler.
    seen_row = conn.execute("SELECT reviews_seen_until FROM users WHERE id = ?", (user["id"],)).fetchone()
    seen = seen_row["reviews_seen_until"] if seen_row and seen_row["reviews_seen_until"] else ""
    projects = []
    for r in rows:
        d = dict(r)
        d["display_name"] = (d.get("name") or d.get("filename") or ("Projekt " + str(d["id"])))
        pid = d["id"]
        # Status pro Rolle: solange EINE Freigabe der Rolle aktiv ist, gilt die
        # Rolle als "in Pruefung"; abgeschlossen erst, wenn alle durch sind.
        d["rollen"] = [dict(x) for x in conn.execute(
            """SELECT role,
                      CASE WHEN SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) > 0
                           THEN 'active' ELSE 'completed' END AS status,
                      MAX(completed_at) AS completed_at
               FROM shares WHERE project_id = ? AND status IN ('active','completed')
               GROUP BY role""", (pid,)).fetchall()]
        # Detail-Zeilen fuer die Unterseite: wer (Name, sonst E-Mail) pro Freigabe,
        # mit Einladungs- und Abschluss-Zeitpunkt (UTC, Frontend rechnet lokal um).
        d["freigaben"] = [dict(x) for x in conn.execute(
            """SELECT role, COALESCE(NULLIF(guest_name,''), guest_email) AS wer,
                      guest_email, email_confirmed_at,
                      status, completed_at, created_at
               FROM shares WHERE project_id = ? AND status IN ('active','completed')
               ORDER BY CASE role WHEN 'lektorat' THEN 0 ELSE 1 END, created_at DESC""",
            (pid,)).fetchall()]
        # "Neu seit letztem Besuch": eigene Besitzer-Nachrichten zaehlen nicht mit;
        # leerer Lese-Stand (Erstnutzung) zaehlt alles Bisherige als neu.
        if seen:
            neu_msgs = conn.execute(
                "SELECT COUNT(*) FROM messages WHERE project_id = ? AND msg_type = 'chat' "
                "AND COALESCE(sender_role,'') <> 'besitzer' AND created_at > ?", (pid, seen)).fetchone()[0]
            neu_reviews = conn.execute(
                "SELECT COUNT(*) FROM image_reviews r JOIN images i ON i.id = r.image_id "
                "WHERE i.project_id = ? AND r.reviewed_at > ?", (pid, seen)).fetchone()[0]
            neu_reviews += conn.execute(
                "SELECT COUNT(*) FROM feld_reviews r JOIN formularfelder f ON f.id = r.feld_id "
                "WHERE f.project_id = ? AND r.reviewed_at > ?", (pid, seen)).fetchone()[0]
            neu_abschluesse = conn.execute(
                "SELECT COUNT(*) FROM shares WHERE project_id = ? AND completed_at IS NOT NULL "
                "AND completed_at > ?", (pid, seen)).fetchone()[0]
        else:
            neu_msgs = conn.execute(
                "SELECT COUNT(*) FROM messages WHERE project_id = ? AND msg_type = 'chat' "
                "AND COALESCE(sender_role,'') <> 'besitzer'", (pid,)).fetchone()[0]
            neu_reviews = conn.execute(
                "SELECT COUNT(*) FROM image_reviews r JOIN images i ON i.id = r.image_id "
                "WHERE i.project_id = ?", (pid,)).fetchone()[0]
            neu_reviews += conn.execute(
                "SELECT COUNT(*) FROM feld_reviews r JOIN formularfelder f ON f.id = r.feld_id "
                "WHERE f.project_id = ?", (pid,)).fetchone()[0]
            neu_abschluesse = conn.execute(
                "SELECT COUNT(*) FROM shares WHERE project_id = ? AND status = 'completed'", (pid,)).fetchone()[0]
        d["neu"] = {"nachrichten": neu_msgs, "pruefungen": neu_reviews, "abschluesse": neu_abschluesse}
        projects.append(d)
    conn.close()
    return {"projects": projects, "seen_until": seen}


@app.post("/api/review-overview/seen")
async def review_overview_seen(user: dict = Depends(get_current_user)):
    """Markiert den Geteilte-Projekte-Kasten als gelesen (UTC-Zeitstempel,
    vergleichbar mit den created_at/reviewed_at/completed_at-Spalten)."""
    conn = get_db()
    conn.execute("UPDATE users SET reviews_seen_until = datetime('now') WHERE id = ?", (user["id"],))
    conn.commit()
    row = conn.execute("SELECT reviews_seen_until FROM users WHERE id = ?", (user["id"],)).fetchone()
    conn.close()
    return {"seen_until": row["reviews_seen_until"] if row else ""}

# Rechtsseiten (Impressum, Datenschutz, Nutzungsbedingungen, Widerrufsbelehrung):
# bis 25.08.2026 rohe HTML-Dateien aus frontend/, seitdem Templates auf dem
# oeffentlichen Seitengeruest base_oeffentlich.html (Skip-Link, Seitenleiste,
# H1 = Seitenthema, einheitliche Fusszeile). Der Rechtstext selbst ist
# unveraendert und bleibt deutsch; der .legal-container darin ist die EINE
# Quelle fuer die In-App-Sichten (_serve_protected_page) und die Demo, die ihn
# per fetch holen. Doku: docs/SEITENGERUEST.md
def _rechtsseite(request: Request, template: str):
    lang = detect_language(request)
    return templates.TemplateResponse(
        template,
        template_context(request, lang, is_staging="staging" in BASE_URL),
    )


@app.get("/impressum", response_class=HTMLResponse)
async def impressum_page(request: Request):
    return _rechtsseite(request, "impressum.html")


@app.get("/datenschutz", response_class=HTMLResponse)
async def datenschutz_page(request: Request):
    return _rechtsseite(request, "datenschutz.html")


@app.get("/nutzungsbedingungen", response_class=HTMLResponse)
async def nutzungsbedingungen_page(request: Request):
    return _rechtsseite(request, "nutzungsbedingungen.html")


@app.get("/widerruf", response_class=HTMLResponse)
async def widerruf_page(request: Request):
    """Widerrufsbelehrung und Muster-Widerrufsformular (08.08.2026)."""
    return _rechtsseite(request, "widerruf.html")


@app.get("/kuendigen", response_class=HTMLResponse)
async def kuendigen_page(request: Request):
    """Kuendigungsknopf nach § 312k BGB — bewusst ohne Anmeldung erreichbar.

    Anders als die uebrigen Rechtsseiten ein Template, weil hier ein Formular
    steht: Angaben, Bestaetigungsseite, Bestaetigung in Textform.
    """
    lang = detect_language(request)
    return templates.TemplateResponse(
        "kuendigen.html",
        template_context(request, lang, is_staging="staging" in BASE_URL),
    )


@app.get("/avv", response_class=HTMLResponse)
async def avv_page(request: Request):
    """Auftragsverarbeitung nach Art. 28 DSGVO (24.08.2026): erfuellt das
    Angebot aus AGB Ziffer 18; Vertrag als barrierefreies PDF zum Download."""
    lang = detect_language(request)
    return templates.TemplateResponse(
        "avv.html",
        template_context(request, lang, is_staging="staging" in BASE_URL),
    )


@app.get("/ueber-uns", response_class=HTMLResponse)
async def ueber_uns_page(request: Request):
    """Ueber-uns-Seite (Steve + Michael, 24.08.2026): InkluTec + Actino."""
    lang = detect_language(request)
    return templates.TemplateResponse(
        "ueber_uns.html",
        template_context(request, lang, is_staging="staging" in BASE_URL),
    )


@app.get("/kontakt", response_class=HTMLResponse)
async def kontakt_page(request: Request):
    """Kontaktseite (24.08.2026): zweiter Kommunikationsweg nach § 5 DDG
    (E-Mail + Kontaktformular statt Telefonnummer), ohne Anmeldung."""
    lang = detect_language(request)
    return templates.TemplateResponse(
        "kontakt.html",
        template_context(request, lang, is_staging="staging" in BASE_URL),
    )


@app.get("/widerrufen", response_class=HTMLResponse)
async def widerrufen_page(request: Request):
    """Widerrufsfunktion nach § 356a BGB (Pflicht seit 19.06.2026; gebaut
    24.08.2026): gut sichtbar, ohne Anmeldung, zweistufig — erst die
    Widerrufserklaerung ("Vertrag widerrufen"), dann die Bestaetigung
    ("Widerruf bestätigen"). Die Erklaerung geht denselben Weg wie der
    Kuendigungsknopf (/api/kuendigung mit art='widerruf'): protokolliert,
    sofort wirksam zum Abo-Ende-Verhalten, Bestaetigung in Textform."""
    lang = detect_language(request)
    return templates.TemplateResponse(
        "widerrufen.html",
        template_context(request, lang, is_staging="staging" in BASE_URL),
    )
